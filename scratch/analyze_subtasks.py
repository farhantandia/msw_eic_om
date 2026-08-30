import openpyxl
import xlrd
import os
import re

def analyze_unit1():
    print("==================================================")
    print("ANALYSIS: UNIT 1 WORK ORDERS, ACTUATORS & INSTRUMENTS")
    print("==================================================")
    
    # 1. Read WO and Subtasks from Progress Outage Unit 1 EIC.xlsx
    wo_file = "Unit 1/Progress Outage Unit 1 EIC.xlsx"
    wb_wo = openpyxl.load_workbook(wo_file, data_only=True)
    ws_wo = wb_wo["UPDATE U1"]
    wo_rows = list(ws_wo.iter_rows(values_only=True))

    wos = []
    current_wo = None
    all_subtasks_list = [] # list of dicts

    for r in wo_rows[11:]:
        no = r[1]
        no_wo = r[2]
        job_desc = str(r[3] or '').strip()
        if no is not None and no_wo is not None:
            current_wo = {
                'no': no,
                'wo_no': str(no_wo).strip(),
                'job_desc': job_desc,
                'subtasks': []
            }
            wos.append(current_wo)
        elif current_wo and job_desc:
            current_wo['subtasks'].append(job_desc)
            all_subtasks_list.append({
                'wo_no': current_wo['wo_no'],
                'wo_desc': current_wo['job_desc'],
                'subtask': job_desc
            })

    print(f"Total Work Orders (WO): {len(wos)}")
    print(f"Total Sub-tasks in WO List: {len(all_subtasks_list)}")

    # 2. Read Actuators from AMP-MSW-Progress Actuator Unit 1 2026.xlsx
    act_file = "Unit 1/AMP-MSW-Progress Actuator Unit 1 2026.xlsx"
    wb_act = openpyxl.load_workbook(act_file, data_only=True)
    ws_act = wb_act["Actuator Valve"]
    act_rows = list(ws_act.iter_rows(values_only=True))
    
    actuators = []
    current_act = None
    for r in act_rows[13:]:
        area = str(r[0] or '').strip()
        desc = str(r[1] or '').strip()
        status_or_sub = str(r[4] or '').strip()
        if desc and ("ACTUATOR" in desc.upper() or area):
            if current_act:
                actuators.append(current_act)
            current_act = {
                'area': area,
                'desc': desc,
                'subitems': []
            }
        elif current_act and (status_or_sub or desc):
            sub_text = status_or_sub if status_or_sub else desc
            current_act['subitems'].append(sub_text)
    if current_act:
        actuators.append(current_act)

    print(f"Total Actuator entries in Actuator file: {len(actuators)}")

    # 3. Read Instruments from JAPA-MSW-Progress Transmitter & Switch Unit 2 2026.xls
    inst_file = "Unit 1/JAPA-MSW-Progress Transmitter & Switch Unit 2 2026.xls"
    wb_inst = xlrd.open_workbook(inst_file)
    
    pt_items = []
    tt_items = []
    ps_items = []
    
    for sname in wb_inst.sheet_names():
        ws = wb_inst.sheet_by_name(sname)
        if "PRESSURE TRANSMITTER" in sname.upper() or "PT" in sname.upper():
            for row_idx in range(12, ws.nrows):
                vals = ws.row_values(row_idx)
                if len(vals) > 3 and (vals[2] or vals[3]):
                    pt_items.append({'area': str(vals[1]).strip(), 'eq': str(vals[2]).strip(), 'kks': str(vals[3]).strip(), 'type': 'PT', 'row': row_idx+1})
        elif "TEMPERATURE" in sname.upper() or "TT" in sname.upper():
            for row_idx in range(12, ws.nrows):
                vals = ws.row_values(row_idx)
                if len(vals) > 3 and (vals[2] or vals[3]):
                    tt_items.append({'area': str(vals[1]).strip(), 'eq': str(vals[2]).strip(), 'kks': str(vals[3]).strip(), 'type': 'TT', 'row': row_idx+1})
        elif "PRESSURE SWITCH" in sname.upper() or "PS" in sname.upper() or "SWITCH" in sname.upper():
            for row_idx in range(12, ws.nrows):
                vals = ws.row_values(row_idx)
                if len(vals) > 3 and (vals[2] or vals[3]):
                    ps_items.append({'area': str(vals[1]).strip(), 'eq': str(vals[2]).strip(), 'kks': str(vals[3]).strip(), 'type': 'PS', 'row': row_idx+1})

    print(f"Total Instruments in JAPA file: PT={len(pt_items)}, TT={len(tt_items)}, PS={len(ps_items)} (Total: {len(pt_items)+len(tt_items)+len(ps_items)})")

    # =========================================================================
    # MATCHING ANALYSIS: ACTUATORS vs WORK ORDER SUB-TASKS
    # =========================================================================
    print("\n" + "="*70)
    print("--- 1. ACTUATOR MATCHING ANALYSIS ---")
    print("="*70)
    
    act_matched = []
    act_unmatched = []
    
    for act in actuators:
        desc = act['desc']
        # normalize spaces (\xa0 to space)
        desc_norm = desc.replace('\xa0', ' ').strip()
        # extract KKS candidate
        kks_matches = re.findall(r'[0-9]{1,2}[A-Z]{3}[0-9]{2}[A-Z]{2}[0-9]{3}', desc_norm)
        kks = kks_matches[0] if kks_matches else ""
        
        matches = []
        for st in all_subtasks_list:
            st_clean = st['subtask'].replace('\xa0', ' ').upper()
            if kks and kks.upper() in st_clean:
                matches.append((st['wo_no'], st['wo_desc'], st['subtask'], f"KKS match: {kks}"))
            elif desc_norm.upper() in st_clean:
                matches.append((st['wo_no'], st['wo_desc'], st['subtask'], "Exact description match"))
            else:
                # check keywords
                core_desc = re.sub(r'^(ACTUATOR|VALVE|UNIT \d+|MSW)\s*', '', desc_norm, flags=re.I).strip()
                if len(core_desc) > 10 and core_desc.upper() in st_clean:
                    matches.append((st['wo_no'], st['wo_desc'], st['subtask'], f"Core match: {core_desc}"))
                    
        if matches:
            act_matched.append({'act': act, 'desc_norm': desc_norm, 'kks': kks, 'matches': matches})
        else:
            act_unmatched.append({'act': act, 'desc_norm': desc_norm, 'kks': kks})
            
    print(f"Total Actuators: {len(actuators)}")
    print(f"  -> Matched in WO Sub-tasks: {len(act_matched)}")
    print(f"  -> Unmatched in WO Sub-tasks: {len(act_unmatched)}")
    
    if act_unmatched:
        print(f"\nUNMATCHED ACTUATORS ({len(act_unmatched)} items):")
        for idx, u in enumerate(act_unmatched, 1):
            print(f"  {idx:2d}. Area: {u['act']['area']:<15} | Desc: {u['desc_norm']} | KKS: {u['kks']}")

    # =========================================================================
    # MATCHING ANALYSIS: INSTRUMENTS vs WORK ORDER SUB-TASKS
    # =========================================================================
    print("\n" + "="*70)
    print("--- 2. INSTRUMENT MATCHING ANALYSIS (PT, TT, PS) ---")
    print("="*70)
    
    all_instruments = pt_items + tt_items + ps_items
    inst_matched = []
    inst_unmatched = []
    
    for inst in all_instruments:
        kks = inst['kks'].replace('\xa0', ' ').strip()
        eq = inst['eq'].replace('\xa0', ' ').strip()
        itype = inst['type']
        
        matches = []
        for st in all_subtasks_list:
            st_clean = st['subtask'].replace('\xa0', ' ').upper()
            if kks and len(kks) >= 5 and kks.upper() in st_clean:
                matches.append((st['wo_no'], st['wo_desc'], st['subtask'], f"KKS match: {kks}"))
            elif eq and len(eq) > 12 and (eq.upper() in st_clean or st_clean in eq.upper()):
                matches.append((st['wo_no'], st['wo_desc'], st['subtask'], "Description match"))
        
        if matches:
            inst_matched.append({'inst': inst, 'matches': matches})
        else:
            inst_unmatched.append({'inst': inst})
            
    print(f"Total Instruments: {len(all_instruments)} (PT: {len(pt_items)}, TT: {len(tt_items)}, PS: {len(ps_items)})")
    print(f"  -> Matched in WO Sub-tasks: {len(inst_matched)}")
    print(f"  -> Unmatched in WO Sub-tasks: {len(inst_unmatched)}")
    
    if inst_unmatched:
        print(f"\nUNMATCHED INSTRUMENTS ({len(inst_unmatched)} items):")
        for idx, u in enumerate(inst_unmatched, 1):
            print(f"  {idx:2d}. [{u['inst']['type']}] Area: {u['inst']['area']:<15} | KKS: {u['inst']['kks']:<15} | Eq: {u['inst']['eq']}")

    # =========================================================================
    # 3. BREAKDOWN OF WO SUB-TASKS: What are the 721 subtasks composed of?
    # =========================================================================
    print("\n" + "="*70)
    print("--- 3. BREAKDOWN OF ALL 721 SUB-TASKS IN WORK ORDER LIST ---")
    print("="*70)
    
    wo_map = {}
    for st in all_subtasks_list:
        wo_no = st['wo_no']
        if wo_no not in wo_map:
            wo_map[wo_no] = {'desc': st['wo_desc'], 'subtasks': []}
        wo_map[wo_no]['subtasks'].append(st['subtask'])
        
    for wo_no, info in wo_map.items():
        st_list = info['subtasks']
        # classify subtasks
        act_cnt = sum(1 for s in st_list if any(k in s.upper() for k in ['ACTUATOR', 'MOV', 'AOV', 'GATE', 'VALVE', 'DAMPER', 'IGV']))
        inst_cnt = sum(1 for s in st_list if any(k in s.upper() for k in ['TRANSMITTER', 'SWITCH', 'CP', 'CT', 'CALIBRATION', 'MEASUREMENT', 'PT', 'TT', 'RTD', 'THERMOCOUPLE']))
        motor_cnt = sum(1 for s in st_list if any(k in s.upper() for k in ['MOTOR', 'BEARING', 'WINDING', 'ISOLASI', 'SOLO RUN', 'FRAME MOTOR', 'REGREASING', 'COOLING FAN', 'TERMINASI', 'GROUNDING']))
        
        print(f"[{wo_no}] {info['desc'][:50]:<50} | Total: {len(st_list):2d} | Actuator: {act_cnt:2d} | Instrument: {inst_cnt:2d} | Motor/Elec: {motor_cnt:2d} | Other: {len(st_list)-act_cnt-inst_cnt-motor_cnt:2d}")

if __name__ == '__main__':
    analyze_unit1()
