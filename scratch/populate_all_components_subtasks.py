import openpyxl
import re
import os
import difflib

def clean_str(val):
    if val is None:
        return ""
    return str(val).replace('\xa0', ' ').strip()

def extract_kks(text):
    if not text:
        return ""
    # Standard KKS patterns
    matches = re.findall(r'[0-9]{1,2}[A-Z]{3}[0-9]{2}[A-Z]{2}[0-9]{3}', text)
    if matches:
        return matches[0]
    matches2 = re.findall(r'[0-9]{1,2}[A-Z]{3}[0-9]{2}[A-Z]{1,2}[0-9]{1,3}', text)
    if matches2:
        return matches2[0]
    return ""

def is_same_equipment(act_desc, act_kks, st_desc):
    act_kks = act_kks.strip().upper()
    st_desc_upper = st_desc.strip().upper()
    act_desc_upper = act_desc.strip().upper()
    
    st_kks = extract_kks(st_desc_upper)
    
    # 1. KKS match
    if act_kks and st_kks:
        # Match if identical or same base (e.g. 10HNA61AA001 vs 20HNA61AA001 if unit specific)
        if act_kks == st_kks:
            return True
        if act_kks[2:] == st_kks[2:]:
            return True
            
    if act_kks and act_kks in st_desc_upper:
        return True
        
    # 2. Exact or near-identical text match
    if act_desc_upper == st_desc_upper:
        return True
        
    # Compare normalized strings without prefixes
    norm_act = re.sub(r'^(ACTUATOR|VALVE|MOV|AOV|UNIT \d+|MSW)\s*', '', act_desc_upper).strip()
    norm_st = re.sub(r'^(ACTUATOR|VALVE|MOV|AOV|UNIT \d+|MSW|[0-9]{1,2}[A-Z]{3}[0-9]{2}[A-Z]{1,2}[0-9]{1,3}[:\s]*)\s*', '', st_desc_upper).strip()
    
    if norm_act and norm_st and norm_act == norm_st:
        return True
        
    ratio = difflib.SequenceMatcher(None, norm_act, norm_st).ratio()
    if ratio > 0.88:
        return True
        
    return False

def map_equipment_to_wo(eq_desc, area, kks, wos):
    desc_upper = eq_desc.upper()
    area_upper = area.upper()
    kks_upper = kks.upper()
    
    # ID FAN
    if 'ID FAN 1' in desc_upper or 'ID FAN 1' in area_upper or '10HNA61' in kks_upper or '20HNA61' in kks_upper or '10HNA71' in kks_upper or '20HNA71' in kks_upper:
        for wn, w in wos.items():
            if 'ID FAN 1' in w['job_description'].upper() or 'INDUCED DRAUGHT FAN 1' in w['job_description'].upper():
                return wn
    if 'ID FAN 2' in desc_upper or 'ID FAN 2' in area_upper or '10HNA62' in kks_upper or '20HNA62' in kks_upper or '10HNA72' in kks_upper or '20HNA72' in kks_upper:
        for wn, w in wos.items():
            if 'ID FAN 2' in w['job_description'].upper() or 'INDUCED DRAUGHT FAN 2' in w['job_description'].upper():
                return wn
                
    # PA FAN
    if 'PA FAN 1' in desc_upper or 'PRIMARY AIR FAN 1' in desc_upper:
        if 'LUBE' in desc_upper or 'LO' in desc_upper or 'HLB01' in kks_upper:
            for wn, w in wos.items():
                if 'PRIMARY AIR FAN 1 LUBE' in w['job_description'].upper(): return wn
        for wn, w in wos.items():
            if 'PRIMARY AIR FAN 1' in w['job_description'].upper() and 'LUBE' not in w['job_description'].upper(): return wn
    if 'PA FAN 2' in desc_upper or 'PRIMARY AIR FAN 2' in desc_upper:
        if 'LUBE' in desc_upper or 'LO' in desc_upper or 'HLB02' in kks_upper:
            for wn, w in wos.items():
                if 'PRIMARY AIR FAN 2 LUBE' in w['job_description'].upper(): return wn
        for wn, w in wos.items():
            if 'PRIMARY AIR FAN 2' in w['job_description'].upper() and 'LUBE' not in w['job_description'].upper(): return wn

    # SA FAN
    if 'SA FAN 1' in desc_upper or 'SECONDARY AIR FAN 1' in desc_upper:
        if 'LUBE' in desc_upper or 'LO' in desc_upper or 'HLB05' in kks_upper:
            for wn, w in wos.items():
                if 'SECONDARY AIR FAN 1 LU' in w['job_description'].upper(): return wn
        for wn, w in wos.items():
            if 'SECONDARY AIR FAN 1' in w['job_description'].upper() and 'LU' not in w['job_description'].upper(): return wn
    if 'SA FAN 2' in desc_upper or 'SECONDARY AIR FAN 2' in desc_upper:
        if 'LUBE' in desc_upper or 'LO' in desc_upper or 'HLB06' in kks_upper:
            for wn, w in wos.items():
                if 'SECONDARY AIR FAN 2 LU' in w['job_description'].upper(): return wn
        for wn, w in wos.items():
            if 'SECONDARY AIR FAN 2' in w['job_description'].upper() and 'LU' not in w['job_description'].upper(): return wn

    # COAL FEEDER
    if 'COAL FEEDER 1' in desc_upper or 'COAL FEEDER NO 1' in desc_upper or '10HHE11' in kks_upper or '20HHE11' in kks_upper:
        for wn, w in wos.items():
            if 'COAL FEEDER NO 1' in w['job_description'].upper() or 'COAL FEEDER 1' in w['job_description'].upper(): return wn
    if 'COAL FEEDER 2' in desc_upper or 'COAL FEEDER NO 2' in desc_upper or '10HHE12' in kks_upper or '20HHE12' in kks_upper:
        for wn, w in wos.items():
            if 'COAL FEEDER NO 2' in w['job_description'].upper() or 'COAL FEEDER 2' in w['job_description'].upper(): return wn

    # BURNER / LANCES / LDO
    if 'BURNER 1' in desc_upper or 'LANCES' in desc_upper or '10HJF11' in kks_upper or '20HJF11' in kks_upper or '10HJF13' in kks_upper or '20HJF13' in kks_upper:
        for wn, w in wos.items():
            if 'BURNER 1' in w['job_description'].upper() or 'LDO START-UP BURNER 1' in w['job_description'].upper(): return wn
    if 'BURNER 2' in desc_upper or '10HJF12' in kks_upper or '20HJF12' in kks_upper:
        for wn, w in wos.items():
            if 'BURNER 2' in w['job_description'].upper() or 'LDO START-UP BURNER 2' in w['job_description'].upper(): return wn

    # BOILER FEED PUMP (BFP)
    if 'BFP 1' in desc_upper or 'FEED PUMP 1' in desc_upper or '10LAC10' in kks_upper or '20LAC10' in kks_upper or '20LAC20' in kks_upper:
        for wn, w in wos.items():
            if 'BOILER FEED PUMP 1' in w['job_description'].upper(): return wn
    if 'BFP 2' in desc_upper or 'FEED PUMP 2' in desc_upper or '10LAC30' in kks_upper or '20LAC30' in kks_upper:
        for wn, w in wos.items():
            if 'BOILER FEED PUMP 2' in w['job_description'].upper(): return wn

    # STEAM DRUM
    if 'DRUM' in desc_upper or 'DRUM' in area_upper or '10HAD' in kks_upper or '20HAD' in kks_upper:
        for wn, w in wos.items():
            if 'BOILER DRUM' in w['job_description'].upper() or 'DRUM' in w['job_description'].upper(): return wn

    # DESUPERHEATER / DESH
    if 'DESH' in desc_upper or 'DESUPERHEATER' in desc_upper or '10LAE' in kks_upper or '20LAE' in kks_upper or '10HAH' in kks_upper or '20HAH' in kks_upper:
        if 'DESH 2' in desc_upper or 'DESH-2' in desc_upper or 'HAH46' in kks_upper or 'HAH36' in kks_upper:
            for wn, w in wos.items():
                if 'DESH SUPERHEATER LINE 2' in w['job_description'].upper() or 'DESH 2' in w['job_description'].upper(): return wn
        for wn, w in wos.items():
            if 'DESH SUPERHEATER LINE 1' in w['job_description'].upper() or 'DESH 1' in w['job_description'].upper() or 'DESH SUPERHEATER' in w['job_description'].upper(): return wn

    # VACUUM PUMP
    if 'VACUUM PUMP 1' in desc_upper or 'VACUUM PUMP A' in desc_upper or 'VACUM PUMP-1' in desc_upper:
        for wn, w in wos.items():
            if 'VACUUM PUMP A' in w['job_description'].upper() or 'VACUUM PUMP 1' in w['job_description'].upper(): return wn
    if 'VACUUM PUMP 2' in desc_upper or 'VACUUM PUMP B' in desc_upper or 'VACUM PUMP-2' in desc_upper:
        for wn, w in wos.items():
            if 'VACUUM PUMP B' in w['job_description'].upper() or 'VACUUM PUMP 2' in w['job_description'].upper(): return wn

    # CONDENSATE PUMP (CEP)
    if 'CEP 1' in desc_upper or 'CEP-1' in desc_upper or 'CONDENSATE PUMP A' in desc_upper:
        for wn, w in wos.items():
            if 'MAIN CONDENSATE PUMP A' in w['job_description'].upper() or 'CONDENSATE PUMP 1' in w['job_description'].upper(): return wn
    if 'CEP 2' in desc_upper or 'CEP-2' in desc_upper or 'CONDENSATE PUMP B' in desc_upper:
        for wn, w in wos.items():
            if 'MAIN CONDENSATE PUMP B' in w['job_description'].upper() or 'CONDENSATE PUMP 2' in w['job_description'].upper(): return wn

    # CYCLONE
    if 'CYCLONE' in desc_upper or '10HNA11' in kks_upper or '20HNA11' in kks_upper:
        for wn, w in wos.items():
            if 'CYCLONE' in w['job_description'].upper(): return wn

    # SOOT BLOWER
    if 'SOOT BLOWER' in desc_upper or '10HCB' in kks_upper or '20HCB' in kks_upper:
        for wn, w in wos.items():
            if 'SOOT BLOWER' in w['job_description'].upper(): return wn

    # HP BYPASS / TURBINE
    if 'HP BYPASS' in desc_upper or '10LBF' in kks_upper or '20LBF' in kks_upper or '10LCE' in kks_upper or '20LCE' in kks_upper:
        for wn, w in wos.items():
            if 'HP BYPASS' in w['job_description'].upper(): return wn

    # TURBINE / STG
    if 'TURBINE' in desc_upper or 'STG' in area_upper or 'EXTRACTION' in desc_upper or '10MAA' in kks_upper or '20MAA' in kks_upper or '10MAX' in kks_upper or '20MAX' in kks_upper or '10LBS' in kks_upper or '20LBS' in kks_upper or '10LBD' in kks_upper or '20LBD' in kks_upper:
        for wn, w in wos.items():
            if 'MAIN TURBINE' in w['job_description'].upper() or 'TURBINE CONTROL' in w['job_description'].upper() or 'TURBINE' in w['job_description'].upper(): return wn

    # COMBUSTOR
    if 'COMBUSTOR' in desc_upper or 'COMBUSTER' in desc_upper or 'BED MATERIAL' in desc_upper or 'LIMESTONE' in desc_upper:
        for wn, w in wos.items():
            if 'COMBUSTOR' in w['job_description'].upper(): return wn

    # FEED WATER / ECONOMIZER
    if 'ECONOMIZER' in desc_upper or 'FEED WATER' in desc_upper or '10LAB' in kks_upper or '20LAB' in kks_upper:
        for wn, w in wos.items():
            if 'FEED WATER LINE' in w['job_description'].upper() or 'ECONOMIZER' in w['job_description'].upper(): return wn

    # COOLING WATER / PUMPS
    if 'MAIN COOLING' in desc_upper:
        for wn, w in wos.items():
            if 'MAIN COOLING WATER PUMP 1' in w['job_description'].upper(): return wn
    if 'AUXILIARY COOLING' in desc_upper:
        for wn, w in wos.items():
            if 'AUXILIARY COOLING WATER PUMP' in w['job_description'].upper(): return wn

    for wn, w in wos.items():
        if 'BOILER' in w['job_description'].upper() and 'DISTRIBUTED' not in w['job_description'].upper():
            return wn

    return list(wos.keys())[0]

def reconcile_unit_perfect(unit_num):
    file_path = f"Template_Outage_EIC_Monitoring_unit {unit_num}.xlsx"
    wb = openpyxl.load_workbook(file_path)
    
    ws_wo = wb['WorkOrder']
    wos = {}
    for r in ws_wo.iter_rows(min_row=2):
        if r[1].value:
            wn = str(r[1].value).strip()
            wos[wn] = {
                'row_idx': r[0].row,
                'no': r[0].value,
                'no_wo': wn,
                'job_description': clean_str(r[3].value),
                'area': clean_str(r[4].value),
                'pic': clean_str(r[9].value),
                'status': clean_str(r[8].value)
            }

    ws_chk = wb['WorkOrder_Checklist']
    subtasks = []
    for r in ws_chk.iter_rows(min_row=2):
        if r[0].value and r[1].value:
            subtasks.append({
                'no_wo': str(r[0].value).strip(),
                'sub_task': clean_str(r[1].value),
                'tanggal': clean_str(r[2].value),
                'pic_task': clean_str(r[3].value),
                'selesai': bool(r[4].value),
                'temuan': clean_str(r[5].value) if len(r)>5 else '',
                'tindak_lanjut': clean_str(r[6].value) if len(r)>6 else '',
                'jumlah_foto': r[7].value if len(r)>7 and r[7].value else 0
            })

    # Reconcile Actuators
    ws_act = wb['ActuatorValve']
    act_added = 0
    for r in ws_act.iter_rows(min_row=2):
        if r[0].value or r[2].value:
            eq_id = clean_str(r[0].value)
            area = clean_str(r[1].value)
            desc = clean_str(r[2].value)
            kks = clean_str(r[3].value)
            pic = clean_str(r[5].value)
            st = clean_str(r[6].value)
            pct = r[7].value
            dt = clean_str(r[8].value)
            gen = bool(r[9].value)
            func = bool(r[10].value)
            is_done = (gen and func) or (pct == 100) or (st == 'FINISH')
            effective_kks = kks or extract_kks(desc)
            
            matched = False
            for st_item in subtasks:
                if is_same_equipment(desc, effective_kks, st_item['sub_task']):
                    matched = True
                    if is_done and not st_item['selesai']:
                        st_item['selesai'] = True
                        st_item['tanggal'] = dt or "22/08/2026"
                    elif st_item['selesai'] and not is_done:
                        r[9].value = True
                        r[10].value = True
                        r[7].value = 100
                        r[6].value = "FINISH"
                        r[8].value = st_item['tanggal'] or "22/08/2026"
                    break
                    
            if not matched:
                target_wo = map_equipment_to_wo(desc, area, effective_kks, wos)
                kks_suffix = f" {effective_kks}" if effective_kks and effective_kks not in desc else ""
                title = f"{desc}{kks_suffix}"
                subtasks.append({
                    'no_wo': target_wo,
                    'sub_task': title,
                    'tanggal': dt if is_done else None,
                    'pic_task': pic or "AMP",
                    'selesai': is_done,
                    'temuan': '',
                    'tindak_lanjut': '',
                    'jumlah_foto': 0
                })
                act_added += 1

    # Reconcile Instruments
    inst_added = 0
    inst_sheets = [
        ('Instrument_PressureTX', 'PT'),
        ('Instrument_TemperatureTX', 'TT'),
        ('Instrument_PressureSwitch', 'PS')
    ]
    for sname, itype in inst_sheets:
        if sname not in wb.sheetnames: continue
        ws_inst = wb[sname]
        for r in ws_inst.iter_rows(min_row=2):
            if r[2].value:
                area = clean_str(r[1].value)
                desc = clean_str(r[2].value)
                kks = clean_str(r[3].value)
                if sname == 'Instrument_PressureSwitch':
                    is_done = bool(r[13].value)
                    dt = clean_str(r[14].value or r[15].value)
                else:
                    is_done = bool(r[7].value)
                    dt = clean_str(r[6].value)
                effective_kks = kks or extract_kks(desc)
                
                matched = False
                for st_item in subtasks:
                    if is_same_equipment(desc, effective_kks, st_item['sub_task']):
                        matched = True
                        if is_done and not st_item['selesai']:
                            st_item['selesai'] = True
                            st_item['tanggal'] = dt or "22/08/2026"
                        elif st_item['selesai'] and not is_done:
                            if sname == 'Instrument_PressureSwitch':
                                r[13].value = True
                                r[14].value = st_item['tanggal'] or "22/08/2026"
                                r[15].value = st_item['tanggal'] or "22/08/2026"
                            else:
                                r[7].value = True
                                r[6].value = st_item['tanggal'] or "22/08/2026"
                        break
                        
                if not matched:
                    target_wo = map_equipment_to_wo(desc, area, effective_kks, wos)
                    prefix = f"{effective_kks}: " if effective_kks else ""
                    title = f"{prefix}{desc}"
                    subtasks.append({
                        'no_wo': target_wo,
                        'sub_task': title,
                        'tanggal': dt if is_done else None,
                        'pic_task': "JAPA",
                        'selesai': is_done,
                        'temuan': '',
                        'tindak_lanjut': '',
                        'jumlah_foto': 0
                    })
                    inst_added += 1

    # Write back
    while ws_chk.max_row > 1:
        ws_chk.delete_rows(2)

    for st in subtasks:
        ws_chk.append([
            st['no_wo'],
            st['sub_task'],
            st['tanggal'],
            st['pic_task'],
            st['selesai'],
            st.get('temuan') or None,
            st.get('tindak_lanjut') or None,
            st.get('jumlah_foto') or 0
        ])

    wo_stats = {}
    for st in subtasks:
        wn = st['no_wo']
        if wn not in wo_stats:
            wo_stats[wn] = {'total': 0, 'done': 0}
        wo_stats[wn]['total'] += 1
        if st['selesai']:
            wo_stats[wn]['done'] += 1

    for r in ws_wo.iter_rows(min_row=2):
        if r[1].value:
            wn = str(r[1].value).strip()
            stat = wo_stats.get(wn, {'total': 0, 'done': 0})
            total_st = stat['total']
            done_st = stat['done']
            pct = round((done_st / total_st) * 100, 1) if total_st > 0 else 0.0
            r[10].value = total_st
            r[11].value = pct
            if done_st == total_st and total_st > 0:
                r[8].value = "FINISH"
            elif done_st > 0:
                r[8].value = "IN PROGRESS"
            else:
                r[8].value = "SCHED-OK"

    wb.save(file_path)
    print(f"Unit {unit_num}: Added {act_added} actuators, {inst_added} instruments. Total subtasks: {len(subtasks)}")

if __name__ == '__main__':
    reconcile_unit_perfect(1)
    reconcile_unit_perfect(2)
