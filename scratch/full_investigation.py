import openpyxl
import re
import difflib

def full_investigation(template_path, unit_num):
    wb = openpyxl.load_workbook(template_path, data_only=True)
    
    # 1. WorkOrder & Checklist
    ws_wo = wb['WorkOrder']
    ws_chk = wb['WorkOrder_Checklist']
    
    wos = {}
    for r in list(ws_wo.iter_rows(min_row=2, values_only=True)):
        if r and r[1]:
            wos[str(r[1]).strip()] = {
                'no': r[0],
                'wo_no': str(r[1]).strip(),
                'desc': str(r[3] or '').strip(),
                'area': str(r[4] or '').strip(),
                'pic': str(r[9] or '').strip()
            }
            
    subtasks = []
    for r in list(ws_chk.iter_rows(min_row=2, values_only=True)):
        if r and r[0] and r[1]:
            subtasks.append({
                'wo_no': str(r[0]).strip(),
                'subtask': str(r[1]).strip(),
                'tanggal': str(r[2] or '').strip(),
                'pic': str(r[3] or '').strip(),
                'selesai': bool(r[4])
            })
            
    print(f"\n================================================================================")
    print(f"=== UNIT {unit_num} COMPREHENSIVE RECONCILIATION ANALYSIS ===")
    print(f"================================================================================")
    print(f"Work Orders: {len(wos)}")
    print(f"Total Checklist Sub-tasks: {len(subtasks)}")

    # 2. Actuators
    ws_act = wb['ActuatorValve']
    actuators = []
    for r in list(ws_act.iter_rows(min_row=2, values_only=True)):
        if r and any(r):
            actuators.append({
                'eq_id': str(r[0] or '').strip(),
                'area': str(r[1] or '').strip(),
                'desc': str(r[2] or '').strip().replace('\xa0', ' '),
                'kks': str(r[3] or '').strip().replace('\xa0', ' '),
                'pic': str(r[5] or '').strip(),
                'status': str(r[6] or '').strip(),
                'progress': r[7],
                'gen_insp': r[9],
                'func_test': r[10]
            })

    # 3. Instruments
    def get_inst(sheet_name, type_label):
        ws = wb[sheet_name]
        items = []
        for r in list(ws.iter_rows(min_row=2, values_only=True)):
            if r and any(r):
                items.append({
                    'no': r[0],
                    'area': str(r[1] or '').strip(),
                    'desc': str(r[2] or '').strip().replace('\xa0', ' '),
                    'kks': str(r[3] or '').strip().replace('\xa0', ' '),
                    'type': type_label,
                    'status': r[7] if len(r) > 7 else None
                })
        return items

    pt_items = get_inst('Instrument_PressureTX', 'Pressure Transmitter')
    tt_items = get_inst('Instrument_TemperatureTX', 'Temperature Transmitter')
    ps_items = get_inst('Instrument_PressureSwitch', 'Pressure Switch')
    
    print(f"Actuators in ActuatorValve sheet: {len(actuators)}")
    print(f"Instruments in Sheets: PT={len(pt_items)}, TT={len(tt_items)}, PS={len(ps_items)} (Total: {len(pt_items)+len(tt_items)+len(ps_items)})")

    # =========================================================================
    # RECONCILIATION: ACTUATORS -> SUBTASKS
    # =========================================================================
    print(f"\n--- [A] RECONCILIATION: ACTUATOR LIST vs WO SUBTASKS ---")
    
    act_found = []
    act_missing = []
    
    for act in actuators:
        kks = act['kks']
        desc = act['desc']
        # Extract potential KKS from desc if kks is empty
        kks_in_desc = re.findall(r'[0-9]{1,2}[A-Z]{3}[0-9]{2}[A-Z]{2}[0-9]{3}', desc)
        effective_kks = kks if kks else (kks_in_desc[0] if kks_in_desc else "")
        
        matches = []
        for st in subtasks:
            st_text = st['subtask'].replace('\xa0', ' ').upper()
            
            # Exact KKS match
            if effective_kks and effective_kks.upper() in st_text:
                matches.append((st['wo_no'], st['subtask'], f"KKS match ({effective_kks})"))
            # High string similarity or core substring match
            elif len(desc) > 10 and (desc.upper() in st_text or st_text in desc.upper()):
                matches.append((st['wo_no'], st['subtask'], "Direct Text Match"))
            else:
                # Try normalized core name
                clean_desc = re.sub(r'^(ACTUATOR|VALVE|MOV|AOV|UNIT \d+)\s*', '', desc, flags=re.I).strip().upper()
                if len(clean_desc) > 12 and clean_desc in st_text:
                    matches.append((st['wo_no'], st['subtask'], f"Core Name Match ({clean_desc})"))
                    
        if matches:
            act_found.append({'act': act, 'effective_kks': effective_kks, 'matches': matches})
        else:
            # Let's find best fuzzy match just in case
            best_fuzzy = []
            for st in subtasks:
                ratio = difflib.SequenceMatcher(None, desc.upper(), st['subtask'].upper()).ratio()
                if ratio > 0.6:
                    best_fuzzy.append((st['wo_no'], st['subtask'], f"Fuzzy {ratio:.2f}"))
            act_missing.append({'act': act, 'effective_kks': effective_kks, 'fuzzy': best_fuzzy})

    print(f"Total Actuators: {len(actuators)}")
    print(f"  [OK] Matched / Included in Subtasks: {len(act_found)} ({len(act_found)/len(actuators)*100:.1f}%)")
    print(f"  [MISSING/SEPARATE] Not matched directly in Subtasks: {len(act_missing)} ({len(act_missing)/len(actuators)*100:.1f}%)")
    
    print("\nDetailed breakdown of Actuators NOT matched in Subtasks:")
    for idx, m in enumerate(act_missing, 1):
        act = m['act']
        fuzzy_str = f" -> Close subtask candidate: {m['fuzzy'][0][1]} (WO: {m['fuzzy'][0][0]})" if m['fuzzy'] else " -> No close subtask"
        print(f"  {idx:2d}. [{act['eq_id']}] Area: {act['area']:<15} | KKS: {m['effective_kks']:<15} | Desc: {act['desc']}{fuzzy_str}")

    # =========================================================================
    # RECONCILIATION: INSTRUMENTS -> SUBTASKS
    # =========================================================================
    print(f"\n--- [B] RECONCILIATION: INSTRUMENT LIST vs WO SUBTASKS ---")
    all_instruments = pt_items + tt_items + ps_items
    
    inst_found = []
    inst_missing = []
    
    for inst in all_instruments:
        kks = inst['kks']
        desc = inst['desc']
        itype = inst['type']
        
        matches = []
        for st in subtasks:
            st_text = st['subtask'].replace('\xa0', ' ').upper()
            if kks and len(kks) >= 6 and kks.upper() in st_text:
                matches.append((st['wo_no'], st['subtask'], f"KKS match ({kks})"))
            elif len(desc) > 12 and (desc.upper() in st_text or st_text in desc.upper()):
                matches.append((st['wo_no'], st['subtask'], "Direct Text Match"))
            else:
                clean_desc = re.sub(r'^(PRESSURE|TEMPERATURE|TRANSMITTER|SWITCH|UNIT \d+)\s*', '', desc, flags=re.I).strip().upper()
                if len(clean_desc) > 12 and clean_desc in st_text:
                    matches.append((st['wo_no'], st['subtask'], f"Core Name Match ({clean_desc})"))
                    
        if matches:
            inst_found.append({'inst': inst, 'matches': matches})
        else:
            best_fuzzy = []
            for st in subtasks:
                ratio = difflib.SequenceMatcher(None, desc.upper(), st['subtask'].upper()).ratio()
                if ratio > 0.65:
                    best_fuzzy.append((st['wo_no'], st['subtask'], f"Fuzzy {ratio:.2f}"))
            inst_missing.append({'inst': inst, 'fuzzy': best_fuzzy})

    print(f"Total Instruments: {len(all_instruments)} (PT: {len(pt_items)}, TT: {len(tt_items)}, PS: {len(ps_items)})")
    print(f"  [OK] Matched / Included in Subtasks: {len(inst_found)} ({len(inst_found)/len(all_instruments)*100:.1f}%)")
    print(f"  [MISSING/SEPARATE] Not matched directly in Subtasks: {len(inst_missing)} ({len(inst_missing)/len(all_instruments)*100:.1f}%)")
    
    print("\nDetailed breakdown of Instruments NOT matched in Subtasks:")
    for idx, m in enumerate(inst_missing, 1):
        inst = m['inst']
        fuzzy_str = f" -> Close subtask candidate: {m['fuzzy'][0][1]} (WO: {m['fuzzy'][0][0]})" if m['fuzzy'] else " -> No close subtask"
        print(f"  {idx:2d}. [{inst['type']}] Area: {inst['area']:<15} | KKS: {inst['kks']:<15} | Desc: {inst['desc']}{fuzzy_str}")

    # =========================================================================
    # RECONCILIATION: HOW MANY SUBTASKS ARE ACTUATORS / INSTRUMENTS / MOTORS?
    # =========================================================================
    print(f"\n--- [C] SUBTASK BREAKDOWN BY WORK ORDER ---")
    wo_map = {}
    for st in subtasks:
        wn = st['wo_no']
        if wn not in wo_map:
            wo_map[wn] = {'wo_info': wos.get(wn, {'desc': 'UNKNOWN', 'area': ''}), 'subtasks': []}
        wo_map[wn]['subtasks'].append(st)
        
    print(f"Total Active WOs with Subtasks: {len(wo_map)}")
    
    # Check which WOs contain what kind of subtasks
    category_counts = {'Actuators': 0, 'Instruments': 0, 'Electrical/Motor': 0, 'Generic/Other': 0}
    for wn, data in wo_map.items():
        for st in data['subtasks']:
            st_text = st['subtask'].upper()
            if any(k in st_text for k in ['ACTUATOR', 'MOV', 'AOV', 'GATE', 'DAMPER', 'IGV']):
                category_counts['Actuators'] += 1
            elif any(k in st_text for k in ['TRANSMITTER', 'SWITCH', 'CALIBRATION', 'MEASUREMENT', 'PT', 'TT', 'RTD', 'THERMOCOUPLE', 'CP', 'CT']):
                category_counts['Instruments'] += 1
            elif any(k in st_text for k in ['MOTOR', 'BEARING', 'WINDING', 'ISOLASI', 'SOLO RUN', 'FRAME MOTOR', 'REGREASING', 'COOLING FAN', 'TERMINASI', 'GROUNDING', 'HEATER', 'BREAKER']):
                category_counts['Electrical/Motor'] += 1
            else:
                category_counts['Generic/Other'] += 1
                
    print("\nOverall Subtask Categories in WorkOrder_Checklist (721 total):")
    for cat, cnt in category_counts.items():
        print(f"  - {cat:<20}: {cnt:3d} sub-tasks ({cnt/len(subtasks)*100:.1f}%)")

if __name__ == '__main__':
    full_investigation('Template_Outage_EIC_Monitoring_unit 1.xlsx', 1)
    full_investigation('Template_Outage_EIC_Monitoring_unit 2.xlsx', 2)
