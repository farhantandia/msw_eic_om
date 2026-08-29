import openpyxl
import xlrd
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = "d:/msw/msw_eic_om"

def clean_val(val):
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
    return val

def populate_unit(unit_num):
    print(f"=== POPULATING UNIT {unit_num} ===")
    template_path = os.path.join(BASE_DIR, f"Template_Outage_EIC_Monitoring_unit {unit_num}.xlsx")
    unit_dir = os.path.join(BASE_DIR, f"Unit {unit_num}")
    
    wb_dest = openpyxl.load_workbook(template_path)
    
    # 1. POPULATE WorkOrder and WorkOrder_Checklist
    wo_src_path = os.path.join(unit_dir, f"Progress Outage Unit {unit_num} EIC.xlsx")
    if os.path.exists(wo_src_path):
        wb_src = openpyxl.load_workbook(wo_src_path, data_only=True)
        sheet_name = f"UPDATE U{unit_num}"
        if sheet_name in wb_src.sheetnames:
            ws_src = wb_src[sheet_name]
            rows = list(ws_src.iter_rows(values_only=True))
            
            ws_wo = wb_dest["WorkOrder"]
            ws_chk = wb_dest["WorkOrder_Checklist"]
            
            # Clear existing data rows (keep header row 1)
            ws_wo.delete_rows(2, ws_wo.max_row + 1)
            ws_chk.delete_rows(2, ws_chk.max_row + 1)
            
            current_wo = None
            wo_count = 0
            chk_count = 0
            
            for r in rows[11:]:
                no = r[1]
                no_wo = r[2]
                job_desc = clean_val(r[3])
                sched = clean_val(r[4])
                act_start = clean_val(r[5])
                finish = clean_val(r[6])
                status = clean_val(r[7])
                pic = clean_val(r[8])
                n_task = r[9]
                pct = r[10]
                scope = clean_val(r[11])
                remarks = clean_val(r[12])
                
                if no is not None and no_wo is not None:
                    current_wo = str(no_wo).strip()
                    area = ""
                    if job_desc:
                        # Extract basic area from job description if possible
                        if "COOLING TOWER" in job_desc.upper():
                            area = "COOLING TOWER"
                        elif "ID FAN" in job_desc.upper():
                            area = "ID FAN"
                        elif "BOILER" in job_desc.upper():
                            area = "BOILER"
                        elif "TURBINE" in job_desc.upper() or "STG" in job_desc.upper():
                            area = "TURBINE"
                        elif "COAL" in job_desc.upper() or "CHP" in job_desc.upper():
                            area = "COAL HANDLING"
                        elif "ESP" in job_desc.upper():
                            area = "ESP"
                        elif "WTP" in job_desc.upper() or "WATER" in job_desc.upper():
                            area = "WATER TREATMENT"
                        else:
                            area = "GENERAL"
                    
                    calc_pct = float(pct) if isinstance(pct, (int, float)) else 0.0
                    if calc_pct > 1.0:
                        calc_pct = round(calc_pct, 2)
                    else:
                        calc_pct = round(calc_pct * 100, 2)
                        
                    ws_wo.append([
                        no, current_wo, unit_num, job_desc, area, sched, act_start, finish,
                        status or "SCHED-OK", pic, n_task or 0, calc_pct, scope, remarks,
                        None, None, 0
                    ])
                    wo_count += 1
                elif current_wo and job_desc:
                    done = bool(pct) if isinstance(pct, bool) else (pct == True or str(pct).upper() == "TRUE")
                    ws_chk.append([
                        current_wo, job_desc, sched, pic, done, None, None, 0
                    ])
                    chk_count += 1
                    
            print(f"  WorkOrder populated: {wo_count} rows, WorkOrder_Checklist: {chk_count} rows")

    # 2. POPULATE ActuatorValve
    act_src_path = os.path.join(unit_dir, f"AMP-MSW-Progress Actuator Unit {unit_num} 2026.xlsx")
    if os.path.exists(act_src_path):
        wb_src = openpyxl.load_workbook(act_src_path, data_only=True)
        ws_src = wb_src["Actuator Valve"]
        rows = list(ws_src.iter_rows(values_only=True))
        
        ws_act = wb_dest["ActuatorValve"]
        ws_act.delete_rows(2, ws_act.max_row + 1)
        
        current_act = None
        actuators = []
        eq_id_counter = 1
        
        for r in rows[13:]:
            area = clean_val(r[0])
            desc = clean_val(r[1])
            finish_date = clean_val(r[2])
            pic = clean_val(r[3])
            status_or_sub = clean_val(r[4])
            pct_or_done = r[5]
            remarks = clean_val(r[6])
            
            if desc and ("ACTUATOR" in str(desc).upper() or area):
                if current_act:
                    actuators.append(current_act)
                eq_id = f"AV-{eq_id_counter:03d}"
                eq_id_counter += 1
                
                # Extract KKS if present at end of description
                kks = ""
                parts = desc.split()
                if len(parts) > 1 and len(parts[-1]) >= 8 and any(c.isdigit() for c in parts[-1]):
                    kks = parts[-1]
                
                pct = float(pct_or_done) if isinstance(pct_or_done, (int, float)) else 0.0
                current_act = {
                    "eq_id": eq_id,
                    "area": area or "GENERAL",
                    "desc": desc,
                    "kks": kks,
                    "unit": unit_num,
                    "pic": pic or "AMP",
                    "status": status_or_sub or "SCHED-OK",
                    "pct": pct,
                    "finish_date": finish_date,
                    "gen_insp": False,
                    "func_test": False,
                    "remarks": remarks
                }
            elif current_act and desc:
                sub = str(desc).upper()
                done = bool(pct_or_done)
                if "INSPECTION" in sub or "CLEANING" in sub:
                    current_act["gen_insp"] = done
                elif "FUNCTION" in sub or "TEST" in sub:
                    current_act["func_test"] = done
                    
        if current_act:
            actuators.append(current_act)
            
        for a in actuators:
            ws_act.append([
                a["eq_id"], a["area"], a["desc"], a["kks"], a["unit"], a["pic"],
                a["status"], a["pct"], a["finish_date"], a["gen_insp"], a["func_test"],
                a["remarks"], None, None, 0
            ])
        print(f"  ActuatorValve populated: {len(actuators)} rows")

    # 3. POPULATE Instruments
    inst_src_path = os.path.join(unit_dir, f"JAPA-MSW-Progress Transmitter & Switch Unit 2 2026.xls")
    if os.path.exists(inst_src_path):
        book = xlrd.open_workbook(inst_src_path)
        
        # Pressure TX
        if "PRESSURE TRANSMITTER" in book.sheet_names():
            sh = book.sheet_by_name("PRESSURE TRANSMITTER")
            ws_ptx = wb_dest["Instrument_PressureTX"]
            ws_ptx.delete_rows(2, ws_ptx.max_row + 1)
            ptx_count = 0
            for r in range(8, sh.nrows):
                vals = sh.row_values(r)
                no = vals[0]
                area = clean_val(vals[1])
                eq = clean_val(vals[2])
                kks = clean_val(vals[3])
                rng = clean_val(vals[4])
                tgl = clean_val(vals[5])
                wdone = bool(vals[6]) if len(vals) > 6 else False
                rem = clean_val(vals[7]) if len(vals) > 7 else None
                if eq:
                    ws_ptx.append([
                        no or (ptx_count + 1), area, eq, kks, unit_num, rng, tgl, wdone, rem, None, None, 0
                    ])
                    ptx_count += 1
            print(f"  Instrument_PressureTX populated: {ptx_count} rows")

        # Temperature TX
        if "TEMPERATURE TRANSMITTER" in book.sheet_names():
            sh = book.sheet_by_name("TEMPERATURE TRANSMITTER")
            ws_ttx = wb_dest["Instrument_TemperatureTX"]
            ws_ttx.delete_rows(2, ws_ttx.max_row + 1)
            ttx_count = 0
            for r in range(8, sh.nrows):
                vals = sh.row_values(r)
                no = vals[0]
                area = clean_val(vals[1])
                eq = clean_val(vals[2])
                kks = clean_val(vals[3])
                rng = clean_val(vals[4])
                finish_date = clean_val(vals[5])
                wdone = bool(vals[6]) if len(vals) > 6 else False
                rem = clean_val(vals[7]) if len(vals) > 7 else None
                if eq:
                    ws_ttx.append([
                        no or (ttx_count + 1), area, eq, kks, unit_num, rng, finish_date, wdone, rem, None, None, 0
                    ])
                    ttx_count += 1
            print(f"  Instrument_TemperatureTX populated: {ttx_count} rows")

        # Pressure Switch
        if "PRESSURE SWITCH" in book.sheet_names():
            sh = book.sheet_by_name("PRESSURE SWITCH")
            ws_psw = wb_dest["Instrument_PressureSwitch"]
            ws_psw.delete_rows(2, ws_psw.max_row + 1)
            psw_count = 0
            for r in range(11, sh.nrows):
                vals = sh.row_values(r)
                no = vals[0]
                area = clean_val(vals[1])
                desc = clean_val(vals[2])
                kks = clean_val(vals[3])
                sub_area = clean_val(vals[4]) if len(vals) > 4 else None
                status_ok = clean_val(vals[5]) if len(vals) > 5 else None
                set_pt = clean_val(vals[6]) if len(vals) > 6 else None
                contact = clean_val(vals[7]) if len(vals) > 7 else None
                
                if desc and no:
                    psw_count += 1
                    ws_psw.append([
                        no, area, desc, kks, unit_num, sub_area, set_pt, contact,
                        None, None, None, None, status_ok, False, None, None, None, None, None, 0
                    ])
            print(f"  Instrument_PressureSwitch populated: {psw_count} rows")

    # 4. POPULATE PIC_Scope_Master
    pic_src_path = os.path.join(BASE_DIR, "PIC Outage Unit 1 2026.xlsx")
    if os.path.exists(pic_src_path):
        wb_pic = openpyxl.load_workbook(pic_src_path, data_only=True)
        ws_scope = wb_dest["PIC_Scope_Master"]
        ws_scope.delete_rows(2, ws_scope.max_row + 1)
        scope_count = 0
        
        for sheet_name, scope_type in [("Vendor Scope", "Vendor"), ("MSW SCOPE", "MSW")]:
            if sheet_name in wb_pic.sheetnames:
                ws_s = wb_pic[sheet_name]
                curr_kat = ""
                for row in list(ws_s.iter_rows(values_only=True))[4:]:
                    col0 = clean_val(row[0])
                    eq_name = clean_val(row[1])
                    scope_me = clean_val(row[2])
                    act_desc = clean_val(row[3])
                    pic_name = clean_val(row[4])
                    
                    if col0 and any(col0.startswith(prefix) for prefix in ['A.', 'B.', 'C.', 'D.', 'E.']):
                        curr_kat = col0
                    
                    if eq_name and eq_name != "NAMA EQUIPMENT / SCOPE PEKERJAAN":
                        ws_scope.append([
                            curr_kat, eq_name, scope_type, scope_me, act_desc, pic_name, unit_num
                        ])
                        scope_count += 1
        print(f"  PIC_Scope_Master populated: {scope_count} rows")
        
    wb_dest.save(template_path)
    print(f"Successfully saved {template_path}\n")

populate_unit(1)
populate_unit(2)
