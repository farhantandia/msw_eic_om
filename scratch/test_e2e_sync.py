import json
import sys
import os
sys.path.insert(0, os.path.abspath('.'))

import openpyxl
from server import (
    load_master_components,
    save_quick_subtask_toggle,
    save_quick_actuator_toggle,
    save_quick_instrument_toggle,
    get_excel_path
)

def test_all():
    print("==================================================")
    print("E2E SYNC AND COMPONENT TEST")
    print("==================================================")
    
    # 1. Master components test
    comps = load_master_components(1)
    print(f"Master components Unit 1: {len(comps['actuators'])} actuators, {len(comps['instruments'])} instruments")
    assert len(comps['actuators']) > 0, "No actuators loaded"
    assert len(comps['instruments']) > 0, "No instruments loaded"
    print("✅ Master components loaded successfully.")

    # 2. Test WO Subtask -> Actuator Sync
    wb_path = get_excel_path(1)
    
    # Let's test Actuator MOV 10HNA61AA001 (AV-011) in WO-100826-0160
    act_wo = "WO-100826-0160"
    act_desc = "UNIT 1 INDUCED DRAUGHT FAN 1 INLET GATE ACTUATOR MOV 20HNA61AA001"
    act_id = "AV-011"
    
    print(f"\n[Test 1] WO Subtask -> Actuator Sync: Subtask '{act_desc}' (Done=True)")
    res1 = save_quick_subtask_toggle({
        "unit": 1,
        "no_wo": act_wo,
        "sub_task": act_desc,
        "selesai": True
    })
    print("Toggle Done Result:", res1)
    
    # Verify in ActuatorValve sheet
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    ws_act = wb["ActuatorValve"]
    matched_act = None
    for r in ws_act.iter_rows(min_row=2, values_only=True):
        if r[0] == act_id:
            matched_act = r
            break
            
    assert matched_act is not None, f"Actuator {act_id} not found"
    print(f"Verified Actuator in Sheet: ID={matched_act[0]}, Desc={matched_act[2]}, Status={matched_act[6]}, Progress={matched_act[7]}%")
    assert matched_act[6] == "FINISH" and matched_act[7] == 100, f"Actuator {act_id} not synced to FINISH/100%"
    print("✅ Subtask -> Actuator Sync: SUCCESS (FINISH / 100%)")

    # 3. Test Actuator -> Subtask Sync
    print(f"\n[Test 2] Actuator -> Subtask Sync: Setting Actuator [{act_id}] to 0% (SCHED-OK)")
    res_act = save_quick_actuator_toggle({
        "unit": 1,
        "equipment_id": act_id,
        "field": "general_inspection",
        "value": False
    })
    # Also set function_test to False
    save_quick_actuator_toggle({
        "unit": 1,
        "equipment_id": act_id,
        "field": "function_test",
        "value": False
    })
    print("Toggle Actuator Result:", res_act)
    
    # Verify in WorkOrder_Checklist
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    ws_chk = wb["WorkOrder_Checklist"]
    found_sub = None
    for r in ws_chk.iter_rows(min_row=2, values_only=True):
        if r[0] == act_wo and r[1] == act_desc:
            found_sub = r
            break
            
    assert found_sub is not None, "Subtask not found in checklist"
    print(f"Verified Subtask in Checklist: WO={found_sub[0]}, Subtask={found_sub[1]}, Selesai={found_sub[4]}")
    assert found_sub[4] == False, "Subtask should be False after actuator reset"
    print("✅ Actuator -> Subtask Sync: SUCCESS (Subtask reset to False)")

    # 4. Test Instrument Sync: Instrument -> Subtask & Subtask -> Instrument
    ws_inst = wb["Instrument_PressureTX"]
    sample_ptx = list(ws_inst.iter_rows(min_row=2, values_only=True))[0]
    sample_kks = str(sample_ptx[3])
    print(f"\n[Test 3] Instrument -> Subtask Sync: PTX KKS [{sample_kks}] -> status_wdone=True")
    res_inst = save_quick_instrument_toggle({
        "unit": 1,
        "type": "pressure_tx",
        "key": sample_kks,
        "status_wdone": True
    })
    print("Toggle Instrument Result:", res_inst)
    
    # Verify matching subtask in checklist
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    ws_chk = wb["WorkOrder_Checklist"]
    inst_sub = None
    for r in ws_chk.iter_rows(min_row=2, values_only=True):
        if sample_kks in str(r[1]):
            inst_sub = r
            break
            
    assert inst_sub is not None, f"Subtask with KKS {sample_kks} not found"
    print(f"Verified Instrument Subtask in Checklist: WO={inst_sub[0]}, Subtask={inst_sub[1]}, Selesai={inst_sub[4]}")
    assert inst_sub[4] == True, "Instrument subtask should be True"
    print("✅ Instrument -> Subtask Sync: SUCCESS (Subtask marked True)")

    # 5. Subtask -> Instrument Sync: Uncheck subtask
    print(f"\n[Test 4] Subtask -> Instrument Sync: Unchecking Subtask [{inst_sub[0]}] '{inst_sub[1]}'")
    save_quick_subtask_toggle({
        "unit": 1,
        "no_wo": str(inst_sub[0]),
        "sub_task": str(inst_sub[1]),
        "selesai": False
    })
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    ws_inst = wb["Instrument_PressureTX"]
    for r in ws_inst.iter_rows(min_row=2, values_only=True):
        if str(r[3]) == sample_kks:
            print(f"Verified Instrument in Sheet: KKS={r[3]}, Status_WDone={r[7]}")
            assert r[7] == False, "Instrument status_wdone should be False after subtask uncheck"
            print("✅ Subtask -> Instrument Sync: SUCCESS (Instrument reset to False)")
            break

    print("\n==================================================")
    print("🎉 ALL 4 TESTS PASSED! FULL BIDIRECTIONAL SYNC IS 100% OPERATIONAL.")
    print("==================================================")

if __name__ == '__main__':
    test_all()
