import http.server
import socketserver
import json
import os
import shutil
import urllib.parse
import sys
import threading
import mimetypes
import datetime
import base64
import openpyxl
import difflib
import re

sys.stdout.reconfigure(encoding='utf-8')

PORT = 8000
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(os.path.abspath(sys.executable))
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

FILE_LOCK = threading.Lock()

def safe_save_workbook(wb, path):
    try:
        wb.save(path)
        return True, ""
    except PermissionError:
        return False, "The Excel file is currently open in Microsoft Excel! Please close the Excel file on your computer first so that the application can save changes."
    except Exception as e:
        return False, f"Failed to save Excel file: {str(e)}"

def get_excel_path(unit):
    candidates = [
        os.path.join(BASE_DIR, f"Template_Outage_EIC_Monitoring_unit {unit}.xlsx"),
        os.path.join(BASE_DIR, f"Template_Outage_EIC_Monitoring_unit_{unit}.xlsx"),
        os.path.join(BASE_DIR, f"Template_Outage_EIC_Monitoring_Unit {unit}.xlsx"),
        os.path.join(os.getcwd(), f"Template_Outage_EIC_Monitoring_unit {unit}.xlsx")
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    return candidates[0]

def clean_val(val):
    if val is None:
        return ""
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, float):
        if val.is_integer():
            return int(val)
    return str(val).strip() if isinstance(val, str) else val

PIC_MAP = {
    "m toher": "M TOHER",
    "m. toher": "M TOHER",
    "m.toher": "M TOHER",
    "toher": "M TOHER",
    "m zaini": "M ZAINI",
    "m. zaini": "M ZAINI",
    "m.zaini": "M ZAINI",
    "zaini": "M ZAINI",
    "zaini y": "ZAINI Y",
    "m iqbal": "M IQBAL",
    "m. iqbal": "M IQBAL",
    "m.iqbal": "M IQBAL",
    "iqbal": "M IQBAL",
    "aidhiy": "H AIDI",
    "h aidi": "H AIDI",
    "h. aidi": "H AIDI",
    "aidi": "H AIDI",
    "jhon fm": "JHON",
    "jhon": "JHON",
    "john": "JHON",
    "ferryus": "FERRYUS",
    "feryus": "FERRYUS",
    "yoga": "YOGA",
    "farhan": "FARHAN",
    "m farhan": "FARHAN",
    "m. farhan": "FARHAN",
    "majid": "MAJID",
    "dede": "DEDE",
    "amp": "AMP",
    "japa": "JAPA",
    "msw": "MSW",
    "bagus": "BAGUS"
}

def normalize_pic(val):
    if not val:
        return ""
    p_clean = str(val).strip().lower()
    return PIC_MAP.get(p_clean, str(val).strip().upper())

def get_unique_pics():
    pics = set()
    for unit in [1, 2]:
        path = get_excel_path(unit)
        if os.path.exists(path):
            wb = openpyxl.load_workbook(path, data_only=True)
            if "PIC_Scope_Master" in wb.sheetnames:
                ws = wb["PIC_Scope_Master"]
                for r in list(ws.iter_rows(values_only=True))[1:]:
                    if len(r) > 5 and r[5]:
                        pstr = normalize_pic(r[5])
                        if pstr and pstr not in ["PIC", "FINISH", "IN PROGRESS", "SCHED-OK", "NONE"]:
                            pics.add(pstr)
    pic_list = sorted(list(pics))
    if not pic_list:
        pic_list = ["AMP", "BAGUS", "DEDE", "FARHAN", "FERRYUS", "H AIDI", "JAPA", "JHON", "M IQBAL", "M TOHER", "M ZAINI", "MAJID", "MSW", "YOGA", "ZAINI Y"]
    return pic_list

def save_add_pic(data):
    pic_name = str(data.get("pic_name", "")).strip()
    if not pic_name:
        return {"status": "error", "message": "Nama PIC tidak boleh kosong."}
    
    with FILE_LOCK:
        for unit in [1, 2]:
            path = get_excel_path(unit)
            if os.path.exists(path):
                wb = openpyxl.load_workbook(path)
                if "PIC_Scope_Master" in wb.sheetnames:
                    ws = wb["PIC_Scope_Master"]
                    exists = False
                    for r in ws.iter_rows(min_row=2):
                        if len(r) > 5 and r[5].value and str(r[5].value).strip().upper() == pic_name.upper():
                            exists = True
                            break
                    if not exists:
                        ws.append(["A. Master PIC EIC", f"Job Scope {pic_name}", "MSW", "ME/SI/SE", "EIC Personnel", pic_name, unit])
                        ok, err = safe_save_workbook(wb, path)
                        if not ok: return {"status": "error", "message": err}
                        
    return {"status": "success", "message": f"PIC '{pic_name}' successfully added to Master PIC (Unit 1 & Unit 2)!"}

def save_delete_pic(data):
    pic_name = str(data.get("pic_name", "")).strip()
    if not pic_name:
        return {"status": "error", "message": "PIC name is empty."}
    
    with FILE_LOCK:
        for unit in [1, 2]:
            path = get_excel_path(unit)
            if os.path.exists(path):
                wb = openpyxl.load_workbook(path)
                if "PIC_Scope_Master" in wb.sheetnames:
                    ws = wb["PIC_Scope_Master"]
                    delete_indices = []
                    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                        if len(row) > 5 and row[5].value and str(row[5].value).strip().upper() == pic_name.upper():
                            delete_indices.append(idx)
                    for idx in reversed(delete_indices):
                        ws.delete_rows(idx, 1)
                    ok, err = safe_save_workbook(wb, path)
                    if not ok: return {"status": "error", "message": err}
    return {"status": "success", "message": f"PIC '{pic_name}' successfully deleted from Master PIC."}

def save_scope_update(data):
    unit = data.get("unit", 1)
    idx = data.get("row_index")
    path = get_excel_path(unit)
    
    with FILE_LOCK:
        wb = openpyxl.load_workbook(path)
        if "PIC_Scope_Master" in wb.sheetnames:
            ws = wb["PIC_Scope_Master"]
            data_rows = list(ws.iter_rows(min_row=2))
            if idx is not None and 0 <= idx < len(data_rows):
                target_row = data_rows[idx]
                if "pic" in data:
                    pic_val = normalize_pic(data["pic"]) if data["pic"] else None
                    if len(target_row) > 5:
                        target_row[5].value = pic_val
                    else:
                        ws.cell(row=idx+2, column=6, value=pic_val)
                if "kategori" in data and len(target_row) > 0: target_row[0].value = data["kategori"]
                if "nama_equipment" in data and len(target_row) > 1: target_row[1].value = data["nama_equipment"]
                if "tipe_scope" in data and len(target_row) > 2: target_row[2].value = data["tipe_scope"]
                ok, err = safe_save_workbook(wb, path)
                if not ok: return {"status": "error", "message": err}
                return {"status": "success", "message": "Master Scope updated successfully."}
    return {"status": "error", "message": "Master Scope row not found."}

def save_delete_scope(data):
    unit = data.get("unit", 1)
    idx = data.get("row_index")
    path = get_excel_path(unit)
    
    with FILE_LOCK:
        wb = openpyxl.load_workbook(path)
        if "PIC_Scope_Master" in wb.sheetnames:
            ws = wb["PIC_Scope_Master"]
            data_rows = list(ws.iter_rows(min_row=2))
            if idx is not None and 0 <= idx < len(data_rows):
                ws.delete_rows(idx + 2, 1)
                ok, err = safe_save_workbook(wb, path)
                if not ok: return {"status": "error", "message": err}
                return {"status": "success", "message": "Master Scope row deleted successfully."}
    return {"status": "error", "message": "Master Scope row not found."}

def save_add_scope(data):
    unit = data.get("unit", 1)
    kategori = str(data.get("kategori", "")).strip() or "GENERAL"
    eq = str(data.get("nama_equipment", "")).strip()
    tipe = str(data.get("tipe_scope", "MSW")).strip()
    pic = normalize_pic(str(data.get("pic", "")).strip()) if data.get("pic") else None
    if not eq:
        return {"status": "error", "message": "Equipment / Scope description is required."}
    path = get_excel_path(unit)
    with FILE_LOCK:
        wb = openpyxl.load_workbook(path)
        if "PIC_Scope_Master" in wb.sheetnames:
            ws = wb["PIC_Scope_Master"]
            ws.append([kategori, eq, tipe, "", "", pic, unit])
            ok, err = safe_save_workbook(wb, path)
            if not ok: return {"status": "error", "message": err}
            return {"status": "success", "message": f"Master Scope '{eq}' successfully added!"}
    return {"status": "error", "message": "Sheet PIC_Scope_Master not found."}
    return {"status": "error", "message": "Sheet PIC_Scope_Master tidak ditemukan."}

def save_add_wo(data):
    unit = data.get("unit", 1)
    no_wo = str(data.get("no_wo", "")).strip()
    job_desc = str(data.get("job_description", "")).strip()
    area = str(data.get("area", "")).strip() or "GENERAL"
    pic = str(data.get("pic", "")).strip()
    sched = str(data.get("tanggal_schedule", "")).strip()
    status = str(data.get("status", "SCHED-OK")).strip()
    checklist_str = str(data.get("checklist_str", "")).strip()
    
    if not no_wo or not job_desc:
        return {"status": "error", "message": "WO Number and Job Description are required."}
        
    path = get_excel_path(unit)
    with FILE_LOCK:
        wb = openpyxl.load_workbook(path)
        if "WorkOrder" in wb.sheetnames:
            ws = wb["WorkOrder"]
            for r in ws.iter_rows(min_row=2):
                if r[1].value and str(r[1].value).strip().upper() == no_wo.upper():
                    return {"status": "error", "message": f"Work Order {no_wo} already exists."}
            
            max_no = 0
            for r in ws.iter_rows(min_row=2):
                if r[0].value and isinstance(r[0].value, (int, float)):
                    max_no = max(max_no, int(r[0].value))
                    
            subtasks = []
            if checklist_str:
                raw_items = [item.strip() for line in checklist_str.splitlines() for item in line.split(',')]
                subtasks = [it for it in raw_items if it]
                
            if "WorkOrder_Checklist" in wb.sheetnames:
                ws_chk = wb["WorkOrder_Checklist"]
                for st_text in subtasks:
                    ws_chk.append([no_wo, st_text, None, pic, False, None, None, 0])
                    
            n_task = len(subtasks)
            ws.append([max_no + 1, no_wo, unit, job_desc, area, sched, None, None, status, pic, n_task, 0.0, "SUPERVISE, QC", None, None, None, 0])
            ok, err = safe_save_workbook(wb, path)
            if not ok: return {"status": "error", "message": err}
    return {"status": "success", "message": f"Work Order '{no_wo}' with {len(subtasks)} sub-tasks successfully added!"}

def save_delete_wo(data):
    unit = data.get("unit", 1)
    no_wo = str(data.get("no_wo", "")).strip()
    path = get_excel_path(unit)
    
    with FILE_LOCK:
        wb = openpyxl.load_workbook(path)
        if "WorkOrder" in wb.sheetnames:
            ws = wb["WorkOrder"]
            for idx, r in enumerate(ws.iter_rows(min_row=2), start=2):
                if r[1].value and str(r[1].value).strip() == no_wo:
                    ws.delete_rows(idx, 1)
                    break
        if "WorkOrder_Checklist" in wb.sheetnames:
            ws_chk = wb["WorkOrder_Checklist"]
            del_rows = []
            for idx, r in enumerate(ws_chk.iter_rows(min_row=2), start=2):
                if r[0].value and str(r[0].value).strip() == no_wo:
                    del_rows.append(idx)
            for idx in reversed(del_rows):
                ws_chk.delete_rows(idx, 1)
        ok, err = safe_save_workbook(wb, path)
        if not ok: return {"status": "error", "message": err}
    return {"status": "success", "message": f"Work Order '{no_wo}' successfully deleted."}

def save_add_subtask(data):
    unit = data.get("unit", 1)
    no_wo = str(data.get("no_wo", "")).strip()
    sub_task = str(data.get("sub_task", "")).strip()
    pic = str(data.get("pic", "")).strip()
    
    if not no_wo or not sub_task:
        return {"status": "error", "message": "Sub-task description cannot be empty."}
        
    path = get_excel_path(unit)
    with FILE_LOCK:
        wb = openpyxl.load_workbook(path)
        if "WorkOrder_Checklist" in wb.sheetnames:
            ws_chk = wb["WorkOrder_Checklist"]
            ws_chk.append([no_wo, sub_task, None, pic, False, None, None, 0])
            
        total_cnt = 0
        done_cnt = 0
        if "WorkOrder_Checklist" in wb.sheetnames:
            ws_chk = wb["WorkOrder_Checklist"]
            for r in ws_chk.iter_rows(min_row=2):
                if r[0].value and str(r[0].value).strip() == no_wo:
                    total_cnt += 1
                    if r[4].value == True:
                        done_cnt += 1
                        
        if "WorkOrder" in wb.sheetnames:
            ws_wo = wb["WorkOrder"]
            for r in ws_wo.iter_rows(min_row=2):
                if r[1].value and str(r[1].value).strip() == no_wo:
                    r[10].value = total_cnt
                    if total_cnt > 0:
                        pct = round((done_cnt / total_cnt) * 100, 1)
                        r[11].value = pct
                    break
        ok, err = safe_save_workbook(wb, path)
        if not ok: return {"status": "error", "message": err}
    return {"status": "success", "message": f"Sub-task successfully added to WO {no_wo}!"}

def save_delete_subtask(data):
    unit = data.get("unit", 1)
    no_wo = str(data.get("no_wo", "")).strip()
    sub_task = str(data.get("sub_task", "")).strip()
    path = get_excel_path(unit)
    
    with FILE_LOCK:
        wb = openpyxl.load_workbook(path)
        if "WorkOrder_Checklist" in wb.sheetnames:
            ws_chk = wb["WorkOrder_Checklist"]
            for idx, r in enumerate(ws_chk.iter_rows(min_row=2), start=2):
                if r[0].value and str(r[0].value).strip() == no_wo and r[1].value and str(r[1].value).strip() == sub_task:
                    ws_chk.delete_rows(idx, 1)
                    break
                    
        total_cnt = 0
        done_cnt = 0
        if "WorkOrder_Checklist" in wb.sheetnames:
            ws_chk = wb["WorkOrder_Checklist"]
            for r in ws_chk.iter_rows(min_row=2):
                if r[0].value and str(r[0].value).strip() == no_wo:
                    total_cnt += 1
                    if r[4].value == True:
                        done_cnt += 1
                        
        if "WorkOrder" in wb.sheetnames:
            ws_wo = wb["WorkOrder"]
            for r in ws_wo.iter_rows(min_row=2):
                if r[1].value and str(r[1].value).strip() == no_wo:
                    r[10].value = total_cnt
                    pct = round((done_cnt / total_cnt) * 100, 1) if total_cnt > 0 else 0.0
                    r[11].value = pct
                    break
        ok, err = safe_save_workbook(wb, path)
        if not ok: return {"status": "error", "message": err}
    return {"status": "success", "message": f"Sub-task '{sub_task}' successfully deleted."}


def extract_kks(text):
    if not text:
        return ""
    matches = re.findall(r'[0-9]{1,2}[A-Z]{3}[0-9]{2}[A-Z]{2}[0-9]{3}', str(text))
    if matches:
        return matches[0]
    matches2 = re.findall(r'[0-9]{1,2}[A-Z]{3}[0-9]{2}[A-Z]{1,2}[0-9]{1,3}', str(text))
    if matches2:
        return matches2[0]
    return ""

def is_same_component(comp_desc, comp_kks, st_desc):
    comp_kks = str(comp_kks or '').replace('\xa0', ' ').strip().upper()
    st_desc_upper = str(st_desc or '').replace('\xa0', ' ').strip().upper()
    comp_desc_upper = str(comp_desc or '').replace('\xa0', ' ').strip().upper()
    
    st_kks = extract_kks(st_desc_upper)
    comp_extracted_kks = extract_kks(comp_desc_upper) or comp_kks
    
    # 1. KKS Match (Direct or ignoring leading Unit digit 10 vs 20)
    if comp_kks and st_kks:
        if comp_kks == st_kks or comp_kks[2:] == st_kks[2:]:
            return True
    if comp_extracted_kks and st_kks:
        if comp_extracted_kks == st_kks or comp_extracted_kks[2:] == st_kks[2:]:
            return True
    if comp_kks and len(comp_kks) >= 6 and (comp_kks in st_desc_upper or comp_kks[2:] in st_desc_upper):
        return True
    if st_kks and len(st_kks) >= 6 and (st_kks in comp_desc_upper or st_kks[2:] in comp_desc_upper):
        return True
        
    # 2. Text Match (normalizing keywords like DRAUGHT -> DRAFT, etc.)
    def normalize_text(t):
        t = re.sub(r'[\xa0\s]+', ' ', t).strip()
        t = t.replace('DRAUGHT', 'DRAFT')
        t = re.sub(r'^(ACTUATOR|VALVE|MOV|AOV|UNIT \d+|MSW|[0-9]{1,2}[A-Z]{3}[0-9]{2}[A-Z]{1,2}[0-9]{1,3}[:\s]*)\s*', '', t).strip()
        t = re.sub(r'\s*(ACTUATOR|VALVE|MOV|AOV|[0-9]{1,2}[A-Z]{3}[0-9]{2}[A-Z]{1,2}[0-9]{1,3})\s*$', '', t).strip()
        return t

    norm_comp = normalize_text(comp_desc_upper)
    norm_st = normalize_text(st_desc_upper)
    
    if norm_comp and norm_st:
        if norm_comp == norm_st or norm_comp in norm_st or norm_st in norm_comp:
            return True
        if len(norm_comp) > 8 and len(norm_st) > 8:
            ratio = difflib.SequenceMatcher(None, norm_comp, norm_st).ratio()
            if ratio > 0.8:
                return True
            
    return False

def sync_subtask_to_components(wb, no_wo, subtask_desc, is_done, date_str):
    today_str = date_str or datetime.date.today().strftime("%d/%m/%Y")
    
    # 1. Check ActuatorValve
    if "ActuatorValve" in wb.sheetnames:
        ws_act = wb["ActuatorValve"]
        for row in ws_act.iter_rows(min_row=2):
            eq_id = str(row[0].value or '').strip()
            desc = str(row[2].value or '').strip()
            kks = str(row[3].value or '').strip()
            if is_same_component(desc, kks or eq_id, subtask_desc):
                row[9].value = is_done
                row[10].value = is_done
                row[7].value = 100 if is_done else 0
                row[6].value = "FINISH" if is_done else "SCHED-OK"
                row[8].value = today_str if is_done else None

    # 2. Check Instruments
    inst_sheets = [
        "Instrument_PressureTX",
        "Instrument_TemperatureTX",
        "Instrument_PressureSwitch"
    ]
    for sname in inst_sheets:
        if sname in wb.sheetnames:
            ws_inst = wb[sname]
            for row in ws_inst.iter_rows(min_row=2):
                desc = str(row[2].value or '').strip()
                kks = str(row[3].value or '').strip()
                if is_same_component(desc, kks, subtask_desc):
                    if sname == "Instrument_PressureSwitch":
                        row[13].value = is_done
                        row[14].value = today_str if is_done else None
                        row[15].value = today_str if is_done else None
                        if len(row) > 20: row[20].value = is_done
                    else:
                        row[7].value = is_done
                        row[6].value = today_str if is_done else None
                        if len(row) > 12: row[12].value = is_done

def sync_actuator_to_subtasks(wb, eq_id, kks, desc, is_done, date_str):
    today_str = date_str or datetime.date.today().strftime("%d/%m/%Y")
    affected_wos = set()
    
    if "WorkOrder_Checklist" in wb.sheetnames:
        ws_chk = wb["WorkOrder_Checklist"]
        for row in ws_chk.iter_rows(min_row=2):
            no_wo = str(row[0].value or '').strip()
            st_desc = str(row[1].value or '').strip()
            if is_same_component(desc, kks or eq_id, st_desc):
                row[4].value = is_done
                row[2].value = today_str if is_done else None
                if no_wo:
                    affected_wos.add(no_wo)
                    
    recalculate_wos_progress(wb, affected_wos, today_str)

def sync_instrument_to_subtasks(wb, inst_type, kks, no_val, desc, is_done, date_str):
    today_str = date_str or datetime.date.today().strftime("%d/%m/%Y")
    affected_wos = set()
    
    if "WorkOrder_Checklist" in wb.sheetnames:
        ws_chk = wb["WorkOrder_Checklist"]
        for row in ws_chk.iter_rows(min_row=2):
            no_wo = str(row[0].value or '').strip()
            st_desc = str(row[1].value or '').strip()
            if is_same_component(desc, kks or no_val, st_desc):
                row[4].value = is_done
                row[2].value = today_str if is_done else None
                if no_wo:
                    affected_wos.add(no_wo)
                    
    recalculate_wos_progress(wb, affected_wos, today_str)

def recalculate_wos_progress(wb, affected_wos, date_str):
    if not affected_wos:
        return
    if "WorkOrder_Checklist" not in wb.sheetnames or "WorkOrder" not in wb.sheetnames:
        return
    
    ws_chk = wb["WorkOrder_Checklist"]
    ws_wo = wb["WorkOrder"]
    
    wo_stats = {wn: {'total': 0, 'done': 0} for wn in affected_wos}
    for row in ws_chk.iter_rows(min_row=2):
        wn = str(row[0].value or '').strip()
        if wn in wo_stats:
            wo_stats[wn]['total'] += 1
            if row[4].value == True:
                wo_stats[wn]['done'] += 1
                
    for row in ws_wo.iter_rows(min_row=2):
        wn = str(row[1].value or '').strip()
        if wn in wo_stats:
            total_st = wo_stats[wn]['total']
            done_st = wo_stats[wn]['done']
            row[10].value = total_st
            pct = round((done_st / total_st) * 100, 1) if total_st > 0 else 0.0
            row[11].value = pct
            if done_st == total_st and total_st > 0:
                row[8].value = "FINISH"
                if not row[7].value:
                    row[7].value = date_str
            elif done_st > 0:
                row[8].value = "IN PROGRESS"
                if row[8].value == "FINISH":
                    row[7].value = None
            else:
                row[8].value = "SCHED-OK"
                row[7].value = None

def load_master_components(unit):
    path = get_excel_path(unit)
    if not os.path.exists(path):
        return {"actuators": [], "instruments": []}
    
    with FILE_LOCK:
        wb = openpyxl.load_workbook(path, data_only=True)
        actuators = []
        if "ActuatorValve" in wb.sheetnames:
            ws = wb["ActuatorValve"]
            for r in list(ws.iter_rows(values_only=True))[1:]:
                if r[0] or r[2]:
                    actuators.append({
                        "equipment_id": clean_val(r[0]),
                        "area": clean_val(r[1]) or "BOILER",
                        "equipment_description": clean_val(r[2]),
                        "kks": clean_val(r[3]),
                        "pic": normalize_pic(clean_val(r[5])),
                        "status": clean_val(r[6]) or "SCHED-OK",
                        "persen_progress": clean_val(r[7]) or 0
                    })
                    
        instruments = []
        inst_sheets = [
            ("Instrument_PressureTX", "pressure_tx", "Pressure Transmitter (PT)"),
            ("Instrument_TemperatureTX", "temperature_tx", "Temperature Transmitter (TT)"),
            ("Instrument_PressureSwitch", "pressure_switch", "Pressure Switch (PS)")
        ]
        for sname, itype, ilabel in inst_sheets:
            if sname in wb.sheetnames:
                ws = wb[sname]
                for r in list(ws.iter_rows(values_only=True))[1:]:
                    if r[2]:
                        instruments.append({
                            "no": clean_val(r[0]),
                            "area": clean_val(r[1]) or "GENERAL",
                            "equipment": clean_val(r[2]),
                            "kks": clean_val(r[3]),
                            "type": itype,
                            "type_label": ilabel,
                            "range": clean_val(r[5]) if len(r)>5 else "",
                            "status_wdone": bool(r[13] if sname=="Instrument_PressureSwitch" else r[7])
                        })
                        
    return {"actuators": actuators, "instruments": instruments}

def save_quick_subtask_toggle(data):
    unit = data.get("unit", 1)
    no_wo = str(data.get("no_wo", "")).strip()
    sub_task = str(data.get("sub_task", "")).strip()
    sub_idx = data.get("sub_idx")
    selesai = bool(data.get("selesai", False))
    path = get_excel_path(unit)
    today_str = datetime.datetime.now().strftime("%d/%m/%Y")

    with FILE_LOCK:
        wb = openpyxl.load_workbook(path)
        if "WorkOrder_Checklist" in wb.sheetnames:
            ws_chk = wb["WorkOrder_Checklist"]
            match_idx = 0
            for r in ws_chk.iter_rows(min_row=2):
                if r[0].value and str(r[0].value).strip() == no_wo:
                    if (sub_task and str(r[1].value).strip() == sub_task) or (sub_idx is not None and match_idx == int(sub_idx)):
                        r[4].value = selesai
                        actual_sub_desc = str(r[1].value).strip()
                        if selesai:
                            if not r[2].value:
                                r[2].value = today_str
                        else:
                            r[2].value = None
                        # Auto-sync to Actuators and Instruments
                        sync_subtask_to_components(wb, no_wo, actual_sub_desc, selesai, today_str)
                        break
                    match_idx += 1

        total_cnt = 0
        done_cnt = 0
        if "WorkOrder_Checklist" in wb.sheetnames:
            ws_chk = wb["WorkOrder_Checklist"]
            for r in ws_chk.iter_rows(min_row=2):
                if r[0].value and str(r[0].value).strip() == no_wo:
                    total_cnt += 1
                    if r[4].value == True:
                        done_cnt += 1

        new_pct = 0.0
        new_status = "SCHED-OK"
        if "WorkOrder" in wb.sheetnames:
            ws_wo = wb["WorkOrder"]
            for r in ws_wo.iter_rows(min_row=2):
                if r[1].value and str(r[1].value).strip() == no_wo:
                    r[10].value = total_cnt
                    if total_cnt > 0:
                        new_pct = round((done_cnt / total_cnt) * 100, 1)
                        r[11].value = new_pct
                        if done_cnt == total_cnt:
                            new_status = "FINISH"
                            if not r[7].value:
                                r[7].value = today_str
                        elif done_cnt > 0:
                            new_status = "IN PROGRESS"
                            if r[8].value == "FINISH":
                                r[7].value = None
                        else:
                            new_status = "SCHED-OK"
                            r[7].value = None
                        r[8].value = new_status
                    break

        ok, err = safe_save_workbook(wb, path)
        if not ok: return {"status": "error", "message": err}

    return {
        "status": "success",
        "message": f"Sub-task updated ({new_pct}%)",
        "persen_progress": new_pct,
        "wo_status": new_status,
        "done_cnt": done_cnt,
        "total_cnt": total_cnt
    }

def save_batch_subtask_toggle(data):
    unit = data.get("unit", 1)
    no_wo = str(data.get("no_wo", "")).strip()
    action = data.get("action", "mark_all_done")
    is_done = (action == "mark_all_done")
    today_str = datetime.date.today().strftime("%d/%m/%Y") if is_done else None
    path = get_excel_path(unit)

    with FILE_LOCK:
        wb = openpyxl.load_workbook(path)
        total_cnt = 0
        if "WorkOrder_Checklist" in wb.sheetnames:
            ws_chk = wb["WorkOrder_Checklist"]
            for r in ws_chk.iter_rows(min_row=2):
                if r[0].value and str(r[0].value).strip() == no_wo:
                    total_cnt += 1
                    r[4].value = is_done
                    r[2].value = today_str
                    actual_sub_desc = str(r[1].value or '').strip()
                    if actual_sub_desc:
                        sync_subtask_to_components(wb, no_wo, actual_sub_desc, is_done, today_str)

        new_pct = 100.0 if (is_done and total_cnt > 0) else 0.0
        new_status = "FINISH" if (is_done and total_cnt > 0) else "SCHED-OK"
        if "WorkOrder" in wb.sheetnames:
            ws_wo = wb["WorkOrder"]
            for r in ws_wo.iter_rows(min_row=2):
                if r[1].value and str(r[1].value).strip() == no_wo:
                    r[10].value = total_cnt
                    r[11].value = new_pct
                    r[8].value = new_status
                    r[7].value = today_str if is_done else None
                    break

        ok, err = safe_save_workbook(wb, path)
        if not ok: return {"status": "error", "message": err}

    return {
        "status": "success",
        "message": f"All sub-tasks of WO {no_wo} {'marked as completed (100%)' if is_done else 'reset to initial status'}",
        "persen_progress": new_pct,
        "wo_status": new_status,
        "done_cnt": total_cnt if is_done else 0,
        "total_cnt": total_cnt
    }

def save_quick_actuator_toggle(data):
    unit = data.get("unit", 1)
    eq_id = str(data.get("equipment_id", "")).strip()
    field = data.get("field")
    val = bool(data.get("value", False))
    path = get_excel_path(unit)

    with FILE_LOCK:
        wb = openpyxl.load_workbook(path)
        if "ActuatorValve" in wb.sheetnames:
            ws = wb["ActuatorValve"]
            for r in ws.iter_rows(min_row=2):
                if r[0].value and str(r[0].value).strip() == eq_id:
                    if field == "general_inspection":
                        r[9].value = val
                    elif field == "function_test":
                        r[10].value = val
                    
                    gen = bool(r[9].value)
                    func = bool(r[10].value)
                    today_str = datetime.date.today().strftime("%d/%m/%Y")
                    if gen and func:
                        r[7].value = 100
                        r[6].value = "FINISH"
                        r[8].value = today_str
                    elif gen or func:
                        r[7].value = 50
                        r[6].value = "IN PROGRESS"
                        r[8].value = None
                    else:
                        r[7].value = 0
                        r[6].value = "SCHED-OK"
                        r[8].value = None
                    
                    new_pct = r[7].value
                    new_status = r[6].value
                    act_desc = str(r[2].value or '').strip()
                    act_kks = str(r[3].value or '').strip()
                    is_finish = (new_status == "FINISH" or new_pct == 100)
                    sync_actuator_to_subtasks(wb, eq_id, act_kks, act_desc, is_finish, today_str)
                    break
        ok, err = safe_save_workbook(wb, path)
        if not ok: return {"status": "error", "message": err}

    return {
        "status": "success",
        "message": f"Actuator {eq_id} updated ({new_pct}%)",
        "persen_progress": new_pct,
        "act_status": new_status
    }

def save_quick_instrument_toggle(data):
    unit = data.get("unit", 1)
    inst_type = data.get("type", "pressure_tx")
    key = str(data.get("key", "")).strip()
    status_wdone = bool(data.get("status_wdone", False))
    path = get_excel_path(unit)

    with FILE_LOCK:
        wb = openpyxl.load_workbook(path)
        sheet_map = {
            "pressure_tx": "Instrument_PressureTX",
            "temperature_tx": "Instrument_TemperatureTX",
            "pressure_switch": "Instrument_PressureSwitch",
            "ptx": "Instrument_PressureTX",
            "ttx": "Instrument_TemperatureTX",
            "psw": "Instrument_PressureSwitch"
        }
        sname = sheet_map.get(inst_type)
        if sname and sname in wb.sheetnames:
            ws = wb[sname]
            today_str = datetime.date.today().strftime("%d/%m/%Y")
            for row in ws.iter_rows(min_row=2):
                row_kks = str(row[3].value).strip() if row[3].value else ""
                row_no = str(row[0].value).strip() if row[0].value else ""
                if row_kks == key or row_no == key:
                    if sname == "Instrument_PressureSwitch":
                        row[13].value = status_wdone
                        row[14].value = today_str if status_wdone else None
                        row[15].value = today_str if status_wdone else None
                    else:
                        row[7].value = status_wdone
                        row[6].value = today_str if status_wdone else None

                    inst_desc = str(row[2].value or '').strip()
                    inst_kks = str(row[3].value or '').strip()
                    inst_no = str(row[0].value or '').strip()
                    sync_instrument_to_subtasks(wb, inst_type, inst_kks, inst_no, inst_desc, status_wdone, today_str)
                    break
        ok, err = safe_save_workbook(wb, path)
        if not ok: return {"status": "error", "message": err}

    return {
        "status": "success",
        "message": f"Instrument status updated to {'DONE' if status_wdone else 'SCHEDULED'}",
        "status_wdone": status_wdone
    }

def save_add_actuator(data):
    unit = data.get("unit", 1)
    eq_id = str(data.get("equipment_id", "")).strip()
    desc = str(data.get("equipment_description", "")).strip()
    area = str(data.get("area", "")).strip() or "BOILER"
    kks = str(data.get("kks", "")).strip()
    pic = str(data.get("pic", "")).strip()
    
    if not eq_id or not desc:
        return {"status": "error", "message": "Equipment ID dan Description wajib diisi."}
        
    path = get_excel_path(unit)
    with FILE_LOCK:
        wb = openpyxl.load_workbook(path)
        if "ActuatorValve" in wb.sheetnames:
            ws = wb["ActuatorValve"]
            ws.append([eq_id, area, desc, kks, unit, pic, "SCHED-OK", 0, None, False, False, None, None, None, 0])
            ok, err = safe_save_workbook(wb, path)
            if not ok: return {"status": "error", "message": err}
    return {"status": "success", "message": f"Actuator Valve '{eq_id}' successfully added!"}

def save_delete_actuator(data):
    unit = data.get("unit", 1)
    eq_id = str(data.get("equipment_id", "")).strip()
    path = get_excel_path(unit)
    
    with FILE_LOCK:
        wb = openpyxl.load_workbook(path)
        if "ActuatorValve" in wb.sheetnames:
            ws = wb["ActuatorValve"]
            for idx, r in enumerate(ws.iter_rows(min_row=2), start=2):
                if r[0].value and str(r[0].value).strip() == eq_id:
                    ws.delete_rows(idx, 1)
                    break
            ok, err = safe_save_workbook(wb, path)
            if not ok: return {"status": "error", "message": err}
    return {"status": "success", "message": f"Actuator '{eq_id}' successfully deleted."}

def save_bulk_instrument_update(data):
    unit = data.get("unit", 1)
    items = data.get("items", [])
    status_wdone = bool(data.get("status_wdone", True))
    path = get_excel_path(unit)
    
    if not items:
        return {"status": "error", "message": "No instruments selected."}
        
    with FILE_LOCK:
        wb = openpyxl.load_workbook(path)
        sheet_name_map = {
            "pressure_tx": "Instrument_PressureTX",
            "temperature_tx": "Instrument_TemperatureTX",
            "pressure_switch": "Instrument_PressureSwitch",
            "ptx": "Instrument_PressureTX",
            "ttx": "Instrument_TemperatureTX",
            "psw": "Instrument_PressureSwitch"
        }
        updated_count = 0
        for itm in items:
            itype = itm.get("type")
            key = str(itm.get("kks") or itm.get("no") or "").strip()
            sname = sheet_name_map.get(itype)
            if sname and sname in wb.sheetnames:
                ws = wb[sname]
                for row in ws.iter_rows(min_row=2):
                    row_kks = str(row[3].value).strip() if row[3].value else ""
                    row_no = str(row[0].value).strip() if row[0].value else ""
                    if row_kks == key or row_no == key:
                        today_str = datetime.date.today().strftime("%d/%m/%Y")
                        if itype in ["pressure_tx", "temperature_tx", "ptx", "ttx"]:
                            row[7].value = status_wdone
                            row[6].value = today_str if status_wdone else None
                        elif itype in ["pressure_switch", "psw"]:
                            row[13].value = status_wdone
                            row[14].value = today_str if status_wdone else None
                            row[15].value = today_str if status_wdone else None

                        inst_desc = str(row[2].value or '').strip()
                        inst_kks = str(row[3].value or '').strip()
                        inst_no = str(row[0].value or '').strip()
                        sync_instrument_to_subtasks(wb, itype, inst_kks, inst_no, inst_desc, status_wdone, today_str)
                        updated_count += 1
                        break
        ok, err = safe_save_workbook(wb, path)
        if not ok: return {"status": "error", "message": err}
        
    action_str = "FINISHED (DONE)" if status_wdone else "SCHEDULED"
    return {"status": "success", "message": f"Successfully updated {updated_count} instruments to {action_str}!"}

def save_add_instrument(data):
    unit = data.get("unit", 1)
    inst_type = data.get("type", "pressure_tx")
    eq = str(data.get("equipment", "")).strip()
    kks = str(data.get("kks", "")).strip()
    area = str(data.get("area", "")).strip() or "GENERAL"
    sub_area = str(data.get("sub_area", "")).strip()
    set_point = str(data.get("set_point", "")).strip()
    contact_type = str(data.get("contact_type", "NO")).strip() or "NO"
    rng = str(data.get("range", "")).strip()
    
    if not eq:
        return {"status": "error", "message": "Equipment description is required."}
        
    path = get_excel_path(unit)
    with FILE_LOCK:
        wb = openpyxl.load_workbook(path)
        sheet_map = {
            "pressure_tx": "Instrument_PressureTX",
            "temperature_tx": "Instrument_TemperatureTX",
            "pressure_switch": "Instrument_PressureSwitch",
            "ptx": "Instrument_PressureTX",
            "ttx": "Instrument_TemperatureTX",
            "psw": "Instrument_PressureSwitch"
        }
        sname = sheet_map.get(inst_type)
        if sname and sname in wb.sheetnames:
            ws = wb[sname]
            max_no = ws.max_row
            if inst_type in ["pressure_switch", "psw"]:
                sp_val = set_point or rng
                ws.append([max_no, area, eq, kks, unit, sub_area or area, sp_val, contact_type, None, None, None, None, "OK", False, None, None, None, None, None, 0, False])
            else:
                ws.append([max_no, area, eq, kks, unit, rng or set_point, None, False, None, None, None, 0, False])
            ok, err = safe_save_workbook(wb, path)
            if not ok: return {"status": "error", "message": err}
    return {"status": "success", "message": f"Instrument '{eq}' successfully added!"}

def save_delete_instrument(data):
    unit = data.get("unit", 1)
    inst_type = data.get("type", "pressure_tx")
    no = data.get("no")
    kks = data.get("kks")
    path = get_excel_path(unit)
    
    with FILE_LOCK:
        wb = openpyxl.load_workbook(path)
        sheet_map = {
            "pressure_tx": "Instrument_PressureTX",
            "temperature_tx": "Instrument_TemperatureTX",
            "pressure_switch": "Instrument_PressureSwitch",
            "ptx": "Instrument_PressureTX",
            "ttx": "Instrument_TemperatureTX",
            "psw": "Instrument_PressureSwitch"
        }
        sname = sheet_map.get(inst_type)
        if sname and sname in wb.sheetnames:
            ws = wb[sname]
            for idx, r in enumerate(ws.iter_rows(min_row=2), start=2):
                r_no = str(r[0].value).strip() if r[0].value else ""
                r_kks = str(r[3].value).strip() if r[3].value else ""
                if (no and r_no == str(no)) or (kks and r_kks == str(kks)):
                    ws.delete_rows(idx, 1)
                    break
            ok, err = safe_save_workbook(wb, path)
            if not ok: return {"status": "error", "message": err}
    return {"status": "success", "message": "Instrument successfully deleted."}

def load_unit_data(unit):
    path = get_excel_path(unit)
    if not os.path.exists(path):
        return {"error": f"Excel template file for Unit {unit} not found: {os.path.basename(path)}"}
    
    with FILE_LOCK:
        wb = openpyxl.load_workbook(path, data_only=True)
        
        # 1. WorkOrder & Checklist
        wo_map = {}
        wo_list = []
        if "WorkOrder" in wb.sheetnames:
            ws = wb["WorkOrder"]
            for r in list(ws.iter_rows(values_only=True))[1:]:
                if r[1]:
                    no_wo = str(r[1]).strip()
                    item = {
                        "no": clean_val(r[0]),
                        "no_wo": no_wo,
                        "unit": r[2] or unit,
                        "job_description": clean_val(r[3]),
                        "area": clean_val(r[4]) or "GENERAL",
                        "tanggal_schedule": clean_val(r[5]),
                        "tanggal_actual_start": clean_val(r[6]),
                        "tanggal_finish": clean_val(r[7]),
                        "status": clean_val(r[8]) or "SCHED-OK",
                        "pic": normalize_pic(clean_val(r[9])),
                        "n_task": clean_val(r[10]) or 0,
                        "persen_progress": clean_val(r[11]) or 0,
                        "scope": clean_val(r[12]),
                        "remarks": clean_val(r[13]),
                        "temuan": clean_val(r[14]),
                        "tindak_lanjut": clean_val(r[15]),
                        "jumlah_foto": clean_val(r[16]) or 0,
                        "checklist": []
                    }
                    wo_map[no_wo] = item
                    wo_list.append(item)
                    
        if "WorkOrder_Checklist" in wb.sheetnames:
            ws = wb["WorkOrder_Checklist"]
            for r in list(ws.iter_rows(values_only=True))[1:]:
                no_wo = clean_val(r[0])
                if no_wo in wo_map:
                    wo_map[no_wo]["checklist"].append({
                        "sub_task": clean_val(r[1]),
                        "tanggal": clean_val(r[2]),
                        "pic_task": normalize_pic(clean_val(r[3])),
                        "selesai": bool(r[4]),
                        "temuan": clean_val(r[5]),
                        "tindak_lanjut": clean_val(r[6]),
                        "jumlah_foto": clean_val(r[7]) or 0
                    })
                    
        # 2. ActuatorValve
        act_list = []
        if "ActuatorValve" in wb.sheetnames:
            ws = wb["ActuatorValve"]
            for r in list(ws.iter_rows(values_only=True))[1:]:
                if r[0] or r[2]:
                    raw_st = clean_val(r[6])
                    if isinstance(raw_st, bool):
                        clean_st = "FINISH" if raw_st else "SCHED-OK"
                    else:
                        clean_st = str(raw_st or "SCHED-OK")
                    act_list.append({
                        "equipment_id": clean_val(r[0]),
                        "area": clean_val(r[1]) or "BOILER",
                        "equipment_description": clean_val(r[2]),
                        "kks": clean_val(r[3]),
                        "unit": r[4] or unit,
                        "pic": normalize_pic(clean_val(r[5])),
                        "status": clean_st,
                        "persen_progress": clean_val(r[7]) or 0,
                        "finish_date": clean_val(r[8]),
                        "general_inspection": bool(r[9]),
                        "function_test": bool(r[10]),
                        "remarks": clean_val(r[11]),
                        "temuan": clean_val(r[12]),
                        "tindak_lanjut": clean_val(r[13]),
                        "jumlah_foto": clean_val(r[14]) or 0
                    })
                    
        # 3. Instruments (2 Checklists: Kalibrasi & Verifikasi, finish determiner = Verifikasi)
        ptx_list = []
        if "Instrument_PressureTX" in wb.sheetnames:
            ws = wb["Instrument_PressureTX"]
            for r in list(ws.iter_rows(values_only=True))[1:]:
                if r[2]:
                    is_verif = bool(r[7])
                    is_calib = bool(r[12]) if len(r) > 12 and r[12] is not None else is_verif
                    ptx_list.append({
                        "no": clean_val(r[0]),
                        "area": clean_val(r[1]) or "GENERAL",
                        "equipment": clean_val(r[2]),
                        "kks": clean_val(r[3]),
                        "unit": r[4] or unit,
                        "range": clean_val(r[5]),
                        "tanggal": clean_val(r[6]),
                        "kalibrasi": is_calib,
                        "verifikasi": is_verif,
                        "status_wdone": is_verif,
                        "remarks": clean_val(r[8]),
                        "temuan": clean_val(r[9]),
                        "tindak_lanjut": clean_val(r[10]),
                        "jumlah_foto": clean_val(r[11]) or 0
                    })
                    
        ttx_list = []
        if "Instrument_TemperatureTX" in wb.sheetnames:
            ws = wb["Instrument_TemperatureTX"]
            for r in list(ws.iter_rows(values_only=True))[1:]:
                if r[2]:
                    is_verif = bool(r[7])
                    is_calib = bool(r[12]) if len(r) > 12 and r[12] is not None else is_verif
                    ttx_list.append({
                        "no": clean_val(r[0]),
                        "area": clean_val(r[1]) or "BOILER",
                        "equipment": clean_val(r[2]),
                        "kks": clean_val(r[3]),
                        "unit": r[4] or unit,
                        "range": clean_val(r[5]),
                        "finish_date": clean_val(r[6]),
                        "kalibrasi": is_calib,
                        "verifikasi": is_verif,
                        "status_wdone": is_verif,
                        "remarks": clean_val(r[8]),
                        "temuan": clean_val(r[9]),
                        "tindak_lanjut": clean_val(r[10]),
                        "jumlah_foto": clean_val(r[11]) or 0
                    })
                    
        psw_list = []
        if "Instrument_PressureSwitch" in wb.sheetnames:
            ws = wb["Instrument_PressureSwitch"]
            for r in list(ws.iter_rows(values_only=True))[1:]:
                if r[2]:
                    is_verif = bool(r[13])
                    is_calib = bool(r[20]) if len(r) > 20 and r[20] is not None else is_verif
                    psw_list.append({
                        "no": clean_val(r[0]),
                        "area": clean_val(r[1]) or "BOILER",
                        "description": clean_val(r[2]),
                        "equipment": clean_val(r[2]),
                        "kks": clean_val(r[3]),
                        "unit": r[4] or unit,
                        "sub_area": clean_val(r[5]),
                        "set_point": clean_val(r[6]),
                        "contact_type": clean_val(r[7]) or "NO",
                        "asfound_set": clean_val(r[8]),
                        "asfound_reset": clean_val(r[9]),
                        "asleft_set": clean_val(r[10]),
                        "asleft_reset": clean_val(r[11]),
                        "status_ok_notok": clean_val(r[12]) or "OK",
                        "kalibrasi": is_calib,
                        "verifikasi": is_verif,
                        "status_wdone": is_verif,
                        "dated": clean_val(r[14]),
                        "finish_date": clean_val(r[15]),
                        "remarks": clean_val(r[16]),
                        "temuan": clean_val(r[17]),
                        "tindak_lanjut": clean_val(r[18]),
                        "jumlah_foto": clean_val(r[19]) or 0
                    })
                    
        # 4. Scope Master
        scope_list = []
        if "PIC_Scope_Master" in wb.sheetnames:
            ws = wb["PIC_Scope_Master"]
            for idx, r in enumerate(list(ws.iter_rows(values_only=True))[1:]):
                if r[1] or (len(r) > 5 and r[5]):
                    scope_list.append({
                        "row_index": idx,
                        "kategori": clean_val(r[0]),
                        "nama_equipment": clean_val(r[1]),
                        "tipe_scope": clean_val(r[2]),
                        "scope_kerja": clean_val(r[3]),
                        "deskripsi_aktivitas": clean_val(r[4]),
                        "pic": clean_val(r[5]),
                        "unit": r[6] if len(r) > 6 and r[6] else unit
                    })

        # Summary Stats (Based on Subtasks / Checklists)
        wo_finish = sum(1 for w in wo_list if w["status"] == "FINISH")
        wo_inprog = sum(1 for w in wo_list if w["status"] == "IN PROGRESS")
        wo_sched = sum(1 for w in wo_list if w["status"] == "SCHED-OK")
        wo_subtask_total = sum(len(w.get("checklist", [])) for w in wo_list)
        wo_subtask_done = sum(sum(1 for c in w.get("checklist", []) if c.get("selesai")) for w in wo_list)
        wo_pct = round((wo_subtask_done / wo_subtask_total * 100), 1) if wo_subtask_total > 0 else (round((wo_finish / len(wo_list) * 100), 1) if len(wo_list) else 0)
        
        act_finish = sum(1 for a in act_list if a["status"] == "FINISH")
        act_inprog = sum(1 for a in act_list if a["status"] == "IN PROGRESS")
        act_sched = sum(1 for a in act_list if a["status"] == "SCHED-OK")
        act_subtask_total = len(act_list) * 2
        act_subtask_done = sum((1 if a.get("general_inspection") else 0) + (1 if a.get("function_test") else 0) for a in act_list)
        act_pct = round((act_subtask_done / act_subtask_total * 100), 1) if act_subtask_total > 0 else (round((act_finish / len(act_list) * 100), 1) if len(act_list) else 0)
        
        # Instrument finish status is solely determined by Verifikasi
        ptx_done = sum(1 for p in ptx_list if p["verifikasi"])
        ttx_done = sum(1 for t in ttx_list if t["verifikasi"])
        psw_done = sum(1 for s in psw_list if s["verifikasi"])
        inst_total = len(ptx_list) + len(ttx_list) + len(psw_list)
        inst_done = ptx_done + ttx_done + psw_done
        inst_pct = round((inst_done / inst_total * 100), 1) if inst_total else 0
        
        findings_count = sum(1 for w in wo_list if w["temuan"] or w["jumlah_foto"] > 0) + \
                         sum(1 for a in act_list if a["temuan"] or a["jumlah_foto"] > 0) + \
                         sum(1 for p in ptx_list if p["temuan"] or p["jumlah_foto"] > 0) + \
                         sum(1 for t in ttx_list if t["temuan"] or t["jumlah_foto"] > 0) + \
                         sum(1 for s in psw_list if s["temuan"] or s["jumlah_foto"] > 0)

        # Area Breakdown
        area_map = {}
        def add_area_stat(area_name, is_done):
            aname = (area_name or "GENERAL").strip().upper()
            if aname.startswith("BOILER"): aname = "BOILER"
            elif "COOLING TOWER" in aname: aname = "COOLING TOWER"
            elif "ESP" in aname or "ID FAN" in aname: aname = "ESP & ID FAN"
            elif "PA FAN" in aname or "FD FAN" in aname or "SA FAN" in aname: aname = "AIR FANS (PA/FD/SA)"
            elif "TURBINE" in aname or "STG" in aname or "GENERATOR" in aname: aname = "TURBINE & STG"
            elif "FEED WATER" in aname or "BFP" in aname: aname = "FEED WATER SYSTEM"
            elif "COAL" in aname or "ASH" in aname: aname = "COAL & ASH HANDLING"
            
            if aname not in area_map:
                area_map[aname] = {"total": 0, "done": 0}
            area_map[aname]["total"] += 1
            if is_done:
                area_map[aname]["done"] += 1

        for w in wo_list: add_area_stat(w["area"], w["status"] == "FINISH")
        for a in act_list: add_area_stat(a["area"], a["status"] == "FINISH")
        for p in ptx_list: add_area_stat(p["area"], p["status_wdone"])
        for t in ttx_list: add_area_stat(t["area"], t["status_wdone"])
        for s in psw_list: add_area_stat(s["area"], s["status_wdone"])

        area_stats = []
        for aname, st in sorted(area_map.items(), key=lambda x: x[1]["total"], reverse=True):
            pct = round((st["done"] / st["total"] * 100), 1) if st["total"] > 0 else 0
            area_stats.append({"area": aname, "total": st["total"], "done": st["done"], "pct": pct})

        # Calculate Grand Outage Progress based on total Work Order subtasks
        grand_total = wo_subtask_total
        grand_done = wo_subtask_done
        grand_pct = round((grand_done / grand_total * 100), 1) if grand_total > 0 else 0

        summary = {
            "grand_pct": grand_pct,
            "grand_total": grand_total,
            "grand_done": grand_done,
            "wo": {
                "total": len(wo_list),
                "finish": wo_finish,
                "in_progress": wo_inprog,
                "sched": wo_sched,
                "subtask_total": wo_subtask_total,
                "subtask_done": wo_subtask_done,
                "pct": wo_pct
            },
            "actuator": {
                "total": len(act_list),
                "finish": act_finish,
                "in_progress": act_inprog,
                "sched": act_sched,
                "subtask_total": act_subtask_total,
                "subtask_done": act_subtask_done,
                "pct": act_pct
            },
            "instrument": {
                "total": inst_total,
                "done": inst_done,
                "pct": inst_pct
            },
            "findings_count": findings_count,
            "area_stats": area_stats
        }

        return {
            "unit": unit,
            "summary": summary,
            "work_orders": wo_list,
            "actuators": act_list,
            "pressure_tx": ptx_list,
            "temperature_tx": ttx_list,
            "pressure_switch": psw_list,
            "scope_master": scope_list,
            "pics": get_unique_pics()
        }

def load_actuator_matrix():
    data_u1 = load_unit_data(1)
    data_u2 = load_unit_data(2)

    acts_u1 = {a["equipment_id"]: a for a in data_u1.get("actuators", [])}
    acts_u2 = {a["equipment_id"]: a for a in data_u2.get("actuators", [])}

    all_ids = list(dict.fromkeys(list(acts_u1.keys()) + list(acts_u2.keys())))
    matrix = []
    for eq_id in all_ids:
        item_u1 = acts_u1.get(eq_id, {})
        item_u2 = acts_u2.get(eq_id, {})
        desc = item_u1.get("equipment_description") or item_u2.get("equipment_description") or eq_id
        area = item_u1.get("area") or item_u2.get("area") or "BOILER"
        kks = item_u1.get("kks") or item_u2.get("kks") or ""
        matrix.append({
            "equipment_id": eq_id,
            "equipment_description": desc,
            "area": area,
            "kks": kks,
            "u1": {
                "status": item_u1.get("status", "SCHED-OK"),
                "pct": item_u1.get("persen_progress", 0),
                "pic": item_u1.get("pic", "-"),
                "gen": item_u1.get("general_inspection", False),
                "func": item_u1.get("function_test", False),
                "temuan": item_u1.get("temuan", ""),
                "jumlah_foto": item_u1.get("jumlah_foto", 0)
            },
            "u2": {
                "status": item_u2.get("status", "SCHED-OK"),
                "pct": item_u2.get("persen_progress", 0),
                "pic": item_u2.get("pic", "-"),
                "gen": item_u2.get("general_inspection", False),
                "func": item_u2.get("function_test", False),
                "temuan": item_u2.get("temuan", ""),
                "jumlah_foto": item_u2.get("jumlah_foto", 0)
            }
        })
    return matrix

def save_wo_update(data):
    unit = data.get("unit", 1)
    no_wo = data.get("no_wo")
    path = get_excel_path(unit)
    
    with FILE_LOCK:
        wb = openpyxl.load_workbook(path)
        if "WorkOrder" in wb.sheetnames:
            ws = wb["WorkOrder"]
            for row in ws.iter_rows(min_row=2):
                if row[1].value and str(row[1].value).strip() == no_wo:
                    if "pic" in data: row[9].value = normalize_pic(data["pic"])
                    if "persen_progress" in data: row[11].value = data["persen_progress"]
                    if "tanggal_schedule" in data: row[5].value = data["tanggal_schedule"]
                    if "tanggal_actual_start" in data: row[6].value = data["tanggal_actual_start"]
                    if "tanggal_finish" in data: row[7].value = data["tanggal_finish"]
                    if "remarks" in data: row[13].value = data["remarks"]
                    if "temuan" in data: row[14].value = data["temuan"]
                    if "tindak_lanjut" in data: row[15].value = data["tindak_lanjut"]
                    if "jumlah_foto" in data: row[16].value = data["jumlah_foto"]
                    break

        if "checklist" in data and "WorkOrder_Checklist" in wb.sheetnames:
            ws_chk = wb["WorkOrder_Checklist"]
            chk_items = {item["sub_task"]: item for item in data["checklist"]}
            for row in ws_chk.iter_rows(min_row=2):
                if row[0].value and str(row[0].value).strip() == no_wo:
                    sub_desc = str(row[1].value).strip() if row[1].value else ""
                    if sub_desc in chk_items:
                        c_data = chk_items[sub_desc]
                        is_done = bool(c_data.get("selesai"))
                        row[4].value = is_done
                        if is_done:
                            row[2].value = c_data.get("tanggal") or datetime.date.today().strftime("%d/%m/%Y")
                        else:
                            row[2].value = None
                        if "pic_task" in c_data: row[3].value = normalize_pic(c_data["pic_task"])
                        # Sync sub-task to ActuatorValve and Instruments
                        sync_subtask_to_components(wb, no_wo, sub_desc, is_done, row[2].value)

        if "WorkOrder_Checklist" in wb.sheetnames:
            ws_chk = wb["WorkOrder_Checklist"]
            total_cnt = 0
            done_cnt = 0
            for row in ws_chk.iter_rows(min_row=2):
                if row[0].value and str(row[0].value).strip() == no_wo:
                    total_cnt += 1
                    if row[4].value == True:
                        done_cnt += 1
            if total_cnt > 0:
                new_pct = round((done_cnt / total_cnt) * 100, 1)
                ws_wo = wb["WorkOrder"]
                for row in ws_wo.iter_rows(min_row=2):
                    if row[1].value and str(row[1].value).strip() == no_wo:
                        row[11].value = new_pct
                        if done_cnt == total_cnt:
                            row[8].value = "FINISH"
                            if not row[7].value: row[7].value = datetime.date.today().strftime("%d/%m/%Y")
                        elif done_cnt > 0:
                            row[8].value = "IN PROGRESS"
                        else:
                            row[8].value = "SCHED-OK"
                        break

        ok, err = safe_save_workbook(wb, path)
        if not ok: return {"status": "error", "message": err}
    return {"status": "success", "message": f"WO {no_wo} updated successfully."}

def save_actuator_update(data):
    unit = data.get("unit", 1)
    eq_id = str(data.get("equipment_id", "")).strip()
    desc = str(data.get("equipment_description", "")).strip()
    path = get_excel_path(unit)
    
    with FILE_LOCK:
        wb = openpyxl.load_workbook(path)
        if "ActuatorValve" in wb.sheetnames:
            ws = wb["ActuatorValve"]
            for row in ws.iter_rows(min_row=2):
                row_eq = str(row[0].value).strip() if row[0].value else ""
                row_desc = str(row[2].value).strip() if row[2].value else ""
                if (eq_id and row_eq == eq_id) or (desc and row_desc == desc):
                    if "pic" in data: row[5].value = normalize_pic(data["pic"])
                    if "finish_date" in data: row[8].value = data["finish_date"]
                    if "general_inspection" in data: row[9].value = bool(data["general_inspection"])
                    if "function_test" in data: row[10].value = bool(data["function_test"])
                    if "remarks" in data: row[11].value = data["remarks"]
                    if "temuan" in data: row[12].value = data["temuan"]
                    if "tindak_lanjut" in data: row[13].value = data["tindak_lanjut"]
                    if "jumlah_foto" in data: row[14].value = data["jumlah_foto"]
                    
                    g_done = bool(row[9].value)
                    f_done = bool(row[10].value)
                    today_str = datetime.date.today().strftime("%d/%m/%Y")
                    if g_done and f_done:
                        row[6].value = "FINISH"
                        row[7].value = 100
                        row[8].value = data.get("finish_date") or today_str
                    elif g_done or f_done:
                        row[6].value = "IN PROGRESS"
                        row[7].value = 50
                        row[8].value = None
                    else:
                        row[6].value = "SCHED-OK"
                        row[7].value = 0
                        row[8].value = None
                    break
        ok, err = safe_save_workbook(wb, path)
        if not ok: return {"status": "error", "message": err}
    return {"status": "success", "message": f"Actuator {eq_id or desc} updated successfully."}

def save_instrument_update(data):
    unit = data.get("unit", 1)
    inst_type = data.get("type", "pressure_tx")
    kks = data.get("kks")
    no = data.get("no")
    path = get_excel_path(unit)
    
    with FILE_LOCK:
        wb = openpyxl.load_workbook(path)
        sheet_name_map = {
            "pressure_tx": "Instrument_PressureTX",
            "temperature_tx": "Instrument_TemperatureTX",
            "pressure_switch": "Instrument_PressureSwitch",
            "ptx": "Instrument_PressureTX",
            "ttx": "Instrument_TemperatureTX",
            "psw": "Instrument_PressureSwitch"
        }
        sname = sheet_name_map.get(inst_type)
        if sname and sname in wb.sheetnames:
            ws = wb[sname]
            for row in ws.iter_rows(min_row=2):
                row_kks = str(row[3].value).strip() if row[3].value else ""
                row_no = str(row[0].value).strip() if row[0].value else ""
                if (kks and row_kks == str(kks)) or (no and row_no == str(no)):
                    today_str = datetime.date.today().strftime("%d/%m/%Y")
                    if inst_type in ["pressure_tx", "temperature_tx", "ptx", "ttx"]:
                        is_verif = bool(data.get("verifikasi")) if "verifikasi" in data else bool(data.get("status_wdone"))
                        row[7].value = is_verif
                        if is_verif:
                            row[6].value = today_str
                        else:
                            row[6].value = None
                        if "kalibrasi" in data:
                            row_idx = row[0].row
                            ws.cell(row=row_idx, column=13, value=bool(data["kalibrasi"]))
                        if "remarks" in data: row[8].value = data["remarks"]
                        if "temuan" in data: row[9].value = data["temuan"]
                        if "tindak_lanjut" in data: row[10].value = data["tindak_lanjut"]
                        if "jumlah_foto" in data: row[11].value = data["jumlah_foto"]
                    elif inst_type in ["pressure_switch", "psw"]:
                        is_verif = bool(data.get("verifikasi")) if "verifikasi" in data else bool(data.get("status_wdone"))
                        row[13].value = is_verif
                        if is_verif:
                            row[14].value = today_str
                            row[15].value = today_str
                        else:
                            row[14].value = None
                            row[15].value = None
                        if "sub_area" in data: row[5].value = data["sub_area"]
                        if "set_point" in data: row[6].value = data["set_point"]
                        if "contact_type" in data: row[7].value = data["contact_type"]
                        if "asfound_set" in data: row[8].value = data["asfound_set"]
                        if "asfound_reset" in data: row[9].value = data["asfound_reset"]
                        if "asleft_set" in data: row[10].value = data["asleft_set"]
                        if "asleft_reset" in data: row[11].value = data["asleft_reset"]
                        if "status_ok_notok" in data: row[12].value = data["status_ok_notok"]
                        if "kalibrasi" in data:
                            row_idx = row[0].row
                            ws.cell(row=row_idx, column=21, value=bool(data["kalibrasi"]))
                        if "remarks" in data: row[16].value = data["remarks"]
                        if "temuan" in data: row[17].value = data["temuan"]
                        if "tindak_lanjut" in data: row[18].value = data["tindak_lanjut"]
                        if "jumlah_foto" in data: row[19].value = data["jumlah_foto"]

                    is_verif = bool(data.get("verifikasi")) if "verifikasi" in data else bool(data.get("status_wdone", True))
                    inst_desc = str(row[2].value or '').strip()
                    inst_kks = str(row[3].value or '').strip()
                    inst_no = str(row[0].value or '').strip()
                    sync_instrument_to_subtasks(wb, inst_type, inst_kks, inst_no, inst_desc, is_verif, today_str)
                    break
        ok, err = safe_save_workbook(wb, path)
        if not ok: return {"status": "error", "message": err}
    return {"status": "success", "message": "Instrument updated successfully."}

def sanitize_folder_name(name):
    clean = "".join(c if c.isalnum() or c in ['-', '_'] else '_' for c in str(name).strip())
    return clean or "ITEM"

def handle_get_finding_data(eq_id):
    folder_name = sanitize_folder_name(eq_id)
    finding_dir = os.path.join(BASE_DIR, "Finding", folder_name)
    res = {"photos": [], "description": "", "temuan": "", "tindak_lanjut": "", "folder": folder_name}
    if os.path.exists(finding_dir):
        desc_path = os.path.join(finding_dir, "deskripsi.txt")
        if os.path.exists(desc_path):
            with open(desc_path, "r", encoding="utf-8") as f:
                content = f.read()
                res["description"] = content
                if "TEMUAN:" in content and "TINDAK LANJUT:" in content:
                    parts = content.split("TINDAK LANJUT:")
                    res["temuan"] = parts[0].replace("TEMUAN:", "").strip()
                    res["tindak_lanjut"] = parts[1].strip() if len(parts) > 1 else ""
                else:
                    res["temuan"] = content
        photos = [f for f in os.listdir(finding_dir) if f.lower().endswith(('.jpg', '.jpeg', '.png', '.webp'))]
        res["photos"] = [{"filename": p, "url": f"/findings_media/{folder_name}/{p}"} for p in sorted(photos)]
    return res

def handle_finding_photo_save(data):
    eq_id = data.get("id")
    unit = data.get("unit", 1)
    item_type = data.get("type", "wo")
    inst_type = data.get("inst_type", "pressure_tx")
    image_base64 = data.get("image_base64")
    photo_source_path = data.get("photo_path")
    temuan_text = data.get("temuan", "")
    tindak_lanjut_text = data.get("tindak_lanjut", "")

    folder_name = sanitize_folder_name(eq_id)
    finding_dir = os.path.join(BASE_DIR, "Finding", folder_name)
    os.makedirs(finding_dir, exist_ok=True)

    desc_file = os.path.join(finding_dir, "deskripsi.txt")
    with open(desc_file, "w", encoding="utf-8") as f:
        f.write(f"TEMUAN:\n{temuan_text}\n\nTINDAK LANJUT:\n{tindak_lanjut_text}\n")

    if image_base64:
        try:
            if "," in image_base64:
                header, encoded = image_base64.split(",", 1)
            else:
                encoded = image_base64
            img_data = base64.b64decode(encoded)
            existing_photos = [f for f in os.listdir(finding_dir) if f.lower().endswith(('.jpg', '.jpeg', '.png', '.webp'))]
            ext = ".jpg"
            if "png" in str(data.get("filename", "")).lower() or (image_base64 and "image/png" in image_base64):
                ext = ".png"
            dest_filename = f"foto_{len(existing_photos) + 1}{ext}"
            with open(os.path.join(finding_dir, dest_filename), "wb") as img_file:
                img_file.write(img_data)
        except Exception as e:
            return {"status": "error", "message": f"Failed to process image: {str(e)}"}

    elif photo_source_path and os.path.exists(photo_source_path):
        ext = os.path.splitext(photo_source_path)[1] or ".jpg"
        existing_photos = [f for f in os.listdir(finding_dir) if f.lower().endswith(('.jpg', '.jpeg', '.png', '.webp'))]
        dest_filename = f"foto_{len(existing_photos) + 1}{ext}"
        shutil.copy(photo_source_path, os.path.join(finding_dir, dest_filename))

    photo_files = [f for f in os.listdir(finding_dir) if f.lower().endswith(('.jpg', '.jpeg', '.png', '.webp'))]
    num_photos = len(photo_files)

    if item_type == "wo":
        save_wo_update({"unit": unit, "no_wo": eq_id, "temuan": temuan_text, "tindak_lanjut": tindak_lanjut_text, "jumlah_foto": num_photos})
    elif item_type == "actuator":
        save_actuator_update({"unit": unit, "equipment_id": eq_id, "temuan": temuan_text, "tindak_lanjut": tindak_lanjut_text, "jumlah_foto": num_photos})
    elif item_type == "instrument":
        save_instrument_update({"unit": unit, "type": inst_type, "no": eq_id, "kks": eq_id, "temuan": temuan_text, "tindak_lanjut": tindak_lanjut_text, "jumlah_foto": num_photos})

    return {
        "status": "success",
        "message": "Photo and field findings saved successfully!",
        "jumlah_foto": num_photos,
        "folder": folder_name,
        "photos": [{"filename": p, "url": f"/findings_media/{folder_name}/{p}"} for p in sorted(photo_files)]
    }

def handle_finding_photo_delete(data):
    eq_id = data.get("id")
    unit = data.get("unit", 1)
    filename = data.get("filename")
    item_type = data.get("type", "wo")
    inst_type = data.get("inst_type", "pressure_tx")

    folder_name = sanitize_folder_name(eq_id)
    finding_dir = os.path.join(BASE_DIR, "Finding", folder_name)

    if os.path.exists(finding_dir) and filename:
        target_file = os.path.join(finding_dir, filename)
        if os.path.exists(target_file):
            os.remove(target_file)

    photo_files = [f for f in os.listdir(finding_dir) if f.lower().endswith(('.jpg', '.jpeg', '.png', '.webp'))]
    num_photos = len(photo_files)

    if item_type == "wo":
        save_wo_update({"unit": unit, "no_wo": eq_id, "jumlah_foto": num_photos})
    elif item_type == "actuator":
        save_actuator_update({"unit": unit, "equipment_id": eq_id, "jumlah_foto": num_photos})
    elif item_type == "instrument":
        save_instrument_update({"unit": unit, "type": inst_type, "no": eq_id, "kks": eq_id, "jumlah_foto": num_photos})

    return {
        "status": "success",
        "message": "Photo deleted successfully.",
        "jumlah_foto": num_photos,
        "photos": [{"filename": p, "url": f"/findings_media/{folder_name}/{p}"} for p in sorted(photo_files)]
    }

class EICMonitoringHandler(http.server.SimpleHTTPRequestHandler):
    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path
        query = urllib.parse.parse_qs(parsed.query)
        
        if path == "/":
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
            self.send_header("Pragma", "no-cache")
            self.send_header("Expires", "0")
            self.end_headers()
            self.wfile.write(HTML_TEMPLATE.encode('utf-8'))
            return
            
        elif path == "/api/data":
            unit = int(query.get("unit", [1])[0])
            data = load_unit_data(unit)
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
            self.end_headers()
            self.wfile.write(json.dumps(data, ensure_ascii=False).encode('utf-8'))
            return

        elif path == "/api/actuator_matrix":
            matrix = load_actuator_matrix()
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
            self.end_headers()
            self.wfile.write(json.dumps(matrix, ensure_ascii=False).encode('utf-8'))
            return

        elif path == "/api/master_components":
            unit = int(query.get("unit", [1])[0])
            data = load_master_components(unit)
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
            self.end_headers()
            self.wfile.write(json.dumps(data, ensure_ascii=False).encode('utf-8'))
            return

        elif path == "/api/findings":
            eq_id = query.get("id", [""])[0]
            res = handle_get_finding_data(eq_id)
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.end_headers()
            self.wfile.write(json.dumps(res, ensure_ascii=False).encode('utf-8'))
            return
            
        elif path.startswith("/findings_media/"):
            rel_path = path.replace("/findings_media/", "")
            full_path = os.path.join(BASE_DIR, "Finding", rel_path)
            if os.path.exists(full_path) and os.path.isfile(full_path):
                mime, _ = mimetypes.guess_type(full_path)
                self.send_response(200)
                self.send_header("Content-Type", mime or "image/jpeg")
                self.end_headers()
                with open(full_path, "rb") as f:
                    self.wfile.write(f.read())
                return

        elif parsed.path == "/api/export_excel":
            unit = 1
            qs = urllib.parse.parse_qs(parsed.query)
            if "unit" in qs:
                try: unit = int(qs["unit"][0])
                except: unit = 1
            path = get_excel_path(unit)
            if not os.path.exists(path):
                self.send_response(404)
                self.end_headers()
                return
            with FILE_LOCK:
                with open(path, "rb") as f:
                    file_bytes = f.read()
            self.send_response(200)
            today_fn = datetime.date.today().strftime("%Y%m%d")
            filename = f"Outage_EIC_Monitoring_Report_Unit_{unit}_{today_fn}.xlsx"
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
            self.send_header("Content-Length", str(len(file_bytes)))
            self.end_headers()
            self.wfile.write(file_bytes)
            return
                
        return super().do_GET()

    def do_POST(self):
        parsed = urllib.parse.urlparse(self.path)
        content_length = int(self.headers.get('Content-Length', 0))
        post_data = self.rfile.read(content_length) if content_length > 0 else b'{}'
        
        try:
            body = json.loads(post_data.decode('utf-8'))
        except Exception:
            body = {}

        routes = {
            "/api/update_wo": save_wo_update,
            "/api/quick_toggle_subtask": save_quick_subtask_toggle,
            "/api/batch_toggle_subtasks": save_batch_subtask_toggle,
            "/api/update_actuator": save_actuator_update,
            "/api/quick_toggle_actuator": save_quick_actuator_toggle,
            "/api/update_instrument": save_instrument_update,
            "/api/quick_toggle_instrument": save_quick_instrument_toggle,
            "/api/bulk_update_instruments": save_bulk_instrument_update,
            "/api/upload_finding_photo": handle_finding_photo_save,
            "/api/delete_finding_photo": handle_finding_photo_delete,
            "/api/add_pic": save_add_pic,
            "/api/delete_pic": save_delete_pic,
            "/api/update_scope": save_scope_update,
            "/api/delete_scope": save_delete_scope,
            "/api/add_scope": save_add_scope,
            "/api/add_wo": save_add_wo,
            "/api/delete_wo": save_delete_wo,
            "/api/add_subtask": save_add_subtask,
            "/api/delete_subtask": save_delete_subtask,
            "/api/add_actuator": save_add_actuator,
            "/api/delete_actuator": save_delete_actuator,
            "/api/add_instrument": save_add_instrument,
            "/api/delete_instrument": save_delete_instrument
        }

        if parsed.path in routes:
            try:
                res = routes[parsed.path](body)
            except Exception as e:
                import traceback
                traceback.print_exc()
                res = {"status": "error", "message": f"Server Error: {str(e)}"}
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.end_headers()
            self.wfile.write(json.dumps(res, ensure_ascii=False).encode('utf-8'))
            return

        self.send_response(404)
        self.end_headers()

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="id">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no, viewport-fit=cover">
    <title>Outage EIC Work Order Monitoring System - PLTU MSW</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=JetBrains+Mono:wght@400;600&display=swap" rel="stylesheet">
    <script>
        // Instant Theme Init to prevent screen flicker (FOUC)
        (function() {
            const saved = localStorage.getItem('eic_theme') || 'dark';
            document.documentElement.setAttribute('data-theme', saved);
        })();
    </script>
    <style>
        :root {
            --safe-top: env(safe-area-inset-top, 0px);
            --safe-bottom: env(safe-area-inset-bottom, 0px);
            --safe-left: env(safe-area-inset-left, 0px);
            --safe-right: env(safe-area-inset-right, 0px);
            --bg-body: #090d16;
            --bg-card: #131d31;
            --bg-card-hover: #1a2844;
            --bg-glass: rgba(19, 29, 49, 0.85);
            --bg-sub: #0b1120;
            --bg-body-dark: #0d1527;
            --bg-header: linear-gradient(180deg, #131d31 0%, rgba(9, 13, 22, 0.95) 100%);
            --bg-banner: linear-gradient(135deg, #16223b 0%, #0d1527 100%);
            --border-color: #22334f;
            --border-highlight: #38bdf8;
            --primary: #38bdf8;
            --primary-hover: #0284c7;
            --primary-glow: rgba(56, 189, 248, 0.25);
            --accent: #818cf8;
            --text-main: #f8fafc;
            --text-muted: #94a3b8;
            --text-sub: #64748b;
            --status-finish: #10b981;
            --status-inprog: #f59e0b;
            --status-sched: #64748b;
            --status-alert: #f43f5e;
            --shadow-card: 0 10px 30px -10px rgba(0,0,0,0.5);
            --shadow-glow: 0 0 20px rgba(56, 189, 248, 0.2);
            --input-bg: #0b1120;
            --stat-hover-bg: #111a2e;
            --date-badge-bg: #1e293b;
            --radius-sm: 6px;
            --radius-md: 10px;
            --radius-lg: 14px;
        }

        [data-theme="light"] {
            --bg-body: #f8fafc;
            --bg-card: #ffffff;
            --bg-card-hover: #f1f5f9;
            --bg-glass: rgba(255, 255, 255, 0.92);
            --bg-sub: #f1f5f9;
            --bg-body-dark: #f8fafc;
            --bg-header: linear-gradient(180deg, #ffffff 0%, rgba(248, 250, 252, 0.98) 100%);
            --bg-banner: linear-gradient(135deg, #e0f2fe 0%, #f0fdf4 100%);
            --border-color: #cbd5e1;
            --border-highlight: #0284c7;
            --primary: #0284c7;
            --primary-hover: #0369a1;
            --primary-glow: rgba(2, 132, 199, 0.2);
            --accent: #4f46e5;
            --text-main: #0f172a;
            --text-muted: #475569;
            --text-sub: #64748b;
            --status-finish: #059669;
            --status-inprog: #d97706;
            --status-sched: #64748b;
            --status-alert: #e11d48;
            --shadow-card: 0 4px 20px -4px rgba(0,0,0,0.08);
            --shadow-glow: 0 0 15px rgba(2, 132, 199, 0.15);
            --input-bg: #ffffff;
            --stat-hover-bg: #f1f5f9;
            --date-badge-bg: #e2e8f0;
        }

        * { box-sizing: border-box; margin: 0; padding: 0; font-family: 'Inter', sans-serif; }
        body { background-color: var(--bg-body); color: var(--text-main); min-height: 100vh; padding-bottom: 70px; transition: background-color 0.25s, color 0.25s; }

        /* Custom Scrollbar */
        ::-webkit-scrollbar { width: 8px; height: 8px; }
        ::-webkit-scrollbar-track { background: var(--bg-body); }
        ::-webkit-scrollbar-thumb { background: var(--border-color); border-radius: 4px; }
        ::-webkit-scrollbar-thumb:hover { background: var(--primary); }

        /* Header */
        header {
            background: var(--bg-header);
            border-bottom: 1px solid var(--border-color);
            padding: 14px 32px;
            display: grid;
            grid-template-columns: 1fr auto 1fr;
            align-items: center;
            gap: 16px;
            position: sticky; top: 0; z-index: 100; backdrop-filter: blur(12px);
            transition: background 0.25s, border-color 0.25s;
        }
        .logo-area { display: flex; align-items: center; justify-self: start; }
        .logo-badge {
            background: linear-gradient(135deg, #38bdf8 0%, #2563eb 100%);
            padding: 8px 14px; border-radius: var(--radius-md); font-weight: 800; color: #fff;
            font-size: 0.92rem; letter-spacing: 0.5px; box-shadow: var(--shadow-glow);
            display: flex; align-items: center; gap: 8px; white-space: nowrap;
            cursor: pointer; user-select: none; transition: transform 0.18s, box-shadow 0.18s, opacity 0.18s;
        }
        .logo-badge:hover {
            transform: translateY(-1px) scale(1.02);
            box-shadow: 0 4px 16px rgba(56, 189, 248, 0.4);
            opacity: 0.95;
        }
        .logo-badge:active {
            transform: scale(0.97);
        }
        .title-group { text-align: center; justify-self: center; }
        .title-group h1 { font-size: 1.25rem; font-weight: 800; color: var(--text-main); letter-spacing: -0.3px; margin: 0; }
        .title-group p { font-size: 0.8rem; color: var(--text-muted); font-weight: 500; margin: 2px 0 0 0; }

        .header-controls { display: flex; align-items: center; gap: 10px; justify-self: end; }
        .unit-switcher { display: flex; background: var(--bg-sub); border-radius: var(--radius-md); padding: 4px; border: 1px solid var(--border-color); }
        .unit-btn { padding: 8px 22px; border: none; background: transparent; color: var(--text-muted); font-weight: 700; border-radius: var(--radius-sm); cursor: pointer; transition: all 0.2s; font-size: 0.88rem; }
        .unit-btn.active { background: var(--primary); color: #fff; box-shadow: 0 2px 12px var(--primary-glow); }
        
        .theme-toggle-btn {
            padding: 8px 12px;
            background: var(--bg-sub);
            border: 1px solid var(--border-color);
            color: var(--text-main);
            border-radius: var(--radius-md);
            font-weight: 700;
            font-size: 1rem;
            cursor: pointer;
            transition: all 0.2s;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            user-select: none;
        }
        .theme-toggle-btn:hover {
            border-color: var(--primary);
            color: var(--primary);
            box-shadow: 0 2px 10px var(--primary-glow);
        }

        .btn-print {
            padding: 8px 16px; background: rgba(56, 189, 248, 0.1); border: 1px solid rgba(56, 189, 248, 0.3);
            color: var(--primary); border-radius: var(--radius-md); font-weight: 600; font-size: 0.85rem;
            cursor: pointer; transition: all 0.2s; display: flex; align-items: center; gap: 6px;
        }
        .btn-print:hover { background: var(--primary); color: #fff; }

        .container { max-width: 1440px; margin: 24px auto; padding: 0 24px; }

        /* Outage D-Day & Progress Banner */
        .outage-banner {
            background: var(--bg-banner);
            border: 1px solid var(--border-color);
            border-radius: var(--radius-lg);
            padding: 18px 24px;
            margin-bottom: 24px;
            display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 16px;
            box-shadow: var(--shadow-card);
            transition: background 0.25s, border-color 0.25s;
        }
        .outage-info { display: flex; align-items: center; gap: 16px; }
        .outage-badge { background: rgba(16, 185, 129, 0.15); border: 1px solid rgba(16, 185, 129, 0.3); color: var(--status-finish); font-weight: 700; font-size: 0.82rem; padding: 6px 14px; border-radius: 20px; text-transform: uppercase; letter-spacing: 0.5px; }
        .outage-title { font-size: 1.1rem; font-weight: 700; color: var(--text-main); }
        .outage-progress-box { display: flex; align-items: center; gap: 16px; min-width: 320px; }
        .outage-pct-huge { font-size: 1.9rem; font-weight: 800; color: var(--primary); font-family: 'JetBrains Mono', monospace; }
        .outage-bar-wrap { flex-grow: 1; }
        .outage-bar-bg { height: 12px; background: var(--bg-sub); border-radius: 6px; overflow: hidden; border: 1px solid var(--border-color); }
        .outage-bar-fill { height: 100%; background: linear-gradient(90deg, #38bdf8 0%, #10b981 100%); border-radius: 6px; transition: width 0.6s cubic-bezier(0.4, 0, 0.2, 1); }

        /* 2-Column Dashboard Main Layout */
        .dashboard-main-layout {
            display: grid;
            grid-template-columns: 290px 1fr;
            gap: 24px;
            align-items: start;
        }

        .sidebar-stats {
            position: sticky;
            top: 20px;
            display: flex;
            flex-direction: column;
            gap: 14px;
        }

        .sidebar-section-title {
            font-size: 0.85rem;
            font-weight: 800;
            color: var(--primary);
            text-transform: uppercase;
            letter-spacing: 0.5px;
            padding-bottom: 6px;
            border-bottom: 1px solid var(--border-color);
            display: flex;
            align-items: center;
            gap: 6px;
        }

        .stats-vertical-list {
            display: flex;
            flex-direction: column;
            gap: 12px;
        }

        .stat-card {
            background: var(--bg-card); border-radius: var(--radius-md); padding: 16px 18px;
            border: 1px solid var(--border-color); box-shadow: var(--shadow-card); position: relative; overflow: hidden;
            transition: transform 0.2s, border-color 0.2s, background 0.2s; cursor: pointer;
        }
        .stat-card:hover { transform: translateY(-2px); border-color: var(--primary); background: var(--stat-hover-bg); }
        .stat-card::before { content: ''; position: absolute; top: 0; left: 0; width: 4px; height: 100%; background: var(--primary); }
        .stat-card.finish::before { background: var(--status-finish); }
        .stat-card.inprog::before { background: var(--status-inprog); }
        .stat-card.findings::before { background: var(--status-alert); }
        .stat-title { font-size: 0.8rem; color: var(--text-muted); font-weight: 700; text-transform: uppercase; letter-spacing: 0.5px; }
        .stat-value { font-size: 1.7rem; font-weight: 800; margin: 6px 0 2px 0; color: var(--text-main); font-family: 'JetBrains Mono', monospace; display: flex; align-items: baseline; justify-content: space-between; }
        .stat-sub { font-size: 0.76rem; color: var(--text-muted); font-weight: 500; }

        .main-content-panel {
            min-width: 0;
            display: flex;
            flex-direction: column;
        }

        @media (max-width: 1024px) {
            .dashboard-main-layout { grid-template-columns: 1fr; }
            .sidebar-stats { position: static; }
            .stats-vertical-list { display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); }
        }

        /* Nav Tabs & Toolbars */
        .nav-tabs { display: flex; gap: 8px; border-bottom: 1px solid var(--border-color); margin-bottom: 20px; overflow-x: auto; }
        .tab-btn {
            padding: 12px 24px; border: none; background: transparent; color: var(--text-muted);
            font-weight: 700; cursor: pointer; border-bottom: 3px solid transparent; transition: all 0.2s;
            white-space: nowrap; font-size: 0.92rem; display: flex; align-items: center; gap: 8px;
        }
        .tab-btn:hover { color: var(--text-main); }
        .tab-btn.active { color: var(--primary); border-bottom-color: var(--primary); }
        .tab-count { background: var(--bg-sub); padding: 2px 8px; border-radius: 12px; font-size: 0.75rem; border: 1px solid var(--border-color); }

        /* Filter Toolbar */
        .filter-toolbar {
            background: var(--bg-card); padding: 16px; border-radius: var(--radius-lg); border: 1px solid var(--border-color);
            margin-bottom: 20px; display: flex; flex-direction: column; gap: 14px;
        }
        .filter-top-row { display: flex; flex-wrap: wrap; gap: 10px; align-items: center; }
        .filter-input { background: var(--input-bg); border: 1px solid var(--border-color); color: var(--text-main); padding: 9px 14px; border-radius: var(--radius-sm); font-size: 0.88rem; outline: none; transition: border-color 0.2s; }
        .filter-input:focus { border-color: var(--primary); }
        .filter-input.search { flex-grow: 1; min-width: 220px; }
        input[type="date"].filter-input { font-family: 'Inter', sans-serif; cursor: pointer; }
        
        .filter-pills { display: flex; flex-wrap: wrap; gap: 8px; align-items: center; }
        .pill-btn {
            padding: 6px 14px; border-radius: 20px; font-size: 0.8rem; font-weight: 600;
            background: var(--bg-sub); border: 1px solid var(--border-color); color: var(--text-muted);
            cursor: pointer; transition: all 0.2s; display: flex; align-items: center; gap: 6px;
        }
        .pill-btn:hover { color: var(--text-main); border-color: var(--primary); }
        .pill-btn.active { background: rgba(56, 189, 248, 0.15); color: var(--primary); border-color: var(--primary); font-weight: 700; }

        .view-switcher { display: flex; background: var(--bg-sub); border: 1px solid var(--border-color); border-radius: var(--radius-sm); margin-left: auto; }
        .view-btn { padding: 6px 14px; border: none; background: transparent; color: var(--text-muted); font-size: 0.82rem; font-weight: 600; cursor: pointer; border-radius: var(--radius-sm); }
        .view-btn.active { background: var(--primary); color: #fff; font-weight: 700; }

        /* Pagination */
        .pagination-bar {
            display: flex; justify-content: space-between; align-items: center; background: var(--bg-card);
            padding: 12px 20px; border-radius: var(--radius-md); border: 1px solid var(--border-color); margin: 16px 0; font-size: 0.85rem;
        }
        .page-btn { background: var(--bg-sub); border: 1px solid var(--border-color); color: var(--text-main); padding: 6px 14px; border-radius: var(--radius-sm); cursor: pointer; font-weight: 600; transition: all 0.2s; }
        .page-btn:disabled { opacity: 0.35; cursor: not-allowed; }
        .page-btn:hover:not(:disabled) { border-color: var(--primary); color: var(--primary); }

        /* Card List Styles */
        .card-list { display: flex; flex-direction: column; gap: 14px; }
        .item-card { background: var(--bg-card); border-radius: var(--radius-md); border: 1px solid var(--border-color); transition: border-color 0.2s, box-shadow 0.2s; overflow: hidden; }
        .item-card:hover { border-color: var(--border-highlight); }
        
        .item-header { padding: 16px 20px; display: flex; align-items: center; justify-content: space-between; cursor: pointer; user-select: none; gap: 16px; }
        .item-header:hover { background: var(--bg-card-hover); }
        .item-title-box { display: flex; flex-direction: column; gap: 4px; flex-grow: 1; }
        .item-code { font-size: 0.8rem; font-weight: 700; color: var(--primary); letter-spacing: 0.5px; font-family: 'JetBrains Mono', monospace; }
        .item-name { font-size: 0.98rem; font-weight: 600; color: var(--text-main); line-height: 1.4; }
        
        .header-actions { display: flex; align-items: center; gap: 12px; }

        /* Status Badges */
        .status-badge { padding: 5px 12px; border-radius: 20px; font-size: 0.75rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.5px; }
        .badge-FINISH { background: rgba(16, 185, 129, 0.15); color: var(--status-finish); border: 1px solid rgba(16, 185, 129, 0.3); }
        .badge-INPROGRESS, .badge-IN_PROGRESS, .badge-IN-PROGRESS { background: rgba(245, 158, 11, 0.15); color: var(--status-inprog); border: 1px solid rgba(245, 158, 11, 0.3); }
        .badge-SCHED-OK, .badge-SCHED_OK { background: rgba(100, 116, 139, 0.15); color: var(--status-sched); border: 1px solid rgba(100, 116, 139, 0.3); }
        .badge-findings { background: rgba(244, 63, 94, 0.15); color: var(--status-alert); border: 1px solid rgba(244, 63, 94, 0.3); }

        .progress-box { display: flex; align-items: center; gap: 8px; min-width: 140px; justify-content: flex-end; }
        .progress-bar-bg { width: 85px; height: 8px; background: var(--bg-sub); border-radius: 4px; overflow: hidden; display: inline-block; border: 1px solid var(--border-color); }
        .progress-bar-fill { height: 100%; background: linear-gradient(90deg, #38bdf8, #10b981); border-radius: 4px; transition: width 0.3s; }

        /* Card Body & Collapsible Forms */
        .item-body { padding: 22px; border-top: 1px solid var(--border-color); background: var(--bg-body-dark); display: none; }
        .item-body.open { display: block; animation: fadeIn 0.25s ease-in-out; }
        .accordion-form { display: none; margin-top: 15px; border-top: 1px solid var(--border-color); padding-top: 15px; }
        .accordion-form.open { display: block !important; animation: fadeIn 0.25s ease-in-out; }
        @keyframes fadeIn { from { opacity: 0; transform: translateY(-4px); } to { opacity: 1; transform: translateY(0); } }

        .section-h4 { font-size: 0.88rem; font-weight: 700; color: var(--primary); margin-bottom: 12px; display: flex; align-items: center; gap: 8px; }
        
        /* Checklist Grid & Subtask Cards */
        .checklist-grid {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(310px, 1fr));
            gap: 12px;
            margin-bottom: 18px;
        }
        .checklist-item {
            background: var(--bg-sub);
            padding: 12px 14px;
            border-radius: var(--radius-sm);
            border: 1px solid var(--border-color);
            display: flex;
            flex-direction: column;
            justify-content: space-between;
            gap: 10px;
            font-size: 0.86rem;
            transition: all 0.2s cubic-bezier(0.4, 0, 0.2, 1);
            position: relative;
            min-height: 85px;
        }
        .checklist-item:hover {
            border-color: rgba(56, 189, 248, 0.45);
            background: var(--bg-card-hover);
            box-shadow: 0 4px 12px rgba(0, 0, 0, 0.2);
        }
        .checklist-item.done {
            background: rgba(16, 185, 129, 0.06);
            border-color: rgba(16, 185, 129, 0.35);
        }
        .checklist-item-body {
            display: flex;
            align-items: flex-start;
            gap: 10px;
            cursor: pointer;
            width: 100%;
            margin: 0;
            user-select: none;
        }
        .checklist-item-body input[type="checkbox"] {
            width: 18px;
            height: 18px;
            margin-top: 2px;
            flex-shrink: 0;
            accent-color: var(--status-finish);
            cursor: pointer;
        }
        .checklist-item-body span {
            word-break: break-word;
            font-size: 0.86rem;
            font-weight: 600;
            color: var(--text-main);
            line-height: 1.45;
            flex-grow: 1;
            transition: color 0.2s, text-decoration 0.2s;
        }
        .checklist-item.done .checklist-item-body span {
            text-decoration: line-through;
            color: var(--text-muted);
            opacity: 0.65;
        }
        .checklist-item-footer {
            display: flex;
            justify-content: space-between;
            align-items: center;
            width: 100%;
            padding-top: 6px;
            border-top: 1px solid rgba(255, 255, 255, 0.05);
            margin-top: auto;
            min-height: 24px;
        }
        .checklist-footer-left {
            display: flex;
            align-items: center;
            gap: 6px;
            flex-wrap: wrap;
        }
        .checklist-footer-right {
            display: flex;
            align-items: center;
            gap: 8px;
            margin-left: auto;
        }

        /* Dense Table View */
        .table-wrap { overflow-x: auto; background: var(--bg-card); border-radius: var(--radius-md); border: 1px solid var(--border-color); }
        .dense-table { width: 100%; border-collapse: collapse; font-size: 0.86rem; text-align: left; }
        .dense-table th { background: var(--bg-sub); padding: 12px 14px; color: var(--text-muted); font-weight: 700; border-bottom: 1px solid var(--border-color); white-space: nowrap; }
        .dense-table td { padding: 12px 14px; border-bottom: 1px solid var(--border-color); vertical-align: middle; }
        .dense-table tr:hover td { background: var(--bg-card-hover); }

        /* Form Controls */
        .form-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 14px; margin-bottom: 18px; }
        .form-group { display: flex; flex-direction: column; gap: 6px; }
        .form-group label { font-size: 0.78rem; font-weight: 600; color: var(--text-muted); }
        .textarea-full { width: 100%; background: var(--input-bg); border: 1px solid var(--border-color); color: var(--text-main); padding: 10px; border-radius: var(--radius-sm); font-size: 0.88rem; resize: vertical; min-height: 65px; outline: none; }
        .textarea-full:focus { border-color: var(--primary); }

        /* Buttons */
        .btn-save {
            background: var(--primary); color: #fff; border: none; padding: 9px 20px; font-weight: 700;
            border-radius: var(--radius-sm); cursor: pointer; transition: all 0.2s; display: inline-flex; align-items: center; gap: 8px; font-size: 0.85rem;
        }
        .btn-save:hover { background: var(--primary-hover); color: #fff; box-shadow: 0 2px 10px var(--primary-glow); }
        .btn-edit-toggle {
            padding: 7px 14px; border-radius: var(--radius-sm); font-size: 0.82rem; font-weight: 700;
            background: rgba(56, 189, 248, 0.12); color: var(--primary); border: 1px solid rgba(56, 189, 248, 0.3); cursor: pointer; transition: all 0.2s;
        }
        .btn-edit-toggle.active { background: var(--primary); color: #fff; }
        .btn-finding {
            padding: 6px 14px; border-radius: 20px; font-size: 0.8rem; font-weight: 700;
            background: rgba(244, 63, 94, 0.15); color: var(--status-alert); border: 1px solid rgba(244, 63, 94, 0.3);
            cursor: pointer; transition: all 0.2s; display: inline-flex; align-items: center; gap: 6px;
        }
        .btn-finding:hover { background: #f43f5e; color: #fff; }
        .btn-danger { background: rgba(244, 63, 94, 0.15); color: var(--status-alert); border: 1px solid rgba(244, 63, 94, 0.3); padding: 7px 14px; font-weight: 700; border-radius: var(--radius-sm); cursor: pointer; font-size: 0.82rem; }

        /* Calibration Box */
        .calib-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 14px; background: var(--bg-sub); border-radius: var(--radius-md); padding: 14px; border: 1px solid var(--border-color); margin-bottom: 16px; }
        .calib-col h5 { font-size: 0.82rem; font-weight: 700; color: var(--accent); margin-bottom: 8px; text-transform: uppercase; }
        .calib-fields { display: grid; grid-template-columns: 1fr 1fr; gap: 8px; }

        /* Toast Notifications */
        #toast-container { position: fixed; bottom: 24px; right: 24px; z-index: 10000; display: flex; flex-direction: column; gap: 10px; pointer-events: none; }
        .toast {
            pointer-events: auto; min-width: 280px; max-width: 420px; background: var(--bg-card); border: 1px solid var(--border-color);
            padding: 12px 18px; border-radius: var(--radius-md); color: var(--text-main); font-size: 0.88rem; font-weight: 600;
            box-shadow: var(--shadow-card); display: flex; align-items: center; gap: 12px;
            animation: toastSlideIn 0.3s cubic-bezier(0.16, 1, 0.3, 1);
        }
        @keyframes toastSlideIn { from { transform: translateY(30px); opacity: 0; } to { transform: translateY(0); opacity: 1; } }
        .toast.success { border-left: 4px solid var(--status-finish); }
        .toast.error { border-left: 4px solid var(--status-alert); }
        .toast.info { border-left: 4px solid var(--primary); }

        /* Modal Styles */
        .modal-overlay { position: fixed; inset: 0; background: rgba(0,0,0,0.75); backdrop-filter: blur(6px); z-index: 9999; display: none; justify-content: center; align-items: center; padding: 20px; }
        .modal-overlay.open, .modal-overlay.active { display: flex !important; animation: fadeIn 0.2s ease; }
        .modal-content { background: var(--bg-card); border: 1px solid var(--border-color); border-radius: var(--radius-lg); width: 100%; max-width: 760px; max-height: 90vh; overflow-y: auto; padding: 24px; box-shadow: var(--shadow-card); position: relative; }
        .modal-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px; border-bottom: 1px solid var(--border-color); padding-bottom: 14px; }
        .modal-close { background: none; border: none; color: var(--text-muted); font-size: 1.5rem; cursor: pointer; font-weight: 700; }
        .modal-close:hover { color: var(--text-main); }

        /* Finding Photos Grid in Modal */
        .photo-dropzone { border: 2px dashed var(--border-color); border-radius: var(--radius-md); padding: 20px; text-align: center; background: var(--bg-sub); cursor: pointer; transition: all 0.2s; margin-bottom: 18px; }
        .photo-dropzone:hover, .photo-dropzone.dragover { border-color: var(--primary); background: var(--stat-hover-bg); }
        .photo-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(130px, 1fr)); gap: 12px; margin-bottom: 18px; }
        .photo-thumb-box { position: relative; border-radius: var(--radius-sm); overflow: hidden; border: 1px solid var(--border-color); aspect-ratio: 4/3; background: var(--bg-sub); }
        .photo-thumb-box img { width: 100%; height: 100%; object-fit: cover; cursor: pointer; transition: transform 0.2s; }
        .photo-thumb-box img:hover { transform: scale(1.05); }
        .photo-delete-btn { position: absolute; top: 4px; right: 4px; background: rgba(0,0,0,0.7); border: none; color: #f43f5e; border-radius: 4px; padding: 3px 6px; cursor: pointer; font-size: 0.75rem; }

        /* Lightbox Image Preview */
        #lightbox-modal img { max-width: 90vw; max-height: 85vh; object-fit: contain; border-radius: 8px; }

        /* Report Paper & Print Styles */
        .report-section-title { font-size: 0.95rem; font-weight: 800; color: var(--primary); text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 10px; border-left: 3px solid var(--primary); padding-left: 8px; }
        .report-kpi-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(190px, 1fr)); gap: 10px; margin-bottom: 15px; }
        .report-kpi-card { background: var(--bg-sub); border: 1px solid var(--border-color); border-radius: var(--radius-sm); padding: 12px; }
        .report-kpi-card.highlight { border-color: var(--primary); background: var(--stat-hover-bg); }
        .report-kpi-lbl { font-size: 0.72rem; font-weight: 700; color: var(--text-muted); text-transform: uppercase; }
        .report-kpi-val { font-size: 1.4rem; font-weight: 800; color: var(--primary); margin: 4px 0 2px 0; font-family: 'JetBrains Mono', monospace; }
        .report-kpi-sub { font-size: 0.75rem; color: var(--text-muted); }
        .report-table { width: 100%; border-collapse: collapse; margin-bottom: 10px; font-size: 0.84rem; background: var(--bg-sub); border-radius: var(--radius-sm); overflow: hidden; border: 1px solid var(--border-color); }
        .report-table th { background: var(--border-color); color: var(--text-muted); font-size: 0.75rem; font-weight: 700; text-transform: uppercase; padding: 9px 10px; text-align: left; border-bottom: 1px solid var(--border-color); }
        .report-table td { padding: 9px 10px; border-bottom: 1px solid var(--border-color); vertical-align: middle; }
        .report-badge { display: inline-block; padding: 2px 6px; border-radius: 4px; font-size: 0.7rem; font-weight: 700; background: rgba(56,189,248,0.15); color: var(--primary); }
        .report-badge.alert { background: rgba(244,63,94,0.15); color: #fda4af; }

        @media print {
            body { background: #fff !important; color: #000 !important; font-size: 9.5pt; font-family: Arial, sans-serif !important; }
            header, .container, .outage-banner, .dashboard-main-layout, .sidebar-stats, .stats-vertical-list, .main-content-panel, .nav-tabs, .filter-toolbar, .card-list, .table-wrap, .pagination-bar, #toast-container, .modal-header, .modal-close, .filter-input, .page-btn, .btn-save, .btn-print, #photo-dropzone, #modal-photos-container, .form-grid, #tab-content, .sub-nav, .view-switcher, .filter-pills, .btn-add-wo, #finding-modal, #lightbox-modal, .form-collapsible, .collapsible-btn, .subtask-add-bar, .stats-grid { display: none !important; }
            .modal-overlay { display: none !important; }
            .modal-overlay#report-modal { position: static !important; background: transparent !important; display: block !important; padding: 0 !important; width: 100% !important; height: auto !important; }
            #report-modal .modal-content { max-width: 100% !important; width: 100% !important; border: none !important; background: transparent !important; box-shadow: none !important; padding: 0 !important; max-height: none !important; }
            #report-modal-scroll { overflow: visible !important; }
            .report-paper { color: #000 !important; }
            .report-header-box div, .report-header-box h2, .report-header-box span { color: #000 !important; }
            .report-section-title { color: #000 !important; border-left: 4px solid #000 !important; font-size: 11pt !important; margin-top: 15px !important; margin-bottom: 8px !important; }
            .report-kpi-grid { grid-template-columns: repeat(4, 1fr) !important; gap: 8px !important; margin-bottom: 12px !important; }
            .report-kpi-card { background: #f8fafc !important; border: 1px solid #94a3b8 !important; color: #000 !important; padding: 8px !important; }
            .report-kpi-lbl { color: #475569 !important; font-size: 8pt !important; }
            .report-kpi-val { color: #000 !important; font-size: 14pt !important; }
            .report-kpi-sub { color: #475569 !important; font-size: 7.5pt !important; }
            .report-table { background: #fff !important; border: 1px solid #94a3b8 !important; width: 100% !important; }
            .report-table th { background: #e2e8f0 !important; color: #0f172a !important; border: 1px solid #94a3b8 !important; padding: 6px 8px !important; font-size: 8.5pt !important; }
            .report-table td { border: 1px solid #cbd5e1 !important; color: #0f172a !important; padding: 5px 8px !important; font-size: 8pt !important; }
            .report-badge { background: #e2e8f0 !important; color: #0f172a !important; border: 1px solid #94a3b8 !important; }
            .report-badge.alert { background: #fee2e2 !important; color: #991b1b !important; }
        }

        /* Sticky Summary Bar */
        .sticky-summary-bar {
            position: fixed;
            top: 0;
            left: 0;
            right: 0;
            background: var(--bg-glass);
            backdrop-filter: blur(12px);
            border-bottom: 1px solid var(--border-color);
            padding: 9px 28px;
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 16px;
            z-index: 99;
            transform: translateY(-100%);
            transition: transform 0.25s ease, background 0.25s;
            box-shadow: 0 4px 20px rgba(0,0,0,0.15);
        }
        .sticky-summary-bar.visible {
            transform: translateY(0);
        }
        .sticky-stats-group {
            display: flex;
            align-items: center;
            gap: 10px;
            flex-wrap: wrap;
        }
        .sticky-stat-pill {
            font-size: 0.8rem;
            font-weight: 700;
            background: var(--bg-sub);
            border: 1px solid var(--border-color);
            padding: 4px 10px;
            border-radius: 20px;
            display: inline-flex;
            align-items: center;
            gap: 6px;
            color: var(--text-main);
        }
        .sticky-actions-group {
            display: flex;
            align-items: center;
            gap: 8px;
        }
        .sticky-action-btn {
            padding: 5px 12px;
            font-size: 0.78rem;
            font-weight: 700;
            border-radius: var(--radius-sm);
            border: 1px solid var(--border-color);
            background: var(--bg-sub);
            color: var(--text-main);
            cursor: pointer;
            display: inline-flex;
            align-items: center;
            gap: 5px;
            transition: all 0.2s;
        }
        .sticky-action-btn:hover {
            background: var(--primary);
            color: #fff;
            border-color: var(--primary);
        }

        /* Back to Top Button */
        .back-to-top-btn {
            position: fixed;
            bottom: 24px;
            right: 24px;
            width: 44px;
            height: 44px;
            border-radius: 50%;
            background: var(--primary);
            color: #fff;
            border: none;
            box-shadow: 0 4px 15px rgba(0,0,0,0.3);
            cursor: pointer;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 1.1rem;
            font-weight: 800;
            z-index: 95;
            opacity: 0;
            visibility: hidden;
            transform: translateY(15px);
            transition: opacity 0.25s, transform 0.25s, visibility 0.25s, background 0.2s;
        }
        .back-to-top-btn.visible {
            opacity: 1;
            visibility: visible;
            transform: translateY(0);
        }
        .back-to-top-btn:hover {
            background: var(--primary-hover);
            transform: translateY(-3px);
        }

        /* Batch Actions */
        .btn-batch-check {
            padding: 3px 8px;
            font-size: 0.72rem;
            font-weight: 700;
            border-radius: 4px;
            border: 1px solid var(--status-finish);
            background: rgba(16, 185, 129, 0.12);
            color: var(--status-finish);
            cursor: pointer;
            transition: all 0.15s;
        }
        .btn-batch-check:hover {
            background: var(--status-finish);
            color: #fff;
        }
        .btn-batch-reset {
            padding: 3px 8px;
            font-size: 0.72rem;
            font-weight: 700;
            border-radius: 4px;
            border: 1px solid var(--border-color);
            background: var(--bg-sub);
            color: var(--text-muted);
            cursor: pointer;
            transition: all 0.15s;
        }
        .btn-batch-reset:hover {
            background: var(--status-alert);
            border-color: var(--status-alert);
            color: #fff;
        }

        /* WhatsApp & SCurve Preview Box */
        .wa-box {
            background: var(--bg-sub);
            border: 1px solid var(--border-color);
            border-radius: var(--radius-md);
            padding: 16px;
            font-family: 'JetBrains Mono', monospace;
            font-size: 0.84rem;
            line-height: 1.6;
            white-space: pre-wrap;
            color: var(--text-main);
            max-height: 480px;
            overflow-y: auto;
            border-left: 4px solid #25D366;
        }

        @media (max-width: 900px) {
            header { padding: 12px 20px; }
            .container { padding: 0 16px; margin: 18px auto; }
            .dashboard-main-layout { grid-template-columns: 1fr; gap: 16px; }
            .sidebar-stats { position: static; }
            .stats-vertical-list { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); }
            .outage-banner { flex-direction: column; align-items: flex-start; }
            .outage-progress-box { width: 100%; }
            .sticky-summary-bar { padding: 10px 18px; }
        }

        /* Smartphone Adaptive Styling (6" - 7" Screens & Viewports 360px - 430px) */
        @media (max-width: 640px) {
            body {
                padding-bottom: calc(75px + var(--safe-bottom));
                font-size: 0.88rem;
            }
            .container {
                padding: 0 10px;
                margin: 10px auto;
            }

            /* Compact Mobile Header / App Bar */
            header {
                padding: 10px 12px;
                padding-top: calc(10px + var(--safe-top));
                display: flex;
                flex-direction: row;
                justify-content: space-between;
                align-items: center;
                gap: 8px;
                position: sticky;
                top: 0;
                z-index: 100;
            }
            .logo-area {
                flex-shrink: 0;
            }
            .logo-badge {
                padding: 6px 10px;
                font-size: 0.78rem;
                letter-spacing: 0.2px;
                border-radius: var(--radius-sm);
            }
            .title-group {
                display: none !important; /* Hide long title from top header on mobile to conserve screen space */
            }
            .header-controls {
                display: flex;
                align-items: center;
                gap: 6px;
                width: auto;
                justify-content: flex-end;
            }
            .unit-switcher {
                padding: 2px;
                border-radius: var(--radius-sm);
            }
            .unit-btn {
                padding: 5px 10px;
                font-size: 0.76rem;
            }
            .theme-toggle-btn {
                padding: 5px 8px;
                font-size: 0.88rem;
                border-radius: var(--radius-sm);
            }
            .btn-print {
                padding: 5px 10px;
                font-size: 0.76rem;
                border-radius: var(--radius-sm);
            }

            /* Outage Banner for Mobile */
            .outage-banner {
                padding: 12px 14px;
                margin-bottom: 12px;
                border-radius: var(--radius-md);
                flex-direction: column;
                align-items: stretch;
                gap: 10px;
            }
            .outage-info {
                flex-direction: column;
                align-items: flex-start;
                gap: 6px;
                width: 100%;
            }
            .outage-title {
                font-size: 0.95rem;
                line-height: 1.35;
            }
            .outage-progress-box {
                min-width: 0;
                width: 100%;
                gap: 10px;
            }
            .outage-pct-huge {
                font-size: 1.45rem;
            }

            /* KPI Stats: 2x2 Grid Layout */
            .dashboard-main-layout {
                grid-template-columns: 1fr;
                gap: 12px;
            }
            .sidebar-stats {
                position: static;
                gap: 6px;
            }
            .sidebar-section-title {
                font-size: 0.76rem;
                padding-bottom: 4px;
            }
            .stats-vertical-list {
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 8px;
            }
            .stat-card {
                padding: 10px 12px;
                border-radius: var(--radius-sm);
            }
            .stat-title {
                font-size: 0.68rem;
                letter-spacing: 0.2px;
                white-space: nowrap;
                overflow: hidden;
                text-overflow: ellipsis;
            }
            .stat-value {
                font-size: 1.25rem;
                margin: 4px 0 2px 0;
                flex-direction: column;
                align-items: flex-start;
                gap: 2px;
            }
            .stat-value span:last-child {
                font-size: 0.72rem !important;
            }
            .stat-sub {
                font-size: 0.65rem;
                white-space: nowrap;
                overflow: hidden;
                text-overflow: ellipsis;
            }

            /* Nav Tabs: Touch Scrollable Pills */
            .nav-tabs {
                gap: 6px;
                margin-bottom: 12px;
                overflow-x: auto;
                scrollbar-width: none;
                -webkit-overflow-scrolling: touch;
                padding-bottom: 2px;
                border-bottom: 1px solid var(--border-color);
            }
            .nav-tabs::-webkit-scrollbar { display: none; }
            .tab-btn {
                padding: 8px 14px;
                font-size: 0.8rem;
                flex-shrink: 0;
                gap: 6px;
            }
            .tab-count {
                padding: 1px 6px;
                font-size: 0.68rem;
            }

            /* Filter & Control Toolbar */
            .filter-toolbar {
                padding: 12px;
                border-radius: var(--radius-md);
                margin-bottom: 12px;
                gap: 8px;
            }
            .filter-top-row {
                display: flex;
                flex-direction: column;
                gap: 8px;
                width: 100%;
            }
            .filter-input {
                width: 100% !important;
                min-width: 0 !important;
                font-size: 16px !important; /* Standard 16px prevents iOS Safari auto-zoom */
                padding: 8px 12px;
                border-radius: var(--radius-sm);
            }
            .view-switcher {
                width: 100%;
                margin-left: 0;
                justify-content: center;
            }
            .view-btn {
                flex: 1;
                text-align: center;
                padding: 7px 10px;
            }
            .filter-pills {
                display: flex;
                flex-wrap: nowrap !important;
                overflow-x: auto;
                scrollbar-width: none;
                -webkit-overflow-scrolling: touch;
                gap: 6px;
                padding-bottom: 4px;
                width: 100%;
            }
            .filter-pills::-webkit-scrollbar { display: none; }
            .pill-btn {
                flex-shrink: 0;
                padding: 5px 11px;
                font-size: 0.74rem;
            }
            #side-pagination-counter {
                width: 100%;
                margin-left: 0 !important;
                justify-content: space-between;
                font-size: 0.74rem;
                padding: 4px 8px;
            }

            /* Pagination Bar */
            .pagination-bar {
                flex-direction: column;
                gap: 10px;
                align-items: center;
                text-align: center;
                padding: 10px 12px;
                font-size: 0.78rem;
            }
            .page-btn {
                padding: 6px 10px;
                font-size: 0.76rem;
            }

            /* Work Order, Actuator & Instrument Cards */
            .item-card {
                border-radius: var(--radius-sm);
            }
            .item-header {
                padding: 12px 14px;
                flex-direction: column;
                align-items: flex-start;
                gap: 8px;
            }
            .item-title-box {
                width: 100%;
            }
            .item-code {
                font-size: 0.74rem;
            }
            .item-name {
                font-size: 0.9rem;
                line-height: 1.35;
            }
            .header-actions {
                width: 100%;
                display: flex;
                flex-wrap: wrap;
                justify-content: space-between;
                align-items: center;
                gap: 8px;
                border-top: 1px solid rgba(255,255,255,0.06);
                padding-top: 8px;
            }
            .header-actions span {
                font-size: 0.76rem !important;
            }
            .btn-finding {
                padding: 4px 10px;
                font-size: 0.72rem;
            }
            .progress-box {
                min-width: 0;
                gap: 6px;
            }
            .progress-bar-bg {
                width: 60px;
                height: 7px;
            }
            .item-body {
                padding: 12px 14px;
            }

            /* 1-Column Subtask Checklist with Thumb-Friendly Checkbox */
            .checklist-grid {
                grid-template-columns: 1fr !important;
                gap: 8px;
            }
            .checklist-item {
                padding: 10px 12px;
                min-height: auto;
                gap: 8px;
            }
            .checklist-item-body {
                gap: 10px;
                align-items: center;
            }
            .checklist-item-body input[type="checkbox"] {
                width: 22px;
                height: 22px;
                min-width: 22px;
                min-height: 22px;
                margin-top: 0;
                cursor: pointer;
            }
            .checklist-item-body span {
                font-size: 0.86rem;
                line-height: 1.35;
            }
            .btn-batch-check, .btn-batch-reset {
                padding: 4px 8px;
                font-size: 0.7rem;
            }
            .btn-del-subtask-cross {
                font-size: 1.3rem;
                padding: 2px 6px;
                opacity: 0.8;
            }
            .subtask-add-bar {
                flex-direction: column;
                align-items: stretch;
                gap: 8px;
            }
            .subtask-add-bar input {
                width: 100%;
            }

            /* Forms & Textarea inside Cards & Modals */
            .form-grid {
                grid-template-columns: 1fr !important;
                gap: 10px;
            }
            .textarea-full {
                font-size: 16px !important;
                min-height: 60px;
            }
            .calib-grid {
                grid-template-columns: 1fr !important;
                gap: 10px;
                padding: 10px;
            }
            .calib-fields {
                grid-template-columns: 1fr 1fr;
                gap: 6px;
            }

            /* Fullscreen Mobile Modal Sheets */
            .modal-overlay {
                padding: 0;
                align-items: flex-end;
            }
            .modal-content {
                width: 100% !important;
                max-width: 100% !important;
                height: 100% !important;
                max-height: 100vh !important;
                border-radius: 0 !important;
                padding: 14px !important;
                padding-top: calc(12px + var(--safe-top)) !important;
                padding-bottom: calc(14px + var(--safe-bottom)) !important;
                border: none !important;
                display: flex !important;
                flex-direction: column !important;
            }
            .modal-header {
                margin-bottom: 10px;
                padding-bottom: 8px;
                flex-shrink: 0;
            }
            .modal-header h3 {
                font-size: 1rem !important;
            }
            .modal-close {
                font-size: 1.8rem;
                line-height: 1;
                padding: 2px 8px;
            }
            .photo-dropzone {
                padding: 14px;
                margin-bottom: 10px;
            }
            .photo-grid {
                grid-template-columns: repeat(3, 1fr);
                gap: 6px;
                margin-bottom: 10px;
            }
            .photo-thumb-box {
                aspect-ratio: 1/1;
            }

            /* Report Modal Mobile Styles */
            .report-type-selector {
                overflow-x: auto;
                flex-wrap: nowrap !important;
                scrollbar-width: none;
                -webkit-overflow-scrolling: touch;
                padding-bottom: 4px;
                gap: 6px;
            }
            .report-type-selector::-webkit-scrollbar { display: none; }
            .report-type-selector .unit-btn {
                flex-shrink: 0;
                padding: 6px 12px;
                font-size: 0.74rem;
            }
            #report-date-bar {
                padding: 8px 10px;
                gap: 8px;
                flex-direction: column;
                align-items: stretch;
            }
            .report-kpi-grid {
                grid-template-columns: 1fr 1fr;
                gap: 8px;
            }
            .report-table {
                font-size: 0.74rem;
            }
            .report-table th, .report-table td {
                padding: 6px 8px;
            }

            /* Sticky Summary Bar: Mobile Bottom Position */
            .sticky-summary-bar {
                top: auto;
                bottom: 0;
                left: 0;
                right: 0;
                padding: 8px 12px;
                padding-bottom: calc(8px + var(--safe-bottom));
                border-top: 1px solid var(--border-color);
                border-bottom: none;
                transform: translateY(100%);
                flex-direction: row;
                justify-content: space-between;
                align-items: center;
                gap: 8px;
                box-shadow: 0 -4px 20px rgba(0,0,0,0.3);
            }
            .sticky-summary-bar.visible {
                transform: translateY(0);
            }
            .sticky-stats-group {
                gap: 4px;
                overflow-x: auto;
                scrollbar-width: none;
                -webkit-overflow-scrolling: touch;
                flex-wrap: nowrap;
                max-width: calc(100vw - 115px);
            }
            .sticky-stats-group::-webkit-scrollbar { display: none; }
            .sticky-stat-pill {
                padding: 3px 7px;
                font-size: 0.68rem;
                flex-shrink: 0;
            }
            .sticky-actions-group {
                gap: 4px;
                flex-shrink: 0;
            }
            .sticky-action-btn {
                padding: 4px 8px;
                font-size: 0.72rem;
            }

            /* Floating Back to Top Button */
            .back-to-top-btn {
                bottom: calc(56px + var(--safe-bottom));
                right: 12px;
                width: 38px;
                height: 38px;
                font-size: 0.9rem;
            }

            /* Toast Notification on Mobile */
            #toast-container {
                bottom: calc(60px + var(--safe-bottom));
                right: 12px;
                left: 12px;
                align-items: stretch;
            }
            .toast {
                max-width: 100%;
                min-width: auto;
                font-size: 0.8rem;
                padding: 10px 14px;
            }
        }
    
        /* Component Badges on Subtask Checklist */
        .badge-tag-comp {
            display: inline-flex;
            align-items: center;
            font-size: 0.65rem;
            font-weight: 700;
            padding: 2px 7px;
            border-radius: 4px;
            letter-spacing: 0.04em;
            font-family: 'JetBrains Mono', monospace;
            text-transform: uppercase;
            white-space: nowrap;
        }
        .badge-tag-act {
            background: rgba(245, 158, 11, 0.15);
            color: #fbbf24;
            border: 1px solid rgba(245, 158, 11, 0.35);
        }
        .badge-tag-inst {
            background: rgba(6, 182, 212, 0.15);
            color: #38bdf8;
            border: 1px solid rgba(6, 182, 212, 0.35);
        }
        .badge-tag-elec {
            background: rgba(139, 92, 246, 0.15);
            color: #a78bfa;
            border: 1px solid rgba(139, 92, 246, 0.35);
        }

        /* Mode Selector Buttons for Add Subtask */
        .comp-mode-btn {
            background: var(--bg-card);
            border: 1px solid var(--border-color);
            color: var(--text-muted);
            padding: 5px 12px;
            border-radius: 6px;
            font-size: 0.78rem;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.2s ease;
            white-space: nowrap;
        }
        .comp-mode-btn:hover {
            border-color: var(--primary);
            color: var(--primary);
            background: rgba(99, 102, 241, 0.08);
        }
        .comp-mode-btn.active {
            background: var(--primary);
            color: #090d16;
            border-color: var(--primary);
            font-weight: 700;
            box-shadow: 0 0 8px rgba(99, 102, 241, 0.35);
        }

        /* Clean Checklist Item Layout */
        .checklist-item-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            width: 100%;
            min-height: 20px;
            margin-bottom: 2px;
        }
        .header-left {
            display: flex;
            align-items: center;
            gap: 6px;
            flex-wrap: wrap;
        }
        .header-right {
            display: flex;
            align-items: center;
            gap: 8px;
            margin-left: auto;
        }
        .btn-del-subtask-cross {
            background: transparent;
            border: none;
            color: var(--text-muted);
            font-size: 1.2rem;
            line-height: 1;
            cursor: pointer;
            padding: 0 4px;
            border-radius: 4px;
            transition: all 0.2s;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            opacity: 0.5;
        }
        .checklist-item:hover .btn-del-subtask-cross {
            opacity: 0.9;
        }
        .btn-del-subtask-cross:hover {
            color: #f43f5e !important;
            background: rgba(244, 63, 94, 0.15);
        }

        /* Pressure Switch Set Point & Contact Styling */
        .setpoint-cell {
            display: inline-flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            min-width: 95px;
            padding: 3px 8px;
            background: var(--bg-sub);
            border: 1px solid var(--border-color);
            border-radius: var(--radius-sm);
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        }
        .setpoint-val {
            font-size: 0.84rem;
            font-weight: 700;
            color: var(--text-main);
            font-family: 'JetBrains Mono', monospace;
            line-height: 1.2;
        }
        .setpoint-divider {
            width: 100%;
            height: 1px;
            background: var(--border-color);
            margin: 2px 0;
        }
        .setpoint-dir {
            font-size: 0.68rem;
            font-weight: 800;
            letter-spacing: 0.06em;
            text-transform: uppercase;
            font-family: 'Inter', sans-serif;
            line-height: 1.1;
        }
        .setpoint-dir.high {
            color: #38bdf8;
        }
        .setpoint-dir.low {
            color: #f59e0b;
        }

        .contact-badge {
            display: inline-block;
            padding: 4px 10px;
            border-radius: 6px;
            font-size: 0.78rem;
            font-weight: 800;
            font-family: 'JetBrains Mono', monospace;
            letter-spacing: 0.03em;
            text-align: center;
            min-width: 40px;
        }
        .contact-no {
            background: rgba(16, 185, 129, 0.15);
            color: var(--status-finish);
            border: 1px solid rgba(16, 185, 129, 0.35);
        }
        .contact-nc {
            background: rgba(245, 158, 11, 0.15);
            color: #fbbf24;
            border: 1px solid rgba(245, 158, 11, 0.35);
        }

        /* Small In-Place Refresh Button */
        .btn-refresh-icon {
            display: inline-flex;
            align-items: center;
            justify-content: center;
            gap: 5px;
            background: rgba(99, 102, 241, 0.12);
            color: var(--primary);
            border: 1px solid rgba(99, 102, 241, 0.3);
            border-radius: 6px;
            padding: 3px 8px;
            font-size: 0.76rem;
            cursor: pointer;
            transition: all 0.2s cubic-bezier(0.4, 0, 0.2, 1);
            user-select: none;
            line-height: 1;
            font-weight: 600;
        }
        .btn-refresh-icon:hover {
            background: var(--primary);
            color: #ffffff;
            border-color: var(--primary);
            box-shadow: 0 2px 8px rgba(99, 102, 241, 0.35);
        }
        .btn-refresh-icon:active {
            transform: scale(0.96);
        }
        .btn-refresh-icon.is-loading {
            pointer-events: none;
            opacity: 0.8;
            cursor: wait;
        }
        .btn-refresh-icon .refresh-icon {
            display: inline-block;
            font-size: 0.82rem;
            transform-origin: center center;
        }
        .btn-refresh-icon.is-loading .refresh-icon {
            animation: spin-loading-icon 0.8s linear infinite;
        }
        @keyframes spin-loading-icon {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }

        /* Scope Master Security & Protection Styles */
        .scope-lock-banner {
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding: 14px 20px;
            border-radius: var(--radius-md);
            margin-bottom: 6px;
            gap: 14px;
            flex-wrap: wrap;
            transition: all 0.3s;
        }
        .scope-lock-banner.locked {
            background: rgba(30, 41, 59, 0.7);
            border: 1px solid rgba(148, 163, 184, 0.25);
        }
        .scope-lock-banner.unlocked {
            background: rgba(16, 185, 129, 0.1);
            border: 1px solid rgba(16, 185, 129, 0.35);
        }
        .lock-icon-circle {
            width: 38px;
            height: 38px;
            border-radius: 50%;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 1.2rem;
            background: rgba(255, 255, 255, 0.06);
            border: 1px solid var(--border-color);
            flex-shrink: 0;
        }
        .lock-icon-circle.open {
            background: rgba(16, 185, 129, 0.2);
            border-color: rgba(16, 185, 129, 0.4);
        }
        .scope-type-badge {
            display: inline-block;
            padding: 3px 10px;
            border-radius: 12px;
            font-size: 0.78rem;
            font-weight: 700;
            background: rgba(99, 102, 241, 0.15);
            color: var(--primary);
            border: 1px solid rgba(99, 102, 241, 0.3);
        }
        @keyframes shake {
            0%, 100% { transform: translateX(0); }
            20%, 60% { transform: translateX(-6px); }
            40%, 80% { transform: translateX(6px); }
        }
        .shake-error {
            animation: shake 0.4s ease-in-out;
            border-color: #f43f5e !important;
        }

        .ui-icon {
            display: inline-flex;
            align-items: center;
            justify-content: center;
            vertical-align: -0.125em;
            width: 1em;
            height: 1em;
            stroke-width: 1.8;
            stroke: currentColor;
            fill: none;
            stroke-linecap: round;
            stroke-linejoin: round;
            flex-shrink: 0;
        }
        .ui-icon.sm { width: 14px; height: 14px; }
        .ui-icon.md { width: 16px; height: 16px; }
        .ui-icon.lg { width: 20px; height: 20px; }
        .stat-title-wrap {
            display: inline-flex;
            align-items: center;
            gap: 6px;
        }

    </style>
</head>
<body>
    <!-- Sticky Summary Bar -->
    <div id="sticky-summary-bar" class="sticky-summary-bar">
        <div class="sticky-stats-group">
            <span class="sticky-stat-pill" style="color:var(--primary); font-family:'JetBrains Mono'; font-weight:800;">UNIT <span id="sticky-unit-num">1</span></span>
            <span class="sticky-stat-pill">Total: <strong style="color:var(--primary);" id="sticky-grand-pct">0%</strong></span>
            <span class="sticky-stat-pill">WO: <strong id="sticky-wo-pct">0%</strong></span>
            <span class="sticky-stat-pill">Valves: <strong id="sticky-act-pct">0%</strong></span>
            <span class="sticky-stat-pill">Inst: <strong id="sticky-inst-pct">0%</strong></span>
        </div>
        <div class="sticky-actions-group">
            <button class="sticky-action-btn" onclick="openReportModal()" style="display:inline-flex; align-items:center; gap:5px;">
                <svg class="ui-icon sm" viewBox="0 0 24 24"><polyline points="6 9 6 2 18 2 18 9"></polyline><path d="M6 18H4a2 2 0 0 1-2-2v-5a2 2 0 0 1 2-2h16a2 2 0 0 1 2 2v5a2 2 0 0 1-2 2h-2"></path><rect x="6" y="14" width="12" height="8"></rect></svg> Reports & Export
            </button>
            <button class="sticky-action-btn" onclick="scrollToTop()" style="display:inline-flex; align-items:center; gap:4px;">
                <svg class="ui-icon sm" viewBox="0 0 24 24"><line x1="12" y1="19" x2="12" y2="5"></line><polyline points="5 12 12 5 19 12"></polyline></svg> Top
            </button>
        </div>
    </div>

    <!-- Floating Back to Top Button -->
    <button id="back-to-top-btn" class="back-to-top-btn" onclick="scrollToTop()" title="Back to Top">
        <svg class="ui-icon md" viewBox="0 0 24 24"><line x1="12" y1="19" x2="12" y2="5"></line><polyline points="5 12 12 5 19 12"></polyline></svg>
    </button>
    <!-- Toast Notification Container -->
    <div id="toast-container"></div>

    <!-- Lightbox Modal -->
    <div class="modal-overlay" id="lightbox-modal" onclick="closeLightbox()">
        <img id="lightbox-img" src="" alt="Finding Preview" onclick="event.stopPropagation()">
    </div>

    <!-- Finding & Photo Management Modal -->
    <div class="modal-overlay" id="finding-modal">
        <div class="modal-content">
            <div class="modal-header">
                <div>
                    <h3 style="color:var(--primary); font-size:1.15rem; font-weight:800; display:flex; align-items:center; gap:8px;" id="modal-finding-title">
                        <svg class="ui-icon lg" viewBox="0 0 24 24"><path d="M23 19a2 2 0 0 1-2 2H3a2 2 0 0 1-2-2V8a2 2 0 0 1 2-2h4l2-3h6l2 3h4a2 2 0 0 1 2 2z"></path><circle cx="12" cy="13" r="4"></circle></svg>
                        Field Findings & Photos
                    </h3>
                    <div style="font-size:0.8rem; color:var(--text-muted);" id="modal-finding-subtitle">Equipment Code</div>
                </div>
                <button class="modal-close" onclick="closeFindingModal()">&times;</button>
            </div>

            <!-- Upload Dropzone -->
            <div class="photo-dropzone" id="photo-dropzone">
                <div style="margin-bottom:6px;">
                    <svg class="ui-icon" style="width:34px; height:34px; color:var(--primary); opacity:0.85;" viewBox="0 0 24 24"><path d="M23 19a2 2 0 0 1-2 2H3a2 2 0 0 1-2-2V8a2 2 0 0 1 2-2h4l2-3h6l2 3h4a2 2 0 0 1 2 2z"></path><circle cx="12" cy="13" r="4"></circle></svg>
                </div>
                <strong style="color:var(--primary); font-size:0.9rem;">Upload Field Photo Documentation</strong>
                <p style="font-size:0.78rem; color:var(--text-muted); margin:4px 0 10px 0;">Supports multiple photos (JPG, PNG, WebP). Automatically saved to <code>Finding/</code> folder and synced with Excel.</p>
                <div style="display:flex; justify-content:center; gap:8px; flex-wrap:wrap;">
                    <button type="button" class="btn-save" style="padding:7px 14px; font-size:0.8rem; display:inline-flex; align-items:center; gap:6px;" onclick="document.getElementById('camera-input-modal').click()">
                        <svg class="ui-icon sm" viewBox="0 0 24 24"><path d="M23 19a2 2 0 0 1-2 2H3a2 2 0 0 1-2-2V8a2 2 0 0 1 2-2h4l2-3h6l2 3h4a2 2 0 0 1 2 2z"></path><circle cx="12" cy="13" r="4"></circle></svg> Take Photo
                    </button>
                    <button type="button" class="page-btn" style="padding:7px 14px; font-size:0.8rem; display:inline-flex; align-items:center; gap:6px;" onclick="document.getElementById('file-input-modal').click()">
                        <svg class="ui-icon sm" viewBox="0 0 24 24"><path d="M22 19a2 2 0 0 1-2 2H4a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h5l2 3h9a2 2 0 0 1 2 2z"></path></svg> Select Files
                    </button>
                </div>
                <input type="file" id="file-input-modal" accept="image/*" multiple style="display:none;" onchange="handleModalFileSelect(this.files)">
                <input type="file" id="camera-input-modal" accept="image/*" capture="environment" style="display:none;" onchange="handleModalFileSelect(this.files)">
            </div>

            <!-- Photo Gallery Grid -->
            <div id="modal-photos-container">
                <div style="font-size:0.82rem; font-weight:700; color:var(--text-muted); margin-bottom:8px;">Saved Photos (<span id="modal-photo-count">0</span>):</div>
                <div class="photo-grid" id="modal-photo-grid"></div>
            </div>

            <!-- Finding & Follow-up Text -->
            <div class="form-grid" style="margin-top:16px;">
                <div class="form-group">
                    <label>Field Finding / Defect Details</label>
                    <textarea id="modal-finding-text" class="textarea-full" placeholder="Describe abnormal conditions, wear, leaks, or calibration deviations..."></textarea>
                </div>
                <div class="form-group">
                    <label>Corrective Action Plan / Recommendations</label>
                    <textarea id="modal-tl-text" class="textarea-full" placeholder="Repair plans, spare part replacements, recalibration actions..."></textarea>
                </div>
            </div>

            <div style="display:flex; justify-content:flex-end; gap:10px; margin-top:16px; border-top:1px solid var(--border-color); padding-top:14px;">
                <button class="page-btn" onclick="closeFindingModal()">Close</button>
                <button class="btn-save" onclick="saveFindingModalData()" style="display:inline-flex; align-items:center; gap:6px;">
                    <svg class="ui-icon sm" viewBox="0 0 24 24"><path d="M19 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h11l5 5v11a2 2 0 0 1-2 2z"></path><polyline points="17 21 17 13 7 13 7 21"></polyline><polyline points="7 3 7 8 15 8"></polyline></svg> Save Findings & Photos
                </button>
            </div>
        </div>
    </div>

    <!-- Dedicated Report Modal with 4 Options -->
    <div class="modal-overlay" id="report-modal">
        <div class="modal-content" style="max-width: 1050px; max-height: 92vh; display: flex; flex-direction: column;">
            <div class="modal-header" style="flex-shrink: 0;">
                <div>
                    <h3 style="color:var(--primary); font-size:1.15rem; font-weight:800; display:flex; align-items:center; gap:8px;" id="report-modal-title">
                        <svg class="ui-icon lg" viewBox="0 0 24 24"><polyline points="6 9 6 2 18 2 18 9"></polyline><path d="M6 18H4a2 2 0 0 1-2-2v-5a2 2 0 0 1 2-2h16a2 2 0 0 1 2 2v5a2 2 0 0 1-2 2h-2"></path><rect x="6" y="14" width="12" height="8"></rect></svg>
                        Outage EIC Report & Print Center
                    </h3>
                    <div style="font-size:0.8rem; color:var(--text-muted);">Select one of 4 official report formats to export and print to PDF (Unit 1 & 2).</div>
                </div>
                <button class="modal-close" onclick="closeReportModal()">&times;</button>
            </div>

            <!-- 4 Report Option Selector Tabs -->
            <div class="report-type-selector" style="display:flex; gap:8px; margin-bottom:12px; flex-wrap:wrap; flex-shrink:0;">
                <button class="unit-btn active" id="reptab-harian" onclick="setReportType('harian')">1. Daily Progress & Findings</button>
                <button class="unit-btn" id="reptab-wo_detail" onclick="setReportType('wo_detail')">2. Full WO & Sub-Tasks</button>
                <button class="unit-btn" id="reptab-actuator" onclick="setReportType('actuator')">3. Actuator Valves</button>
                <button class="unit-btn" id="reptab-instruments" onclick="setReportType('instruments')">4. Instruments (TX & PSW)</button>
            </div>

            <!-- Date Range Filter Bar for Harian Report -->
            <div id="report-date-bar" style="display:flex; flex-wrap:wrap; align-items:center; gap:10px; background:var(--bg-sub); padding:10px 16px; border-radius:var(--radius-sm); border:1px solid var(--border-color); margin-bottom:14px; flex-shrink: 0;">
                <span style="font-size:0.82rem; font-weight:700; color:var(--text-main);">Task Period:</span>
                <input type="date" id="report-start-date" class="filter-input" style="padding:5px 10px; font-size:0.82rem;">
                <span style="font-size:0.82rem; color:var(--text-muted);">to</span>
                <input type="date" id="report-end-date" class="filter-input" style="padding:5px 10px; font-size:0.82rem;">
                <button class="page-btn" style="padding:5px 12px; font-size:0.8rem;" onclick="generateReportContent()">Apply Filter</button>
                <div style="margin-left:auto; display:flex; gap:8px;">
                    <button class="btn-save" style="padding:6px 16px; font-size:0.85rem; display:inline-flex; align-items:center; gap:6px;" onclick="printReportModal()">
                        <svg class="ui-icon sm" viewBox="0 0 24 24"><polyline points="6 9 6 2 18 2 18 9"></polyline><path d="M6 18H4a2 2 0 0 1-2-2v-5a2 2 0 0 1 2-2h16a2 2 0 0 1 2 2v5a2 2 0 0 1-2 2h-2"></path><rect x="6" y="14" width="12" height="8"></rect></svg> Print / Export PDF
                    </button>
                </div>
            </div>

            <!-- Report Body (Scrollable in Modal, Full Width when Printing) -->
            <div style="overflow-y:auto; flex-grow:1; padding-right:6px;" id="report-modal-scroll">
                <div id="report-printable-content">
                    <!-- Dynamic Report Content Generated by JS -->
                </div>
            </div>

            <div style="display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:10px; margin-top:14px; border-top:1px solid var(--border-color); padding-top:12px; flex-shrink: 0;">
                <div style="display:flex; gap:8px; flex-wrap:wrap;">
                    <button class="page-btn" onclick="openSCurveModal()" style="font-weight:700; display:inline-flex; align-items:center; gap:6px;">
                        <svg class="ui-icon sm" viewBox="0 0 24 24"><polyline points="23 6 13.5 15.5 8.5 10.5 1 18"></polyline><polyline points="17 6 23 6 23 12"></polyline></svg> S-Curve & Trends
                    </button>
                    <button class="page-btn" onclick="openWaSummaryModal()" style="font-weight:700; color:#10b981; border-color:rgba(16,185,129,0.4); display:inline-flex; align-items:center; gap:6px;">
                        <svg class="ui-icon sm" viewBox="0 0 24 24"><path d="M21 15a2 2 0 0 1-2 2H7l-4 4V5a2 2 0 0 1 2-2h14a2 2 0 0 1 2 2z"></path></svg> WhatsApp Summary
                    </button>
                    <button class="page-btn" onclick="downloadExcel()" style="color:var(--primary); font-weight:700; display:inline-flex; align-items:center; gap:6px;">
                        <svg class="ui-icon sm" viewBox="0 0 24 24"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"></path><polyline points="7 10 12 15 17 10"></polyline><line x1="12" y1="15" x2="12" y2="3"></line></svg> Download Excel (.xlsx)
                    </button>
                </div>
                <div style="display:flex; gap:8px;">
                    <button class="page-btn" onclick="closeReportModal()">Close</button>
                    <button class="btn-save" onclick="printReportModal()" style="display:inline-flex; align-items:center; gap:6px;">
                        <svg class="ui-icon sm" viewBox="0 0 24 24"><polyline points="6 9 6 2 18 2 18 9"></polyline><path d="M6 18H4a2 2 0 0 1-2-2v-5a2 2 0 0 1 2-2h16a2 2 0 0 1 2 2v5a2 2 0 0 1-2 2h-2"></path><rect x="6" y="14" width="12" height="8"></rect></svg> Print / Export PDF
                    </button>
                </div>
            </div>
        </div>
    </div>

    <!-- S-Curve & Daily Progress Trend Modal -->
    <div class="modal-overlay" id="scurve-modal">
        <div class="modal-content" style="max-width: 880px; max-height: 90vh; display: flex; flex-direction: column;">
            <div class="modal-header" style="margin-bottom: 12px;">
                <div>
                    <h3 style="color:var(--primary); font-size:1.15rem; font-weight:800; display:flex; align-items:center; gap:8px;">
                        <svg class="ui-icon lg" viewBox="0 0 24 24"><polyline points="23 6 13.5 15.5 8.5 10.5 1 18"></polyline><polyline points="17 6 23 6 23 12"></polyline></svg>
                        S-Curve & Progress Trends Unit <span id="scurve-unit-label">1</span>
                    </h3>
                    <div style="font-size:0.8rem; color:var(--text-muted);">Cumulative progress trajectory and daily completion trends (Work Orders, Valves, & Instruments).</div>
                </div>
                <button class="modal-close" onclick="closeSCurveModal()">&times;</button>
            </div>

            <!-- Outage Schedule Range Control Bar -->
            <div style="display:flex; flex-wrap:wrap; align-items:center; justify-content:space-between; gap:10px; background:var(--bg-sub); padding:10px 14px; border-radius:var(--radius-sm); border:1px solid var(--border-color); margin-bottom:12px; flex-shrink:0;">
                <div style="display:flex; align-items:center; gap:8px; flex-wrap:wrap;">
                    <span style="font-size:0.82rem; font-weight:700; color:var(--text-main);">Outage Period:</span>
                    <input type="date" id="scurve-start-date" class="filter-input" style="padding:4px 8px; font-size:0.8rem;" onchange="saveAndRenderSCurve()" title="Outage Start Date">
                    <span style="font-size:0.8rem; color:var(--text-muted);">to</span>
                    <input type="date" id="scurve-end-date" class="filter-input" style="padding:4px 8px; font-size:0.8rem;" onchange="saveAndRenderSCurve()" title="Outage Target Date">
                    <button class="page-btn" style="padding:4px 12px; font-size:0.8rem; font-weight:700;" onclick="renderSCurveChart()">Calculate S-Curve</button>
                </div>
                <div id="scurve-kpi-badge" style="display:flex; align-items:center; gap:6px; font-size:0.8rem; font-weight:700;"></div>
            </div>

            <div style="overflow-y:auto; flex-grow:1; padding-right:4px;">
                <div id="scurve-chart-container"></div>
            </div>
            <div style="display:flex; justify-content:flex-end; gap:10px; margin-top:14px; border-top:1px solid var(--border-color); padding-top:12px; flex-shrink:0;">
                <button class="page-btn" onclick="closeSCurveModal()">Close</button>
            </div>
        </div>
    </div>

    <!-- WhatsApp Summary Generator Modal -->
    <div class="modal-overlay" id="wa-modal">
        <div class="modal-content" style="max-width: 680px; max-height: 90vh; display: flex; flex-direction: column;">
            <div class="modal-header">
                <div>
                    <h3 style="color:#10b981; font-size:1.15rem; font-weight:800; display:flex; align-items:center; gap:8px;">
                        <svg class="ui-icon lg" viewBox="0 0 24 24"><path d="M21 15a2 2 0 0 1-2 2H7l-4 4V5a2 2 0 0 1 2-2h14a2 2 0 0 1 2 2z"></path></svg>
                        WhatsApp Summary Generator
                    </h3>
                    <div style="font-size:0.8rem; color:var(--text-muted);">Structured summary message format formatted for team WhatsApp coordination.</div>
                </div>
                <button class="modal-close" onclick="closeWaSummaryModal()">&times;</button>
            </div>
            <div style="overflow-y:auto; flex-grow:1; padding-right:4px;">
                <textarea id="wa-text-box" class="wa-box" style="width:100%; height:320px; resize:vertical; font-family:'JetBrains Mono', monospace; font-size:0.82rem;" readonly></textarea>
            </div>
            <div style="display:flex; justify-content:space-between; align-items:center; margin-top:14px; border-top:1px solid var(--border-color); padding-top:12px;">
                <button class="page-btn" onclick="closeWaSummaryModal()">Close</button>
                <div style="display:flex; gap:8px;">
                    <button class="btn-save" onclick="copyWaText()" style="background:#10b981; display:inline-flex; align-items:center; gap:6px;">
                        <svg class="ui-icon sm" viewBox="0 0 24 24"><path d="M16 4h2a2 2 0 0 1 2 2v14a2 2 0 0 1-2 2H6a2 2 0 0 1-2-2V6a2 2 0 0 1 2-2h2"></path><rect x="8" y="2" width="8" height="4" rx="1" ry="1"></rect></svg> Copy to Clipboard
                    </button>
                </div>
            </div>
        </div>
    </div>

    <!-- Master EIC Authorization Password Modal -->
    <div class="modal-overlay" id="master-pwd-modal">
        <div class="modal-content" style="max-width: 440px; padding: 24px;">
            <div class="modal-header" style="margin-bottom: 14px;">
                <div>
                    <h3 style="color:var(--primary); font-size:1.15rem; font-weight:800; display:flex; align-items:center; gap:8px;">
                        <svg class="ui-icon lg" viewBox="0 0 24 24"><rect x="3" y="11" width="18" height="11" rx="2" ry="2"></rect><path d="M7 11V7a5 5 0 0 1 10 0v4"></path></svg>
                        Master EIC Authorization
                    </h3>
                    <div style="font-size:0.8rem; color:var(--text-muted); margin-top:2px;">Enter master authorization password to unlock Master PIC & Scope editing.</div>
                </div>
                <button class="modal-close" onclick="closeMasterPasswordModal()">&times;</button>
            </div>
            
            <form onsubmit="submitMasterPassword(event)" style="margin-top: 10px;">
                <div class="form-group" style="margin-bottom: 14px;">
                    <label style="font-size:0.82rem; font-weight:700; color:var(--text-main);">Master Password:</label>
                    <div style="position:relative; display:flex; align-items:center;">
                        <input type="password" id="master-pwd-input" class="filter-input" placeholder="Enter master password..." style="width:100%; padding:10px 38px 10px 12px; font-size:0.95rem; font-family:'Inter', sans-serif;" autocomplete="current-password">
                        <button type="button" onclick="toggleMasterPwdVisibility()" style="position:absolute; right:10px; background:none; border:none; color:var(--text-muted); cursor:pointer; font-size:1rem; display:flex; align-items:center;" title="Show / Hide Password" id="pwd-eye-btn">
                            <svg class="ui-icon md" viewBox="0 0 24 24"><path d="M1 12s4-8 11-8 11 8 11 8-4 8-11 8-11-8-11-8z"></path><circle cx="12" cy="12" r="3"></circle></svg>
                        </button>
                    </div>
                    <div id="master-pwd-error" style="color:#f43f5e; font-size:0.78rem; font-weight:700; margin-top:6px; display:none;">Incorrect password! Please try again.</div>
                </div>
                
                <div style="display:flex; justify-content:flex-end; gap:10px; margin-top:18px;">
                    <button type="button" class="page-btn" onclick="closeMasterPasswordModal()">Cancel</button>
                    <button type="submit" class="btn-save" style="padding:9px 20px; display:inline-flex; align-items:center; gap:6px;">
                        <svg class="ui-icon sm" viewBox="0 0 24 24"><rect x="3" y="11" width="18" height="11" rx="2" ry="2"></rect><path d="M7 11V7a5 5 0 0 1 9.9-1"></path></svg> Unlock Edit Access
                    </button>
                </div>
            </form>
        </div>
    </div>

    <!-- Header -->
    <header>
        <div class="logo-area">
            <div class="logo-badge" onclick="window.location.reload()" title="Click to Reload Page" style="display:inline-flex; align-items:center; gap:6px;">
                <svg class="ui-icon sm" viewBox="0 0 24 24"><polygon points="13 2 3 14 12 14 11 22 21 10 12 10 13 2"></polygon></svg>
                PLTU MSW EIC
            </div>
        </div>
        <div class="title-group">
            <h1>Outage Work Order Monitoring System</h1>
            <p>Section Electric, Instrument & Control &bull; Real-Time Dashboard</p>
        </div>
        <div class="header-controls">
            <button class="theme-toggle-btn" id="theme-toggle-btn" onclick="toggleTheme()" title="Toggle Light / Dark Mode">
                <span id="theme-icon" style="display:inline-flex; align-items:center; justify-content:center;"></span>
            </button>
            <button class="btn-print" onclick="openReportModal()" title="Open Reports / Export PDF" style="display:inline-flex; align-items:center; gap:6px;">
                <svg class="ui-icon sm" viewBox="0 0 24 24"><polyline points="6 9 6 2 18 2 18 9"></polyline><path d="M6 18H4a2 2 0 0 1-2-2v-5a2 2 0 0 1 2-2h16a2 2 0 0 1 2 2v5a2 2 0 0 1-2 2h-2"></path><rect x="6" y="14" width="12" height="8"></rect></svg> Reports
            </button>
        </div>
    </header>


    <div class="container">
        <!-- Outage Banner & Progress -->
        <div class="outage-banner">
            <div class="outage-info">
                <div class="unit-switcher">
                    <button class="unit-btn active" id="btn-unit-1" onclick="switchUnit(1)">UNIT 1</button>
                    <button class="unit-btn" id="btn-unit-2" onclick="switchUnit(2)">UNIT 2</button>
                </div>
                <div>
                    <div class="outage-title" id="outage-unit-title">Outage EIC Monitoring Progress Unit 1</div>
                    <div style="font-size:0.8rem; color:var(--text-muted);">Seamlessly synchronized with Excel Templates & Findings Media</div>
                </div>
            </div>
            <div class="outage-progress-box">
                <div class="outage-pct-huge" id="grand-pct">0%</div>
                <div class="outage-bar-wrap">
                    <div style="display:flex; justify-content:space-between; align-items:center; gap:8px; font-size:0.78rem; font-weight:700; margin-bottom:4px; white-space:nowrap;">
                        <span>Total Completed: </span>
                        <span id="grand-counts">0 / 0 Items</span>
                    </div>
                    <div class="outage-bar-bg"><div class="outage-bar-fill" id="grand-bar-fill" style="width:0%;"></div></div>
                </div>
            </div>
        </div>

        <!-- 2-Column Dashboard Main Layout -->
        <div class="dashboard-main-layout">
            <!-- Left Column / Sidebar: KPI Stats Cards -->
            <aside class="sidebar-stats">
                <div class="sidebar-section-title">Progress Status</div>
                <div class="stats-vertical-list">
                    <div class="stat-card" onclick="switchTab('wo')" title="Click to view Work Orders">
                        <div class="stat-title">
                            <span class="stat-title-wrap">
                                <svg class="ui-icon sm" style="color:var(--primary);" viewBox="0 0 24 24"><path d="M16 4h2a2 2 0 0 1 2 2v14a2 2 0 0 1-2 2H6a2 2 0 0 1-2-2V6a2 2 0 0 1 2-2h2"></path><rect x="8" y="2" width="8" height="4" rx="1" ry="1"></rect></svg>
                                Work Orders (WO)
                            </span>
                        </div>
                        <div class="stat-value"><span id="wo-pct">0%</span> <span style="font-size:0.85rem; color:var(--text-muted);" id="wo-counts">0 / 0</span></div>
                        <div class="stat-sub" id="wo-sub">Finish: 0 | In-Progress: 0</div>
                    </div>
                    <div class="stat-card finish" onclick="switchTab('actuator')" title="Click to view Actuator Valves">
                        <div class="stat-title">
                            <span class="stat-title-wrap">
                                <svg class="ui-icon sm" style="color:var(--status-finish);" viewBox="0 0 24 24"><line x1="4" y1="21" x2="4" y2="14"></line><line x1="4" y1="10" x2="4" y2="3"></line><line x1="12" y1="21" x2="12" y2="12"></line><line x1="12" y1="8" x2="12" y2="3"></line><line x1="20" y1="21" x2="20" y2="16"></line><line x1="20" y1="12" x2="20" y2="3"></line><line x1="1" y1="14" x2="7" y2="14"></line><line x1="9" y1="8" x2="15" y2="8"></line><line x1="17" y1="16" x2="23" y2="16"></line></svg>
                                Actuator Valves
                            </span>
                        </div>
                        <div class="stat-value"><span id="act-pct">0%</span> <span style="font-size:0.85rem; color:var(--text-muted);" id="act-counts">0 / 0</span></div>
                        <div class="stat-sub" id="act-sub">Finish: 0 | In-Progress: 0</div>
                    </div>
                    <div class="stat-card inprog" onclick="switchTab('instrument')" title="Click to view Instruments">
                        <div class="stat-title">
                            <span class="stat-title-wrap">
                                <svg class="ui-icon sm" style="color:var(--status-inprog);" viewBox="0 0 24 24"><path d="m12 14 4-4"></path><path d="M3.34 19a10 10 0 1 1 17.32 0"></path></svg>
                                Instruments
                            </span>
                        </div>
                        <div class="stat-value"><span id="inst-pct">0%</span> <span style="font-size:0.85rem; color:var(--text-muted);" id="inst-counts">0 / 0</span></div>
                        <div class="stat-sub" id="inst-sub">PTX: 0 | TTX: 0 | PSW: 0</div>
                    </div>
                    <div class="stat-card findings" onclick="setQuickFilter('findings')" title="Click to filter logged findings">
                        <div class="stat-title">
                            <span class="stat-title-wrap">
                                <svg class="ui-icon sm" style="color:var(--status-alert);" viewBox="0 0 24 24"><path d="m21.73 18-8-14a2 2 0 0 0-3.48 0l-8 14A2 2 0 0 0 4 21h16a2 2 0 0 0 1.73-3Z"></path><line x1="12" y1="9" x2="12" y2="13"></line><line x1="12" y1="17" x2="12.01" y2="17"></line></svg>
                                Active Findings
                            </span>
                        </div>
                        <div class="stat-value" id="findings-count" style="color:var(--status-alert);">0</div>
                        <div class="stat-sub">Items with logged findings/photos</div>
                    </div>
                </div>
            </aside>

            <!-- Right Column: Main Content (Tabs, Filters, Table/Cards) -->
            <main class="main-content-panel">
                <!-- Main Navigation Tabs -->
                <div class="nav-tabs">
                    <button class="tab-btn active" onclick="switchTab('wo')" style="display:inline-flex; align-items:center; gap:6px;">
                        <svg class="ui-icon sm" viewBox="0 0 24 24"><path d="M16 4h2a2 2 0 0 1 2 2v14a2 2 0 0 1-2 2H6a2 2 0 0 1-2-2V6a2 2 0 0 1 2-2h2"></path><rect x="8" y="2" width="8" height="4" rx="1" ry="1"></rect></svg>
                        Work Orders <span class="tab-count" id="tab-cnt-wo">0</span>
                    </button>
                    <button class="tab-btn" onclick="switchTab('actuator')" style="display:inline-flex; align-items:center; gap:6px;">
                        <svg class="ui-icon sm" viewBox="0 0 24 24"><line x1="4" y1="21" x2="4" y2="14"></line><line x1="4" y1="10" x2="4" y2="3"></line><line x1="12" y1="21" x2="12" y2="12"></line><line x1="12" y1="8" x2="12" y2="3"></line><line x1="20" y1="21" x2="20" y2="16"></line><line x1="20" y1="12" x2="20" y2="3"></line><line x1="1" y1="14" x2="7" y2="14"></line><line x1="9" y1="8" x2="15" y2="8"></line><line x1="17" y1="16" x2="23" y2="16"></line></svg>
                        Actuator Valves <span class="tab-count" id="tab-cnt-act">0</span>
                    </button>
                    <button class="tab-btn" onclick="switchTab('instrument')" style="display:inline-flex; align-items:center; gap:6px;">
                        <svg class="ui-icon sm" viewBox="0 0 24 24"><path d="m12 14 4-4"></path><path d="M3.34 19a10 10 0 1 1 17.32 0"></path></svg>
                        Instruments <span class="tab-count" id="tab-cnt-inst">0</span>
                    </button>
                    <button class="tab-btn" onclick="switchTab('scope')" style="display:inline-flex; align-items:center; gap:6px;">
                        <svg class="ui-icon sm" viewBox="0 0 24 24"><path d="M12 22s8-4 8-10V5l-8-3-8 3v7c0 6 8 10 8 10z"></path></svg>
                        Scope & Master EIC
                    </button>
                </div>

                <!-- Filter & Control Toolbar -->
                <div class="filter-toolbar">
                    <div class="filter-top-row">
                        <input type="text" id="search-input" class="filter-input search" placeholder="Search WO, KKS, Equipment, Area, Findings, or PIC..." oninput="applyFilters()">
                        <select id="filter-status" class="filter-input" onchange="applyFilters()">
                            <option value="">All Statuses</option>
                            <option value="FINISH">FINISH</option>
                            <option value="IN PROGRESS">IN PROGRESS</option>
                            <option value="SCHED-OK">SCHED-OK</option>
                        </select>
                        <select id="filter-pic" class="filter-input" onchange="applyFilters()">
                            <option value="">All PICs</option>
                        </select>
                        <select id="filter-area" class="filter-input" onchange="applyFilters()">
                            <option value="">All Areas</option>
                        </select>

                        <div class="view-switcher" id="view-switcher-box">
                            <button class="view-btn active" id="btn-view-cards" onclick="switchViewMode('cards')" style="display:inline-flex; align-items:center; gap:5px;">
                                <svg class="ui-icon sm" viewBox="0 0 24 24"><rect x="3" y="3" width="7" height="7"></rect><rect x="14" y="3" width="7" height="7"></rect><rect x="14" y="14" width="7" height="7"></rect><rect x="3" y="14" width="7" height="7"></rect></svg> Cards
                            </button>
                            <button class="view-btn" id="btn-view-table" onclick="switchViewMode('table')" style="display:inline-flex; align-items:center; gap:5px;">
                                <svg class="ui-icon sm" viewBox="0 0 24 24"><line x1="3" y1="6" x2="21" y2="6"></line><line x1="3" y1="12" x2="21" y2="12"></line><line x1="3" y1="18" x2="21" y2="18"></line></svg> Table
                            </button>
                        </div>
                    </div>

                    <!-- Quick Filter Chips & Side Item Counter -->
                    <div class="filter-pills" style="display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:8px;">
                        <div style="display:flex; align-items:center; gap:6px; flex-wrap:wrap;">
                            <button class="pill-btn active" id="pill-all" onclick="setQuickFilter('all')">All Items</button>
                            <button class="pill-btn" id="pill-findings" onclick="setQuickFilter('findings')" style="display:inline-flex; align-items:center; gap:5px;">
                                <svg class="ui-icon sm" style="color:var(--status-alert);" viewBox="0 0 24 24"><path d="m21.73 18-8-14a2 2 0 0 0-3.48 0l-8 14A2 2 0 0 0 4 21h16a2 2 0 0 0 1.73-3Z"></path><line x1="12" y1="9" x2="12" y2="13"></line><line x1="12" y1="17" x2="12.01" y2="17"></line></svg>
                                Active Findings / Photos
                            </button>
                            <button class="pill-btn" id="pill-inprog" onclick="setQuickFilter('inprog')">In Progress</button>
                            <button class="pill-btn" id="pill-finish" onclick="setQuickFilter('finish')">Completed</button>
                        </div>
                        <div id="side-pagination-counter" style="margin-left:auto; display:flex; align-items:center; gap:8px; font-size:0.82rem; font-weight:600; color:var(--text-muted); background:var(--bg-sub); padding:4px 10px; border-radius:20px; border:1px solid var(--border-color);">
                            <span>Showing <strong style="color:var(--primary);" id="side-item-range">0 - 0</strong> of <strong style="color:var(--text-main);" id="side-item-total">0</strong> items</span>
                            <select id="side-page-size-select" class="filter-input" style="padding:2px 6px; font-size:0.75rem; border-radius:4px; margin-left:2px;" onchange="changePageSize(this.value)">
                                <option value="20" selected>20 / page</option>
                                <option value="40">40 / page</option>
                                <option value="99999">All</option>
                            </select>
                            <button class="btn-refresh-icon" onclick="refreshDataSilently()" title="Refresh Latest Data (In-place update without page reload)">
                                <span class="refresh-icon">
                                    <svg class="ui-icon sm" viewBox="0 0 24 24"><polyline points="23 4 23 10 17 10"></polyline><polyline points="1 20 1 14 7 14"></polyline><path d="M3.51 9a9 9 0 0 1 14.85-3.36L23 10M1 14l4.64 4.36A9 9 0 0 0 20.49 15"></path></svg>
                                </span>
                                <span class="refresh-text">Refresh</span>
                            </button>
                        </div>
                    </div>
                </div>

                <!-- Dynamic Tab Content Container -->
                <div id="tab-content">
                    <div style="text-align:center; padding:60px; color:var(--text-muted);">Loading outage data...</div>
                </div>
            </main>
        </div>
    </div>

    <script>
        const Icons = {
            clipboard: `<svg class="ui-icon sm" viewBox="0 0 24 24"><path d="M16 4h2a2 2 0 0 1 2 2v14a2 2 0 0 1-2 2H6a2 2 0 0 1-2-2V6a2 2 0 0 1 2-2h2"></path><rect x="8" y="2" width="8" height="4" rx="1" ry="1"></rect></svg>`,
            sliders: `<svg class="ui-icon sm" viewBox="0 0 24 24"><line x1="4" y1="21" x2="4" y2="14"></line><line x1="4" y1="10" x2="4" y2="3"></line><line x1="12" y1="21" x2="12" y2="12"></line><line x1="12" y1="8" x2="12" y2="3"></line><line x1="20" y1="21" x2="20" y2="16"></line><line x1="20" y1="12" x2="20" y2="3"></line><line x1="1" y1="14" x2="7" y2="14"></line><line x1="9" y1="8" x2="15" y2="8"></line><line x1="17" y1="16" x2="23" y2="16"></line></svg>`,
            gauge: `<svg class="ui-icon sm" viewBox="0 0 24 24"><path d="m12 14 4-4"></path><path d="M3.34 19a10 10 0 1 1 17.32 0"></path></svg>`,
            shield: `<svg class="ui-icon sm" viewBox="0 0 24 24"><path d="M12 22s8-4 8-10V5l-8-3-8 3v7c0 6 8 10 8 10z"></path></svg>`,
            users: `<svg class="ui-icon sm" viewBox="0 0 24 24"><path d="M17 21v-2a4 4 0 0 0-4-4H5a4 4 0 0 0-4 4v2"></path><circle cx="9" cy="7" r="4"></circle><path d="M23 21v-2a4 4 0 0 0-3-3.87"></path><path d="M16 3.13a4 4 0 0 1 0 7.75"></path></svg>`,
            user: `<svg class="ui-icon sm" viewBox="0 0 24 24"><path d="M20 21v-2a4 4 0 0 0-4-4H8a4 4 0 0 0-4 4v2"></path><circle cx="12" cy="7" r="4"></circle></svg>`,
            plus: `<svg class="ui-icon sm" viewBox="0 0 24 24"><line x1="12" y1="5" x2="12" y2="19"></line><line x1="5" y1="12" x2="19" y2="12"></line></svg>`,
            trash: `<svg class="ui-icon sm" viewBox="0 0 24 24"><polyline points="3 6 5 6 21 6"></polyline><path d="M19 6v14a2 2 0 0 1-2 2H7a2 2 0 0 1-2-2V6m3 0V4a2 2 0 0 1 2-2h4a2 2 0 0 1 2 2v2"></path></svg>`,
            save: `<svg class="ui-icon sm" viewBox="0 0 24 24"><path d="M19 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h11l5 5v11a2 2 0 0 1-2 2z"></path><polyline points="17 21 17 13 7 13 7 21"></polyline><polyline points="7 3 7 8 15 8"></polyline></svg>`,
            camera: `<svg class="ui-icon sm" viewBox="0 0 24 24"><path d="M23 19a2 2 0 0 1-2 2H3a2 2 0 0 1-2-2V8a2 2 0 0 1 2-2h4l2-3h6l2 3h4a2 2 0 0 1 2 2z"></path><circle cx="12" cy="13" r="4"></circle></svg>`,
            check: `<svg class="ui-icon sm" viewBox="0 0 24 24"><polyline points="20 6 9 17 4 12"></polyline></svg>`,
            checkCircle: `<svg class="ui-icon sm" viewBox="0 0 24 24"><path d="M22 11.08V12a10 10 0 1 1-5.93-9.14"></path><polyline points="22 4 12 14.01 9 11.01"></polyline></svg>`,
            alertTriangle: `<svg class="ui-icon sm" viewBox="0 0 24 24"><path d="m21.73 18-8-14a2 2 0 0 0-3.48 0l-8 14A2 2 0 0 0 4 21h16a2 2 0 0 0 1.73-3Z"></path><line x1="12" y1="9" x2="12" y2="13"></line><line x1="12" y1="17" x2="12.01" y2="17"></line></svg>`,
            activity: `<svg class="ui-icon sm" viewBox="0 0 24 24"><polyline points="22 12 18 12 15 21 9 3 6 12 2 12"></polyline></svg>`,
            lock: `<svg class="ui-icon sm" viewBox="0 0 24 24"><rect x="3" y="11" width="18" height="11" rx="2" ry="2"></rect><path d="M7 11V7a5 5 0 0 1 10 0v4"></path></svg>`,
            unlock: `<svg class="ui-icon sm" viewBox="0 0 24 24"><rect x="3" y="11" width="18" height="11" rx="2" ry="2"></rect><path d="M7 11V7a5 5 0 0 1 9.9-1"></path></svg>`,
            refresh: `<svg class="ui-icon sm" viewBox="0 0 24 24"><polyline points="23 4 23 10 17 10"></polyline><polyline points="1 20 1 14 7 14"></polyline><path d="M3.51 9a9 9 0 0 1 14.85-3.36L23 10M1 14l4.64 4.36A9 9 0 0 0 20.49 15"></path></svg>`,
            sun: `<svg class="ui-icon md" viewBox="0 0 24 24"><circle cx="12" cy="12" r="5"></circle><line x1="12" y1="1" x2="12" y2="3"></line><line x1="12" y1="21" x2="12" y2="23"></line><line x1="4.22" y1="4.22" x2="5.64" y2="5.64"></line><line x1="18.36" y1="18.36" x2="19.78" y2="19.78"></line><line x1="1" y1="12" x2="3" y2="12"></line><line x1="21" y1="12" x2="23" y2="12"></line><line x1="4.22" y1="19.78" x2="5.64" y2="18.36"></line><line x1="18.36" y1="5.64" x2="19.78" y2="4.22"></line></svg>`,
            moon: `<svg class="ui-icon md" viewBox="0 0 24 24"><path d="M21 12.79A9 9 0 1 1 11.21 3 7 7 0 0 0 21 12.79z"></path></svg>`,
            eye: `<svg class="ui-icon md" viewBox="0 0 24 24"><path d="M1 12s4-8 11-8 11 8 11 8-4 8-11 8-11-8-11-8z"></path><circle cx="12" cy="12" r="3"></circle></svg>`,
            eyeOff: `<svg class="ui-icon md" viewBox="0 0 24 24"><path d="M17.94 17.94A10.07 10.07 0 0 1 12 20c-7 0-11-8-11-8a18.45 18.45 0 0 1 5.06-5.94M9.9 4.24A9.12 9.12 0 0 1 12 4c7 0 11 8 11 8a18.5 18.5 0 0 1-2.16 3.19m-6.72-1.07a3 3 0 1 1-4.24-4.24"></path><line x1="1" y1="1" x2="23" y2="23"></line></svg>`
        };
        let currentUnit = 1;
        let currentTab = 'wo';
        let currentViewMode = 'cards';
        let actuatorViewMode = 'unit';
        let instSubtab = 'ptx';
        let quickFilter = 'all';
        let fullData = null;
        let matrixData = null;
        let currentPage = 1;
        let pageSize = 20;
        let editModeState = {};
        let activeFinding = null;
        let isScopeMasterUnlocked = sessionStorage.getItem('eic_scope_unlocked') === 'true';
        const MASTER_PASSWORDS = ['eic123', 'admin123', 'msweic', '123456', 'eic2026'];

        // Date Helpers for Calendar Pickers
        function formatDateForInput(dateVal) {
            if(!dateVal) return '';
            const str = String(dateVal).trim();
            if (/^\d{4}-\d{2}-\d{2}/.test(str)) {
                return str.substring(0, 10);
            }
            if (/^\d{1,2}\/\d{1,2}\/\d{4}/.test(str)) {
                const parts = str.split('/');
                const day = parts[0].padStart(2, '0');
                const month = parts[1].padStart(2, '0');
                const year = parts[2].substring(0, 4);
                return `${year}-${month}-${day}`;
            }
            const d = new Date(str);
            if (!isNaN(d.getTime())) {
                const year = d.getFullYear();
                const month = String(d.getMonth() + 1).padStart(2, '0');
                const day = String(d.getDate()).padStart(2, '0');
                return `${year}-${month}-${day}`;
            }
            return '';
        }

        function formatDateForStorage(inputVal) {
            if(!inputVal) return '';
            const str = String(inputVal).trim();
            if (/^\d{4}-\d{2}-\d{2}/.test(str)) {
                const parts = str.split('-');
                return `${parts[2]}/${parts[1]}/${parts[0]}`;
            }
            return str;
        }

        // Toast System
        function showToast(message, type = 'success', duration = 3000) {
            const container = document.getElementById('toast-container');
            const toast = document.createElement('div');
            toast.className = `toast ${type}`;
            const iconSvg = type === 'success' ? Icons.checkCircle : (type === 'error' ? Icons.alertTriangle : Icons.activity);
            toast.innerHTML = `<span style="display:inline-flex;">${iconSvg}</span><div>${message}</div>`;
            container.appendChild(toast);
            setTimeout(() => {
                toast.style.opacity = '0';
                toast.style.transform = 'translateY(20px)';
                toast.style.transition = 'all 0.3s';
                setTimeout(() => toast.remove(), 300);
            }, duration);
        }

        async function loadData() {
            try {
                const [res, compRes] = await Promise.all([
                    fetch(`/api/data?unit=${currentUnit}`),
                    fetch(`/api/master_components?unit=${currentUnit}`).catch(() => null)
                ]);
                if(!res.ok) throw new Error(`HTTP ${res.status} Server Error`);
                fullData = await res.json();
                if(compRes && compRes.ok) {
                    const compData = await compRes.json();
                    fullData.master_actuators = compData.actuators || [];
                    fullData.master_instruments = compData.instruments || [];
                }
                const titleEl = document.getElementById('outage-unit-title');
                if(titleEl) titleEl.innerText = `Outage EIC Monitoring Progress Unit ${currentUnit}`;
                
                renderStats();
                populateFilterDropdowns();
                renderTabContent();
            } catch(e) {
                console.error("Error loading data:", e);
                document.getElementById('tab-content').innerHTML = `
                <div style="color:#ef4444; padding:24px; background:var(--bg-card); border-radius:12px; border:1px solid rgba(239,68,68,0.3);">
                    <strong style="font-size:1.1rem;">Failed to load Unit ${currentUnit} data</strong>
                    <div style="margin-top:8px; font-size:0.88rem; color:var(--text-main);">Error Detail: ${e.message}</div>
                    <div style="margin-top:12px; font-size:0.82rem; color:var(--text-muted);">Troubleshooting: Ensure server.exe or start_app.bat is running and the file <code>Template_Outage_EIC_Monitoring_unit ${currentUnit}.xlsx</code> exists in the root directory.</div>
                </div>`;
            }
        }

        async function refreshDataSilently() {
            const btns = document.querySelectorAll('.btn-refresh-icon');
            btns.forEach(b => {
                b.classList.add('is-loading');
                const txt = b.querySelector('.refresh-text');
                if(txt) txt.innerText = 'Loading...';
            });
            try {
                await loadData();
                showToast('✓ Data refreshed successfully!', 'info');
            } catch(e) {
                showToast('Failed to load latest data: ' + e.message, 'error');
            } finally {
                btns.forEach(b => {
                    b.classList.remove('is-loading');
                    const txt = b.querySelector('.refresh-text');
                    if(txt) txt.innerText = 'Refresh';
                });
            }
        }

        function switchUnit(unit) {
            currentUnit = unit;
            currentPage = 1;
            document.getElementById('btn-unit-1').classList.toggle('active', unit === 1);
            document.getElementById('btn-unit-2').classList.toggle('active', unit === 2);
            loadData();
        }

        function switchTab(tab) {
            currentTab = tab;
            currentPage = 1;
            document.querySelectorAll('.nav-tabs .tab-btn').forEach((btn, i) => {
                btn.classList.toggle('active', ['wo', 'actuator', 'instrument', 'scope'][i] === tab);
            });
            document.getElementById('view-switcher-box').style.display = (tab === 'scope') ? 'none' : 'flex';
            renderTabContent();
        }

        function switchViewMode(mode) {
            currentViewMode = mode;
            document.getElementById('btn-view-cards').classList.toggle('active', mode === 'cards');
            document.getElementById('btn-view-table').classList.toggle('active', mode === 'table');
            renderTabContent();
        }

        function setQuickFilter(f) {
            quickFilter = f;
            ['all', 'findings', 'inprog', 'finish'].forEach(p => {
                const el = document.getElementById(`pill-${p}`);
                if(el) el.classList.toggle('active', p === f);
            });
            currentPage = 1;
            renderTabContent();
        }

        function renderStats() {
            if(!fullData || !fullData.summary) return;
            const s = fullData.summary;

            document.getElementById('grand-pct').innerText = `${s.grand_pct}%`;
            document.getElementById('grand-counts').innerText = `${s.grand_done} / ${s.grand_total} Sub-tasks / Items`;
            document.getElementById('grand-bar-fill').style.width = `${s.grand_pct}%`;

            document.getElementById('wo-pct').innerText = `${s.wo.pct}%`;
            document.getElementById('wo-counts').innerText = `${s.wo.subtask_done} / ${s.wo.subtask_total} Sub-tasks`;
            document.getElementById('wo-sub').innerText = `Finish WO: ${s.wo.finish} / ${s.wo.total} | In-Prog: ${s.wo.in_progress}`;

            document.getElementById('act-pct').innerText = `${s.actuator.pct}%`;
            document.getElementById('act-counts').innerText = `${s.actuator.subtask_done} / ${s.actuator.subtask_total} Sub-tasks`;
            document.getElementById('act-sub').innerText = `Finish Valves: ${s.actuator.finish} / ${s.actuator.total} | In-Prog: ${s.actuator.in_progress}`;

            document.getElementById('inst-pct').innerText = `${s.instrument.pct}%`;
            document.getElementById('inst-counts').innerText = `${s.instrument.done} / ${s.instrument.total}`;
            document.getElementById('inst-sub').innerText = `PTX (${(fullData.pressure_tx||[]).length}) | TTX (${(fullData.temperature_tx||[]).length}) | PSW (${(fullData.pressure_switch||[]).length})`;

            document.getElementById('findings-count').innerText = s.findings_count;

            document.getElementById('tab-cnt-wo').innerText = (fullData.work_orders || []).length;
            document.getElementById('tab-cnt-act').innerText = (fullData.actuators || []).length;
            document.getElementById('tab-cnt-inst').innerText = s.instrument.total;

            // Update Sticky Summary Bar
            const uNum = document.getElementById('sticky-unit-num');
            const gPct = document.getElementById('sticky-grand-pct');
            const wPct = document.getElementById('sticky-wo-pct');
            const aPct = document.getElementById('sticky-act-pct');
            const iPct = document.getElementById('sticky-inst-pct');
            if(uNum) uNum.innerText = currentUnit;
            if(gPct) gPct.innerText = `${s.grand_pct}%`;
            if(wPct) wPct.innerText = `${s.wo.pct}%`;
            if(aPct) aPct.innerText = `${s.actuator.pct}%`;
            if(iPct) iPct.innerText = `${s.instrument.pct}%`;
        }

        function populateFilterDropdowns() {
            const picSelect = document.getElementById('filter-pic');
            if(picSelect) {
                const current = picSelect.value;
                picSelect.innerHTML = '<option value="">All PICs</option>';
                (fullData.pics || []).forEach(p => {
                    if(p) picSelect.innerHTML += `<option value="${p}" ${current===p?'selected':''}>${p}</option>`;
                });
            }

            const areaSelect = document.getElementById('filter-area');
            if(areaSelect) {
                const current = areaSelect.value;
                const areas = new Set();
                (fullData.work_orders || []).forEach(w => { if(w && w.area) areas.add(w.area); });
                (fullData.actuators || []).forEach(a => { if(a && a.area) areas.add(a.area); });
                (fullData.pressure_tx || []).forEach(p => { if(p && p.area) areas.add(p.area); });
                areaSelect.innerHTML = '<option value="">All Areas</option>';
                Array.from(areas).sort().forEach(a => {
                    areaSelect.innerHTML += `<option value="${a}" ${current===a?'selected':''}>${a}</option>`;
                });
            }
        }

        function applyFilters() {
            currentPage = 1;
            renderTabContent();
        }

        function toggleEditMode(itemId) {
            editModeState[itemId] = !editModeState[itemId];
            renderTabContent();
        }

        function updateSidePaginationCounter(totalItems) {
            const rangeEl = document.getElementById('side-item-range');
            const totalEl = document.getElementById('side-item-total');
            const selEl = document.getElementById('side-page-size-select');
            
            if(rangeEl && totalEl) {
                const from = totalItems === 0 ? 0 : (currentPage - 1) * pageSize + 1;
                const to = Math.min(currentPage * pageSize, totalItems);
                rangeEl.innerText = `${from} - ${to}`;
                totalEl.innerText = `${totalItems}`;
            }
            if(selEl) {
                selEl.value = String(pageSize);
            }
        }

        function renderPaginationControls(totalItems) {
            updateSidePaginationCounter(totalItems);
            const totalPages = Math.ceil(totalItems / pageSize) || 1;
            if(currentPage > totalPages) currentPage = totalPages;

            if(totalPages <= 1) {
                return '';
            }

            return `
            <div class="pagination-bar" style="margin-top:20px;">
                <div style="display:flex; align-items:center; gap:8px;">
                    <span>Showing <strong>${totalItems === 0 ? 0 : (currentPage - 1) * pageSize + 1} - ${Math.min(currentPage * pageSize, totalItems)}</strong> of <strong>${totalItems}</strong> items</span>
                    <button class="btn-refresh-icon" onclick="refreshDataSilently()" title="Refresh Latest Data (In-place update without page reload)" style="padding:2px 6px; font-size:0.74rem;">
                        <span class="refresh-icon">${Icons.refresh}</span><span class="refresh-text">Refresh</span>
                    </button>
                </div>
                <div style="display:flex; align-items:center; gap:6px;">
                    <button class="page-btn" onclick="changePage(1)" ${currentPage===1?'disabled':''}>First</button>
                    <button class="page-btn" onclick="changePage(${currentPage-1})" ${currentPage===1?'disabled':''}>Prev</button>
                    <span style="font-weight:700; margin:0 6px; font-size:0.85rem;">Page ${currentPage} / ${totalPages}</span>
                    <button class="page-btn" onclick="changePage(${currentPage+1})" ${currentPage===totalPages?'disabled':''}>Next</button>
                    <button class="page-btn" onclick="changePage(${totalPages})" ${currentPage===totalPages?'disabled':''}>Last</button>
                </div>
            </div>`;
        }

        function changePage(page) {
            currentPage = page;
            renderTabContent();
            window.scrollTo({top: 350, behavior: 'smooth'});
        }

        function changePageSize(size) {
            pageSize = parseInt(size);
            currentPage = 1;
            renderTabContent();
        }

        function filterItem(item, searchStr, statusFilter, picFilter, areaFilter, codeKey, descKey, statusKey, picKey, areaKey) {
            const code = (item[codeKey] || '').toLowerCase();
            const desc = (item[descKey] || '').toLowerCase();
            const remarks = (item.remarks || '').toLowerCase();
            const matchSearch = !searchStr || code.includes(searchStr) || desc.includes(searchStr) || remarks.includes(searchStr);
            
            const matchStatus = !statusFilter || item[statusKey] === statusFilter;
            const matchPic = !picFilter || item[picKey] === picFilter;
            const matchArea = !areaFilter || item[areaKey] === areaFilter;

            let matchQuick = true;
            if(quickFilter === 'findings') {
                matchQuick = !!item.temuan || (item.jumlah_foto > 0);
            } else if(quickFilter === 'inprog') {
                matchQuick = item[statusKey] !== 'FINISH' && !item.status_wdone;
            } else if(quickFilter === 'finish') {
                matchQuick = item[statusKey] === 'FINISH' || item.status_wdone === true;
            }

            return matchSearch && matchStatus && matchPic && matchArea && matchQuick;
        }

        /* ---------------- WORK ORDER RENDERING ---------------- */
        function renderWorkOrders(container) {
            const searchStr = document.getElementById('search-input').value.toLowerCase();
            const statusFilter = document.getElementById('filter-status').value;
            const picFilter = document.getElementById('filter-pic').value;
            const areaFilter = document.getElementById('filter-area').value;

            let filteredItems = (fullData.work_orders || []).filter(w => 
                filterItem(w, searchStr, statusFilter, picFilter, areaFilter, 'no_wo', 'job_description', 'status', 'pic', 'area')
            );

            const isAddWoOpen = openCardIds.has('add-wo-form');
            let html = '';

            html += `
            <div style="margin-bottom:20px; background:var(--bg-card); border-radius:var(--radius-md); padding:16px; border:1px solid var(--border-color);">
                <div style="display:flex; justify-content:space-between; align-items:center; cursor:pointer;" onclick="toggleAccordion('add-wo-form')">
                    <h3 style="font-size:0.95rem; color:var(--primary); font-weight:700; display:inline-flex; align-items:center; gap:6px;">${Icons.plus} Add New Work Order (WO)</h3>
                    <span id="arrow-add-wo-form" style="font-weight:700; color:var(--text-muted); font-size:0.85rem;">${isAddWoOpen ? 'Close Form' : 'Add'}</span>
                </div>
                <div id="add-wo-form" class="accordion-form ${isAddWoOpen ? 'open' : ''}" style="${isAddWoOpen ? 'display:block;' : ''}">
                    <div class="form-grid">
                        <div class="form-group">
                            <label>WO Number <span style="color:#f43f5e;">*</span></label>
                            <input type="text" id="new-wo-code" class="filter-input" placeholder="e.g. WO-100826-0099">
                        </div>
                        <div class="form-group">
                            <label>Job Description <span style="color:#f43f5e;">*</span></label>
                            <input type="text" id="new-wo-desc" class="filter-input" placeholder="WO task description...">
                        </div>
                        <div class="form-group">
                            <label>System Area</label>
                            <input type="text" id="new-wo-area" class="filter-input" placeholder="BOILER, ID FAN, COOLING TOWER...">
                        </div>
                        <div class="form-group">
                            <label>Person in Charge (PIC)</label>
                            <select id="new-wo-pic" class="filter-input">
                                <option value="">Select PIC...</option>
                                ${(fullData.pics || []).map(p => `<option value="${p}">${p}</option>`).join('')}
                            </select>
                        </div>
                        <div class="form-group">
                            <label>Schedule Date</label>
                            <input type="date" id="new-wo-sched" class="filter-input">
                        </div>
                        <div class="form-group" style="grid-column: span 2;">
                            <label>Sub-Task Checklist (Separate by comma or new line)</label>
                            <textarea id="new-wo-checklist" class="textarea-full" placeholder="e.g. General Inspection, Cleaning Contact, Function Test, Tightening Bolt"></textarea>
                        </div>
                    </div>
                    <div style="display:flex; gap:10px; margin-top:14px;">
                        <button class="btn-save" onclick="saveNewWO()" style="display:inline-flex; align-items:center; gap:6px;">${Icons.save} Save New Work Order</button>
                        <button class="page-btn" onclick="toggleAccordion('add-wo-form')">Cancel</button>
                    </div>
                </div>
            </div>`;

            updateSidePaginationCounter(filteredItems.length);

            if(filteredItems.length === 0) {
                html += '<div style="text-align:center; padding:50px; color:var(--text-muted); background:var(--bg-card); border-radius:12px;">No Work Orders match the current filter criteria.</div>';
                container.innerHTML = html;
                return;
            }

            const startIndex = (currentPage - 1) * pageSize;
            const pageItems = filteredItems.slice(startIndex, startIndex + pageSize);

            if(currentViewMode === 'cards') {
                html += '<div class="card-list">';
                pageItems.forEach((item, idx) => {
                    const st = String(item.status || 'SCHED-OK').replace(/\s+/g, '_');
                    const hasFindings = !!item.temuan || (item.jumlah_foto > 0);
                    const doneCount = (item.checklist || []).filter(c => c.selesai).length;
                    const totalCount = item.checklist ? item.checklist.length : 0;
                    const bodyId = getCardBodyId('wo', item.no_wo || idx);
                    const isOpen = openCardIds.has(bodyId);

                    html += `
                    <div class="item-card" id="card-wo-${item.no_wo}">
                        <div class="item-header" onclick="toggleAccordion('${bodyId}')">
                            <div class="item-title-box">
                                <div class="item-code">${item.no_wo} &bull; ${item.area || 'GENERAL'} ${item.tanggal_finish ? '&bull; Finished: ' + item.tanggal_finish : ''}</div>
                                <div class="item-name">${item.job_description}</div>
                            </div>
                            <div class="header-actions">
                                <button class="btn-finding ${hasFindings?'active':''}" onclick="event.stopPropagation(); openFindingModal('wo', '${item.no_wo}', '${item.no_wo} - ${item.job_description.replace(/'/g, "\\'")}', '${item.area}', '${(item.temuan||'').replace(/'/g, "\\'")}', '${(item.tindak_lanjut||'').replace(/'/g, "\\'")}')">
                                    ${Icons.camera} ${item.jumlah_foto > 0 ? item.jumlah_foto + ' Photos' : (hasFindings ? 'Findings' : '+ Finding')}
                                </button>
                                <span style="font-size:0.82rem; color:var(--text-muted); font-weight:600;">${item.pic || '-'}</span>
                                <span class="wo-subtask-progress" style="font-size:0.82rem; font-weight:700; color:var(--text-muted);">${doneCount} / ${totalCount} Sub-tasks</span>
                                <span class="status-badge badge-${st}">${item.status}</span>
                                <div class="progress-box">
                                    <div class="progress-bar-bg"><div class="progress-bar-fill progress-fill" style="width:${item.persen_progress}%;"></div></div>
                                    <span class="progress-text" style="font-size:0.85rem; font-weight:800; font-family:'JetBrains Mono';">${item.persen_progress}%</span>
                                </div>
                            </div>
                        </div>
                        <div class="item-body ${isOpen ? 'open' : ''}" id="${bodyId}" style="${isOpen ? 'display:block;' : ''}">
                            <div class="checklist-section">
                                <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:8px; flex-wrap:wrap; gap:8px;">
                                    <div class="section-h4" style="margin-bottom:0;">Sub-Task Checklist (${doneCount} / ${totalCount} Completed)</div>
                                    ${totalCount > 0 ? `
                                        <div style="display:flex; gap:6px;">
                                            <button class="btn-batch-check" onclick="batchToggleSubtasks('${item.no_wo}', 'mark_all_done')" title="Mark all sub-tasks as completed" style="display:inline-flex; align-items:center; gap:4px;">${Icons.check} Mark All Done</button>
                                            <button class="btn-batch-reset" onclick="batchToggleSubtasks('${item.no_wo}', 'reset_all')" title="Reset all sub-tasks" style="display:inline-flex; align-items:center; gap:4px;">${Icons.refresh} Reset</button>
                                        </div>
                                    ` : ''}
                                </div>
                                <div class="checklist-grid">
                                    ${(item.checklist || []).map((c, cIdx) => {
                                        const typeBadge = getSubtaskTypeBadge(c.sub_task);
                                        return `
                                        <div class="checklist-item ${c.selesai ? 'done' : ''}">
                                            <label class="checklist-item-body">
                                                <input type="checkbox" id="chk-${item.no_wo}-${cIdx}" ${c.selesai ? 'checked' : ''} onchange="toggleLocalSubtask('${item.no_wo}', ${cIdx}, this.checked)">
                                                <span>${c.sub_task}</span>
                                            </label>
                                            <div class="checklist-item-footer">
                                                <div class="checklist-footer-left">
                                                    ${typeBadge}
                                                </div>
                                                <div class="checklist-footer-right">
                                                    ${c.tanggal ? `<span class="date-badge" style="font-size:0.72rem; color:var(--status-finish); font-family:'JetBrains Mono',monospace; background:rgba(16,185,129,0.12); border:1px solid rgba(16,185,129,0.3); padding:2px 7px; border-radius:4px;" title="Date Completed">${c.tanggal}</span>` : ''}
                                                    <button class="btn-del-subtask-cross" title="Delete Subtask" onclick="deleteSubtask('${item.no_wo}', '${(c.sub_task||'').toString().replace(/'/g, "\\'")}')" aria-label="Delete">&times;</button>
                                                </div>
                                            </div>
                                        </div>`;
                                    }).join('')}
                                </div>
                                
                                <div style="margin-top:14px; background:rgba(0,0,0,0.2); border:1px solid var(--border-color); border-radius:8px; padding:10px 12px;">
                                    <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:8px; flex-wrap:wrap; gap:6px;">
                                        <span style="font-size:0.78rem; font-weight:700; color:var(--text-muted);">Add Sub-task to WO:</span>
                                        <div style="display:flex; gap:5px;">
                                            <button type="button" class="comp-mode-btn btn-mode-manual ${(subtaskAddModes[item.no_wo]||'manual')==='manual'?'active':''}" onclick="setSubtaskMode('${item.no_wo}', 'manual')">Manual</button>
                                            <button type="button" class="comp-mode-btn btn-mode-act ${subtaskAddModes[item.no_wo]==='actuator'?'active':''}" onclick="setSubtaskMode('${item.no_wo}', 'actuator')">Select Actuator (${(fullData.master_actuators||fullData.actuators||[]).length})</button>
                                            <button type="button" class="comp-mode-btn btn-mode-inst ${subtaskAddModes[item.no_wo]==='instrument'?'active':''}" onclick="setSubtaskMode('${item.no_wo}', 'instrument')">Select Instrument (${((fullData.pressure_tx||[]).length + (fullData.temperature_tx||[]).length + (fullData.pressure_switch||[]).length)})</button>
                                        </div>
                                    </div>
                                    
                                    <!-- Mode 1: Manual Input -->
                                    <div id="box-subtask-manual-${item.no_wo}" style="display:${(subtaskAddModes[item.no_wo]||'manual')==='manual'?'flex':'none'}; gap:8px;">
                                        <input type="text" id="new-subtask-${item.no_wo}" class="filter-input" placeholder="Type sub-task description (e.g. MOTOR FRAME INSPECTION)..." style="flex-grow:1; font-size:0.82rem;">
                                        <button class="btn-save" style="padding:6px 14px; font-size:0.8rem; white-space:nowrap; display:inline-flex; align-items:center; gap:4px;" onclick="addSubtask('${item.no_wo}', 'manual')">${Icons.plus} Add</button>
                                    </div>

                                    <!-- Mode 2: Actuator Dropdown Picker -->
                                    <div id="box-subtask-act-${item.no_wo}" style="display:${subtaskAddModes[item.no_wo]==='actuator'?'flex':'none'}; gap:8px;">
                                        <select id="new-subtask-act-${item.no_wo}" class="filter-input" style="flex-grow:1; font-size:0.82rem;">
                                            <option value="">-- Select Actuator Valve from Master List --</option>
                                            ${(fullData.master_actuators || fullData.actuators || []).map(a => `
                                                <option value="${a.equipment_description} ${a.kks||''}">[${a.area}] ${a.equipment_description} ${a.kks ? '('+a.kks+')' : ''}</option>
                                            `).join('')}
                                        </select>
                                        <button class="btn-save" style="padding:6px 14px; font-size:0.8rem; white-space:nowrap; background:#f59e0b; border-color:#d97706; color:#000; display:inline-flex; align-items:center; gap:4px;" onclick="addSubtask('${item.no_wo}', 'actuator')">${Icons.plus} Add Actuator</button>
                                    </div>

                                    <!-- Mode 3: Instrument Dropdown Picker -->
                                    <div id="box-subtask-inst-${item.no_wo}" style="display:${subtaskAddModes[item.no_wo]==='instrument'?'flex':'none'}; gap:8px;">
                                        <select id="new-subtask-inst-${item.no_wo}" class="filter-input" style="flex-grow:1; font-size:0.82rem;">
                                            <option value="">-- Select Instrument from Master List (PT/TT/PS) --</option>
                                            <optgroup label="Pressure Transmitter (PTX)">
                                                ${(fullData.pressure_tx || []).map(p => `
                                                    <option value="${p.kks ? p.kks+': ' : ''}${p.equipment}">[PTX - ${p.area}] ${p.kks ? p.kks+' : ' : ''}${p.equipment}</option>
                                                `).join('')}
                                            </optgroup>
                                            <optgroup label="Temperature Transmitter (TTX)">
                                                ${(fullData.temperature_tx || []).map(t => `
                                                    <option value="${t.kks ? t.kks+': ' : ''}${t.equipment}">[TTX - ${t.area}] ${t.kks ? t.kks+' : ' : ''}${t.equipment}</option>
                                                `).join('')}
                                            </optgroup>
                                            <optgroup label="Pressure Switch (PSW)">
                                                ${(fullData.pressure_switch || []).map(s => `
                                                    <option value="${s.kks ? s.kks+': ' : ''}${s.equipment}">[PSW - ${s.area}] ${s.kks ? s.kks+' : ' : ''}${s.equipment}</option>
                                                `).join('')}
                                            </optgroup>
                                        </select>
                                        <button class="btn-save" style="padding:6px 14px; font-size:0.8rem; white-space:nowrap; background:#06b6d4; border-color:#0891b2; color:#000; display:inline-flex; align-items:center; gap:4px;" onclick="addSubtask('${item.no_wo}', 'instrument')">${Icons.plus} Add Instrument</button>
                                    </div>
                                </div>
                            </div>

                            <div class="form-grid">
                                <div class="form-group">
                                    <label>Person in Charge (PIC)</label>
                                    <select id="pic-${item.no_wo}" class="filter-input">
                                        <option value="">Select PIC...</option>
                                        ${fullData.pics.map(p => `<option value="${p}" ${item.pic===p?'selected':''}>${p}</option>`).join('')}
                                    </select>
                                </div>
                                <div class="form-group">
                                    <label>Finish Date</label>
                                    <input type="date" id="finish-wo-${item.no_wo}" class="filter-input" value="${formatDateForInput(item.tanggal_finish)}">
                                </div>
                                <div class="form-group" style="grid-column: span 2;">
                                    <label>Field Remarks / Notes</label>
                                    <input type="text" id="rem-${item.no_wo}" class="filter-input" value="${item.remarks || ''}" placeholder="Field remarks...">
                                </div>
                            </div>

                            <div style="display:flex; justify-content:space-between; align-items:center; margin-top:15px; border-top:1px solid var(--border-color); padding-top:12px;">
                                <div>
                                    <button class="btn-danger" onclick="deleteWO('${item.no_wo}')" style="display:inline-flex; align-items:center; gap:6px;">${Icons.trash} Delete WO</button>
                                </div>
                                <div style="display:flex; gap:10px;">
                                    <button class="btn-save" onclick="saveWorkOrder('${item.no_wo}')" style="display:inline-flex; align-items:center; gap:6px;">${Icons.save} Save Changes</button>
                                </div>
                            </div>
                        </div>
                    </div>`;
                });
                html += '</div>';
            } else {
                html += `
                <div class="table-wrap">
                    <table class="dense-table">
                        <thead>
                            <tr>
                                <th>WO No</th>
                                <th>Job Description</th>
                                <th>Area</th>
                                <th>PIC</th>
                                <th>Subtasks</th>
                                <th>Progress</th>
                                <th>Status</th>
                                <th>Finish Date</th>
                                <th>Findings & Photos</th>
                                <th>Actions</th>
                            </tr>
                        </thead>
                        <tbody>
                            ${pageItems.map(item => {
                                const st = (item.status || 'SCHED-OK').replace(/\s+/g, '_');
                                const doneCount = (item.checklist || []).filter(c => c.selesai).length;
                                const totalCount = item.checklist ? item.checklist.length : 0;
                                const hasFindings = !!item.temuan || (item.jumlah_foto > 0);
                                return `
                                <tr>
                                    <td style="font-family:'JetBrains Mono'; font-weight:700; color:var(--primary);">${item.no_wo}</td>
                                    <td style="font-weight:600;">${item.job_description}</td>
                                    <td><span style="font-size:0.8rem; color:var(--text-muted);">${item.area || 'GENERAL'}</span></td>
                                    <td style="font-weight:600;">${item.pic || '-'}</td>
                                    <td><span style="font-size:0.82rem; font-weight:700;">${doneCount}/${totalCount}</span></td>
                                    <td>
                                        <div style="display:flex; align-items:center; gap:6px;">
                                            <div class="progress-bar-bg" style="width:60px;"><div class="progress-bar-fill" style="width:${item.persen_progress}%;"></div></div>
                                            <span style="font-size:0.8rem; font-weight:700;">${item.persen_progress}%</span>
                                        </div>
                                    </td>
                                    <td><span class="status-badge badge-${st}">${item.status}</span></td>
                                    <td style="font-family:'JetBrains Mono'; font-size:0.8rem; color:var(--text-muted);">${item.tanggal_finish || '-'}</td>
                                    <td>
                                        <button class="btn-finding ${hasFindings?'active':''}" onclick="openFindingModal('wo', '${item.no_wo}', '${item.no_wo} - ${item.job_description.replace(/'/g, "\\'")}', '${item.area}', '${(item.temuan||'').replace(/'/g, "\\'")}', '${(item.tindak_lanjut||'').replace(/'/g, "\\'")}')">
                                            ${Icons.camera} ${item.jumlah_foto > 0 ? item.jumlah_foto + ' Photos' : (hasFindings ? 'Findings' : '+ Photo')}
                                        </button>
                                    </td>
                                    <td>
                                        <button class="page-btn" style="padding:4px 8px; font-size:0.78rem;" onclick="switchViewMode('cards'); toggleAccordion('body-wo-0');">Details</button>
                                    </td>
                                </tr>`;
                            }).join('')}
                        </tbody>
                    </table>
                </div>`;
            }

            html += renderPaginationControls(filteredItems.length);
            container.innerHTML = html;
        }

        /* ---------------- ACTUATOR VALVE RENDERING ---------------- */
        function renderActuators(container) {
            const searchStr = document.getElementById('search-input').value.toLowerCase();
            const statusFilter = document.getElementById('filter-status').value;
            const picFilter = document.getElementById('filter-pic').value;
            const areaFilter = document.getElementById('filter-area').value;

            let html = '';

            let filteredItems = (fullData.actuators || []).filter(a => 
                filterItem(a, searchStr, statusFilter, picFilter, areaFilter, 'equipment_id', 'equipment_description', 'status', 'pic', 'area')
            );

            const isAddActOpen = openCardIds.has('add-act-form');
            html += `
            <div style="margin-bottom:20px; background:var(--bg-card); border-radius:var(--radius-md); padding:16px; border:1px solid var(--border-color);">
                <div style="display:flex; justify-content:space-between; align-items:center; cursor:pointer;" onclick="toggleAccordion('add-act-form')">
                    <h3 style="font-size:0.95rem; color:var(--primary); font-weight:700; display:inline-flex; align-items:center; gap:6px;">${Icons.plus} Add New Actuator Valve</h3>
                    <span id="arrow-add-act-form" style="font-weight:700; color:var(--text-muted); font-size:0.85rem;">${isAddActOpen ? 'Close Form' : 'Add'}</span>
                </div>
                <div id="add-act-form" class="accordion-form ${isAddActOpen ? 'open' : ''}" style="${isAddActOpen ? 'display:block;' : ''}">
                    <div class="form-grid">
                        <div class="form-group">
                            <label>Equipment ID <span style="color:#f43f5e;">*</span></label>
                            <input type="text" id="new-act-id" class="filter-input" placeholder="e.g. AV-099">
                        </div>
                        <div class="form-group">
                            <label>Equipment Description <span style="color:#f43f5e;">*</span></label>
                            <input type="text" id="new-act-desc" class="filter-input" placeholder="FEED WATER ACTUATOR...">
                        </div>
                        <div class="form-group">
                            <label>System Area</label>
                            <input type="text" id="new-act-area" class="filter-input" placeholder="BOILER, ID FAN...">
                        </div>
                        <div class="form-group">
                            <label>KKS Tag</label>
                            <input type="text" id="new-act-kks" class="filter-input" placeholder="10LAB30AA210">
                        </div>
                        <div class="form-group">
                            <label>Person in Charge (PIC)</label>
                            <select id="new-act-pic" class="filter-input">
                                <option value="">Select PIC...</option>
                                ${(fullData.pics || []).map(p => `<option value="${p}">${p}</option>`).join('')}
                            </select>
                        </div>
                    </div>
                    <div style="display:flex; gap:10px; margin-top:14px;">
                        <button class="btn-save" onclick="saveNewActuator()" style="display:inline-flex; align-items:center; gap:6px;">${Icons.save} Save New Actuator</button>
                        <button class="page-btn" onclick="toggleAccordion('add-act-form')">Cancel</button>
                    </div>
                </div>
            </div>`;

            updateSidePaginationCounter(filteredItems.length);

            if(filteredItems.length === 0) {
                html += '<div style="text-align:center; padding:50px; color:var(--text-muted); background:var(--bg-card); border-radius:12px;">No Actuator Valves match the current filter criteria.</div>';
                container.innerHTML = html;
                return;
            }

            const startIndex = (currentPage - 1) * pageSize;
            const pageItems = filteredItems.slice(startIndex, startIndex + pageSize);

            if(currentViewMode === 'cards') {
                html += '<div class="card-list">';
                pageItems.forEach((item, idx) => {
                    const st = String(item.status || 'SCHED-OK').replace(/\s+/g, '_');
                    const hasFindings = !!item.temuan || (item.jumlah_foto > 0);
                    const bodyId = getCardBodyId('act', item.equipment_id || idx);
                    const isOpen = openCardIds.has(bodyId);

                    html += `
                    <div class="item-card">
                        <div class="item-header" onclick="toggleAccordion('${bodyId}')">
                            <div class="item-title-box">
                                <div class="item-code">${item.equipment_id} &bull; ${item.area} ${item.kks ? '&bull; KKS: ' + item.kks : ''} ${item.finish_date ? '&bull; Finished: ' + item.finish_date : ''}</div>
                                <div class="item-name">${item.equipment_description}</div>
                            </div>
                            <div class="header-actions">
                                <button class="btn-finding ${hasFindings?'active':''}" onclick="event.stopPropagation(); openFindingModal('actuator', '${item.equipment_id}', '${item.equipment_id} - ${item.equipment_description.replace(/'/g, "\\'")}', '${item.area}', '${(item.temuan||'').replace(/'/g, "\\'")}', '${(item.tindak_lanjut||'').replace(/'/g, "\\'")}')">
                                    ${Icons.camera} ${item.jumlah_foto > 0 ? item.jumlah_foto + ' Photos' : (hasFindings ? 'Findings' : '+ Finding')}
                                </button>
                                <span style="font-size:0.82rem; color:var(--text-muted); font-weight:600;">${item.pic || '-'}</span>
                                <span class="status-badge badge-${st}">${item.status}</span>
                                <div class="progress-box">
                                    <div class="progress-bar-bg"><div class="progress-bar-fill" style="width:${item.persen_progress}%;"></div></div>
                                    <span style="font-size:0.85rem; font-weight:800; font-family:'JetBrains Mono';">${item.persen_progress}%</span>
                                </div>
                            </div>
                        </div>
                        <div class="item-body ${isOpen ? 'open' : ''}" id="${bodyId}" style="${isOpen ? 'display:block;' : ''}">
                            <div class="checklist-section">
                                <div class="section-h4">Actuator Valve Sub-Tasks</div>
                                <div class="checklist-grid">
                                    <div class="checklist-item ${item.general_inspection ? 'done' : ''}">
                                        <label>
                                            <input type="checkbox" id="gen-${item.equipment_id}" ${item.general_inspection ? 'checked' : ''} onchange="quickToggleActuator('${item.equipment_id}', 'general_inspection', this.checked)">
                                            <span>General Inspection, Cleaning & Calibration (50%)</span>
                                        </label>
                                    </div>
                                    <div class="checklist-item ${item.function_test ? 'done' : ''}">
                                        <label>
                                            <input type="checkbox" id="func-${item.equipment_id}" ${item.function_test ? 'checked' : ''} onchange="quickToggleActuator('${item.equipment_id}', 'function_test', this.checked)">
                                            <span>Function Test & Stroke Check (50%)</span>
                                        </label>
                                    </div>
                                </div>
                            </div>

                            <div class="form-grid">
                                <div class="form-group">
                                    <label>Person in Charge (PIC)</label>
                                    <select id="pic-act-${item.equipment_id}" class="filter-input">
                                        ${fullData.pics.map(p => `<option value="${p}" ${item.pic===p?'selected':''}>${p}</option>`).join('')}
                                    </select>
                                </div>
                                <div class="form-group">
                                    <label>Finish Date</label>
                                    <input type="date" id="finish-act-${item.equipment_id}" class="filter-input" value="${formatDateForInput(item.finish_date)}">
                                </div>
                                <div class="form-group">
                                    <label>Field Remarks / Notes</label>
                                    <input type="text" id="rem-act-${item.equipment_id}" class="filter-input" value="${item.remarks || ''}" placeholder="Remarks...">
                                </div>
                            </div>

                            <div style="display:flex; justify-content:space-between; align-items:center; margin-top:15px; border-top:1px solid var(--border-color); padding-top:12px;">
                                <div>
                                    <button class="btn-danger" onclick="deleteActuator('${item.equipment_id}')" style="display:inline-flex; align-items:center; gap:6px;">${Icons.trash} Delete Actuator</button>
                                </div>
                                <div style="display:flex; gap:10px;">
                                    <button class="btn-save" onclick="saveActuator('${item.equipment_id}', '${item.equipment_description}')" style="display:inline-flex; align-items:center; gap:6px;">${Icons.save} Save Actuator</button>
                                </div>
                            </div>
                        </div>
                    </div>`;
                });
                html += '</div>';
            } else {
                html += `
                <div class="table-wrap">
                    <table class="dense-table">
                        <thead>
                            <tr>
                                <th>Equipment ID</th>
                                <th>Actuator Description</th>
                                <th>KKS</th>
                                <th>Area</th>
                                <th>PIC</th>
                                <th>General Insp</th>
                                <th>Function Test</th>
                                <th>Status</th>
                                <th>Findings & Photos</th>
                            </tr>
                        </thead>
                        <tbody>
                            ${pageItems.map(item => {
                                const st = String(item.status || 'SCHED-OK').replace(/\s+/g, '_');
                                const hasFindings = !!item.temuan || (item.jumlah_foto > 0);
                                return `
                                <tr>
                                    <td style="font-family:'JetBrains Mono'; font-weight:700; color:var(--primary);">${item.equipment_id}</td>
                                    <td style="font-weight:600;">${item.equipment_description}</td>
                                    <td><span style="font-size:0.78rem; font-family:'JetBrains Mono'; color:var(--text-muted);">${item.kks || '-'}</span></td>
                                    <td><span style="font-size:0.8rem; color:var(--text-muted);">${item.area}</span></td>
                                    <td style="font-weight:600;">${item.pic || '-'}</td>
                                    <td>
                                        <input type="checkbox" ${item.general_inspection?'checked':''} onchange="quickToggleActuator('${item.equipment_id}', 'general_inspection', this.checked)" style="width:18px; height:18px; accent-color:var(--primary); cursor:pointer;">
                                    </td>
                                    <td>
                                        <input type="checkbox" ${item.function_test?'checked':''} onchange="quickToggleActuator('${item.equipment_id}', 'function_test', this.checked)" style="width:18px; height:18px; accent-color:var(--primary); cursor:pointer;">
                                    </td>
                                    <td><span class="status-badge badge-${st}">${item.status} (${item.persen_progress}%)</span></td>
                                    <td>
                                        <button class="btn-finding ${hasFindings?'active':''}" onclick="openFindingModal('actuator', '${item.equipment_id}', '${item.equipment_id} - ${item.equipment_description.replace(/'/g, "\\'")}', '${item.area}', '${(item.temuan||'').replace(/'/g, "\\'")}', '${(item.tindak_lanjut||'').replace(/'/g, "\\'")}')">
                                            ${Icons.camera} ${item.jumlah_foto > 0 ? item.jumlah_foto + ' Photos' : (hasFindings ? 'Findings' : '+ Photo')}
                                        </button>
                                    </td>
                                </tr>`;
                            }).join('')}
                        </tbody>
                    </table>
                </div>`;
            }

            html += renderPaginationControls(filteredItems.length);
            container.innerHTML = html;
        }

        /* ---------------- INSTRUMENTS RENDERING ---------------- */
        function switchInstSubtab(sub) {
            instSubtab = sub;
            currentPage = 1;
            renderTabContent();
        }

        function getSetPointDetails(item) {
            let rawSp = String(item.set_point || item.range || '').trim();
            let dir = '';
            
            // Check if explicit direction tag exists in rawSp (e.g. "8 Bar (HIGH)", "2.5 Kg/Cm2 (LOW)", "12.0 kg/cm2 HIGH")
            if(/\bHIGH\b/i.test(rawSp) || /\bHI\b/i.test(rawSp)) {
                dir = 'HIGH';
                rawSp = rawSp.replace(/\s*\(\s*HIGH\s*\)/i, '').replace(/\s*\bHIGH\b/i, '').trim();
            } else if(/\bLOW\b/i.test(rawSp) || /\bLO\b/i.test(rawSp)) {
                dir = 'LOW';
                rawSp = rawSp.replace(/\s*\(\s*LOW\s*\)/i, '').replace(/\s*\bLOW\b/i, '').trim();
            } else {
                // Infer from description / name if available
                const desc = (item.description || item.equipment || '').toUpperCase();
                if(desc.includes('HIGH TRIP') || desc.includes('PRESSURE OK') || desc.includes('PERMISSIVE') || desc.includes('ACRROS') || desc.includes('DP SWITCH') || desc.includes('BEFORE MTV')) {
                    dir = 'HIGH';
                } else if(desc.includes('LOW TRIP') || desc.includes('PRESSURE LOW') || desc.includes('LOW ALARM') || desc.includes('STANDBY PUMP') || desc.includes('INKET PRESSURE LOW') || desc.includes('INLET PRESSURE LOW')) {
                    dir = 'LOW';
                } else if((item.contact_type || '').toUpperCase() === 'NC') {
                    dir = 'LOW';
                } else {
                    dir = 'HIGH';
                }
            }
            return {
                value: rawSp || '-',
                dir: dir || 'HIGH'
            };
        }

        function updateAddInstFormFields(type) {
            const isPsw = type === 'pressure_switch' || type === 'psw';
            const subAreaGroup = document.getElementById('group-inst-subarea');
            const rangeGroup = document.getElementById('group-inst-range');
            const spGroup = document.getElementById('group-inst-sp');
            const dirGroup = document.getElementById('group-inst-dir');
            const contactGroup = document.getElementById('group-inst-contact');
            
            if(subAreaGroup) subAreaGroup.style.display = isPsw ? 'flex' : 'none';
            if(rangeGroup) rangeGroup.style.display = isPsw ? 'none' : 'flex';
            if(spGroup) spGroup.style.display = isPsw ? 'flex' : 'none';
            if(dirGroup) dirGroup.style.display = isPsw ? 'flex' : 'none';
            if(contactGroup) contactGroup.style.display = isPsw ? 'flex' : 'none';
        }

        function renderInstruments(container) {
            const searchStr = document.getElementById('search-input').value.toLowerCase();
            const statusFilter = document.getElementById('filter-status').value;
            const areaFilter = document.getElementById('filter-area').value;

            let items = [];
            if(instSubtab === 'ptx') items = fullData.pressure_tx || [];
            else if(instSubtab === 'ttx') items = fullData.temperature_tx || [];
            else if(instSubtab === 'psw') items = fullData.pressure_switch || [];

            let filteredItems = items.filter(itm => {
                const title = (itm.equipment || itm.description || '').toLowerCase();
                const kks = (itm.kks || '').toLowerCase();
                const area = (itm.area || '').toLowerCase();
                const subArea = (itm.sub_area || '').toLowerCase();
                const setPoint = (itm.set_point || itm.range || '').toLowerCase();
                const matchSearch = !searchStr || title.includes(searchStr) || kks.includes(searchStr) || area.includes(searchStr) || subArea.includes(searchStr) || setPoint.includes(searchStr);
                const matchArea = !areaFilter || itm.area === areaFilter;
                
                let matchStatus = true;
                if(statusFilter === 'FINISH') matchStatus = itm.verifikasi === true;
                else if(statusFilter === 'SCHED-OK' || statusFilter === 'IN PROGRESS') matchStatus = itm.verifikasi === false;

                let matchQuick = true;
                if(quickFilter === 'findings') matchQuick = !!itm.temuan || (itm.jumlah_foto > 0);
                else if(quickFilter === 'inprog') matchQuick = !itm.verifikasi;
                else if(quickFilter === 'finish') matchQuick = !!itm.verifikasi;

                return matchSearch && matchArea && matchStatus && matchQuick;
            });

            const isAddInstOpen = openCardIds.has('add-inst-form');
            let html = `
            <div style="margin-bottom:18px; background:var(--bg-card); border-radius:var(--radius-md); padding:16px; border:1px solid var(--border-color);">
                <div style="display:flex; justify-content:space-between; align-items:center; cursor:pointer;" onclick="toggleAccordion('add-inst-form')">
                    <h3 style="font-size:0.95rem; color:var(--primary); font-weight:700; display:inline-flex; align-items:center; gap:6px;">${Icons.plus} Add New Instrument</h3>
                    <span id="arrow-add-inst-form" style="font-weight:700; color:var(--text-muted); font-size:0.85rem;">${isAddInstOpen ? 'Close Form' : 'Add'}</span>
                </div>
                <div id="add-inst-form" class="accordion-form ${isAddInstOpen ? 'open' : ''}" style="${isAddInstOpen ? 'display:block;' : ''}">
                    <div class="form-grid">
                        <div class="form-group">
                            <label>Instrument Type <span style="color:#f43f5e;">*</span></label>
                            <select id="new-inst-type" class="filter-input" onchange="updateAddInstFormFields(this.value)">
                                <option value="pressure_tx" ${instSubtab==='ptx'?'selected':''}>Pressure Transmitter (PTX)</option>
                                <option value="temperature_tx" ${instSubtab==='ttx'?'selected':''}>Temperature Transmitter (TTX)</option>
                                <option value="pressure_switch" ${instSubtab==='psw'?'selected':''}>Pressure Switch (PSW)</option>
                            </select>
                        </div>
                        <div class="form-group">
                            <label>Equipment Name / Description <span style="color:#f43f5e;">*</span></label>
                            <input type="text" id="new-inst-desc" class="filter-input" placeholder="LDO PRESSURE BEFORE MTV, INLET ID FAN 1...">
                        </div>
                        <div class="form-group">
                            <label>KKS Tag</label>
                            <input type="text" id="new-inst-kks" class="filter-input" placeholder="10HJF11CP301 / 10HNA61CP001">
                        </div>
                        <div class="form-group">
                            <label>System Area</label>
                            <input type="text" id="new-inst-area" class="filter-input" placeholder="BOILER, ESP, TURBINE...">
                        </div>
                        <div class="form-group" id="group-inst-subarea" style="${instSubtab==='psw'?'':'display:none;'}">
                            <label>Sub-Area / System (e.g. SUB 1, Lower Burner, Pri Fan 1)</label>
                            <input type="text" id="new-inst-subarea" class="filter-input" placeholder="SUB 1, Lower Burner, Pri Fan 1, BFP 1...">
                        </div>
                        <div class="form-group" id="group-inst-range" style="${instSubtab==='psw'?'display:none;':''}">
                            <label>Transmitter Range</label>
                            <input type="text" id="new-inst-range" class="filter-input" placeholder="-70 - 70 mbar / 0 - 100 kPa">
                        </div>
                        <div class="form-group" id="group-inst-sp" style="${instSubtab==='psw'?'':'display:none;'}">
                            <label>Set Point Value</label>
                            <input type="text" id="new-inst-sp-val" class="filter-input" placeholder="e.g. 8 Bar, 2.5 Kg/Cm2, 0.4 Kg/Cm2">
                        </div>
                        <div class="form-group" id="group-inst-dir" style="${instSubtab==='psw'?'':'display:none;'}">
                            <label>Set Point Direction</label>
                            <select id="new-inst-sp-dir" class="filter-input">
                                <option value="HIGH">HIGH</option>
                                <option value="LOW">LOW</option>
                            </select>
                        </div>
                        <div class="form-group" id="group-inst-contact" style="${instSubtab==='psw'?'':'display:none;'}">
                            <label>Contact Type (NO / NC)</label>
                            <select id="new-inst-contact" class="filter-input">
                                <option value="NO">NO (Normally Open)</option>
                                <option value="NC">NC (Normally Closed)</option>
                            </select>
                        </div>
                    </div>
                    <div style="display:flex; gap:10px; margin-top:14px;">
                        <button class="btn-save" onclick="saveNewInstrument()" style="display:inline-flex; align-items:center; gap:6px;">${Icons.save} Save New Instrument</button>
                        <button class="page-btn" onclick="toggleAccordion('add-inst-form')">Cancel</button>
                    </div>
                </div>
            </div>

            <!-- Instrument Sub-tabs -->
            <div class="nav-tabs" style="margin-bottom:16px;">
                <button class="tab-btn ${instSubtab==='ptx'?'active':''}" onclick="switchInstSubtab('ptx')">Pressure Transmitter (${(fullData.pressure_tx||[]).length})</button>
                <button class="tab-btn ${instSubtab==='ttx'?'active':''}" onclick="switchInstSubtab('ttx')">Temperature Transmitter (${(fullData.temperature_tx||[]).length})</button>
                <button class="tab-btn ${instSubtab==='psw'?'active':''}" onclick="switchInstSubtab('psw')">Pressure Switch (${(fullData.pressure_switch||[]).length})</button>
            </div>`;

            updateSidePaginationCounter(filteredItems.length);

            if(filteredItems.length === 0) {
                html += '<div style="text-align:center; padding:50px; color:var(--text-muted); background:var(--bg-card); border-radius:12px;">No instruments match the current filter criteria.</div>';
                container.innerHTML = html;
                return;
            }

            const startIndex = (currentPage - 1) * pageSize;
            const pageItems = filteredItems.slice(startIndex, startIndex + pageSize);

            if(currentViewMode === 'cards') {
                html += '<div class="card-list">';
                pageItems.forEach((item, idx) => {
                    const title = (instSubtab==='psw' ? item.description : item.equipment) || `Item #${item.no}`;
                    const isVerif = !!item.verifikasi;
                    const isCalib = !!item.kalibrasi;
                    const hasFindings = !!item.temuan || (item.jumlah_foto > 0);
                    const cardPrefix = instSubtab === 'psw' ? 'psw' : 'inst';
                    const bodyId = getCardBodyId(cardPrefix, item.kks || item.no || idx);
                    const isOpen = openCardIds.has(bodyId);
                    const sp = instSubtab === 'psw' ? getSetPointDetails(item) : null;
                    const contact = (item.contact_type || 'NO').toUpperCase();

                    html += `
                    <div class="item-card">
                        <div class="item-header" onclick="toggleAccordion('${bodyId}')">
                            <div style="display:flex; align-items:center; gap:12px; flex-grow:1;">
                                <div class="item-title-box">
                                    <div class="item-code" style="display:flex; align-items:center; gap:8px; flex-wrap:wrap;">
                                        <span style="font-family:'JetBrains Mono'; font-weight:700; color:var(--primary);">${item.kks || 'No Tag'}</span>
                                        <span>&bull; ${instSubtab === 'psw' ? (item.sub_area || item.area) : item.area}</span>
                                        ${(item.finish_date || item.dated || item.tanggal) ? '<span>&bull; Finished: ' + (item.finish_date || item.dated || item.tanggal) + '</span>' : ''}
                                        ${instSubtab === 'psw' ? `
                                            <div class="setpoint-cell" style="padding:1px 6px; min-width:auto; margin-left:2px;">
                                                <span class="setpoint-val" style="font-size:0.75rem;">${sp.value}</span>
                                                <div class="setpoint-divider"></div>
                                                <span class="setpoint-dir ${sp.dir.toLowerCase()}" style="font-size:0.62rem;">${sp.dir}</span>
                                            </div>
                                            <span class="contact-badge ${contact==='NC'?'contact-nc':'contact-no'}" style="padding:1px 7px; font-size:0.72rem; min-width:auto;">${contact}</span>
                                        ` : ''}
                                    </div>
                                    <div class="item-name">${title} ${instSubtab!=='psw' && item.range ? ' (Range: ' + item.range + ')' : ''}</div>
                                </div>
                            </div>
                            <div class="header-actions">
                                <button class="btn-finding ${hasFindings?'active':''}" onclick="event.stopPropagation(); openFindingModal('instrument', '${item.kks || item.no}', '${item.kks} - ${title.replace(/'/g, "\\'")}', '${item.area}', '${(item.temuan||'').replace(/'/g, "\\'")}', '${(item.tindak_lanjut||'').replace(/'/g, "\\'")}', '${instSubtab==='psw'?'pressure_switch':(instSubtab==='ptx'?'pressure_tx':'temperature_tx')}')">
                                    ${Icons.camera} ${item.jumlah_foto > 0 ? item.jumlah_foto + ' Photos' : (hasFindings ? 'Findings' : '+ Finding')}
                                </button>
                                <span class="status-badge ${isVerif ? 'badge-FINISH' : (isCalib ? 'badge-IN-PROGRESS' : 'badge-SCHED-OK')}" id="badge-${cardPrefix}-${idx}">${isVerif ? 'DONE (100%)' : (isCalib ? 'IN PROGRESS (Calibration OK)' : 'SCHEDULED')}</span>
                            </div>
                        </div>
                        <div class="item-body ${isOpen ? 'open' : ''}" id="${bodyId}" style="${isOpen ? 'display:block;' : ''}">
                            <div class="checklist-section" style="margin-bottom:14px;">
                                <div class="section-h4">Instrument Milestone Checklist (Verification determines Finish)</div>
                                <div class="checklist-grid">
                                    <div class="checklist-item ${isCalib ? 'done' : ''}" id="card-calib-${cardPrefix}-${idx}">
                                        <label style="cursor:pointer; display:flex; align-items:center; gap:8px;">
                                             <input type="checkbox" id="inst-calib-${idx}" ${isCalib ? 'checked' : ''} onchange="toggleLocalInstCheck('${instSubtab}', ${idx}, 'kalibrasi', this.checked)">
                                             <span style="font-weight:700;">1. Calibration Done</span>
                                        </label>
                                    </div>
                                    <div class="checklist-item ${isVerif ? 'done' : ''}" id="card-verif-${cardPrefix}-${idx}" style="${isVerif ? 'border-color:var(--status-finish);' : ''}">
                                        <label style="cursor:pointer; display:flex; align-items:center; gap:8px;">
                                             <input type="checkbox" id="inst-verif-${idx}" ${isVerif ? 'checked' : ''} onchange="toggleLocalInstCheck('${instSubtab}', ${idx}, 'verifikasi', this.checked)">
                                             <span style="font-weight:800; color:${isVerif ? 'var(--status-finish)' : 'var(--primary)'};">2. Verification Finished (100% DONE)</span>
                                        </label>
                                    </div>
                                </div>
                            </div>
                            ${instSubtab === 'psw' ? `
                            <div class="form-grid" style="margin-bottom:14px; background:var(--bg-sub); padding:12px; border-radius:var(--radius-sm); border:1px solid var(--border-color);">
                                <div class="form-group">
                                    <label>Sub-Area / System</label>
                                    <input type="text" id="subarea-psw-${idx}" class="filter-input" value="${item.sub_area || item.area || ''}" placeholder="SUB 1, Lower Burner, Pri Fan 1...">
                                </div>
                                <div class="form-group">
                                    <label>Set Point (Value)</label>
                                    <input type="text" id="sp-val-psw-${idx}" class="filter-input" value="${sp.value}" placeholder="e.g. 8 Bar, 2.5 Kg/Cm2">
                                </div>
                                <div class="form-group">
                                    <label>Set Point Direction</label>
                                    <select id="sp-dir-psw-${idx}" class="filter-input">
                                        <option value="HIGH" ${sp.dir==='HIGH'?'selected':''}>HIGH</option>
                                        <option value="LOW" ${sp.dir==='LOW'?'selected':''}>LOW</option>
                                    </select>
                                </div>
                                <div class="form-group">
                                    <label>Contact Type (NO / NC)</label>
                                    <select id="contact-psw-${idx}" class="filter-input">
                                        <option value="NO" ${contact==='NO'?'selected':''}>NO (Normally Open)</option>
                                        <option value="NC" ${contact==='NC'?'selected':''}>NC (Normally Closed)</option>
                                    </select>
                                </div>
                            </div>
                            <div class="calib-grid">
                                <div class="calib-col">
                                    <h5>AS FOUND (Initial Condition)</h5>
                                    <div class="calib-fields">
                                        <div class="form-group">
                                            <label>Set Point</label>
                                            <input type="text" id="af-set-${idx}" class="filter-input" value="${item.asfound_set || ''}" placeholder="e.g. 7.8 Bar">
                                        </div>
                                        <div class="form-group">
                                            <label>Reset Point</label>
                                            <input type="text" id="af-reset-${idx}" class="filter-input" value="${item.asfound_reset || ''}" placeholder="e.g. 7.2 Bar">
                                        </div>
                                    </div>
                                </div>
                                <div class="calib-col">
                                    <h5>AS LEFT (After Calibration)</h5>
                                    <div class="calib-fields">
                                        <div class="form-group">
                                            <label>Set Point</label>
                                            <input type="text" id="al-set-${idx}" class="filter-input" value="${item.asleft_set || ''}" placeholder="e.g. 8.0 Bar">
                                        </div>
                                        <div class="form-group">
                                            <label>Reset Point</label>
                                            <input type="text" id="al-reset-${idx}" class="filter-input" value="${item.asleft_reset || ''}" placeholder="e.g. 7.5 Bar">
                                        </div>
                                    </div>
                                </div>
                            </div>
                            <div class="form-grid">
                                <div class="form-group">
                                    <label>Calibration Result</label>
                                    <select id="res-psw-${idx}" class="filter-input">
                                        <option value="OK" ${item.status_ok_notok==='OK'?'selected':''}>OK / Within Tolerance</option>
                                        <option value="NOT OK" ${item.status_ok_notok==='NOT OK'?'selected':''}>NOT OK / Deviation</option>
                                    </select>
                                </div>` : `<div class="form-grid">`}
                                <div class="form-group" style="grid-column: span 2;">
                                    <label>Field Remarks / Notes</label>
                                    <input type="text" id="inst-rem-${idx}" class="filter-input" value="${item.remarks || ''}" placeholder="Calibration / verification remarks...">
                                </div>
                            </div>

                            <div style="display:flex; justify-content:space-between; align-items:center; margin-top:15px; border-top:1px solid var(--border-color); padding-top:12px;">
                                <div>
                                    <button class="btn-danger" onclick="deleteInstrument('${instSubtab}', '${item.kks || item.no}')" style="display:inline-flex; align-items:center; gap:6px;">${Icons.trash} Delete Instrument</button>
                                </div>
                                <div style="display:flex; gap:10px;">
                                    <button class="btn-save" onclick="${instSubtab==='psw' ? 'savePressureSwitch' : 'saveTransmitter'}('${instSubtab}', '${item.kks || item.no}', ${idx})" style="display:inline-flex; align-items:center; gap:6px;">${Icons.save} Save Calibration & Verification</button>
                                </div>
                            </div>
                        </div>
                    </div>`;
                });
                html += '</div>';
            } else {
                if(instSubtab === 'psw') {
                    html += `
                    <div class="table-wrap">
                        <table class="dense-table">
                            <thead>
                                <tr>
                                    <th>KKS Tag</th>
                                    <th>Equipment Description</th>
                                    <th>Area / Sub-Area</th>
                                    <th style="text-align:center;">Set Point</th>
                                    <th style="text-align:center;">Contact (NO/NC)</th>
                                    <th style="text-align:center;">Calibration</th>
                                    <th style="text-align:center;">Verification</th>
                                    <th>Finish Status</th>
                                    <th>Findings & Photos</th>
                                </tr>
                            </thead>
                            <tbody>
                                ${pageItems.map((item, idx) => {
                                    const title = item.description || item.equipment || `Item #${item.no}`;
                                    const isVerif = !!item.verifikasi;
                                    const isCalib = !!item.kalibrasi;
                                    const hasFindings = !!item.temuan || (item.jumlah_foto > 0);
                                    const sp = getSetPointDetails(item);
                                    const contact = (item.contact_type || 'NO').toUpperCase();
                                    const areaDisplay = item.sub_area ? `${item.sub_area} <span style="color:var(--text-muted); font-size:0.75rem;">(${item.area || 'BOILER'})</span>` : (item.area || '-');
                                    return `
                                    <tr>
                                        <td style="font-family:'JetBrains Mono'; font-weight:700; color:var(--primary);">${item.kks || '-'}</td>
                                        <td style="font-weight:600;">${title}</td>
                                        <td><span style="font-size:0.82rem; font-weight:600; color:var(--text-main);">${areaDisplay}</span></td>
                                        <td style="text-align:center;">
                                            <div class="setpoint-cell">
                                                <span class="setpoint-val">${sp.value}</span>
                                                <div class="setpoint-divider"></div>
                                                <span class="setpoint-dir ${sp.dir.toLowerCase()}">${sp.dir}</span>
                                            </div>
                                        </td>
                                        <td style="text-align:center;">
                                            <span class="contact-badge ${contact==='NC'?'contact-nc':'contact-no'}">${contact}</span>
                                        </td>
                                        <td style="text-align:center;">
                                            <input type="checkbox" ${isCalib?'checked':''} onchange="toggleDirectInstCheck('psw', '${item.kks || item.no}', 'kalibrasi', this.checked)" style="width:18px; height:18px; accent-color:var(--primary); cursor:pointer;">
                                        </td>
                                        <td style="text-align:center;">
                                            <input type="checkbox" ${isVerif?'checked':''} onchange="toggleDirectInstCheck('psw', '${item.kks || item.no}', 'verifikasi', this.checked)" style="width:18px; height:18px; accent-color:var(--status-finish); cursor:pointer;">
                                        </td>
                                        <td>
                                            <span class="status-badge ${isVerif?'badge-FINISH':(isCalib?'badge-IN-PROGRESS':'badge-SCHED-OK')}">${isVerif?'DONE (100%)':(isCalib?'IN PROGRESS':'SCHEDULED')}</span>
                                        </td>
                                        <td>
                                            <button class="btn-finding ${hasFindings?'active':''}" onclick="openFindingModal('instrument', '${item.kks || item.no}', '${item.kks} - ${title.replace(/'/g, "\\'")}', '${item.area}', '${(item.temuan||'').replace(/'/g, "\\'")}', '${(item.tindak_lanjut||'').replace(/'/g, "\\'")}', 'pressure_switch')">
                                                ${Icons.camera} ${item.jumlah_foto > 0 ? item.jumlah_foto + ' Photos' : (hasFindings ? 'Findings' : '+ Photo')}
                                            </button>
                                        </td>
                                    </tr>`;
                                }).join('')}
                            </tbody>
                        </table>
                    </div>`;
                } else {
                    html += `
                    <div class="table-wrap">
                        <table class="dense-table">
                            <thead>
                                <tr>
                                    <th>KKS Tag</th>
                                    <th>Equipment Description</th>
                                    <th>Area</th>
                                    <th>Range / Units</th>
                                    <th style="text-align:center;">Calibration</th>
                                    <th style="text-align:center;">Verification</th>
                                    <th>Finish Status</th>
                                    <th>Findings & Photos</th>
                                </tr>
                            </thead>
                            <tbody>
                                ${pageItems.map((item, idx) => {
                                    const title = item.equipment || `Item #${item.no}`;
                                    const isVerif = !!item.verifikasi;
                                    const isCalib = !!item.kalibrasi;
                                    const hasFindings = !!item.temuan || (item.jumlah_foto > 0);
                                    return `
                                    <tr>
                                        <td style="font-family:'JetBrains Mono'; font-weight:700; color:var(--primary);">${item.kks || '-'}</td>
                                        <td style="font-weight:600;">${title}</td>
                                        <td><span style="font-size:0.8rem; color:var(--text-muted);">${item.area}</span></td>
                                        <td><span style="font-size:0.82rem; font-family:'JetBrains Mono';">${item.range || '-'}</span></td>
                                        <td style="text-align:center;">
                                            <input type="checkbox" ${isCalib?'checked':''} onchange="toggleDirectInstCheck('${instSubtab}', '${item.kks || item.no}', 'kalibrasi', this.checked)" style="width:18px; height:18px; accent-color:var(--primary); cursor:pointer;">
                                        </td>
                                        <td style="text-align:center;">
                                            <input type="checkbox" ${isVerif?'checked':''} onchange="toggleDirectInstCheck('${instSubtab}', '${item.kks || item.no}', 'verifikasi', this.checked)" style="width:18px; height:18px; accent-color:var(--status-finish); cursor:pointer;">
                                        </td>
                                        <td>
                                            <span class="status-badge ${isVerif?'badge-FINISH':(isCalib?'badge-IN-PROGRESS':'badge-SCHED-OK')}">${isVerif?'DONE (100%)':(isCalib?'IN PROGRESS':'SCHEDULED')}</span>
                                        </td>
                                        <td>
                                            <button class="btn-finding ${hasFindings?'active':''}" onclick="openFindingModal('instrument', '${item.kks || item.no}', '${item.kks} - ${title.replace(/'/g, "\\'")}', '${item.area}', '${(item.temuan||'').replace(/'/g, "\\'")}', '${(item.tindak_lanjut||'').replace(/'/g, "\\'")}', '${instSubtab==='ptx'?'pressure_tx':'temperature_tx'}')">
                                                ${Icons.camera} ${item.jumlah_foto > 0 ? item.jumlah_foto + ' Photos' : (hasFindings ? 'Findings' : '+ Photo')}
                                            </button>
                                        </td>
                                    </tr>`;
                                }).join('')}
                            </tbody>
                        </table>
                    </div>`;
                }
            }

            html += renderPaginationControls(filteredItems.length);
            container.innerHTML = html;
        }

        /* ---------------- SCOPE & PIC MASTER RENDERING ---------------- */
        function renderScopeMaster(container) {
            let html = `
            <div style="display:flex; flex-direction:column; gap:16px;">
                <!-- Master Protection & Security Status Banner -->
                <div class="scope-lock-banner ${isScopeMasterUnlocked ? 'unlocked' : 'locked'}">
                    <div style="display:flex; align-items:center; gap:12px;">
                        <div class="lock-icon-circle ${isScopeMasterUnlocked ? 'open' : ''}">${isScopeMasterUnlocked ? Icons.unlock : Icons.lock}</div>
                        <div>
                            <div style="font-size:0.94rem; font-weight:800; color:${isScopeMasterUnlocked ? '#10b981' : 'var(--text-main)'};">
                                ${isScopeMasterUnlocked ? 'Master Edit Mode Unlocked (Full Access)' : 'Master EIC Protection Mode (Protected)'}
                            </div>
                            <div style="font-size:0.8rem; color:var(--text-muted); margin-top:2px;">
                                ${isScopeMasterUnlocked ? 'You can add, edit, and delete Master PIC personnel and Outage Job Scope definitions.' : 'Master PIC personnel and Master Scope data are protected. Enter authorization password to unlock edit mode.'}
                            </div>
                        </div>
                    </div>
                    <div>
                        ${isScopeMasterUnlocked ? `
                            <button class="page-btn" style="padding:7px 16px; font-size:0.82rem; color:#f43f5e; border-color:rgba(244,63,94,0.4); font-weight:700;" onclick="lockMasterScope()" title="Lock edit access">
                                ${Icons.lock} Lock Mode
                            </button>
                        ` : `
                            <button class="btn-save" style="padding:8px 18px; font-size:0.84rem; background:linear-gradient(135deg, #6366f1, #4f46e5); box-shadow:0 4px 12px rgba(99,102,241,0.3);" onclick="openMasterPasswordModal()" title="Enter authorization password to unlock edit access">
                                ${Icons.unlock} Unlock Edit Mode
                            </button>
                        `}
                    </div>
                </div>

                <!-- Master PIC Card -->
                <div style="background:var(--bg-card); border-radius:var(--radius-lg); padding:22px; border:1px solid var(--border-color);">
                    <div style="margin-bottom:8px;">
                        <h3 style="color:var(--primary); font-size:1.1rem; font-weight:800; margin:0; display:inline-flex; align-items:center; gap:8px;">${Icons.users} Master EIC Team PICs</h3>
                    </div>
                    <p style="font-size:0.82rem; color:var(--text-muted); margin-bottom:16px;">Registered EIC responsible personnel. All PIC dropdown selectors across Work Orders, Actuators, & Instruments automatically sync from this master list.</p>
                    
                    <div style="display:flex; flex-wrap:wrap; gap:8px; margin-bottom:${isScopeMasterUnlocked ? '18px' : '4px'};">
                        ${(fullData.pics || []).map(p => `
                            <span style="padding:6px 14px; background:var(--bg-sub); border:1px solid var(--border-color); border-radius:20px; font-size:0.85rem; font-weight:600; color:var(--text-main); display:inline-flex; align-items:center; gap:8px; box-shadow:0 2px 5px rgba(0,0,0,0.05);">
                                ${p}
                                ${isScopeMasterUnlocked ? `
                                    <button style="background:none; border:none; color:#f43f5e; cursor:pointer; font-weight:700; font-size:0.85rem; padding:0 2px;" title="Delete PIC" onclick="deletePic('${(p||'').toString().replace(/'/g, "\\'")}')">&times;</button>
                                ` : ''}
                            </span>
                        `).join('')}
                    </div>

                    ${isScopeMasterUnlocked ? `
                        <div style="display:flex; gap:10px; max-width:480px;">
                            <input type="text" id="new-pic-input" class="filter-input" placeholder="Enter new PIC personnel / vendor name..." style="flex-grow:1;">
                            <button class="btn-save" onclick="addNewPic()" style="display:inline-flex; align-items:center; gap:6px;">${Icons.plus} Add PIC</button>
                        </div>
                    ` : ''}
                </div>

                <!-- Scope Master Table -->
                <div style="background:var(--bg-card); border-radius:var(--radius-lg); padding:22px; border:1px solid var(--border-color);">
                    <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:14px; flex-wrap:wrap; gap:10px;">
                        <h3 style="color:var(--primary); font-size:1.1rem; font-weight:800; margin:0; display:inline-flex; align-items:center; gap:8px;">${Icons.clipboard} Master Outage Job Scope</h3>
                        ${isScopeMasterUnlocked ? `
                            <button class="page-btn" style="font-size:0.82rem; font-weight:700; display:inline-flex; align-items:center; gap:6px;" onclick="toggleAccordion('add-scope-form')">${openCardIds.has('add-scope-form') ? 'Close Scope Form' : Icons.plus + ' Add Scope'}</button>
                        ` : ''}
                    </div>
                    
                    ${isScopeMasterUnlocked ? `
                        <div id="add-scope-form" class="accordion-form ${openCardIds.has('add-scope-form') ? 'open' : ''}" style="${openCardIds.has('add-scope-form') ? 'display:block;' : ''}; margin-bottom:18px;">
                            <div class="form-grid">
                                <div class="form-group">
                                    <label>Scope Category</label>
                                    <input type="text" id="new-scope-cat" class="filter-input" placeholder="BOILER, TURBINE...">
                                </div>
                                <div class="form-group">
                                    <label>Equipment / Job Scope <span style="color:#f43f5e;">*</span></label>
                                    <input type="text" id="new-scope-eq" class="filter-input" placeholder="Burner Inspection...">
                                </div>
                                <div class="form-group">
                                    <label>Scope Type</label>
                                    <select id="new-scope-type" class="filter-input">
                                        <option value="MSW">MSW</option>
                                        <option value="Vendor">Vendor</option>
                                        <option value="Internal">Internal</option>
                                    </select>
                                </div>
                                <div class="form-group">
                                    <label>Person in Charge (PIC)</label>
                                    <select id="new-scope-pic" class="filter-input">
                                        <option value="">Select PIC...</option>
                                        ${(fullData.pics || []).map(p => `<option value="${p}">${p}</option>`).join('')}
                                    </select>
                                </div>
                            </div>
                            <div style="display:flex; gap:10px; margin-top:12px;">
                                <button class="btn-save" onclick="saveNewScope()" style="display:inline-flex; align-items:center; gap:6px;">${Icons.save} Save New Scope</button>
                                <button class="page-btn" onclick="toggleAccordion('add-scope-form')">Cancel</button>
                            </div>
                        </div>
                    ` : ''}

                    <div class="table-wrap">
                        <table class="dense-table">
                            <thead>
                                <tr>
                                    <th>Scope Category</th>
                                    <th>Equipment / Job Scope</th>
                                    <th>Scope Type</th>
                                    <th>Person in Charge (PIC)</th>
                                    ${isScopeMasterUnlocked ? '<th>Actions</th>' : ''}
                                </tr>
                            </thead>
                            <tbody>
                                ${(fullData.scope_master || []).map((s, sIdx) => `
                                    <tr>
                                        <td style="font-size:0.8rem; color:var(--text-muted); font-weight:600;">${s.kategori || '-'}</td>
                                        <td>
                                            ${isScopeMasterUnlocked ? `
                                                <input type="text" id="scope-eq-${sIdx}" class="filter-input" value="${(s.nama_equipment || '').replace(/"/g, '&quot;')}" style="padding:5px 8px; font-size:0.85rem; width:100%; min-width:200px;" onblur="saveScopeRow(${sIdx})">
                                            ` : `
                                                <div style="font-weight:600; font-size:0.88rem; color:var(--text-main);">${s.nama_equipment || '-'}</div>
                                            `}
                                        </td>
                                        <td>
                                            ${isScopeMasterUnlocked ? `
                                                <select id="scope-type-${sIdx}" class="filter-input" style="padding:5px 8px; font-size:0.85rem;" onchange="saveScopeRow(${sIdx})">
                                                    <option value="Vendor" ${s.tipe_scope==='Vendor'?'selected':''}>Vendor</option>
                                                    <option value="MSW" ${s.tipe_scope==='MSW'?'selected':''}>MSW</option>
                                                    <option value="Internal" ${s.tipe_scope==='Internal'?'selected':''}>Internal</option>
                                                </select>
                                            ` : `
                                                <span class="scope-type-badge">${s.tipe_scope || 'MSW'}</span>
                                            `}
                                        </td>
                                        <td>
                                            ${isScopeMasterUnlocked ? `
                                                <select id="scope-pic-${sIdx}" class="filter-input" style="padding:5px 8px; font-size:0.85rem;" onchange="saveScopeRow(${sIdx})">
                                                    <option value="">Select PIC...</option>
                                                    ${(fullData.pics || []).map(p => {
                                                        const isSel = (s.pic === p || (s.pic||'').toUpperCase() === (p||'').toUpperCase());
                                                        return `<option value="${p}" ${isSel ? 'selected' : ''}>${p}</option>`;
                                                    }).join('')}
                                                </select>
                                            ` : `
                                                <span style="font-size:0.85rem; font-weight:600; color:var(--primary);">${s.pic || '-'}</span>
                                            `}
                                        </td>
                                        ${isScopeMasterUnlocked ? `
                                            <td>
                                                <div style="display:flex; gap:6px;">
                                                    <button class="btn-save" style="padding:5px 12px; font-size:0.78rem; display:inline-flex; align-items:center; gap:4px;" onclick="saveScopeRow(${sIdx})">${Icons.save} Save</button>
                                                    <button class="btn-danger" style="padding:5px 10px; font-size:0.78rem; display:inline-flex; align-items:center; gap:4px;" onclick="deleteScopeRow(${sIdx})">${Icons.trash} Delete</button>
                                                </div>
                                            </td>
                                        ` : ''}
                                    </tr>
                                `).join('')}
                            </tbody>
                        </table>
                    </div>
                </div>
            </div>`;
            container.innerHTML = html;
        }

        function renderTabContent() {
            const container = document.getElementById('tab-content');
            if(!fullData) return;

            if (currentTab === 'wo') renderWorkOrders(container);
            else if (currentTab === 'actuator') renderActuators(container);
            else if (currentTab === 'instrument') renderInstruments(container);
            else if (currentTab === 'scope') renderScopeMaster(container);
        }

        const openCardIds = new Set();
        function getCardBodyId(prefix, key) {
            return 'body-' + prefix + '-' + String(key).replace(/[^a-zA-Z0-9_-]/g, '_');
        }

        function toggleAccordion(id) {
            const el = document.getElementById(id);
            if(!el) return;
            const wasOpen = openCardIds.has(id) || el.classList.contains('open') || (el.style.display === 'block');
            const arrow = document.getElementById(`arrow-${id}`);
            if (wasOpen) {
                openCardIds.delete(id);
                el.classList.remove('open');
                el.style.display = 'none';
                if(arrow) {
                    if(id === 'add-scope-form') arrow.innerText = 'Add Scope';
                    else arrow.innerText = 'Add';
                }
            } else {
                openCardIds.add(id);
                el.classList.add('open');
                el.style.display = 'block';
                if(arrow) {
                    if(id === 'add-scope-form') arrow.innerText = 'Close Scope';
                    else arrow.innerText = 'Close Form';
                }
            }
        }

        /* ---------------- FINDING & PHOTO MODAL ---------------- */
        async function openFindingModal(itemType, id, title, area, temuan, tindakLanjut, instType = '') {
            activeFinding = { itemType, id, title, area, temuan, tindakLanjut, instType };
            document.getElementById('modal-finding-title').innerText = `Findings & Photos: ${id}`;
            document.getElementById('modal-finding-subtitle').innerText = `${title} (${area || 'GENERAL'})`;
            document.getElementById('modal-finding-text').value = temuan || '';
            document.getElementById('modal-tl-text').value = tindakLanjut || '';

            try {
                const res = await fetch(`/api/findings?id=${encodeURIComponent(id)}`);
                const data = await res.json();
                renderModalPhotos(data.photos || []);
                if(data.temuan) document.getElementById('modal-finding-text').value = data.temuan;
                if(data.tindak_lanjut) document.getElementById('modal-tl-text').value = data.tindak_lanjut;
            } catch(e) {
                renderModalPhotos([]);
            }

            document.getElementById('finding-modal').classList.add('open');
        }

        function closeFindingModal() {
            document.getElementById('finding-modal').classList.remove('open');
            activeFinding = null;
        }

        function renderModalPhotos(photos) {
            document.getElementById('modal-photo-count').innerText = photos.length;
            const grid = document.getElementById('modal-photo-grid');
            if(photos.length === 0) {
                grid.innerHTML = '<div style="font-size:0.8rem; color:var(--text-muted); grid-column:span 4; padding:8px 0;">No photos uploaded yet.</div>';
                return;
            }
            grid.innerHTML = photos.map(p => `
                <div class="photo-thumb-box">
                    <img src="${p.url}" alt="${p.filename}" onclick="openLightbox('${p.url}')" title="Click to enlarge">
                    <button class="photo-delete-btn" onclick="deleteModalPhoto('${p.filename}')" title="Delete this photo"><svg class="ui-icon sm" viewBox="0 0 24 24"><polyline points="3 6 5 6 21 6"></polyline><path d="M19 6v14a2 2 0 0 1-2 2H7a2 2 0 0 1-2-2V6m3 0V4a2 2 0 0 1 2-2h4a2 2 0 0 1 2 2v2"></path></svg></button>
                </div>
            `).join('');
        }

        function openLightbox(url) {
            document.getElementById('lightbox-img').src = url;
            document.getElementById('lightbox-modal').classList.add('open');
        }

        function closeLightbox() {
            document.getElementById('lightbox-modal').classList.remove('open');
        }

        async function handleModalFileSelect(files) {
            if(!activeFinding || !files || files.length === 0) return;

            for(let file of files) {
                const reader = new FileReader();
                reader.onload = async function(e) {
                    const base64Data = e.target.result;
                    try {
                        const payload = {
                            id: activeFinding.id,
                            unit: currentUnit,
                            type: activeFinding.itemType,
                            inst_type: activeFinding.instType,
                            filename: file.name,
                            image_base64: base64Data,
                            temuan: document.getElementById('modal-finding-text').value,
                            tindak_lanjut: document.getElementById('modal-tl-text').value
                        };
                        const res = await fetch('/api/upload_finding_photo', {
                            method: 'POST',
                            headers: {'Content-Type': 'application/json'},
                            body: JSON.stringify(payload)
                        });
                        const result = await res.json();
                        if(result.status === 'success') {
                            showToast(`Photo '${file.name}' saved successfully!`, 'success');
                            renderModalPhotos(result.photos || []);
                            loadData();
                        } else {
                            showToast(result.message || 'Failed to save photo', 'error');
                        }
                    } catch(err) {
                        showToast('Error uploading photo: ' + err.message, 'error');
                    }
                };
                reader.readAsDataURL(file);
            }
        }

        async function deleteModalPhoto(filename) {
            if(!activeFinding || !filename) return;
            if(!confirm(`Delete photo ${filename}?`)) return;

            try {
                const payload = {
                    id: activeFinding.id,
                    unit: currentUnit,
                    filename: filename,
                    type: activeFinding.itemType,
                    inst_type: activeFinding.instType
                };
                const res = await fetch('/api/delete_finding_photo', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify(payload)
                });
                const result = await res.json();
                if(result.status === 'success') {
                    showToast('Photo deleted successfully', 'info');
                    renderModalPhotos(result.photos || []);
                    loadData();
                }
            } catch(e) {
                showToast('Failed to delete photo', 'error');
            }
        }

        async function saveFindingModalData() {
            if(!activeFinding) return;
            const temuan = document.getElementById('modal-finding-text').value;
            const tindakLanjut = document.getElementById('modal-tl-text').value;

            try {
                const payload = {
                    id: activeFinding.id,
                    unit: currentUnit,
                    type: activeFinding.itemType,
                    inst_type: activeFinding.instType,
                    temuan: temuan,
                    tindak_lanjut: tindakLanjut
                };
                const res = await fetch('/api/upload_finding_photo', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify(payload)
                });
                const result = await res.json();
                showToast('Field findings and recommendations saved successfully!', 'success');
                closeFindingModal();
                loadData();
            } catch(e) {
                showToast('Failed to save findings', 'error');
            }
        }

        /* ---------------- DEDICATED REPORT & PRINT ---------------- */
        function openReportModal() {
            const today = new Date();
            const yesterday = new Date(Date.now() - 86400000);
            
            const startInput = document.getElementById('report-start-date');
            const endInput = document.getElementById('report-end-date');
            if(startInput && !startInput.value) startInput.value = formatDateForInput(yesterday);
            if(endInput && !endInput.value) endInput.value = formatDateForInput(today);
            
            generateReportContent();
            document.getElementById('report-modal').classList.add('open');
        }

        function closeReportModal() {
            document.getElementById('report-modal').classList.remove('open');
        }

        function printReportModal() {
            window.print();
        }

        function parseDateStrToTime(str) {
            if(!str) return null;
            str = String(str).trim();
            if (/^\d{4}-\d{2}-\d{2}/.test(str)) {
                const p = str.split('-');
                return new Date(parseInt(p[0]), parseInt(p[1]) - 1, parseInt(p[2])).getTime();
            }
            if (/^\d{1,2}\/\d{1,2}\/\d{4}/.test(str)) {
                const parts = str.split('/');
                return new Date(parseInt(parts[2]), parseInt(parts[1]) - 1, parseInt(parts[0])).getTime();
            }
            const d = new Date(str);
            return isNaN(d.getTime()) ? null : d.getTime();
        }

        function isDateInRange(dateStr, startYMD, endYMD) {
            if(!startYMD && !endYMD) return true;
            if(!dateStr) return false;
            const itemTime = parseDateStrToTime(dateStr);
            if(!itemTime) return false;
            
            let startTime = startYMD ? parseDateStrToTime(startYMD) : null;
            let endTime = endYMD ? parseDateStrToTime(endYMD) : null;
            
            if(startTime && itemTime < startTime) return false;
            if(endTime && itemTime > endTime + (24 * 60 * 60 * 1000 - 1)) return false;
            return true;
        }

        let currentReportType = 'harian';

        function generateSCurveHTMLForReport() {
            if(!fullData) return '';

            const dateMap = {};
            let totalTasks = 0;
            let undatedCount = 0;

            (fullData.work_orders || []).forEach(w => {
                const chk = w.checklist || [];
                if(chk.length > 0) {
                    totalTasks += chk.length;
                    chk.forEach(c => {
                        if(c.selesai) {
                            const dStr = c.tanggal || w.tanggal_finish || w.tanggal_actual_start || w.tanggal_schedule;
                            const parsed = extractValidDate(dStr);
                            if(parsed && parsed.ymd) {
                                dateMap[parsed.ymd] = (dateMap[parsed.ymd] || 0) + 1;
                            } else {
                                undatedCount += 1;
                            }
                        }
                    });
                } else {
                    totalTasks += 1;
                    if(w.status === 'FINISH') {
                        const dStr = w.tanggal_finish || w.tanggal_actual_start || w.tanggal_schedule;
                        const parsed = extractValidDate(dStr);
                        if(parsed && parsed.ymd) {
                            dateMap[parsed.ymd] = (dateMap[parsed.ymd] || 0) + 1;
                        } else {
                            undatedCount += 1;
                        }
                    }
                }
            });

            let savedStart = localStorage.getItem(`eic_scurve_start_u${currentUnit}`);
            let savedEnd = localStorage.getItem(`eic_scurve_end_u${currentUnit}`);

            const activeYMDs = Object.keys(dateMap).sort();
            if(!savedStart) savedStart = activeYMDs.length > 0 ? activeYMDs[0] : '2026-08-20';
            if(!savedEnd) {
                const latestActive = activeYMDs.length > 0 ? activeYMDs[activeYMDs.length - 1] : '2026-08-30';
                const d = new Date(latestActive);
                d.setDate(d.getDate() + 5);
                savedEnd = `${d.getFullYear()}-${String(d.getMonth() + 1).padStart(2, '0')}-${String(d.getDate()).padStart(2, '0')}`;
            }

            const startDateObj = new Date(savedStart);
            const endDateObj = new Date(savedEnd);
            let dayDiff = Math.round((endDateObj - startDateObj) / 86400000);
            if(dayDiff < 1) dayDiff = 1;
            const totalDays = dayDiff + 1;

            let cumActualCount = undatedCount;
            Object.keys(dateMap).forEach(ymd => {
                if(ymd < savedStart) {
                    cumActualCount += dateMap[ymd];
                }
            });

            const timeline = [];
            const todayTime = new Date().setHours(23, 59, 59, 999);

            for(let i = 0; i < totalDays; i++) {
                const cur = new Date(startDateObj);
                cur.setDate(cur.getDate() + i);
                const ymd = `${cur.getFullYear()}-${String(cur.getMonth() + 1).padStart(2, '0')}-${String(cur.getDate()).padStart(2, '0')}`;
                const shortLabel = `${cur.getDate()} ${cur.toLocaleString('en-US', { month: 'short' })}`;
                const time = cur.getTime();

                const x = totalDays > 1 ? (i / (totalDays - 1)) : 1.0;
                const k = 7.0;
                const sRaw = 1.0 / (1.0 + Math.exp(-k * (x - 0.5)));
                const sMin = 1.0 / (1.0 + Math.exp(-k * (0 - 0.5)));
                const sMax = 1.0 / (1.0 + Math.exp(-k * (1 - 0.5)));
                const targetPct = Math.round(((sRaw - sMin) / (sMax - sMin)) * 1000) / 10;

                const dailyCount = dateMap[ymd] || 0;
                let actualPct = null;
                if(time <= todayTime) {
                    cumActualCount += dailyCount;
                    actualPct = totalTasks > 0 ? Math.min(100, Math.round((cumActualCount / totalTasks) * 1000) / 10) : 0;
                }

                timeline.push({
                    index: i,
                    ymd: ymd,
                    shortLabel: shortLabel,
                    time: time,
                    daily: dailyCount,
                    cumActual: actualPct !== null ? cumActualCount : null,
                    actualPct: actualPct,
                    targetPct: targetPct
                });
            }

            const actualPoints = timeline.filter(t => t.actualPct !== null);
            const latestActual = actualPoints.length > 0 ? actualPoints[actualPoints.length - 1] : { actualPct: 0, targetPct: 0, cumActual: 0 };
            const variance = Math.round((latestActual.actualPct - latestActual.targetPct) * 10) / 10;

            const svgW = 780;
            const svgH = 200;
            const padL = 45;
            const padR = 25;
            const padT = 20;
            const padB = 32;
            const graphW = svgW - padL - padR;
            const graphH = svgH - padT - padB;

            const n = timeline.length;
            const getX = (i) => padL + (n === 1 ? graphW / 2 : (i / (n - 1)) * graphW);
            const getY = (pct) => padT + graphH - (pct / 100) * graphH;

            const targetPathPoints = timeline.map((pt, i) => `${getX(i)},${getY(pt.targetPct)}`);
            const targetPathD = `M ${targetPathPoints.join(' L ')}`;

            let actualPathD = '';
            let fillD = '';
            if(actualPoints.length > 0) {
                const pts = actualPoints.map(pt => `${getX(pt.index)},${getY(pt.actualPct)}`);
                actualPathD = `M ${pts.join(' L ')}`;
                const lastPt = actualPoints[actualPoints.length - 1];
                const firstPt = actualPoints[0];
                fillD = `${actualPathD} L ${getX(lastPt.index)} ${padT + graphH} L ${getX(firstPt.index)} ${padT + graphH} Z`;
            }

            const labelInterval = Math.max(1, Math.ceil(n / 8));

            return `
            <div style="margin-top:18px; margin-bottom:14px; background:var(--bg-sub); border:1px solid var(--border-color); border-radius:var(--radius-sm); padding:12px; page-break-inside:avoid;">
                <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:8px;">
                    <div style="font-size:0.88rem; font-weight:800; color:var(--primary);">OUTAGE UNIT ${currentUnit} PROGRESS S-CURVE (TARGET VS ACTUAL)</div>
                    <div style="font-size:0.8rem; font-weight:700; color:${variance >= 0 ? '#10b981' : '#f43f5e'};">
                        Actual: ${latestActual.actualPct}% &bull; Target: ${latestActual.targetPct}% (${variance >= 0 ? '+' : ''}${variance}% Variance)
                    </div>
                </div>
                <svg viewBox="0 0 ${svgW} ${svgH}" style="width:100%; height:auto; background:var(--bg-card); border-radius:var(--radius-sm); border:1px solid var(--border-color);">
                    ${[0, 25, 50, 75, 100].map(p => {
                        const y = getY(p);
                        return `
                        <line x1="${padL}" y1="${y}" x2="${svgW - padR}" y2="${y}" stroke="var(--border-color)" stroke-dasharray="4 4" stroke-width="1"/>
                        <text x="${padL - 8}" y="${y + 4}" fill="var(--text-muted)" font-size="9" text-anchor="end" font-family="'JetBrains Mono'">${p}%</text>`;
                    }).join('')}
                    ${fillD ? `<path d="${fillD}" fill="url(#rpt-scurve-grad)" opacity="0.25"/>` : ''}
                    <path d="${targetPathD}" fill="none" stroke="#94a3b8" stroke-dasharray="5 3" stroke-width="2"/>
                    ${actualPathD ? `<path d="${actualPathD}" fill="none" stroke="var(--primary)" stroke-width="3" stroke-linecap="round"/>` : ''}
                    ${actualPoints.map(pt => {
                        const x = getX(pt.index);
                        const y = getY(pt.actualPct);
                        return `
                        <circle cx="${x}" cy="${y}" r="3.5" fill="var(--primary)" stroke="#fff" stroke-width="1.5"/>
                        <text x="${x}" y="${y - 6}" fill="var(--primary)" font-size="8.5" font-weight:800; text-anchor="middle" font-family="'JetBrains Mono'">${pt.actualPct}%</text>`;
                    }).join('')}
                    ${timeline.map((pt, i) => {
                        if(i % labelInterval === 0 || i === n - 1) {
                            const x = getX(i);
                            return `<text x="${x}" y="${svgH - 9}" fill="var(--text-muted)" font-size="8.5" font-weight:600; text-anchor="middle" font-family="'JetBrains Mono'">${pt.shortLabel}</text>`;
                        }
                        return '';
                    }).join('')}
                    <defs>
                        <linearGradient id="rpt-scurve-grad" x1="0" y1="0" x2="0" y2="1">
                            <stop offset="0%" stop-color="var(--primary)" stop-opacity="0.8"/>
                            <stop offset="100%" stop-color="var(--primary)" stop-opacity="0.0"/>
                        </linearGradient>
                    </defs>
                </svg>
                <div style="display:flex; justify-content:space-between; align-items:center; margin-top:6px; font-size:0.75rem; color:var(--text-muted);">
                    <div>Total Sub-Task WO: <strong>${totalTasks} sub-tasks</strong> &bull; Completed Sub-Tasks: <strong style="color:var(--primary);">${latestActual.cumActual} sub-tasks (${latestActual.actualPct}%)</strong></div>
                    <div style="display:flex; gap:12px;">
                        <span><span style="width:10px; height:3px; background:var(--primary); display:inline-block; vertical-align:middle; border-radius:1px;"></span> Actual Realization</span>
                        <span><span style="width:10px; height:2px; background:#94a3b8; border-top:2px dashed #94a3b8; display:inline-block; vertical-align:middle;"></span> S-Curve Target</span>
                    </div>
                </div>
            </div>`;
        }

        function setReportType(type) {
            currentReportType = type;
            ['harian', 'wo_detail', 'actuator', 'instruments'].forEach(t => {
                const btn = document.getElementById(`reptab-${t}`);
                if(btn) btn.classList.toggle('active', t === type);
            });
            const dateBar = document.getElementById('report-date-bar');
            if(dateBar) {
                dateBar.style.display = (type === 'harian') ? 'flex' : 'none';
            }
            generateReportContent();
        }

        function generateReportContent() {
            const container = document.getElementById('report-printable-content');
            if(!container || !fullData || !fullData.summary) return;
            
            const s = fullData.summary;
            const printDateStr = new Date().toLocaleString('en-US');
            
            if(currentReportType === 'harian') {
                renderHarianReport(container, s, printDateStr);
            } else if(currentReportType === 'wo_detail') {
                renderWODetailReport(container, s, printDateStr);
            } else if(currentReportType === 'actuator') {
                renderActuatorReport(container, s, printDateStr);
            } else if(currentReportType === 'instruments') {
                renderInstrumentsReport(container, s, printDateStr);
            }
        }

        /* OPSI 1: Daily Progress & Findings Report */
        function renderHarianReport(container, s, printDateStr) {
            const startYMD = document.getElementById('report-start-date').value;
            const endYMD = document.getElementById('report-end-date').value;
            const startDisp = startYMD ? formatDateForStorage(startYMD) : 'Start';
            const endDisp = endYMD ? formatDateForStorage(endYMD) : 'Today';

            const completedTasks = [];
            
            // 1. Work Orders Updates
            (fullData.work_orders || []).forEach(w => {
                let subtasksAdded = 0;
                (w.checklist || []).forEach(c => {
                    if(c.selesai && isDateInRange(c.tanggal, startYMD, endYMD)) {
                        completedTasks.push({
                            type: 'Work Order',
                            code: w.no_wo,
                            item_name: w.job_description,
                            subtask: `Sub-task: ${c.sub_task}`,
                            area: w.area || 'GENERAL',
                            pic: c.pic_task || w.pic || '-',
                            date: c.tanggal || w.tanggal_finish || '-'
                        });
                        subtasksAdded++;
                    }
                });
                // If WO has no subtask list or entire WO marked finish in range
                if(subtasksAdded === 0 && (w.status === 'FINISH' || (w.persen_progress && w.persen_progress > 0)) && isDateInRange(w.tanggal_finish, startYMD, endYMD)) {
                    completedTasks.push({
                        type: 'Work Order',
                        code: w.no_wo,
                        item_name: w.job_description,
                        subtask: w.status === 'FINISH' ? 'Work Order Completed (100%)' : `WO Progress (${w.persen_progress}%)`,
                        area: w.area || 'GENERAL',
                        pic: w.pic || '-',
                        date: w.tanggal_finish || '-'
                    });
                }
            });

            // 2. Actuator Valves Updates
            (fullData.actuators || []).forEach(a => {
                const isGen = !!a.general_inspection;
                const isFunc = !!a.function_test;
                const isDone = isGen && isFunc;
                if((isGen || isFunc) && isDateInRange(a.finish_date, startYMD, endYMD)) {
                    let desc = isDone ? 'General Inspection & Function Test Completed (100% FINISH)' : (isGen ? 'General Inspection Completed (50%)' : 'Function Test Completed (50%)');
                    if(a.remarks) desc += ` [${a.remarks}]`;
                    completedTasks.push({
                        type: 'Actuator Valve',
                        code: a.equipment_id,
                        item_name: a.equipment_description,
                        subtask: desc,
                        area: a.area || 'BOILER',
                        pic: a.pic || '-',
                        date: a.finish_date || '-'
                    });
                }
            });

            // 3. Instruments Updates (Pressure TX, Temperature TX, Pressure Switch)
            const checkInst = (list, typeName) => {
                (list || []).forEach(inst => {
                    const isVerif = !!inst.verifikasi || !!inst.status_wdone;
                    const isCalib = !!inst.kalibrasi;
                    const instDate = inst.tanggal || inst.finish_date || inst.dated;
                    if((isVerif || isCalib) && isDateInRange(instDate, startYMD, endYMD)) {
                        let desc = isVerif ? 'Calibration & Verification Finished (100% DONE)' : 'Calibration Completed (In Progress)';
                        if(inst.remarks) desc += ` [${inst.remarks}]`;
                        completedTasks.push({
                            type: typeName,
                            code: inst.kks || `Item #${inst.no}`,
                            item_name: inst.equipment || inst.description || typeName,
                            subtask: desc,
                            area: inst.area || 'GENERAL',
                            pic: inst.pic || 'EIC Team',
                            date: instDate || '-'
                        });
                    }
                });
            };
            checkInst(fullData.pressure_tx, 'Pressure TX');
            checkInst(fullData.temperature_tx, 'Temperature TX');
            checkInst(fullData.pressure_switch, 'Pressure Switch');

            // 4. Findings Collection
            const findingsList = [];
            const collectFinding = (list, typeName, codeField, descField) => {
                (list || []).forEach(item => {
                    if(item.temuan || (item.jumlah_foto > 0)) {
                        findingsList.push({
                            type: typeName,
                            code: item[codeField] || 'Item',
                            desc: item[descField] || item.equipment || item.job_description || '-',
                            area: item.area || 'GENERAL',
                            temuan: item.temuan || '(No defect description, field photo documentation logged)',
                            tindak_lanjut: item.tindak_lanjut || 'Pending field verification',
                            foto_count: item.jumlah_foto || 0
                        });
                    }
                });
            };
            collectFinding(fullData.work_orders, 'Work Order', 'no_wo', 'job_description');
            collectFinding(fullData.actuators, 'Actuator Valve', 'equipment_id', 'equipment_description');
            collectFinding(fullData.pressure_tx, 'Pressure TX', 'kks', 'equipment');
            collectFinding(fullData.temperature_tx, 'Temperature TX', 'kks', 'equipment');
            collectFinding(fullData.pressure_switch, 'Pressure Switch', 'kks', 'equipment');

            let html = `
            <div class="report-paper">
                <div class="report-header-box">
                    <div style="display:flex; justify-content:space-between; align-items:flex-start; border-bottom:2px solid #334155; padding-bottom:12px; margin-bottom:16px;">
                        <div>
                            <div style="font-size:1.1rem; font-weight:800; color:var(--primary); letter-spacing:0.5px;">PLTU MSW &bull; SECTION ELECTRIC, INSTRUMENT & CONTROL</div>
                            <h2 style="font-size:1.25rem; margin:4px 0 2px 0; color:#fff;">DAILY OUTAGE PROGRESS REPORT (WO, ACTUATOR & INSTRUMENTS)</h2>
                            <div style="font-size:0.85rem; color:var(--text-muted);">Daily Progress Monitoring & Field Findings Summary Unit ${currentUnit}</div>
                        </div>
                        <div style="text-align:right; font-size:0.8rem; color:var(--text-muted);">
                            <div><strong>Period:</strong> ${startDisp} to ${endDisp}</div>
                            <div><strong>Printed:</strong> ${printDateStr}</div>
                            <div style="color:var(--primary); font-weight:700; margin-top:2px;">UNIT ${currentUnit}</div>
                        </div>
                    </div>
                </div>

                <div class="report-section-title">1. OVERALL PROGRESS SUMMARY (UNIT ${currentUnit})</div>
                <div class="report-kpi-grid" style="grid-template-columns: repeat(3, 1fr);">
                    <div class="report-kpi-card">
                        <div class="report-kpi-lbl">WORK ORDERS (SUB-TASKS)</div>
                        <div class="report-kpi-val">${s.wo.pct}%</div>
                        <div class="report-kpi-sub">${s.wo.subtask_done} / ${s.wo.subtask_total} Sub-tasks (${s.wo.finish}/${s.wo.total} WO Finish)</div>
                    </div>
                    <div class="report-kpi-card">
                        <div class="report-kpi-lbl">ACTUATOR VALVES</div>
                        <div class="report-kpi-val">${s.actuator.pct}%</div>
                        <div class="report-kpi-sub">${s.actuator.subtask_done} / ${s.actuator.subtask_total} Tests (${s.actuator.finish}/${s.actuator.total} Valve Finish)</div>
                    </div>
                    <div class="report-kpi-card">
                        <div class="report-kpi-lbl">INSTRUMENTS (TX & PSW)</div>
                        <div class="report-kpi-val">${s.instrument.pct}%</div>
                        <div class="report-kpi-sub">${s.instrument.done} / ${s.instrument.total} Verification Finished</div>
                    </div>
                </div>

                ${generateSCurveHTMLForReport()}

                <div class="report-section-title" style="margin-top:22px;">
                    2. COMPLETED JOB UPDATES LOG (WO, ACTUATOR & INSTRUMENTS)
                    <span style="font-size:0.8rem; font-weight:600; color:var(--text-muted); float:right;">Total: ${completedTasks.length} Completed Items</span>
                </div>
                ${completedTasks.length === 0 ? `
                    <div style="padding:14px; background:var(--bg-sub); border:1px dashed var(--border-color); border-radius:var(--radius-sm); font-size:0.85rem; color:var(--text-muted); text-align:center;">
                        ℹ️ No sub-tasks, actuator valves, or instruments recorded completed between <strong>${startDisp} to ${endDisp}</strong>.
                    </div>
                ` : `
                    <table class="report-table">
                        <thead>
                            <tr>
                                <th style="width:35px;">No</th>
                                <th style="width:115px;">Category</th>
                                <th style="width:135px;">WO No / KKS Tag</th>
                                <th>Task Description / Sub-Task / Progress Update</th>
                                <th style="width:110px;">Area</th>
                                <th style="width:110px;">PIC</th>
                                <th style="width:90px; text-align:center;">Date Completed</th>
                            </tr>
                        </thead>
                        <tbody>
                            ${completedTasks.map((t, idx) => `
                                <tr>
                                    <td style="text-align:center; font-family:'JetBrains Mono';">${idx + 1}</td>
                                    <td><span class="report-badge">${t.type}</span></td>
                                    <td style="font-family:'JetBrains Mono'; font-weight:700; color:var(--primary);">${t.code}</td>
                                    <td>
                                        <div style="font-weight:600;">${t.subtask}</div>
                                        <div style="font-size:0.75rem; color:var(--text-muted);">${t.item_name}</div>
                                    </td>
                                    <td>${t.area}</td>
                                    <td style="font-weight:600;">${t.pic}</td>
                                    <td style="text-align:center; font-family:'JetBrains Mono'; font-size:0.8rem;">${t.date}</td>
                                </tr>
                            `).join('')}
                        </tbody>
                    </table>
                `}

                <div class="report-section-title" style="margin-top:24px;">
                    3. FIELD FINDINGS & CORRECTIVE ACTION SUMMARY (ACTIVE FINDINGS)
                    <span style="font-size:0.8rem; font-weight:600; color:var(--text-muted); float:right;">Total: ${findingsList.length} Findings</span>
                </div>
                ${findingsList.length === 0 ? `
                    <div style="padding:14px; background:var(--bg-sub); border:1px dashed var(--border-color); border-radius:var(--radius-sm); font-size:0.85rem; color:#10b981; text-align:center;">
                        Nil. All equipment and instrumentation are in normal operating condition with no open findings.
                    </div>
                ` : `
                    <table class="report-table">
                        <thead>
                            <tr>
                                <th style="width:35px;">No</th>
                                <th style="width:110px;">Category</th>
                                <th style="width:130px;">WO / Tag</th>
                                <th style="width:160px;">Equipment</th>
                                <th>Defect / Finding Details</th>
                                <th>Recommended Action Plan</th>
                                <th style="width:75px; text-align:center;">Photos</th>
                            </tr>
                        </thead>
                        <tbody>
                            ${findingsList.map((f, idx) => `
                                <tr>
                                    <td style="text-align:center; font-family:'JetBrains Mono';">${idx + 1}</td>
                                    <td><span class="report-badge alert">${f.type}</span></td>
                                    <td style="font-family:'JetBrains Mono'; font-weight:700; color:var(--status-alert);">${f.code}</td>
                                    <td style="font-weight:600; font-size:0.82rem;">${f.desc}<div style="font-size:0.75rem; color:var(--text-muted);">${f.area}</div></td>
                                    <td style="color:#fca5a5; font-size:0.82rem;">${f.temuan}</td>
                                    <td style="color:#fef08a; font-size:0.82rem;">${f.tindak_lanjut}</td>
                                    <td style="text-align:center; font-size:0.8rem; font-weight:700;">${f.foto_count > 0 ? f.foto_count + ' Photos' : '-'}</td>
                                </tr>
                            `).join('')}
                        </tbody>
                    </table>
                `}
            </div>`;
            container.innerHTML = html;
        }

        /* OPSI 2: Full Work Orders & Detailed Sub-Tasks Report */
        function renderWODetailReport(container, s, printDateStr) {
            const woList = fullData.work_orders || [];
            let html = `
            <div class="report-paper">
                <div class="report-header-box">
                    <div style="display:flex; justify-content:space-between; align-items:flex-start; border-bottom:2px solid #334155; padding-bottom:12px; margin-bottom:16px;">
                        <div>
                            <div style="font-size:1.1rem; font-weight:800; color:var(--primary); letter-spacing:0.5px;">PLTU MSW &bull; SECTION ELECTRIC, INSTRUMENT & CONTROL</div>
                            <h2 style="font-size:1.25rem; margin:4px 0 2px 0; color:#fff;">FULL WORK ORDER STATUS & SUB-TASK DETAILS REPORT</h2>
                            <div style="font-size:0.85rem; color:var(--text-muted);">Summary of All Work Orders & Detailed Sub-Task Checklists for Unit ${currentUnit}</div>
                        </div>
                        <div style="text-align:right; font-size:0.8rem; color:var(--text-muted);">
                            <div><strong>Total WOs:</strong> ${woList.length} Jobs</div>
                            <div><strong>Printed:</strong> ${printDateStr}</div>
                            <div style="color:var(--primary); font-weight:700; margin-top:2px;">UNIT ${currentUnit}</div>
                        </div>
                    </div>
                </div>

                <div class="report-section-title">1. WORK ORDERS PROGRESS SUMMARY</div>
                <div class="report-kpi-grid" style="grid-template-columns: repeat(4, 1fr);">
                    <div class="report-kpi-card">
                        <div class="report-kpi-lbl">TOTAL WORK ORDERS</div>
                        <div class="report-kpi-val">${s.wo.total}</div>
                        <div class="report-kpi-sub">Total WO Items</div>
                    </div>
                    <div class="report-kpi-card">
                        <div class="report-kpi-lbl">WO FINISH (100%)</div>
                        <div class="report-kpi-val" style="color:var(--status-finish);">${s.wo.finish}</div>
                        <div class="report-kpi-sub">Fully Completed WOs</div>
                    </div>
                    <div class="report-kpi-card">
                        <div class="report-kpi-lbl">WO IN PROGRESS</div>
                        <div class="report-kpi-val" style="color:#38bdf8;">${s.wo.in_progress}</div>
                        <div class="report-kpi-sub">In-Progress WOs</div>
                    </div>
                    <div class="report-kpi-card">
                        <div class="report-kpi-lbl">SUB-TASK PROGRESS</div>
                        <div class="report-kpi-val">${s.wo.pct}%</div>
                        <div class="report-kpi-sub">${s.wo.subtask_done} / ${s.wo.subtask_total} Sub-Tasks Done</div>
                    </div>
                </div>

                ${generateSCurveHTMLForReport()}

                <div class="report-section-title" style="margin-top:22px;">
                    2. COMPREHENSIVE WORK ORDER & SUB-TASK DETAILS (${woList.length} WOs)
                </div>

                <div style="display:flex; flex-direction:column; gap:16px;">
                    ${woList.map((w, idx) => {
                        const doneCount = (w.checklist || []).filter(c => c.selesai).length;
                        const totalCount = (w.checklist || []).length;
                        const stBadge = w.status === 'FINISH' ? 'var(--status-finish)' : (w.status === 'IN PROGRESS' ? '#38bdf8' : '#94a3b8');
                        
                        return `
                        <div style="background:var(--bg-sub); border:1px solid var(--border-color); border-radius:var(--radius-sm); padding:12px 14px; page-break-inside:avoid;">
                            <div style="display:flex; justify-content:space-between; align-items:center; border-bottom:1px solid var(--border-color); padding-bottom:8px; margin-bottom:10px; flex-wrap:wrap; gap:8px;">
                                <div>
                                    <span style="font-family:'JetBrains Mono'; font-weight:800; color:var(--primary); font-size:0.95rem;">${idx + 1}. [${w.no_wo}]</span>
                                    <strong style="color:var(--text-main); font-size:0.95rem; margin-left:6px;">${w.job_description}</strong>
                                    <span style="font-size:0.8rem; color:var(--text-muted); margin-left:8px;">&bull; Area: <strong>${w.area || 'GENERAL'}</strong></span>
                                </div>
                                <div style="display:flex; align-items:center; gap:10px;">
                                    <span style="font-size:0.8rem; color:var(--text-muted);">PIC: <strong style="color:var(--text-main);">${w.pic || '-'}</strong></span>
                                    <span style="display:inline-block; padding:2px 8px; border-radius:4px; font-size:0.75rem; font-weight:800; background:rgba(56, 189, 248, 0.1); color:${stBadge}; border:1px solid ${stBadge};">${w.status} (${w.persen_progress}%)</span>
                                </div>
                            </div>

                            ${totalCount === 0 ? `
                                <div style="font-size:0.8rem; color:var(--text-muted); padding:4px 0;">(No sub-task checklist recorded)</div>
                            ` : `
                                <table class="report-table" style="margin-bottom:4px;">
                                    <thead>
                                        <tr>
                                            <th style="width:30px;">No</th>
                                            <th>Sub-Task Description (${doneCount} / ${totalCount} Completed)</th>
                                            <th style="width:110px; text-align:center;">Status</th>
                                            <th style="width:100px; text-align:center;">Finish Date</th>
                                            <th style="width:110px;">Sub-Task PIC</th>
                                        </tr>
                                    </thead>
                                    <tbody>
                                        ${(w.checklist || []).map((c, cIdx) => `
                                            <tr style="${c.selesai ? 'background:rgba(16, 185, 129, 0.03);' : ''}">
                                                <td style="text-align:center; font-family:'JetBrains Mono'; font-size:0.78rem;">${cIdx + 1}</td>
                                                <td style="${c.selesai ? 'color:var(--text-main); font-weight:600;' : 'color:var(--text-muted);'}">
                                                    ${c.sub_task}
                                                </td>
                                                <td style="text-align:center;">
                                                    <span style="font-size:0.75rem; font-weight:700; color:${c.selesai ? 'var(--status-finish)' : 'var(--text-muted)'};">
                                                        ${c.selesai ? 'Done' : 'Pending'}
                                                    </span>
                                                </td>
                                                <td style="text-align:center; font-family:'JetBrains Mono'; font-size:0.78rem; color:var(--text-muted);">${c.tanggal || '-'}</td>
                                                <td style="font-size:0.8rem; font-weight:600;">${c.pic_task || w.pic || '-'}</td>
                                            </tr>
                                        `).join('')}
                                    </tbody>
                                </table>
                            `}

                            ${(w.temuan || (w.jumlah_foto > 0)) ? `
                                <div style="margin-top:8px; padding:6px 10px; background:rgba(244,63,94,0.08); border:1px solid rgba(244,63,94,0.25); border-radius:4px; font-size:0.8rem;">
                                    <strong style="color:#fda4af;">Finding:</strong> ${w.temuan || '(Logged photo documentation)'} &bull; 
                                    <strong style="color:#fef08a;">Action Plan:</strong> ${w.tindak_lanjut || 'Pending verification'} 
                                    ${w.jumlah_foto > 0 ? `(${w.jumlah_foto} Photos)` : ''}
                                </div>
                            ` : ''}
                        </div>`;
                    }).join('')}
                </div>
            </div>`;
            container.innerHTML = html;
        }

        /* OPSI 3: Full Actuator Valves Report */
        function renderActuatorReport(container, s, printDateStr) {
            const actList = fullData.actuators || [];
            let html = `
            <div class="report-paper">
                <div class="report-header-box">
                    <div style="display:flex; justify-content:space-between; align-items:flex-start; border-bottom:2px solid #334155; padding-bottom:12px; margin-bottom:16px;">
                        <div>
                            <div style="font-size:1.1rem; font-weight:800; color:var(--primary); letter-spacing:0.5px;">PLTU MSW &bull; SECTION ELECTRIC, INSTRUMENT & CONTROL</div>
                            <h2 style="font-size:1.25rem; margin:4px 0 2px 0; color:#fff;">COMPREHENSIVE ACTUATOR VALVES MONITORING REPORT</h2>
                            <div style="font-size:0.85rem; color:var(--text-muted);">Status of General Inspection, Function Test & Findings for Unit ${currentUnit}</div>
                        </div>
                        <div style="text-align:right; font-size:0.8rem; color:var(--text-muted);">
                            <div><strong>Total Actuators:</strong> ${actList.length} Valves</div>
                            <div><strong>Printed:</strong> ${printDateStr}</div>
                            <div style="color:var(--primary); font-weight:700; margin-top:2px;">UNIT ${currentUnit}</div>
                        </div>
                    </div>
                </div>

                <div class="report-section-title">1. ACTUATOR VALVES PROGRESS SUMMARY</div>
                <div class="report-kpi-grid" style="grid-template-columns: repeat(4, 1fr);">
                    <div class="report-kpi-card">
                        <div class="report-kpi-lbl">TOTAL ACTUATOR VALVES</div>
                        <div class="report-kpi-val">${s.actuator.total}</div>
                        <div class="report-kpi-sub">Total Scheduled Valves</div>
                    </div>
                    <div class="report-kpi-card">
                        <div class="report-kpi-lbl">VALVES FINISH (100%)</div>
                        <div class="report-kpi-val" style="color:var(--status-finish);">${s.actuator.finish}</div>
                        <div class="report-kpi-sub">Insp & Func Completed</div>
                    </div>
                    <div class="report-kpi-card">
                        <div class="report-kpi-lbl">VALVES IN PROGRESS</div>
                        <div class="report-kpi-val" style="color:#38bdf8;">${s.actuator.in_progress}</div>
                        <div class="report-kpi-sub">One Stage Completed</div>
                    </div>
                    <div class="report-kpi-card">
                        <div class="report-kpi-lbl">OVERALL PROGRESS</div>
                        <div class="report-kpi-val">${s.actuator.pct}%</div>
                        <div class="report-kpi-sub">${s.actuator.subtask_done} / ${s.actuator.subtask_total} Tests Completed</div>
                    </div>
                </div>

                <div class="report-section-title" style="margin-top:22px;">
                    2. FULL ACTUATOR VALVES STATUS TABLE (${actList.length} Valves)
                </div>

                <table class="report-table">
                    <thead>
                        <tr>
                            <th style="width:30px;">No</th>
                            <th style="width:120px;">Equipment ID</th>
                            <th style="width:110px;">KKS Tag</th>
                            <th>Actuator Valve Description</th>
                            <th style="width:85px;">Area</th>
                            <th style="width:75px; text-align:center;">General Insp</th>
                            <th style="width:75px; text-align:center;">Function Test</th>
                            <th style="width:65px; text-align:center;">% Prog</th>
                            <th style="width:85px; text-align:center;">Status</th>
                            <th style="width:85px;">PIC</th>
                            <th style="width:85px; text-align:center;">Finish Date</th>
                        </tr>
                    </thead>
                    <tbody>
                        ${actList.map((a, idx) => `
                            <tr>
                                <td style="text-align:center; font-family:'JetBrains Mono';">${idx + 1}</td>
                                <td style="font-family:'JetBrains Mono'; font-weight:700; color:var(--primary);">${a.equipment_id}</td>
                                <td style="font-family:'JetBrains Mono'; font-size:0.78rem; color:var(--text-muted);">${a.kks || '-'}</td>
                                <td style="font-weight:600; font-size:0.83rem;">
                                    ${a.equipment_description}
                                    ${(a.temuan || (a.jumlah_foto > 0)) ? `
                                        <div style="font-size:0.75rem; color:#fda4af; font-weight:normal; margin-top:2px;">Finding: ${a.temuan || 'Photo logged'}</div>
                                    ` : ''}
                                </td>
                                <td><span style="font-size:0.8rem; color:var(--text-muted);">${a.area || 'BOILER'}</span></td>
                                <td style="text-align:center; font-size:0.8rem; font-weight:700; color:${a.general_inspection ? 'var(--status-finish)' : 'var(--text-muted)'};">
                                    ${a.general_inspection ? 'OK' : 'Pending'}
                                </td>
                                <td style="text-align:center; font-size:0.8rem; font-weight:700; color:${a.function_test ? 'var(--status-finish)' : 'var(--text-muted)'};">
                                    ${a.function_test ? 'OK' : 'Pending'}
                                </td>
                                <td style="text-align:center; font-family:'JetBrains Mono'; font-weight:800;">${a.persen_progress || 0}%</td>
                                <td style="text-align:center;">
                                    <span class="report-badge" style="background:${a.status === 'FINISH' ? 'rgba(16,185,129,0.15)' : (a.status === 'IN PROGRESS' ? 'rgba(56,189,248,0.15)' : 'rgba(148,163,184,0.15)')}; color:${a.status === 'FINISH' ? 'var(--status-finish)' : (a.status === 'IN PROGRESS' ? '#38bdf8' : '#94a3b8')};">
                                        ${a.status}
                                    </span>
                                </td>
                                <td style="font-weight:600; font-size:0.8rem;">${a.pic || '-'}</td>
                                <td style="text-align:center; font-family:'JetBrains Mono'; font-size:0.78rem;">${a.finish_date || '-'}</td>
                            </tr>
                        `).join('')}
                    </tbody>
                </table>
            </div>`;
            container.innerHTML = html;
        }

        /* OPSI 4: Full Instruments (TX & PSW) Report */
        function renderInstrumentsReport(container, s, printDateStr) {
            const ptxList = fullData.pressure_tx || [];
            const ttxList = fullData.temperature_tx || [];
            const pswList = fullData.pressure_switch || [];
            const totalInst = ptxList.length + ttxList.length + pswList.length;

            let html = `
            <div class="report-paper">
                <div class="report-header-box">
                    <div style="display:flex; justify-content:space-between; align-items:flex-start; border-bottom:2px solid #334155; padding-bottom:12px; margin-bottom:16px;">
                        <div>
                            <div style="font-size:1.1rem; font-weight:800; color:var(--primary); letter-spacing:0.5px;">PLTU MSW &bull; SECTION ELECTRIC, INSTRUMENT & CONTROL</div>
                            <h2 style="font-size:1.25rem; margin:4px 0 2px 0; color:#fff;">COMPREHENSIVE INSTRUMENTS MONITORING REPORT</h2>
                            <div style="font-size:0.85rem; color:var(--text-muted);">Status of Calibration & Verification for Transmitters (PTX, TTX) and Pressure Switches Unit ${currentUnit}</div>
                        </div>
                        <div style="text-align:right; font-size:0.8rem; color:var(--text-muted);">
                            <div><strong>Total Instruments:</strong> ${totalInst} Items</div>
                            <div><strong>Printed:</strong> ${printDateStr}</div>
                            <div style="color:var(--primary); font-weight:700; margin-top:2px;">UNIT ${currentUnit}</div>
                        </div>
                    </div>
                </div>

                <div class="report-section-title">1. INSTRUMENTS STATUS SUMMARY</div>
                <div class="report-kpi-grid" style="grid-template-columns: repeat(4, 1fr);">
                    <div class="report-kpi-card">
                        <div class="report-kpi-lbl">TOTAL INSTRUMENTS</div>
                        <div class="report-kpi-val">${s.instrument.total}</div>
                        <div class="report-kpi-sub">PTX, TTX, and Pressure Switches</div>
                    </div>
                    <div class="report-kpi-card">
                        <div class="report-kpi-lbl">VERIFICATION FINISHED (DONE)</div>
                        <div class="report-kpi-val" style="color:var(--status-finish);">${s.instrument.done}</div>
                        <div class="report-kpi-sub">100% Completion Milestone</div>
                    </div>
                    <div class="report-kpi-card">
                        <div class="report-kpi-lbl">IN PROGRESS (CALIBRATION)</div>
                        <div class="report-kpi-val" style="color:#38bdf8;">${s.instrument.calib_done || 0}</div>
                        <div class="report-kpi-sub">Calibration Executed</div>
                    </div>
                    <div class="report-kpi-card">
                        <div class="report-kpi-lbl">VERIFICATION PROGRESS</div>
                        <div class="report-kpi-val">${s.instrument.pct}%</div>
                        <div class="report-kpi-sub">${s.instrument.done} / ${s.instrument.total} Fully Finished</div>
                    </div>
                </div>

                <!-- Part A: Pressure Transmitter -->
                <div class="report-section-title" style="margin-top:22px;">
                    2A. PRESSURE TRANSMITTERS (PTX) &bull; ${ptxList.length} Items
                </div>
                <table class="report-table">
                    <thead>
                        <tr>
                            <th style="width:30px;">No</th>
                            <th style="width:130px;">KKS Tag</th>
                            <th>Equipment Description</th>
                            <th style="width:120px;">Range / Unit</th>
                            <th style="width:90px; text-align:center;">Calibration</th>
                            <th style="width:105px; text-align:center;">Verification (Done)</th>
                            <th style="width:95px; text-align:center;">Status</th>
                            <th>Remarks / Findings</th>
                        </tr>
                    </thead>
                    <tbody>
                        ${ptxList.map((item, idx) => {
                            const isDone = !!item.verifikasi;
                            const isCalib = !!item.kalibrasi;
                            return `
                            <tr>
                                <td style="text-align:center; font-family:'JetBrains Mono';">${idx + 1}</td>
                                <td style="font-family:'JetBrains Mono'; font-weight:700; color:var(--primary);">${item.kks}</td>
                                <td style="font-weight:600;">${item.equipment}</td>
                                <td style="font-size:0.8rem; color:var(--text-muted);">${item.range || '-'} ${item.eng_unit || ''}</td>
                                <td style="text-align:center; font-size:0.8rem; font-weight:700; color:${isCalib ? '#38bdf8' : 'var(--text-muted)'};">
                                    ${isCalib ? 'Done' : 'Pending'}
                                </td>
                                <td style="text-align:center; font-size:0.8rem; font-weight:700; color:${isDone ? 'var(--status-finish)' : 'var(--text-muted)'};">
                                    ${isDone ? 'Done' : 'Pending'}
                                </td>
                                <td style="text-align:center;">
                                    <span class="report-badge" style="background:${isDone ? 'rgba(16,185,129,0.15)' : (isCalib ? 'rgba(56,189,248,0.15)' : 'rgba(148,163,184,0.15)')}; color:${isDone ? 'var(--status-finish)' : (isCalib ? '#38bdf8' : '#94a3b8')};">
                                        ${isDone ? 'DONE' : (isCalib ? 'CALIB OK' : 'SCHEDULED')}
                                    </span>
                                </td>
                                <td style="font-size:0.8rem; color:var(--text-muted);">${item.temuan ? 'Finding: ' + item.temuan : (item.remarks || '-')}</td>
                            </tr>`;
                        }).join('')}
                    </tbody>
                </table>

                <!-- Part B: Temperature Transmitter -->
                <div class="report-section-title" style="margin-top:22px;">
                    2B. TEMPERATURE TRANSMITTERS (TTX) &bull; ${ttxList.length} Items
                </div>
                <table class="report-table">
                    <thead>
                        <tr>
                            <th style="width:30px;">No</th>
                            <th style="width:130px;">KKS Tag</th>
                            <th>Equipment Description</th>
                            <th style="width:120px;">Range / Unit</th>
                            <th style="width:90px; text-align:center;">Calibration</th>
                            <th style="width:105px; text-align:center;">Verification (Done)</th>
                            <th style="width:95px; text-align:center;">Status</th>
                            <th>Remarks / Findings</th>
                        </tr>
                    </thead>
                    <tbody>
                        ${ttxList.map((item, idx) => {
                            const isDone = !!item.verifikasi;
                            const isCalib = !!item.kalibrasi;
                            return `
                            <tr>
                                <td style="text-align:center; font-family:'JetBrains Mono';">${idx + 1}</td>
                                <td style="font-family:'JetBrains Mono'; font-weight:700; color:var(--primary);">${item.kks}</td>
                                <td style="font-weight:600;">${item.equipment}</td>
                                <td style="font-size:0.8rem; color:var(--text-muted);">${item.range || '-'} ${item.eng_unit || ''}</td>
                                <td style="text-align:center; font-size:0.8rem; font-weight:700; color:${isCalib ? '#38bdf8' : 'var(--text-muted)'};">
                                    ${isCalib ? 'Done' : 'Pending'}
                                </td>
                                <td style="text-align:center; font-size:0.8rem; font-weight:700; color:${isDone ? 'var(--status-finish)' : 'var(--text-muted)'};">
                                    ${isDone ? 'Done' : 'Pending'}
                                </td>
                                <td style="text-align:center;">
                                    <span class="report-badge" style="background:${isDone ? 'rgba(16,185,129,0.15)' : (isCalib ? 'rgba(56,189,248,0.15)' : 'rgba(148,163,184,0.15)')}; color:${isDone ? 'var(--status-finish)' : (isCalib ? '#38bdf8' : '#94a3b8')};">
                                        ${isDone ? 'DONE' : (isCalib ? 'CALIB OK' : 'SCHEDULED')}
                                    </span>
                                </td>
                                <td style="font-size:0.8rem; color:var(--text-muted);">${item.temuan ? 'Finding: ' + item.temuan : (item.remarks || '-')}</td>
                            </tr>`;
                        }).join('')}
                    </tbody>
                </table>

                <!-- Part C: Pressure Switch -->
                <div class="report-section-title" style="margin-top:22px;">
                    2C. PRESSURE SWITCHES (PSW) &bull; ${pswList.length} Items
                </div>
                <table class="report-table">
                    <thead>
                        <tr>
                            <th style="width:30px;">No</th>
                            <th style="width:125px;">KKS Tag</th>
                            <th>Description</th>
                            <th style="width:105px;">Sub-Area</th>
                            <th style="width:100px; text-align:center;">Set Point</th>
                            <th style="width:75px; text-align:center;">Contact</th>
                            <th style="width:85px; text-align:center;">Calibration</th>
                            <th style="width:95px; text-align:center;">Verification</th>
                            <th style="width:90px; text-align:center;">Status</th>
                            <th>Remarks / Findings</th>
                        </tr>
                    </thead>
                    <tbody>
                        ${pswList.map((item, idx) => {
                            const isDone = !!item.verifikasi;
                            const isCalib = !!item.kalibrasi;
                            const sp = getSetPointDetails(item);
                            const contact = (item.contact_type || 'NO').toUpperCase();
                            return `
                            <tr>
                                <td style="text-align:center; font-family:'JetBrains Mono';">${idx + 1}</td>
                                <td style="font-family:'JetBrains Mono'; font-weight:700; color:var(--primary);">${item.kks}</td>
                                <td style="font-weight:600;">${item.description}</td>
                                <td style="font-size:0.8rem; color:var(--text-muted);">${item.sub_area || item.area || '-'}</td>
                                <td style="text-align:center;">
                                    <div class="setpoint-cell" style="padding:1px 5px; min-width:auto;">
                                        <span class="setpoint-val" style="font-size:0.75rem;">${sp.value}</span>
                                        <div class="setpoint-divider"></div>
                                        <span class="setpoint-dir ${sp.dir.toLowerCase()}" style="font-size:0.62rem;">${sp.dir}</span>
                                    </div>
                                </td>
                                <td style="text-align:center;">
                                    <span class="contact-badge ${contact==='NC'?'contact-nc':'contact-no'}" style="padding:1px 6px; font-size:0.72rem; min-width:auto;">${contact}</span>
                                </td>
                                <td style="text-align:center; font-size:0.8rem; font-weight:700; color:${isCalib ? '#38bdf8' : 'var(--text-muted)'};">
                                    ${isCalib ? 'Done' : 'Pending'}
                                </td>
                                <td style="text-align:center; font-size:0.8rem; font-weight:700; color:${isDone ? 'var(--status-finish)' : 'var(--text-muted)'};">
                                    ${isDone ? 'Done' : 'Pending'}
                                </td>
                                <td style="text-align:center;">
                                    <span class="report-badge" style="background:${isDone ? 'rgba(16,185,129,0.15)' : (isCalib ? 'rgba(56,189,248,0.15)' : 'rgba(148,163,184,0.15)')}; color:${isDone ? 'var(--status-finish)' : (isCalib ? '#38bdf8' : '#94a3b8')};">
                                        ${isDone ? 'DONE' : (isCalib ? 'CALIB OK' : 'SCHEDULED')}
                                    </span>
                                </td>
                                <td style="font-size:0.8rem; color:var(--text-muted);">${item.temuan ? 'Finding: ' + item.temuan : (item.remarks || '-')}</td>
                            </tr>`;
                        }).join('')}
                    </tbody>
                </table>
            </div>`;
            container.innerHTML = html;
        }

        
        function getSubtaskTypeBadge(desc) {
            if(!desc) return '';
            const s = String(desc).replace(/\xa0/g, ' ').replace(/\s+/g, ' ').trim().toUpperCase();
            
            // 1. Actuator Detection
            if(s.includes('ACTUATOR') || s.includes(' MOV') || s.includes('MOV ') || s.includes(' AOV') || s.includes('AOV ') || s.includes('GATE ACTUATOR') || s.includes('DAMPER ACTUATOR') || s.includes('VALVE ACTUATOR') || s.includes('FEED WATER CONTROL VALVE')) {
                return '<span class="badge-tag-comp badge-tag-act">ACTUATOR</span>';
            }
            const acts = (fullData.master_actuators || fullData.actuators || []);
            for(let i = 0; i < acts.length; i++) {
                const a = acts[i];
                const aKks = (a.kks || '').replace(/\xa0/g, ' ').trim().toUpperCase();
                if(aKks.length >= 6) {
                    const kksCore = aKks.length >= 8 ? aKks.slice(2) : aKks;
                    if(s.includes(aKks) || s.includes(kksCore)) {
                        return '<span class="badge-tag-comp badge-tag-act">ACTUATOR</span>';
                    }
                }
                const aDesc = (a.equipment_description || '').replace(/\xa0/g, ' ').replace(/\s+/g, ' ').trim().toUpperCase();
                if(aDesc.length >= 8 && (s.includes(aDesc) || aDesc.includes(s))) {
                    return '<span class="badge-tag-comp badge-tag-act">ACTUATOR</span>';
                }
            }
            
            // 2. Instrument Detection (PTX, TTX, PSW)
            if(s.includes('TRANSMITTER') || s.includes('PRESSURE TRANSMITTER') || s.includes('TEMP TRANSMITTER') || s.includes('TEMPERATURE TRANSMITTER') || s.includes('PRESSURE SWITCH') || s.includes('TEMP SWITCH') || s.includes('TEMPERATURE SWITCH') || s.includes('KALIBRASI TRANSMITTER') || s.includes('CALIBRATION MEASUREMENT') || s.includes('MEASUREMENT DEVICE')) {
                return '<span class="badge-tag-comp badge-tag-inst">INSTRUMENT</span>';
            }
            const insts = [...(fullData.pressure_tx || []), ...(fullData.temperature_tx || []), ...(fullData.pressure_switch || []), ...(fullData.master_instruments || [])];
            for(let i = 0; i < insts.length; i++) {
                const inst = insts[i];
                const iKks = (inst.kks || '').replace(/\xa0/g, ' ').trim().toUpperCase();
                if(iKks.length >= 6) {
                    const kksCore = iKks.length >= 8 ? iKks.slice(2) : iKks;
                    if(s.includes(iKks) || s.includes(kksCore)) {
                        return '<span class="badge-tag-comp badge-tag-inst">INSTRUMENT</span>';
                    }
                }
                const iDesc = (inst.equipment || inst.description || '').replace(/\xa0/g, ' ').replace(/\s+/g, ' ').trim().toUpperCase();
                if(iDesc.length >= 8 && (s.includes(iDesc) || iDesc.includes(s))) {
                    return '<span class="badge-tag-comp badge-tag-inst">INSTRUMENT</span>';
                }
            }
            
            // Jika bukan Actuator dan bukan Instrument -> TANPA LABEL
            return '';
        }

        const subtaskAddModes = {};

        function setSubtaskMode(noWo, mode) {
            subtaskAddModes[noWo] = mode;
            const cardEl = document.getElementById(`card-wo-${noWo}`);
            if(!cardEl) return;
            
            const btnManual = cardEl.querySelector(`.btn-mode-manual`);
            const btnAct = cardEl.querySelector(`.btn-mode-act`);
            const btnInst = cardEl.querySelector(`.btn-mode-inst`);
            
            if(btnManual) btnManual.classList.toggle('active', mode === 'manual');
            if(btnAct) btnAct.classList.toggle('active', mode === 'actuator');
            if(btnInst) btnInst.classList.toggle('active', mode === 'instrument');
            
            const boxManual = document.getElementById(`box-subtask-manual-${noWo}`);
            const boxAct = document.getElementById(`box-subtask-act-${noWo}`);
            const boxInst = document.getElementById(`box-subtask-inst-${noWo}`);
            
            if(boxManual) boxManual.style.display = mode === 'manual' ? 'flex' : 'none';
            if(boxAct) boxAct.style.display = mode === 'actuator' ? 'flex' : 'none';
            if(boxInst) boxInst.style.display = mode === 'instrument' ? 'flex' : 'none';
        }

        /* ---------------- QUICK ACTIONS (NO AUTO REFRESH) ---------------- */
        async function toggleLocalSubtask(noWo, cIdx, isChecked) {
            const item = (fullData.work_orders || []).find(w => w.no_wo === noWo);
            if(!item || !item.checklist || !item.checklist[cIdx]) return;
            
            const subTaskDesc = item.checklist[cIdx].sub_task;
            item.checklist[cIdx].selesai = isChecked;
            const nowStr = getTodayFormatted();
            item.checklist[cIdx].tanggal = isChecked ? nowStr : '';
            
            const chkItemEl = document.getElementById(`chk-${noWo}-${cIdx}`)?.closest('.checklist-item');
            if(chkItemEl) {
                chkItemEl.classList.toggle('done', isChecked);
                const spanEl = chkItemEl.querySelector('.checklist-item-body span');
                if(spanEl) {
                    spanEl.style.textDecoration = isChecked ? 'line-through' : 'none';
                    spanEl.style.color = isChecked ? 'var(--text-muted)' : 'var(--text-main)';
                    spanEl.style.opacity = isChecked ? '0.65' : '1';
                }
                const rightBox = chkItemEl.querySelector('.checklist-footer-right') || chkItemEl.querySelector('.header-right');
                let dateBadge = rightBox?.querySelector('.date-badge');
                if(isChecked) {
                    if(!dateBadge && rightBox) {
                        const badge = document.createElement('span');
                        badge.className = 'date-badge';
                        badge.title = "Date Completed";
                        badge.style.cssText = "font-size:0.72rem; color:var(--status-finish); font-family:'JetBrains Mono',monospace; background:rgba(16,185,129,0.12); border:1px solid rgba(16,185,129,0.3); padding:2px 7px; border-radius:4px;";
                        badge.innerText = `${nowStr}`;
                        const delBtn = rightBox.querySelector('.btn-del-subtask-cross');
                        rightBox.insertBefore(badge, delBtn);
                    }
                } else {
                    if(dateBadge) dateBadge.remove();
                }
            }
            
            const total = item.checklist.length;
            const done = item.checklist.filter(c => c.selesai).length;
            const pct = total > 0 ? Math.round((done / total) * 100) : 0;
            item.persen_progress = pct;
            item.status = (done === total && total > 0) ? 'FINISH' : (done > 0 ? 'IN PROGRESS' : 'SCHED-OK');
            
            // Update card elements
            const headerEl = document.getElementById(`card-wo-${noWo}`)?.querySelector('.item-header');
            if(headerEl) {
                const subCountEl = headerEl.querySelector('.wo-subtask-progress');
                if(subCountEl) subCountEl.innerText = `${done} / ${total} Sub-task`;
                const barFill = headerEl.querySelector('.progress-fill');
                if(barFill) barFill.style.width = `${pct}%`;
                const pctText = headerEl.querySelector('.progress-text');
                if(pctText) pctText.innerText = `${pct}%`;
                
                const badgeEl = headerEl.querySelector('.status-badge');
                if(badgeEl) {
                    const st = item.status.replace(/\s+/g, '_');
                    badgeEl.className = `status-badge badge-${st}`;
                    badgeEl.innerText = item.status;
                }
            }
            
            // Send background sync request to server so Excel & Actuator/Instrument sync immediately
            try {
                const res = await fetch('/api/quick_toggle_subtask', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({
                        unit: currentUnit,
                        no_wo: noWo,
                        sub_task: subTaskDesc,
                        sub_idx: cIdx,
                        selesai: isChecked
                    })
                });
                const result = await res.json();
                if(result.status === 'success') {
                    showToast(`✓ Sub-task & linked components synchronized (${result.persen_progress}%)`, 'success', 1500);
                }
            } catch(err) {
                console.error("Error saving quick subtask toggle:", err);
            }
        }

        function toggleLocalInstCheck(instSubtab, idx, field, isChecked) {
            let list = instSubtab === 'ptx' ? fullData.pressure_tx : (instSubtab === 'ttx' ? fullData.temperature_tx : fullData.pressure_switch);
            if(!list || !list[idx]) return;
            const item = list[idx];
            item[field] = isChecked;
            if(field === 'verifikasi') {
                item.status_wdone = isChecked;
            }
            
            const isVerif = !!item.verifikasi;
            const isCalib = !!item.kalibrasi;
            
            // Update card elements locally
            const cardPrefix = instSubtab === 'psw' ? 'psw' : 'inst';
            const calibCard = document.getElementById(`card-calib-${cardPrefix}-${idx}`);
            const verifCard = document.getElementById(`card-verif-${cardPrefix}-${idx}`);
            const badgeEl = document.getElementById(`badge-${cardPrefix}-${idx}`);
            
            if(calibCard) calibCard.classList.toggle('done', isCalib);
            if(verifCard) {
                verifCard.classList.toggle('done', isVerif);
                verifCard.style.borderColor = isVerif ? 'var(--status-finish)' : 'var(--border-color)';
                const verifSpan = verifCard.querySelector('label span');
                if(verifSpan) verifSpan.style.color = isVerif ? 'var(--status-finish)' : 'var(--primary)';
            }
            if(badgeEl) {
                badgeEl.className = `status-badge ${isVerif ? 'badge-FINISH' : (isCalib ? 'badge-IN-PROGRESS' : 'badge-SCHED-OK')}`;
                badgeEl.innerText = isVerif ? 'DONE (100%)' : (isCalib ? 'IN PROGRESS (Calibration OK)' : 'SCHEDULED');
            }
        }

        async function toggleDirectInstCheck(instSubtab, key, field, isChecked) {
            const payload = {
                unit: currentUnit,
                type: instSubtab === 'ptx' ? 'pressure_tx' : (instSubtab === 'ttx' ? 'temperature_tx' : 'pressure_switch'),
                kks: key,
                no: key
            };
            payload[field] = isChecked;
            if(field === 'verifikasi') payload.status_wdone = isChecked;
            
            try {
                const res = await fetch('/api/update_instrument', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify(payload)
                });
                const result = await res.json();
                if(result.status === 'success') {
                    showToast(`✓ ${field === 'verifikasi' ? 'Verification' : 'Calibration'} updated!`, 'success', 1800);
                    loadData();
                }
            } catch(e) {
                showToast('Failed to update instrument', 'error');
            }
        }

        /* ---------------- SAVE / EDIT SUBMISSIONS ---------------- */
        function getTodayFormatted() {
            const d = new Date();
            const day = String(d.getDate()).padStart(2, '0');
            const month = String(d.getMonth() + 1).padStart(2, '0');
            const year = d.getFullYear();
            return `${day}/${month}/${year}`;
        }

        async function saveWorkOrder(noWo) {
            const item = (fullData.work_orders || []).find(w => w.no_wo === noWo);
            if (!item) return;

            const picSelect = document.getElementById(`pic-${noWo}`);
            const remInput = document.getElementById(`rem-${noWo}`);

            const checklistPayload = (item.checklist || []).map((c, cIdx) => {
                const chk = document.getElementById(`chk-${noWo}-${cIdx}`);
                const isChecked = chk ? chk.checked : !!c.selesai;
                return {
                    sub_task: c.sub_task,
                    selesai: isChecked,
                    tanggal: isChecked ? (c.tanggal || getTodayFormatted()) : '',
                    pic_task: c.pic_task || ''
                };
            });

            const allDone = checklistPayload.length > 0 && checklistPayload.every(c => c.selesai);
            const autoFinishDate = allDone ? (item.tanggal_finish || getTodayFormatted()) : '';

            const payload = {
                unit: currentUnit,
                no_wo: noWo,
                pic: picSelect ? picSelect.value : (item.pic || ''),
                tanggal_finish: autoFinishDate,
                remarks: remInput ? remInput.value : (item.remarks || ''),
                checklist: checklistPayload
            };

            try {
                const res = await fetch('/api/update_wo', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify(payload)
                });
                const result = await res.json();
                showToast(result.message || 'WO updated successfully!', 'success');
                loadData();
            } catch(e) {
                showToast('Failed to save WO changes', 'error');
            }
        }

        async function quickToggleActuator(eqId, field, isChecked) {
            try {
                const payload = {
                    unit: currentUnit,
                    equipment_id: eqId,
                    field: field,
                    value: isChecked
                };
                const res = await fetch('/api/quick_toggle_actuator', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify(payload)
                });
                const result = await res.json();
                if(result.status === 'success') {
                    showToast(result.message, 'success');
                    loadData();
                } else {
                    showToast(result.message || 'Failed to update actuator status', 'error');
                }
            } catch(e) {
                showToast('Failed to update actuator status', 'error');
            }
        }

        async function saveActuator(eqId, desc) {
            const genChk = document.getElementById(`gen-${eqId}`);
            const funcChk = document.getElementById(`func-${eqId}`);
            const isGen = genChk ? genChk.checked : false;
            const isFunc = funcChk ? funcChk.checked : false;
            const isAllDone = isGen && isFunc;
            const autoFinishDate = isAllDone ? getTodayFormatted() : '';

            const payload = {
                unit: currentUnit,
                equipment_id: eqId,
                equipment_description: desc,
                pic: document.getElementById(`pic-act-${eqId}`).value,
                finish_date: autoFinishDate,
                remarks: document.getElementById(`rem-act-${eqId}`).value,
                general_inspection: isGen,
                function_test: isFunc
            };

            try {
                const res = await fetch('/api/update_actuator', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify(payload)
                });
                const result = await res.json();
                showToast(result.message || 'Actuator updated successfully!', 'success');
                loadData();
            } catch(e) {
                showToast('Failed to save actuator', 'error');
            }
        }

        async function saveTransmitter(type, key, idx) {
            const calibChk = document.getElementById(`inst-calib-${idx}`);
            const verifChk = document.getElementById(`inst-verif-${idx}`);
            const remInput = document.getElementById(`inst-rem-${idx}`);
            
            const payload = {
                unit: currentUnit,
                type: type === 'ptx' ? 'pressure_tx' : 'temperature_tx',
                kks: key,
                no: key,
                kalibrasi: calibChk ? calibChk.checked : false,
                verifikasi: verifChk ? verifChk.checked : false,
                remarks: remInput ? remInput.value : ''
            };

            try {
                const res = await fetch('/api/update_instrument', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify(payload)
                });
                const result = await res.json();
                showToast('Instrument updated successfully!', 'success');
                loadData();
            } catch(e) {
                showToast('Failed to save instrument', 'error');
            }
        }

        async function savePressureSwitch(arg1, arg2, arg3) {
            let type = 'pressure_switch';
            let key = arg1;
            let idx = arg2;
            if(arg3 !== undefined) {
                type = arg1;
                key = arg2;
                idx = arg3;
            }
            
            const calibChk = document.getElementById(`inst-calib-${idx}`);
            const verifChk = document.getElementById(`inst-verif-${idx}`);
            const remInput = document.getElementById(`inst-rem-${idx}`);
            const subAreaInput = document.getElementById(`subarea-psw-${idx}`);
            const spValInput = document.getElementById(`sp-val-psw-${idx}`);
            const spDirInput = document.getElementById(`sp-dir-psw-${idx}`);
            const contactInput = document.getElementById(`contact-psw-${idx}`);
            
            let combinedSp = spValInput ? spValInput.value.trim() : '';
            if(spDirInput && spDirInput.value && combinedSp) {
                if(!combinedSp.toUpperCase().includes(spDirInput.value)) {
                    combinedSp = `${combinedSp} (${spDirInput.value})`;
                }
            }

            const payload = {
                unit: currentUnit,
                type: 'pressure_switch',
                kks: key,
                no: key,
                sub_area: subAreaInput ? subAreaInput.value.trim() : '',
                set_point: combinedSp,
                contact_type: contactInput ? contactInput.value : 'NO',
                asfound_set: document.getElementById(`af-set-${idx}`) ? document.getElementById(`af-set-${idx}`).value : '',
                asfound_reset: document.getElementById(`af-reset-${idx}`) ? document.getElementById(`af-reset-${idx}`).value : '',
                asleft_set: document.getElementById(`al-set-${idx}`) ? document.getElementById(`al-set-${idx}`).value : '',
                asleft_reset: document.getElementById(`al-reset-${idx}`) ? document.getElementById(`al-reset-${idx}`).value : '',
                status_ok_notok: document.getElementById(`res-psw-${idx}`) ? document.getElementById(`res-psw-${idx}`).value : 'OK',
                kalibrasi: calibChk ? calibChk.checked : false,
                verifikasi: verifChk ? verifChk.checked : false,
                remarks: remInput ? remInput.value : ''
            };

            try {
                const res = await fetch('/api/update_instrument', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify(payload)
                });
                const result = await res.json();
                if(result.status === 'success') {
                    showToast('✓ Set Point, Contact & Calibration for Pressure Switch saved successfully!', 'success');
                    loadData();
                } else {
                    showToast(result.message || 'Failed to save Pressure Switch', 'error');
                }
            } catch(e) {
                showToast('Failed to save PSW calibration: ' + e.message, 'error');
            }
        }

        async function saveNewWO() {
            const noWo = (document.getElementById('new-wo-code').value || '').trim();
            const desc = (document.getElementById('new-wo-desc').value || '').trim();
            const area = (document.getElementById('new-wo-area').value || '').trim();
            const pic = document.getElementById('new-wo-pic').value;
            const sched = formatDateForStorage(document.getElementById('new-wo-sched').value);
            const checklistStr = (document.getElementById('new-wo-checklist').value || '').trim();

            if(!noWo || !desc) {
                showToast('WO No and Job Description are required!', 'error');
                return;
            }

            try {
                const res = await fetch('/api/add_wo', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({unit: currentUnit, no_wo: noWo, job_description: desc, area: area, pic: pic, tanggal_schedule: sched, checklist_str: checklistStr})
                });
                const result = await res.json();
                if(result.status === 'success') {
                    showToast(result.message || 'New WO added successfully!', 'success');
                    document.getElementById('new-wo-code').value = '';
                    document.getElementById('new-wo-desc').value = '';
                    document.getElementById('new-wo-area').value = '';
                    document.getElementById('new-wo-sched').value = '';
                    document.getElementById('new-wo-checklist').value = '';
                    toggleAccordion('add-wo-form');
                    loadData();
                } else {
                    showToast(result.message || 'Failed to add WO', 'error');
                }
            } catch(e) {
                showToast('Failed to add new WO: ' + e.message, 'error');
            }
        }

        async function deleteWO(noWo) {
            if(!confirm(`Are you sure you want to delete Work Order ${noWo} and all of its subtasks?`)) return;

            try {
                const res = await fetch('/api/delete_wo', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({unit: currentUnit, no_wo: noWo})
                });
                const result = await res.json();
                showToast(result.message || 'WO deleted successfully!', 'info');
                loadData();
            } catch(e) {
                showToast('Failed to delete WO', 'error');
            }
        }

        async function addSubtask(noWo, mode = 'manual') {
            let subTask = '';
            let defaultPic = '';
            
            if(mode === 'actuator') {
                const sel = document.getElementById(`new-subtask-act-${noWo}`);
                subTask = (sel?.value || '').trim();
                defaultPic = 'AMP';
                if(!subTask) {
                    showToast('Please select an Actuator from the dropdown!', 'error');
                    return;
                }
            } else if(mode === 'instrument') {
                const sel = document.getElementById(`new-subtask-inst-${noWo}`);
                subTask = (sel?.value || '').trim();
                defaultPic = 'JAPA';
                if(!subTask) {
                    showToast('Please select an Instrument from the dropdown!', 'error');
                    return;
                }
            } else {
                const input = document.getElementById(`new-subtask-${noWo}`);
                subTask = (input?.value || '').trim();
                if(!subTask) {
                    showToast('Please enter sub-task description!', 'error');
                    return;
                }
            }

            try {
                const res = await fetch('/api/add_subtask', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({unit: currentUnit, no_wo: noWo, sub_task: subTask, pic: defaultPic})
                });
                const result = await res.json();
                if(result.status === 'success') {
                    showToast(result.message || 'Sub-task added and synchronized successfully!', 'success');
                    const input = document.getElementById(`new-subtask-${noWo}`);
                    if(input) input.value = '';
                    loadData();
                } else {
                    showToast(result.message || 'Failed to add sub-task', 'error');
                }
            } catch(e) {
                showToast('Failed to add sub-task: ' + e.message, 'error');
            }
        }

        async function deleteSubtask(noWo, subTask) {
            if(!confirm(`Delete sub-task "${subTask}" from ${noWo}?`)) return;

            try {
                const res = await fetch('/api/delete_subtask', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({unit: currentUnit, no_wo: noWo, sub_task: subTask})
                });
                const result = await res.json();
                showToast(result.message || 'Sub-task deleted successfully!', 'info');
                loadData();
            } catch(e) {
                showToast('Failed to delete sub-task', 'error');
            }
        }

        async function saveNewActuator() {
            const eqId = (document.getElementById('new-act-id').value || '').trim();
            const desc = (document.getElementById('new-act-desc').value || '').trim();
            const area = (document.getElementById('new-act-area').value || '').trim();
            const kks = (document.getElementById('new-act-kks').value || '').trim();
            const pic = document.getElementById('new-act-pic').value;

            if(!eqId || !desc) {
                showToast('Equipment ID and Description are required!', 'error');
                return;
            }

            try {
                const res = await fetch('/api/add_actuator', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({unit: currentUnit, equipment_id: eqId, equipment_description: desc, area: area, kks: kks, pic: pic})
                });
                const result = await res.json();
                if(result.status === 'success') {
                    showToast(result.message || 'New Actuator added successfully!', 'success');
                    document.getElementById('new-act-id').value = '';
                    document.getElementById('new-act-desc').value = '';
                    document.getElementById('new-act-area').value = '';
                    document.getElementById('new-act-kks').value = '';
                    toggleAccordion('add-act-form');
                    loadData();
                } else {
                    showToast(result.message || 'Failed to add Actuator', 'error');
                }
            } catch(e) {
                showToast('Failed to add new Actuator: ' + e.message, 'error');
            }
        }

        async function deleteActuator(eqId) {
            if(!confirm(`Are you sure you want to delete Actuator Valve ${eqId}?`)) return;

            try {
                const res = await fetch('/api/delete_actuator', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({unit: currentUnit, equipment_id: eqId})
                });
                const result = await res.json();
                showToast(result.message || 'Actuator deleted successfully!', 'info');
                loadData();
            } catch(e) {
                showToast('Failed to delete Actuator', 'error');
            }
        }

        async function saveNewInstrument() {
            const type = document.getElementById('new-inst-type').value;
            const desc = (document.getElementById('new-inst-desc').value || '').trim();
            const kks = (document.getElementById('new-inst-kks').value || '').trim();
            const area = (document.getElementById('new-inst-area').value || '').trim();
            const subArea = (document.getElementById('new-inst-subarea')?.value || '').trim();
            const range = (document.getElementById('new-inst-range')?.value || '').trim();
            const spVal = (document.getElementById('new-inst-sp-val')?.value || '').trim();
            const spDir = (document.getElementById('new-inst-sp-dir')?.value || 'HIGH');
            const contact = (document.getElementById('new-inst-contact')?.value || 'NO');

            if(!desc) {
                showToast('Equipment Name is required!', 'error');
                return;
            }

            let spCombined = spVal;
            if(spVal && spDir && !spVal.toUpperCase().includes(spDir)) {
                spCombined = `${spVal} (${spDir})`;
            }

            try {
                const res = await fetch('/api/add_instrument', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({
                        unit: currentUnit,
                        type: type,
                        equipment: desc,
                        kks: kks,
                        area: area,
                        sub_area: subArea || area,
                        range: range || spCombined,
                        set_point: spCombined || range,
                        contact_type: contact
                    })
                });
                const result = await res.json();
                if(result.status === 'success') {
                    showToast(result.message || 'New Instrument added successfully!', 'success');
                    document.getElementById('new-inst-desc').value = '';
                    document.getElementById('new-inst-kks').value = '';
                    document.getElementById('new-inst-area').value = '';
                    if(document.getElementById('new-inst-subarea')) document.getElementById('new-inst-subarea').value = '';
                    if(document.getElementById('new-inst-range')) document.getElementById('new-inst-range').value = '';
                    if(document.getElementById('new-inst-sp-val')) document.getElementById('new-inst-sp-val').value = '';
                    toggleAccordion('add-inst-form');
                    loadData();
                } else {
                    showToast(result.message || 'Failed to add Instrument', 'error');
                }
            } catch(e) {
                showToast('Failed to add new Instrument: ' + e.message, 'error');
            }
        }

        /* ---------------- MASTER EIC SECURITY & PASSWORD LOGIC ---------------- */
        function openMasterPasswordModal() {
            const modal = document.getElementById('master-pwd-modal');
            const input = document.getElementById('master-pwd-input');
            const err = document.getElementById('master-pwd-error');
            if(err) err.style.display = 'none';
            if(input) {
                input.value = '';
                input.type = 'password';
            }
            if(modal) {
                modal.classList.add('open');
                modal.classList.add('active');
                setTimeout(() => { if(input) input.focus(); }, 150);
            }
        }

        function closeMasterPasswordModal() {
            const modal = document.getElementById('master-pwd-modal');
            if(modal) {
                modal.classList.remove('open');
                modal.classList.remove('active');
            }
        }

        function toggleMasterPwdVisibility() {
            const input = document.getElementById('master-pwd-input');
            const btn = document.getElementById('pwd-eye-btn');
            if(!input) return;
            const isPassword = input.type === 'password';
            input.type = isPassword ? 'text' : 'password';
            if(btn) btn.innerHTML = isPassword ? Icons.eyeOff : Icons.eye;
        }

        function submitMasterPassword(e) {
            if(e) e.preventDefault();
            const input = document.getElementById('master-pwd-input');
            const err = document.getElementById('master-pwd-error');
            const val = (input ? input.value : '').trim();

            if(MASTER_PASSWORDS.includes(val.toLowerCase()) || val === 'eic123' || val === 'admin123') {
                isScopeMasterUnlocked = true;
                sessionStorage.setItem('eic_scope_unlocked', 'true');
                closeMasterPasswordModal();
                showToast('✓ Master EIC Edit Access Unlocked!', 'success');
                renderTabContent();
            } else {
                if(err) err.style.display = 'block';
                if(input) {
                    input.classList.add('shake-error');
                    setTimeout(() => input.classList.remove('shake-error'), 500);
                    input.select();
                }
            }
        }

        function lockMasterScope() {
            isScopeMasterUnlocked = false;
            sessionStorage.removeItem('eic_scope_unlocked');
            showToast('Master EIC Edit Mode Locked', 'info');
            renderTabContent();
        }

        async function saveNewScope() {
            if(!isScopeMasterUnlocked) { openMasterPasswordModal(); return; }
            const cat = (document.getElementById('new-scope-cat').value || '').trim();
            const eq = (document.getElementById('new-scope-eq').value || '').trim();
            const type = document.getElementById('new-scope-type').value;
            const pic = document.getElementById('new-scope-pic').value;

            if(!eq) {
                showToast('Equipment Name / Job Scope is required!', 'error');
                return;
            }

            try {
                const res = await fetch('/api/add_scope', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({unit: currentUnit, kategori: cat, nama_equipment: eq, tipe_scope: type, pic: pic})
                });
                const result = await res.json();
                if(result.status === 'success') {
                    showToast(result.message || 'Scope Master added successfully!', 'success');
                    document.getElementById('new-scope-cat').value = '';
                    document.getElementById('new-scope-eq').value = '';
                    toggleAccordion('add-scope-form');
                    loadData();
                } else {
                    showToast(result.message || 'Failed to add Scope Master', 'error');
                }
            } catch(e) {
                showToast('Failed to add Scope: ' + e.message, 'error');
            }
        }

        async function deleteInstrument(type, key) {
            if(!confirm(`Are you sure you want to delete this Instrument?`)) return;

            try {
                const res = await fetch('/api/delete_instrument', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({unit: currentUnit, type: type, no: key, kks: key})
                });
                const result = await res.json();
                showToast(result.message || 'Instrument deleted successfully!', 'info');
                loadData();
            } catch(e) {
                showToast('Failed to delete Instrument', 'error');
            }
        }

        async function addNewPic() {
            if(!isScopeMasterUnlocked) { openMasterPasswordModal(); return; }
            const input = document.getElementById('new-pic-input');
            const picName = (input ? input.value : '').trim();
            if(!picName) {
                showToast('Please enter new PIC personnel name.', 'error');
                return;
            }

            try {
                const res = await fetch('/api/add_pic', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({unit: currentUnit, pic_name: picName})
                });
                const result = await res.json();
                showToast(result.message || 'PIC added successfully!', 'success');
                if(input) input.value = '';
                loadData();
            } catch(e) {
                showToast('Failed to add PIC', 'error');
            }
        }

        async function deletePic(picName) {
            if(!isScopeMasterUnlocked) { openMasterPasswordModal(); return; }
            if(!confirm(`Are you sure you want to delete PIC "${picName}" from Master PICs?`)) return;

            try {
                const res = await fetch('/api/delete_pic', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({unit: currentUnit, pic_name: picName})
                });
                const result = await res.json();
                showToast(result.message || 'PIC deleted from master!', 'info');
                loadData();
            } catch(e) {
                showToast('Failed to delete PIC', 'error');
            }
        }

        async function saveScopeRow(sIdx) {
            if(!isScopeMasterUnlocked) { openMasterPasswordModal(); return; }
            const item = (fullData.scope_master || [])[sIdx];
            if(!item) return;
            const eqInput = document.getElementById(`scope-eq-${sIdx}`);
            const typeSelect = document.getElementById(`scope-type-${sIdx}`);
            const picSelect = document.getElementById(`scope-pic-${sIdx}`);

            const payload = {
                unit: currentUnit,
                row_index: item.row_index !== undefined ? item.row_index : sIdx,
                nama_equipment: (eqInput ? eqInput.value : item.nama_equipment || '').trim(),
                tipe_scope: typeSelect ? typeSelect.value : (item.tipe_scope || 'Vendor'),
                pic: picSelect ? picSelect.value : (item.pic || '')
            };

            try {
                const res = await fetch('/api/update_scope', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify(payload)
                });
                const result = await res.json();
                showToast(result.message || 'Scope Master updated successfully!', 'success');
                loadData();
            } catch(e) {
                showToast('Failed to update Scope Master', 'error');
            }
        }

        async function deleteScopeRow(sIdx) {
            if(!isScopeMasterUnlocked) { openMasterPasswordModal(); return; }
            const item = (fullData.scope_master || [])[sIdx];
            if(!item) return;
            if(!confirm(`Are you sure you want to delete Master Scope row "${item.nama_equipment || ''}"?`)) return;

            try {
                const res = await fetch('/api/delete_scope', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({unit: currentUnit, row_index: item.row_index !== undefined ? item.row_index : sIdx})
                });
                const result = await res.json();
                showToast(result.message || 'Master Scope row deleted!', 'info');
                loadData();
            } catch(e) {
                showToast('Failed to delete Master Scope row', 'error');
            }
        }

        // Drag and drop dropzone handling
        const dropzone = document.getElementById('photo-dropzone');
        if(dropzone) {
            ['dragenter', 'dragover'].forEach(eventName => {
                dropzone.addEventListener(eventName, (e) => { e.preventDefault(); e.stopPropagation(); dropzone.classList.add('dragover'); }, false);
            });
            ['dragleave', 'drop'].forEach(eventName => {
                dropzone.addEventListener(eventName, (e) => { e.preventDefault(); e.stopPropagation(); dropzone.classList.remove('dragover'); }, false);
            });
            dropzone.addEventListener('drop', (e) => {
                const dt = e.dataTransfer;
                const files = dt.files;
                handleModalFileSelect(files);
            }, false);
        }

        /* ---------------- S-CURVE, WA SUMMARY & EXCEL MODALS ---------------- */
        function extractValidDate(str) {
            if(!str) return null;
            str = String(str).trim();
            if(['amp', 'msw', 'japa', 'farhan', 'sched-ok', 'in progress', 'finish', 'true', 'false'].includes(str.toLowerCase())) {
                return null;
            }
            // YYYY-MM-DD
            let m = str.match(/^(\d{4})[/-](\d{1,2})[/-](\d{1,2})/);
            if(m) {
                let yr = parseInt(m[1]), mo = parseInt(m[2]), da = parseInt(m[3]);
                if(mo >= 1 && mo <= 12 && da >= 1 && da <= 31) {
                    return {
                        ymd: `${yr}-${String(mo).padStart(2, '0')}-${String(da).padStart(2, '0')}`,
                        dmy: `${String(da).padStart(2, '0')}/${String(mo).padStart(2, '0')}/${yr}`,
                        time: new Date(yr, mo - 1, da).getTime()
                    };
                }
            }
            // DD/MM/YYYY or DD-MM-YYYY
            m = str.match(/^(\d{1,2})[/-](\d{1,2})[/-](\d{4})/);
            if(m) {
                let da = parseInt(m[1]), mo = parseInt(m[2]), yr = parseInt(m[3]);
                if(mo >= 1 && mo <= 12 && da >= 1 && da <= 31) {
                    return {
                        ymd: `${yr}-${String(mo).padStart(2, '0')}-${String(da).padStart(2, '0')}`,
                        dmy: `${String(da).padStart(2, '0')}/${String(mo).padStart(2, '0')}/${yr}`,
                        time: new Date(yr, mo - 1, da).getTime()
                    };
                }
            }
            return null;
        }

        function openSCurveModal() {
            const unitLabel = document.getElementById('scurve-unit-label');
            if(unitLabel) unitLabel.innerText = currentUnit;

            // Load saved outage start and end dates from localStorage
            let savedStart = localStorage.getItem(`eic_scurve_start_u${currentUnit}`);
            let savedEnd = localStorage.getItem(`eic_scurve_end_u${currentUnit}`);

            // If not found in localStorage, discover from data
            if(!savedStart || !savedEnd) {
                let allTimes = [];
                const checkItem = (dStr) => {
                    const parsed = extractValidDate(dStr);
                    if(parsed) allTimes.push(parsed.time);
                };
                (fullData.work_orders || []).forEach(w => {
                    (w.checklist || []).forEach(c => { if(c.tanggal) checkItem(c.tanggal); });
                    if(w.tanggal_finish) checkItem(w.tanggal_finish);
                });
                (fullData.actuators || []).forEach(a => { if(a.finish_date) checkItem(a.finish_date); });
                (fullData.pressure_tx || []).forEach(i => { if(i.tanggal || i.finish_date) checkItem(i.tanggal || i.finish_date); });
                (fullData.temperature_tx || []).forEach(i => { if(i.tanggal || i.finish_date) checkItem(i.tanggal || i.finish_date); });
                (fullData.pressure_switch || []).forEach(i => { if(i.dated || i.finish_date) checkItem(i.dated || i.finish_date); });

                let minTime = allTimes.length > 0 ? Math.min(...allTimes) : new Date().getTime();
                let maxTime = allTimes.length > 0 ? Math.max(...allTimes) : new Date().getTime();

                const toYMD = (t) => {
                    const d = new Date(t);
                    return `${d.getFullYear()}-${String(d.getMonth() + 1).padStart(2, '0')}-${String(d.getDate()).padStart(2, '0')}`;
                };

                if(!savedStart) {
                    savedStart = toYMD(minTime);
                }
                if(!savedEnd) {
                    let endT = Math.max(minTime + 14 * 86400000, maxTime + 4 * 86400000);
                    savedEnd = toYMD(endT);
                }
            }

            const startInp = document.getElementById('scurve-start-date');
            const endInp = document.getElementById('scurve-end-date');
            if(startInp) startInp.value = savedStart;
            if(endInp) endInp.value = savedEnd;

            renderSCurveChart();
            document.getElementById('scurve-modal').classList.add('open');
        }

        function closeSCurveModal() {
            document.getElementById('scurve-modal').classList.remove('open');
        }

        function saveAndRenderSCurve() {
            const startVal = document.getElementById('scurve-start-date').value;
            const endVal = document.getElementById('scurve-end-date').value;
            if(startVal) localStorage.setItem(`eic_scurve_start_u${currentUnit}`, startVal);
            if(endVal) localStorage.setItem(`eic_scurve_end_u${currentUnit}`, endVal);
            renderSCurveChart();
        }

        function openWaSummaryModal() {
            generateWaText();
            document.getElementById('wa-modal').classList.add('open');
        }
        function closeWaSummaryModal() {
            document.getElementById('wa-modal').classList.remove('open');
        }

        function downloadExcel() {
            const url = `/api/export_excel?unit=${currentUnit}`;
            showToast(`Downloading Unit ${currentUnit} Excel Report...`, 'info', 2000);
            const a = document.createElement('a');
            a.href = url;
            a.download = `Outage_EIC_Monitoring_Report_Unit_${currentUnit}.xlsx`;
            document.body.appendChild(a);
            a.click();
            a.remove();
        }

        function scrollToTop() {
            window.scrollTo({ top: 0, behavior: 'smooth' });
        }

        // Scroll listener for Sticky Bar & Back to Top Button
        window.addEventListener('scroll', () => {
            const stBar = document.getElementById('sticky-summary-bar');
            const topBtn = document.getElementById('back-to-top-btn');
            const scrollPos = window.scrollY;
            if(stBar) {
                stBar.classList.toggle('visible', scrollPos > 180);
            }
            if(topBtn) {
                topBtn.classList.toggle('visible', scrollPos > 300);
            }
        });

        async function batchToggleSubtasks(noWo, action) {
            const isDone = (action === 'mark_all_done');
            if(!confirm(`Are you sure you want to ${isDone ? 'mark ALL sub-tasks as DONE' : 'reset ALL sub-tasks'} for WO ${noWo}?`)) {
                return;
            }
            try {
                const payload = { unit: currentUnit, no_wo: noWo, action: action };
                const res = await fetch('/api/batch_toggle_subtasks', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify(payload)
                });
                const result = await res.json();
                if(result.status === 'success') {
                    showToast(result.message, 'success');
                    loadData();
                } else {
                    showToast(result.message || 'Failed to update batch sub-tasks', 'error');
                }
            } catch(e) {
                showToast('Failed to process batch checklist', 'error');
            }
        }

        function renderSCurveChart() {
            const chartBox = document.getElementById('scurve-chart-container');
            const kpiBadge = document.getElementById('scurve-kpi-badge');
            if(!chartBox || !fullData) return;

            // Collect all completed Work Order subtasks per normalized date
            const dateMap = {}; // key: YYYY-MM-DD -> count
            let totalTasks = 0;
            let totalCompletedWO = 0;
            let undatedCount = 0;

            (fullData.work_orders || []).forEach(w => {
                const chk = w.checklist || [];
                if(chk.length > 0) {
                    totalTasks += chk.length;
                    chk.forEach(c => {
                        if(c.selesai) {
                            totalCompletedWO += 1;
                            const dStr = c.tanggal || w.tanggal_finish || w.tanggal_actual_start || w.tanggal_schedule;
                            const parsed = extractValidDate(dStr);
                            if(parsed && parsed.ymd) {
                                dateMap[parsed.ymd] = (dateMap[parsed.ymd] || 0) + 1;
                            } else {
                                undatedCount += 1;
                            }
                        }
                    });
                } else {
                    totalTasks += 1;
                    if(w.status === 'FINISH') {
                        totalCompletedWO += 1;
                        const dStr = w.tanggal_finish || w.tanggal_actual_start || w.tanggal_schedule;
                        const parsed = extractValidDate(dStr);
                        if(parsed && parsed.ymd) {
                            dateMap[parsed.ymd] = (dateMap[parsed.ymd] || 0) + 1;
                        } else {
                            undatedCount += 1;
                        }
                    }
                }
            });

            // Read configured start & end date
            const startInput = document.getElementById('scurve-start-date');
            const endInput = document.getElementById('scurve-end-date');
            let startYMD = startInput ? startInput.value : '';
            let endYMD = endInput ? endInput.value : '';

            // Fallback if empty
            const activeYMDs = Object.keys(dateMap).sort();
            if(!startYMD) startYMD = activeYMDs.length > 0 ? activeYMDs[0] : '2026-08-20';
            if(!endYMD) {
                const latestActive = activeYMDs.length > 0 ? activeYMDs[activeYMDs.length - 1] : '2026-08-30';
                const d = new Date(latestActive);
                d.setDate(d.getDate() + 5);
                endYMD = `${d.getFullYear()}-${String(d.getMonth() + 1).padStart(2, '0')}-${String(d.getDate()).padStart(2, '0')}`;
            }

            const startDateObj = new Date(startYMD);
            const endDateObj = new Date(endYMD);
            let dayDiff = Math.round((endDateObj - startDateObj) / 86400000);
            if(dayDiff < 1) dayDiff = 1;
            const totalDays = dayDiff + 1;

            // Baseline count: all completed subtasks dated BEFORE startYMD + undated ones
            let cumActualCount = undatedCount;
            Object.keys(dateMap).forEach(ymd => {
                if(ymd < startYMD) {
                    cumActualCount += dateMap[ymd];
                }
            });

            // Generate daily timeline
            const timeline = [];
            const todayTime = new Date().setHours(23, 59, 59, 999);

            for(let i = 0; i < totalDays; i++) {
                const cur = new Date(startDateObj);
                cur.setDate(cur.getDate() + i);
                const ymd = `${cur.getFullYear()}-${String(cur.getMonth() + 1).padStart(2, '0')}-${String(cur.getDate()).padStart(2, '0')}`;
                const dmy = `${String(cur.getDate()).padStart(2, '0')}/${String(cur.getMonth() + 1).padStart(2, '0')}/${cur.getFullYear()}`;
                const shortLabel = `${cur.getDate()} ${cur.toLocaleString('en-US', { month: 'short' })}`;
                const time = cur.getTime();

                // Target Planned S-Curve using normalized sigmoid
                const x = totalDays > 1 ? (i / (totalDays - 1)) : 1.0;
                const k = 7.0;
                const sRaw = 1.0 / (1.0 + Math.exp(-k * (x - 0.5)));
                const sMin = 1.0 / (1.0 + Math.exp(-k * (0 - 0.5)));
                const sMax = 1.0 / (1.0 + Math.exp(-k * (1 - 0.5)));
                const targetPct = Math.round(((sRaw - sMin) / (sMax - sMin)) * 1000) / 10;

                // Actual Count
                const dailyCount = dateMap[ymd] || 0;
                let actualPct = null;
                if(time <= todayTime) {
                    cumActualCount += dailyCount;
                    actualPct = totalTasks > 0 ? Math.min(100, Math.round((cumActualCount / totalTasks) * 1000) / 10) : 0;
                }

                timeline.push({
                    index: i,
                    ymd: ymd,
                    dmy: dmy,
                    shortLabel: shortLabel,
                    time: time,
                    daily: dailyCount,
                    cumActual: actualPct !== null ? cumActualCount : null,
                    actualPct: actualPct,
                    targetPct: targetPct
                });
            }

            // Latest actual vs plan variance
            const actualPoints = timeline.filter(t => t.actualPct !== null);
            const latestActual = actualPoints.length > 0 ? actualPoints[actualPoints.length - 1] : { actualPct: 0, targetPct: 0, cumActual: 0 };
            const variance = Math.round((latestActual.actualPct - latestActual.targetPct) * 10) / 10;

            if(kpiBadge) {
                if(variance >= 0) {
                    kpiBadge.innerHTML = `<span style="background:rgba(16,185,129,0.15); color:#10b981; border:1px solid rgba(16,185,129,0.4); padding:3px 10px; border-radius:12px;">Ahead (+${variance}%)</span>`;
                } else {
                    kpiBadge.innerHTML = `<span style="background:rgba(244,63,94,0.15); color:#f43f5e; border:1px solid rgba(244,63,94,0.4); padding:3px 10px; border-radius:12px;">Behind (${variance}%)</span>`;
                }
            }

            // SVG dimensions
            const svgW = 780;
            const svgH = 270;
            const padL = 48;
            const padR = 28;
            const padT = 30;
            const padB = 40;
            const graphW = svgW - padL - padR;
            const graphH = svgH - padT - padB;

            const n = timeline.length;
            const getX = (i) => padL + (n === 1 ? graphW / 2 : (i / (n - 1)) * graphW);
            const getY = (pct) => padT + graphH - (pct / 100) * graphH;

            // Target Planned Path (Dashed)
            let targetPathD = '';
            timeline.forEach((pt, i) => {
                const x = getX(i);
                const y = getY(pt.targetPct);
                targetPathD += (i === 0 ? `M ${x} ${y}` : ` L ${x} ${y}`);
            });

            // Actual Progress Path
            let actualPathD = '';
            let fillD = '';
            if(actualPoints.length > 0) {
                actualPoints.forEach((pt, i) => {
                    const x = getX(pt.index);
                    const y = getY(pt.actualPct);
                    actualPathD += (i === 0 ? `M ${x} ${y}` : ` L ${x} ${y}`);
                });
                const lastPt = actualPoints[actualPoints.length - 1];
                const firstPt = actualPoints[0];
                fillD = `${actualPathD} L ${getX(lastPt.index)} ${padT + graphH} L ${getX(firstPt.index)} ${padT + graphH} Z`;
            }

            // X-axis label decimation for neat layout
            const labelInterval = Math.max(1, Math.ceil(n / 8));

            let svgHTML = `
            <svg viewBox="0 0 ${svgW} ${svgH}" style="width:100%; height:auto; background:var(--bg-sub); border-radius:var(--radius-md); border:1px solid var(--border-color);">
                <!-- Grid Lines -->
                ${[0, 25, 50, 75, 100].map(p => {
                    const y = getY(p);
                    return `
                    <line x1="${padL}" y1="${y}" x2="${svgW - padR}" y2="${y}" stroke="var(--border-color)" stroke-dasharray="4 4" stroke-width="1"/>
                    <text x="${padL - 8}" y="${y + 4}" fill="var(--text-muted)" font-size="10" text-anchor="end" font-family="'JetBrains Mono'">${p}%</text>`;
                }).join('')}

                <!-- Fill Area under Actual Curve -->
                ${fillD ? `<path d="${fillD}" fill="url(#scurve-grad)" opacity="0.28"/>` : ''}

                <!-- Target Plan Line (Dashed) -->
                <path d="${targetPathD}" fill="none" stroke="#94a3b8" stroke-dasharray="6 4" stroke-width="2.5"/>

                <!-- Actual Curve Line -->
                ${actualPathD ? `<path d="${actualPathD}" fill="none" stroke="var(--primary)" stroke-width="3.5" stroke-linecap="round" stroke-linejoin="round"/>` : ''}

                <!-- Actual Points & Value Tooltips -->
                ${actualPoints.map(pt => {
                    const x = getX(pt.index);
                    const y = getY(pt.actualPct);
                    return `
                    <circle cx="${x}" cy="${y}" r="5" fill="var(--primary)" stroke="#fff" stroke-width="2">
                        <title>${pt.dmy}: Finished +${pt.daily} tasks | Cumulative: ${pt.cumActual}/${totalTasks} (${pt.actualPct}%) | Target: ${pt.targetPct}%</title>
                    </circle>
                    <text x="${x}" y="${y - 9}" fill="var(--primary)" font-size="10" font-weight="800" text-anchor="middle" font-family="'JetBrains Mono'">${pt.actualPct}%</text>`;
                }).join('')}

                <!-- X Axis Labels -->
                ${timeline.map((pt, i) => {
                    if(i % labelInterval === 0 || i === n - 1) {
                        const x = getX(i);
                        return `<text x="${x}" y="${svgH - 12}" fill="var(--text-muted)" font-size="9.5" font-weight="600" text-anchor="middle" font-family="'JetBrains Mono'">${pt.shortLabel}</text>`;
                    }
                    return '';
                }).join('')}

                <defs>
                    <linearGradient id="scurve-grad" x1="0" y1="0" x2="0" y2="1">
                        <stop offset="0%" stop-color="var(--primary)" stop-opacity="0.85"/>
                        <stop offset="100%" stop-color="var(--primary)" stop-opacity="0.0"/>
                    </linearGradient>
                </defs>
            </svg>`;

            // KPI Stat Cards
            let kpiCardsHTML = `
            <div style="display:grid; grid-template-columns:repeat(auto-fit, minmax(160px, 1fr)); gap:10px; margin-bottom:14px;">
                <div class="report-kpi-card" style="padding:10px;">
                    <div class="report-kpi-lbl">Total Sub-Tasks WO</div>
                    <div class="report-kpi-val" style="font-size:1.3rem;">${totalTasks} <span style="font-size:0.75rem; color:var(--text-muted);">sub-tasks</span></div>
                    <div class="report-kpi-sub">Total Sub-Task Work Orders</div>
                </div>
                <div class="report-kpi-card highlight" style="padding:10px;">
                    <div class="report-kpi-lbl">Actual Realization</div>
                    <div class="report-kpi-val" style="font-size:1.3rem; color:var(--primary);">${latestActual.actualPct}%</div>
                    <div class="report-kpi-sub">${latestActual.cumActual} / ${totalTasks} sub-tasks completed</div>
                </div>
                <div class="report-kpi-card" style="padding:10px;">
                    <div class="report-kpi-lbl">Planned Target</div>
                    <div class="report-kpi-val" style="font-size:1.3rem; color:#94a3b8;">${latestActual.targetPct}%</div>
                    <div class="report-kpi-sub">Today's Baseline S-Curve</div>
                </div>
                <div class="report-kpi-card" style="padding:10px; border-left:3px solid ${variance >= 0 ? '#10b981' : '#f43f5e'};">
                    <div class="report-kpi-lbl">Progress Variance</div>
                    <div class="report-kpi-val" style="font-size:1.3rem; color:${variance >= 0 ? '#10b981' : '#f43f5e'};">${variance >= 0 ? '+' : ''}${variance}%</div>
                    <div class="report-kpi-sub">${variance >= 0 ? 'Ahead of Schedule' : 'Behind Schedule'}</div>
                </div>
            </div>`;

            // Active Daily Breakdown Table (Only days with activity or milestone)
            const activeBreakdown = timeline.filter(t => t.daily > 0 || t.cumActual !== null);
            let tableHTML = `
            <div style="margin-top:16px;">
                <div style="font-size:0.85rem; font-weight:700; color:var(--text-main); margin-bottom:8px;">Daily Progress Breakdown:</div>
                <table class="dense-table" style="font-size:0.8rem;">
                    <thead>
                        <tr>
                            <th>Date</th>
                            <th style="text-align:center;">Tasks Completed Today</th>
                            <th style="text-align:center;">Cumulative Realization</th>
                            <th style="text-align:center;">Actual Progress</th>
                            <th style="text-align:center;">Planned Target</th>
                            <th style="text-align:center;">Status</th>
                        </tr>
                    </thead>
                    <tbody>
                        ${activeBreakdown.map(pt => {
                            const dev = pt.actualPct !== null ? Math.round((pt.actualPct - pt.targetPct) * 10) / 10 : null;
                            return `
                            <tr>
                                <td style="font-family:'JetBrains Mono'; font-weight:700;">${pt.dmy}</td>
                                <td style="text-align:center; color:var(--status-finish); font-weight:700;">${pt.daily > 0 ? '+' + pt.daily + ' tasks' : '-'}</td>
                                <td style="text-align:center; font-family:'JetBrains Mono';">${pt.cumActual !== null ? pt.cumActual + ' / ' + totalTasks : '-'}</td>
                                <td style="text-align:center; font-family:'JetBrains Mono'; font-weight:800; color:var(--primary);">${pt.actualPct !== null ? pt.actualPct + '%' : '-'}</td>
                                <td style="text-align:center; font-family:'JetBrains Mono'; color:var(--text-muted);">${pt.targetPct}%</td>
                                <td style="text-align:center;">
                                    ${dev !== null ? (dev >= 0 ? `<span class="badge badge-success" style="font-size:0.7rem;">+${dev}% Ahead</span>` : `<span class="badge badge-error" style="font-size:0.7rem;">${dev}% Behind</span>`) : '-'}
                                </td>
                            </tr>`;
                        }).join('')}
                    </tbody>
                </table>
            </div>`;

            chartBox.innerHTML = `
            ${kpiCardsHTML}
            <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:8px;">
                <div style="font-size:0.8rem; color:var(--text-muted);">
                    Real-Time S-Curve Outage Unit ${currentUnit} (${startYMD} to ${endYMD})
                </div>
                <div style="display:flex; gap:14px; font-size:0.75rem;">
                    <span style="display:flex; align-items:center; gap:5px;"><span style="width:12px; height:3.5px; background:var(--primary); display:inline-block; border-radius:2px;"></span> Actual Progress</span>
                    <span style="display:flex; align-items:center; gap:5px;"><span style="width:12px; height:2px; background:#94a3b8; border-top:2px dashed #94a3b8; display:inline-block;"></span> S-Curve Planned Target</span>
                </div>
            </div>
            ${svgHTML}
            ${tableHTML}`;
        }

        function generateWaText() {
            if(!fullData || !fullData.summary) return;
            const s = fullData.summary;
            const todayStr = getTodayFormatted();

            // Find completed tasks today
            const todayTasks = [];
            (fullData.work_orders || []).forEach(w => {
                (w.checklist || []).forEach(c => {
                    if(c.selesai && (c.tanggal === todayStr || !c.tanggal)) {
                        todayTasks.push(`[${w.no_wo}] Sub-task: ${c.sub_task} (PIC: ${c.pic_task || w.pic || '-'})`);
                    }
                });
            });
            (fullData.actuators || []).forEach(a => {
                if((a.general_inspection || a.function_test) && a.finish_date === todayStr) {
                    todayTasks.push(`[${a.equipment_id}] ${a.equipment_description} (${a.status})`);
                }
            });

            // Find open findings
            const openFindings = [];
            const collectF = (list, codeF, descF) => {
                (list || []).forEach(item => {
                    if(item.temuan) {
                        openFindings.push(`▪ [${item[codeF]}] ${item[descF] || ''}: ${item.temuan} (Action: ${item.tindak_lanjut || 'Under verification'})`);
                    }
                });
            };
            collectF(fullData.work_orders, 'no_wo', 'job_description');
            collectF(fullData.actuators, 'equipment_id', 'equipment_description');
            collectF(fullData.pressure_tx, 'kks', 'equipment');
            collectF(fullData.temperature_tx, 'kks', 'equipment');
            collectF(fullData.pressure_switch, 'kks', 'equipment');

            let msg = `*EIC OUTAGE PROGRESS REPORT — UNIT ${currentUnit}*\n`;
            msg += `*PLTU MSW — SECTION ELECTRIC, INSTRUMENT & CONTROL*\n`;
            msg += `*Date:* ${todayStr}\n\n`;

            msg += `*PROGRESS SUMMARY:*\n`;
            msg += `▪ *Overall Progress:* *${s.grand_pct}%* (${s.grand_done}/${s.grand_total} Sub-tasks Completed)\n`;
            msg += `▪ *Work Orders:* ${s.wo.pct}% (${s.wo.finish}/${s.wo.total} WO Finish &bull; ${s.wo.subtask_done}/${s.wo.subtask_total} Sub-tasks)\n`;
            msg += `▪ *Actuator Valves:* ${s.actuator.pct}% (${s.actuator.finish}/${s.actuator.total} Valve Finish)\n`;
            msg += `▪ *Instruments:* ${s.instrument.pct}% (${s.instrument.done}/${s.instrument.total} Verification OK)\n\n`;

            msg += `*COMPLETED WORK UPDATES:* (${todayTasks.length} Tasks)\n`;
            if(todayTasks.length > 0) {
                todayTasks.slice(0, 10).forEach((t, i) => {
                    msg += `${i+1}. ${t}\n`;
                });
                if(todayTasks.length > 10) msg += `... and ${todayTasks.length - 10} other items.\n`;
            } else {
                msg += `_No tasks completed on ${todayStr}_\n`;
            }
            msg += `\n`;

            msg += `*ACTIVE FINDINGS & RECOMMENDATIONS:* (${openFindings.length} Findings)\n`;
            if(openFindings.length > 0) {
                openFindings.forEach(f => {
                    msg += `${f}\n`;
                });
            } else {
                msg += `_Nil (Equipment in normal condition with no open findings)_\n`;
            }
            msg += `\n_Report generated automatically from MSW PLTU EIC Monitoring System_`;

            const waBox = document.getElementById('wa-text-box');
            if(waBox) waBox.value = msg;
        }

        async function copyWaText() {
            const waBox = document.getElementById('wa-text-box');
            if(!waBox) return;
            try {
                if(navigator.clipboard && window.isSecureContext) {
                    await navigator.clipboard.writeText(waBox.value);
                } else {
                    waBox.select();
                    document.execCommand('copy');
                }
                showToast('✓ WhatsApp report format copied to clipboard!', 'success', 2500);
            } catch(e) {
                waBox.select();
                document.execCommand('copy');
                showToast('✓ Report format copied!', 'success');
            }
        }

        function initTheme() {
            const saved = localStorage.getItem('eic_theme') || 'dark';
            setTheme(saved);
        }

        function toggleTheme() {
            const current = document.documentElement.getAttribute('data-theme') || 'dark';
            const next = current === 'dark' ? 'light' : 'dark';
            setTheme(next);
        }

        function setTheme(theme) {
            document.documentElement.setAttribute('data-theme', theme);
            localStorage.setItem('eic_theme', theme);
            const icon = document.getElementById('theme-icon');
            if(icon) {
                icon.innerHTML = (theme === 'light') ? Icons.sun : Icons.moon;
            }
            const btn = document.getElementById('theme-toggle-btn');
            if(btn) {
                btn.title = (theme === 'light') ? 'Switch to Dark Mode' : 'Switch to Light Mode';
            }
        }

        window.onload = function() {
            initTheme();
            loadData();
        };
    </script>
</body>
</html>
"""

class ReuseTCPServer(socketserver.ThreadingTCPServer):
    allow_reuse_address = False

def get_local_ip():
    try:
        import socket
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.settimeout(1.0)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"

def run_server():
    os.chdir(BASE_DIR)
    server_address = ('', PORT)
    local_ip = get_local_ip()
    try:
        httpd = ReuseTCPServer(server_address, EICMonitoringHandler)
    except OSError as e:
        print(f"==================================================================")
        print(f" [INFO] Server Outage EIC Monitoring sudah aktif di port {PORT}!")
        print(f" Akses dari PC ini  : http://localhost:{PORT}")
        print(f" Akses dari PC lain : http://{local_ip}:{PORT}")
        print(f"==================================================================")
        return

    print(f"==================================================================")
    print(f" OUTAGE EIC WORK ORDER MONITORING SYSTEM SERVER IS RUNNING!")
    print(f" * Akses dari PC ini (Lokal)     : http://localhost:{PORT}")
    print(f" * Akses dari PC lain (LAN/Wi-Fi): http://{local_ip}:{PORT}")
    print(f" Shared Directory                 : {BASE_DIR}")
    print(f"==================================================================")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\nStopping server...")
        httpd.server_close()

if __name__ == "__main__":
    run_server()
